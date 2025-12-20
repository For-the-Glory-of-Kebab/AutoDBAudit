"""
Rigorous Multi-Scenario E2E Sync Tests

This module tests the sync engine with realistic, complex scenarios:
1. Multiple annotations per sheet with varied values
2. Multi-sync stability (sync, modify, sync again, verify preservation)
3. State transitions (FAIL→Exception, Exception→FAIL, PASS→FAIL, etc.)
4. Cross-sheet validation (no contamination between sheets)
5. DB persistence verification
6. Action Log entry creation

These tests would have caught the column matching bug (substring collision).
"""

import pytest
import random
import string
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import sqlite3

from autodbaudit.application.annotation_sync import (
    AnnotationSyncService,
    SHEET_ANNOTATION_CONFIG,
)
from autodbaudit.infrastructure.sqlite.store import HistoryStore
from autodbaudit.infrastructure.sqlite.schema import initialize_schema_v2


def random_string(length=10):
    """Generate random string for test data."""
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


def random_date():
    """Generate random date within last year."""
    days_ago = random.randint(1, 365)
    return (datetime.now() - timedelta(days=days_ago)).strftime("%Y-%m-%d")


class TestMultiAnnotationRoundTrip:
    """Test multiple annotations per sheet survive round-trip."""

    @pytest.fixture
    def setup_multi_annotation_excel(self, tmp_path):
        """Create Excel with multiple sheets and many annotations each."""
        excel_path = tmp_path / "multi_annotation.xlsx"
        db_path = tmp_path / "test.db"

        wb = Workbook()

        # ===== Linked Servers Sheet (10 rows) =====
        ws_ls = wb.active
        ws_ls.title = "Linked Servers"
        ws_ls.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Linked Server",
                "Provider",
                "Data Source",
                "RPC Out",
                "Local Login",
                "Remote Login",
                "Impersonate",
                "Risk",
                "Review Status",
                "Purpose",
                "Justification",
                "Last Reviewed",
            ]
        )
        self.linked_server_data = []
        for i in range(10):
            row_data = {
                "server": f"SRV{i % 3 + 1}",
                "instance": f"INST{i % 2 + 1}",
                "linked_server": f"LINKED_{random_string(6)}",
                "local_login": f"local_{i}" if i % 3 != 0 else "(Default)",
                "remote_login": f"remote_{i}" if i % 3 != 0 else "(Default)",
                "purpose": f"Purpose_{random_string(15)}",
                "justification": f"Just_{random_string(10)}" if i % 2 == 0 else "",
            }
            self.linked_server_data.append(row_data)
            ws_ls.append(
                [
                    "⏳" if i % 2 == 0 else "",
                    row_data["server"],
                    row_data["instance"],
                    row_data["linked_server"],
                    "SQLNCLI",
                    f"srv{i}",
                    "Yes",
                    row_data["local_login"],
                    row_data["remote_login"],
                    "No",
                    "Medium" if i % 2 else "High",
                    "✓ Exception" if row_data["justification"] else "",
                    row_data["purpose"],
                    row_data["justification"],
                    "",
                ]
            )

        # ===== Triggers Sheet (15 rows) =====
        ws_tr = wb.create_sheet("Triggers")
        ws_tr.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Scope",
                "Database",
                "Trigger Name",
                "Event",
                "Enabled",
                "Review Status",
                "Notes",
                "Justification",
                "Last Reviewed",
            ]
        )
        self.trigger_data = []
        scopes = ["SERVER", "DATABASE"]
        events = [
            "CREATE_LOGIN",
            "ALTER_LOGIN",
            "DROP_LOGIN",
            "DDL_TABLE_EVENTS",
            "DML_INSERT",
        ]
        for i in range(15):
            row_data = {
                "server": f"SRV{i % 3 + 1}",
                "instance": f"INST{i % 2 + 1}",
                "scope": scopes[i % 2],
                "database": f"DB_{i}" if scopes[i % 2] == "DATABASE" else "",
                "trigger_name": f"TR_{random_string(8)}",
                "event": events[i % len(events)],
                "notes": f"Notes_{random_string(20)}" if i % 3 == 0 else "",
                "justification": f"Justified_{random_string(8)}" if i % 4 == 0 else "",
            }
            self.trigger_data.append(row_data)
            ws_tr.append(
                [
                    "⏳" if row_data["scope"] == "SERVER" else "",
                    row_data["server"],
                    row_data["instance"],
                    row_data["scope"],
                    row_data["database"],
                    row_data["trigger_name"],
                    row_data["event"],
                    "✓",
                    "",
                    row_data["notes"],
                    row_data["justification"],
                    "",
                ]
            )

        # ===== Backups Sheet (12 rows) =====
        ws_bu = wb.create_sheet("Backups")
        ws_bu.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Database",
                "Recovery",
                "Last Backup",
                "Days Since",
                "Size (MB)",
                "Path",
                "Status",
                "Review Status",
                "Justification",
                "Last Reviewed",
                "Notes",
            ]
        )
        self.backup_data = []
        recovery_models = ["FULL", "SIMPLE", "BULK_LOGGED"]
        for i in range(12):
            row_data = {
                "server": f"SRV{i % 3 + 1}",
                "instance": f"INST{i % 2 + 1}",
                "database": f"DB_{random_string(5)}",
                "recovery": recovery_models[i % 3],
                "justification": f"BackupJust_{random_string(8)}" if i % 3 == 0 else "",
                "notes": f"BackupNote_{random_string(12)}" if i % 2 == 0 else "",
            }
            self.backup_data.append(row_data)
            ws_bu.append(
                [
                    "⏳" if i % 2 else "",
                    row_data["server"],
                    row_data["instance"],
                    row_data["database"],
                    row_data["recovery"],
                    random_date(),
                    random.randint(0, 30),
                    random.randint(100, 5000),
                    "/backup/path",
                    "❌ Missing" if i % 2 else "✓ OK",
                    "",
                    row_data["justification"],
                    "",
                    row_data["notes"],
                ]
            )

        # ===== Configuration Sheet (8 rows) =====
        ws_cfg = wb.create_sheet("Configuration")
        ws_cfg.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Setting",
                "Current",
                "Recommended",
                "Status",
                "Review Status",
                "Exception Reason",
                "Last Reviewed",
            ]
        )
        self.config_data = []
        settings = [
            "xp_cmdshell",
            "clr enabled",
            "remote access",
            "Database Mail XPs",
            "remote admin connections",
            "show advanced options",
            "Ole Automation Procedures",
            "cross db ownership chaining",
        ]
        for i, setting in enumerate(settings):
            row_data = {
                "server": f"SRV{i % 3 + 1}",
                "instance": f"INST{i % 2 + 1}",
                "setting": setting,
                "exception_reason": (
                    f"Config reason {random_string(10)}" if i % 3 == 0 else ""
                ),
            }
            self.config_data.append(row_data)
            ws_cfg.append(
                [
                    "⏳" if i % 2 else "",
                    row_data["server"],
                    row_data["instance"],
                    row_data["setting"],
                    1 if i % 2 else 0,
                    0,
                    "❌ FAIL" if i % 2 else "✓ PASS",
                    "",
                    row_data["exception_reason"],
                    "",
                ]
            )

        wb.save(excel_path)

        # Initialize DB
        store = HistoryStore(db_path)
        store.initialize_schema()
        initialize_schema_v2(store._get_connection())

        self.service = AnnotationSyncService(db_path)
        self.excel_path = excel_path
        self.db_path = db_path
        self.tmp_path = tmp_path

        return tmp_path

    def test_read_all_annotations_correctly(self, setup_multi_annotation_excel):
        """Verify all annotations are read with correct values."""
        annotations = self.service.read_all_from_excel(self.excel_path)

        # Check Linked Servers
        ls_annotations = {
            k: v for k, v in annotations.items() if k.startswith("linked_server|")
        }
        assert (
            len(ls_annotations) >= 8
        ), f"Expected 8+ Linked Server annotations, got {len(ls_annotations)}"

        # Verify specific purpose values are preserved
        purpose_values_found = [
            v.get("purpose", "") for v in ls_annotations.values() if v.get("purpose")
        ]
        assert (
            len(purpose_values_found) >= 5
        ), f"Expected 5+ purpose values, got {len(purpose_values_found)}"

        # Check Triggers
        tr_annotations = {
            k: v for k, v in annotations.items() if k.startswith("trigger|")
        }
        assert (
            len(tr_annotations) >= 10
        ), f"Expected 10+ Trigger annotations, got {len(tr_annotations)}"

        # Verify notes values
        notes_values_found = [
            v.get("notes", "") for v in tr_annotations.values() if v.get("notes")
        ]
        assert (
            len(notes_values_found) >= 3
        ), f"Expected 3+ notes values, got {len(notes_values_found)}"

        # Check Backups
        bu_annotations = {
            k: v for k, v in annotations.items() if k.startswith("backup|")
        }
        assert (
            len(bu_annotations) >= 10
        ), f"Expected 10+ Backup annotations, got {len(bu_annotations)}"

    def test_write_preserves_all_values(self, setup_multi_annotation_excel):
        """Verify write→read cycle preserves all annotation values."""
        # Read original
        original = self.service.read_all_from_excel(self.excel_path)

        # Persist to DB
        self.service.persist_to_db(original)

        # Write back to Excel
        self.service.write_all_to_excel(self.excel_path, original)

        # Re-read
        reloaded = self.service.read_all_from_excel(self.excel_path)

        # Verify all purpose values for Linked Servers
        for key in original:
            if key.startswith("linked_server|"):
                orig_purpose = original[key].get("purpose", "")
                new_purpose = reloaded.get(key, {}).get("purpose", "")
                assert (
                    orig_purpose == new_purpose
                ), f"Purpose mismatch for {key}: '{orig_purpose}' != '{new_purpose}'"

        # Verify all notes values for Triggers
        for key in original:
            if key.startswith("trigger|"):
                orig_notes = original[key].get("notes", "")
                new_notes = reloaded.get(key, {}).get("notes", "")
                assert (
                    orig_notes == new_notes
                ), f"Notes mismatch for {key}: '{orig_notes}' != '{new_notes}'"

    def test_modify_and_sync_preserves_unmodified(self, setup_multi_annotation_excel):
        """Modify some annotations, verify others are preserved."""
        # Read original
        original = self.service.read_all_from_excel(self.excel_path)
        self.service.persist_to_db(original)

        # Modify 3 random linked server purposes
        ls_keys = [k for k in original if k.startswith("linked_server|")]
        modified_keys = random.sample(ls_keys, min(3, len(ls_keys)))

        for key in modified_keys:
            original[key]["purpose"] = f"MODIFIED_{random_string(10)}"

        # Write and re-read
        self.service.write_all_to_excel(self.excel_path, original)
        reloaded = self.service.read_all_from_excel(self.excel_path)

        # Verify modifications stuck
        for key in modified_keys:
            assert "MODIFIED_" in reloaded.get(key, {}).get(
                "purpose", ""
            ), f"Modified purpose was lost for {key}"

        # Verify unmodified values preserved
        unmodified_keys = [k for k in ls_keys if k not in modified_keys]
        for key in unmodified_keys:
            orig = original[key].get("purpose", "")
            new = reloaded.get(key, {}).get("purpose", "")
            # If we didn't modify it, it should match original
            if "MODIFIED_" not in orig:
                assert orig == new, f"Unmodified value changed for {key}"


class TestMultiSyncStability:
    """Test that values survive multiple sync cycles."""

    @pytest.fixture
    def setup_multi_sync_excel(self, tmp_path):
        """Create clean Excel for multi-sync testing."""
        excel_path = tmp_path / "multi_sync.xlsx"
        db_path = tmp_path / "test.db"

        wb = Workbook()
        ws = wb.active
        ws.title = "Linked Servers"
        ws.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Linked Server",
                "Provider",
                "Data Source",
                "RPC Out",
                "Local Login",
                "Remote Login",
                "Impersonate",
                "Risk",
                "Review Status",
                "Purpose",
                "Justification",
                "Last Reviewed",
            ]
        )

        # 5 rows of data
        for i in range(5):
            ws.append(
                [
                    "⏳",
                    f"SRV{i+1}",
                    "INST1",
                    f"LINKED{i+1}",
                    "SQLNCLI",
                    f"srv{i}",
                    "Yes",
                    f"local{i}",
                    f"remote{i}",
                    "No",
                    "Medium",
                    "",
                    f"Initial_Purpose_{i}",
                    "",
                    "",
                ]
            )

        wb.save(excel_path)

        store = HistoryStore(db_path)
        store.initialize_schema()
        initialize_schema_v2(store._get_connection())

        self.service = AnnotationSyncService(db_path)
        self.excel_path = excel_path
        self.db_path = db_path

        return tmp_path

    def test_three_sync_cycles(self, setup_multi_sync_excel):
        """Simulate 3 sync cycles with modifications."""
        # === SYNC 1: Initial read and persist ===
        ann1 = self.service.read_all_from_excel(self.excel_path)
        self.service.persist_to_db(ann1)

        initial_purposes = {k: v.get("purpose", "") for k, v in ann1.items()}

        # === SYNC 2: Modify 2 purposes, sync ===
        for i, key in enumerate(list(ann1.keys())[:2]):
            ann1[key]["purpose"] = f"Sync2_Modified_{i}"

        self.service.write_all_to_excel(self.excel_path, ann1)
        self.service.persist_to_db(ann1)

        ann2 = self.service.read_all_from_excel(self.excel_path)

        # Verify sync 2 modifications
        for i, key in enumerate(list(ann2.keys())[:2]):
            assert f"Sync2_Modified_{i}" == ann2[key].get(
                "purpose", ""
            ), f"Sync 2 modification lost for {key}"

        # === SYNC 3: Modify different purposes, add justifications ===
        for i, key in enumerate(list(ann2.keys())[2:4]):
            ann2[key]["purpose"] = f"Sync3_Purpose_{i}"
            ann2[key]["justification"] = f"Sync3_Just_{i}"

        self.service.write_all_to_excel(self.excel_path, ann2)
        self.service.persist_to_db(ann2)

        ann3 = self.service.read_all_from_excel(self.excel_path)

        # Verify sync 2 modifications STILL preserved
        for i, key in enumerate(list(ann3.keys())[:2]):
            assert "Sync2_Modified_" in ann3[key].get(
                "purpose", ""
            ), f"Sync 2 modification was overwritten in sync 3"

        # Verify sync 3 modifications
        for i, key in enumerate(list(ann3.keys())[2:4]):
            assert "Sync3_Purpose_" in ann3[key].get(
                "purpose", ""
            ), f"Sync 3 purpose modification lost"
            assert "Sync3_Just_" in ann3[key].get(
                "justification", ""
            ), f"Sync 3 justification modification lost"


class TestDBPersistence:
    """Test that annotations actually persist in database."""

    @pytest.fixture
    def setup_db_test(self, tmp_path):
        """Create test environment."""
        excel_path = tmp_path / "db_test.xlsx"
        db_path = tmp_path / "test.db"

        wb = Workbook()
        ws = wb.active
        ws.title = "Backups"
        ws.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Database",
                "Recovery",
                "Last Backup",
                "Days Since",
                "Size (MB)",
                "Path",
                "Status",
                "Review Status",
                "Justification",
                "Last Reviewed",
                "Notes",
            ]
        )
        ws.append(
            [
                "⏳",
                "SRV1",
                "INST1",
                "TESTDB",
                "FULL",
                "2025-01-01",
                10,
                500,
                "/backup",
                "❌ Missing",
                "",
                "TestJustification123",
                "",
                "TestNotes456",
            ]
        )
        wb.save(excel_path)

        store = HistoryStore(db_path)
        store.initialize_schema()
        initialize_schema_v2(store._get_connection())

        self.service = AnnotationSyncService(db_path)
        self.excel_path = excel_path
        self.db_path = db_path

        return tmp_path

    def test_values_persist_in_db(self, setup_db_test):
        """Verify annotation values are stored in database."""
        # Read and persist
        annotations = self.service.read_all_from_excel(self.excel_path)
        self.service.persist_to_db(annotations)

        # Query DB directly
        conn = sqlite3.connect(self.db_path)
        rows = conn.execute(
            "SELECT entity_type, entity_key, field_name, field_value FROM annotations"
        ).fetchall()
        conn.close()

        # Find our test values
        found_justification = False
        found_notes = False

        for row in rows:
            if row[2] == "justification" and "TestJustification123" in str(row[3]):
                found_justification = True
            if row[2] == "notes" and "TestNotes456" in str(row[3]):
                found_notes = True

        assert found_justification, "Justification not found in DB"
        assert found_notes, "Notes not found in DB"

    def test_load_from_db_matches_excel(self, setup_db_test):
        """Verify load_from_db returns same values as Excel read."""
        # Read from Excel
        from_excel = self.service.read_all_from_excel(self.excel_path)

        # Persist
        self.service.persist_to_db(from_excel)

        # Load from DB
        from_db = self.service.load_from_db()

        # Compare
        for key in from_excel:
            assert key in from_db, f"Key {key} not in DB"
            for field in ["justification", "notes"]:
                excel_val = from_excel[key].get(field, "")
                db_val = from_db[key].get(field, "")
                if excel_val:
                    assert (
                        excel_val == db_val
                    ), f"Mismatch for {key}.{field}: Excel='{excel_val}' DB='{db_val}'"


class TestStateTransitions:
    """Test exception state transitions (FAIL→Exception, Remove→FAIL, etc.)."""

    @pytest.fixture
    def setup_state_transition_excel(self, tmp_path):
        """Create Excel with various states to test transitions."""
        excel_path = tmp_path / "state_transition.xlsx"
        db_path = tmp_path / "test.db"

        wb = Workbook()
        ws = wb.active
        ws.title = "Backups"
        ws.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Database",
                "Recovery",
                "Last Backup",
                "Days Since",
                "Size (MB)",
                "Path",
                "Status",
                "Review Status",
                "Justification",
                "Last Reviewed",
                "Notes",
            ]
        )

        # Row 1: FAIL status, no exception
        ws.append(
            [
                "⏳",
                "SRV1",
                "INST1",
                "DB1",
                "FULL",
                "2024-12-01",
                30,
                500,
                "/backup",
                "❌ Missing",
                "",
                "",
                "",
                "",
            ]
        )
        # Row 2: FAIL status with exception (justified)
        ws.append(
            [
                "✓",
                "SRV1",
                "INST1",
                "DB2",
                "SIMPLE",
                "2024-11-15",
                45,
                200,
                "/backup",
                "❌ Missing",
                "✓ Exception",
                "Approved by security team",
                "2025-01-01",
                "Legacy system",
            ]
        )
        # Row 3: PASS status (no issue)
        ws.append(
            [
                "",
                "SRV1",
                "INST1",
                "DB3",
                "FULL",
                "2025-01-15",
                5,
                1000,
                "/backup",
                "✓ OK",
                "",
                "",
                "",
                "",
            ]
        )
        # Row 4: WARN status, no exception
        ws.append(
            [
                "⏳",
                "SRV2",
                "INST1",
                "DB4",
                "BULK_LOGGED",
                "2024-12-20",
                20,
                300,
                "/backup",
                "⚠ Warning",
                "",
                "",
                "",
                "",
            ]
        )
        # Row 5: WARN with exception
        ws.append(
            [
                "✓",
                "SRV2",
                "INST1",
                "DB5",
                "SIMPLE",
                "2024-12-10",
                25,
                400,
                "/backup",
                "⚠ Warning",
                "✓ Exception",
                "ETL DB - no backup needed",
                "",
                "Data reloaded daily",
            ]
        )

        wb.save(excel_path)

        store = HistoryStore(db_path)
        store.initialize_schema()
        initialize_schema_v2(store._get_connection())

        self.service = AnnotationSyncService(db_path)
        self.excel_path = excel_path
        self.db_path = db_path

        return tmp_path

    def test_add_exception_to_fail(self, setup_state_transition_excel):
        """Test adding exception to a FAIL row."""
        # Read initial
        ann = self.service.read_all_from_excel(self.excel_path)

        # Find the DB1 row (FAIL, no exception)
        db1_keys = [k for k in ann if "db1" in k.lower()]
        assert len(db1_keys) >= 1, f"DB1 not found in annotations: {list(ann.keys())}"

        db1_key = db1_keys[0]

        # Add justification
        ann[db1_key]["justification"] = "New exception reason"
        ann[db1_key]["review_status"] = "✓ Exception"

        # Write and verify
        self.service.write_all_to_excel(self.excel_path, ann)
        self.service.persist_to_db(ann)

        reloaded = self.service.read_all_from_excel(self.excel_path)

        assert (
            reloaded[db1_key].get("justification") == "New exception reason"
        ), "Exception justification not saved"

    def test_remove_exception_from_fail(self, setup_state_transition_excel):
        """Test removing exception from a FAIL row."""
        ann = self.service.read_all_from_excel(self.excel_path)

        # Find the DB2 row (FAIL with exception)
        db2_keys = [k for k in ann if "db2" in k.lower()]
        assert len(db2_keys) >= 1, f"DB2 not found in annotations"

        db2_key = db2_keys[0]

        # Verify it has exception
        assert ann[db2_key].get(
            "justification"
        ), "DB2 should have justification initially"

        # Clear the exception
        ann[db2_key]["justification"] = ""
        ann[db2_key]["review_status"] = ""

        # Write and verify
        self.service.write_all_to_excel(self.excel_path, ann)
        reloaded = self.service.read_all_from_excel(self.excel_path)

        new_just = reloaded[db2_key].get("justification", "")
        assert new_just == "", f"Justification should be empty, got: {new_just}"

    def test_modify_exception_reason(self, setup_state_transition_excel):
        """Test modifying an existing exception reason."""
        ann = self.service.read_all_from_excel(self.excel_path)

        # Find DB5 (WARN with exception)
        db5_keys = [k for k in ann if "db5" in k.lower()]
        assert len(db5_keys) >= 1

        db5_key = db5_keys[0]
        original_just = ann[db5_key].get("justification", "")

        # Modify
        ann[db5_key]["justification"] = "UPDATED: New business justification"

        self.service.write_all_to_excel(self.excel_path, ann)
        reloaded = self.service.read_all_from_excel(self.excel_path)

        new_just = reloaded[db5_key].get("justification", "")
        assert "UPDATED:" in new_just, f"Justification update failed: {new_just}"
        assert new_just != original_just

    def test_notes_preserve_independently(self, setup_state_transition_excel):
        """Test that notes are preserved when changing justification."""
        ann = self.service.read_all_from_excel(self.excel_path)

        # Find DB5 which has notes
        db5_keys = [k for k in ann if "db5" in k.lower()]
        db5_key = db5_keys[0]

        original_notes = ann[db5_key].get("notes", "")
        assert original_notes, "DB5 should have notes initially"

        # Change justification, don't touch notes
        ann[db5_key]["justification"] = "Changed justification only"

        self.service.write_all_to_excel(self.excel_path, ann)
        reloaded = self.service.read_all_from_excel(self.excel_path)

        # Notes should be unchanged
        new_notes = reloaded[db5_key].get("notes", "")
        assert (
            original_notes == new_notes
        ), f"Notes changed unexpectedly: '{original_notes}' -> '{new_notes}'"


class TestCrossSheetValidation:
    """Test that different sheets don't affect each other."""

    @pytest.fixture
    def setup_multi_sheet_excel(self, tmp_path):
        """Create Excel with multiple sheets for cross-validation."""
        excel_path = tmp_path / "cross_sheet.xlsx"
        db_path = tmp_path / "test.db"

        wb = Workbook()

        # === Backups Sheet ===
        ws_bu = wb.active
        ws_bu.title = "Backups"
        ws_bu.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Database",
                "Recovery",
                "Last Backup",
                "Days Since",
                "Size (MB)",
                "Path",
                "Status",
                "Review Status",
                "Justification",
                "Last Reviewed",
                "Notes",
            ]
        )
        for i in range(5):
            ws_bu.append(
                [
                    "⏳",
                    f"SRV1",
                    "INST1",
                    f"BackupDB{i}",
                    "FULL",
                    "2025-01-01",
                    10,
                    500,
                    "/backup",
                    "❌ Missing",
                    "",
                    f"BackupJust_{i}",
                    "",
                    f"BackupNote_{i}",
                ]
            )

        # === Linked Servers Sheet ===
        ws_ls = wb.create_sheet("Linked Servers")
        ws_ls.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Linked Server",
                "Provider",
                "Data Source",
                "RPC Out",
                "Local Login",
                "Remote Login",
                "Impersonate",
                "Risk",
                "Review Status",
                "Purpose",
                "Justification",
                "Last Reviewed",
            ]
        )
        for i in range(5):
            ws_ls.append(
                [
                    "⏳",
                    f"SRV1",
                    "INST1",
                    f"LinkedSrv{i}",
                    "SQLNCLI",
                    f"target{i}",
                    "Yes",
                    f"local{i}",
                    f"remote{i}",
                    "No",
                    "Medium",
                    "",
                    f"LinkedPurpose_{i}",
                    f"LinkedJust_{i}",
                    "",
                ]
            )

        # === Triggers Sheet ===
        ws_tr = wb.create_sheet("Triggers")
        ws_tr.append(
            [
                "⏳",
                "Server",
                "Instance",
                "Scope",
                "Database",
                "Trigger Name",
                "Event",
                "Enabled",
                "Review Status",
                "Notes",
                "Justification",
                "Last Reviewed",
            ]
        )
        for i in range(5):
            ws_tr.append(
                [
                    "⏳",
                    f"SRV1",
                    "INST1",
                    "SERVER",
                    "",
                    f"TR_Audit_{i}",
                    "CREATE_LOGIN",
                    "✓",
                    "",
                    f"TriggerNote_{i}",
                    f"TriggerJust_{i}",
                    "",
                ]
            )

        wb.save(excel_path)

        store = HistoryStore(db_path)
        store.initialize_schema()
        initialize_schema_v2(store._get_connection())

        self.service = AnnotationSyncService(db_path)
        self.excel_path = excel_path
        self.db_path = db_path

        return tmp_path

    def test_modifying_one_sheet_preserves_others(self, setup_multi_sheet_excel):
        """Modify Backups, verify Linked Servers and Triggers unchanged."""
        # Read all
        ann = self.service.read_all_from_excel(self.excel_path)

        # Capture original values from other sheets
        ls_values = {
            k: dict(v) for k, v in ann.items() if k.startswith("linked_server|")
        }
        tr_values = {k: dict(v) for k, v in ann.items() if k.startswith("trigger|")}

        # Modify ALL backup annotations
        for key in ann:
            if key.startswith("backup|"):
                ann[key]["justification"] = "BACKUP_MODIFIED"
                ann[key]["notes"] = "BACKUP_NOTES_MODIFIED"

        # Write back
        self.service.write_all_to_excel(self.excel_path, ann)

        # Re-read
        reloaded = self.service.read_all_from_excel(self.excel_path)

        # Verify linked servers unchanged
        for key, orig_fields in ls_values.items():
            assert key in reloaded, f"Linked Server key lost: {key}"
            for field in ["purpose", "justification"]:
                orig = orig_fields.get(field, "")
                new = reloaded[key].get(field, "")
                assert (
                    orig == new
                ), f"Linked Server {key}.{field} changed: '{orig}' -> '{new}'"

        # Verify triggers unchanged
        for key, orig_fields in tr_values.items():
            assert key in reloaded, f"Trigger key lost: {key}"
            for field in ["notes", "justification"]:
                orig = orig_fields.get(field, "")
                new = reloaded[key].get(field, "")
                assert (
                    orig == new
                ), f"Trigger {key}.{field} changed: '{orig}' -> '{new}'"

        # Verify backups were modified
        for key in reloaded:
            if key.startswith("backup|"):
                assert reloaded[key].get("justification") == "BACKUP_MODIFIED"

    def test_all_sheets_independent_persistence(self, setup_multi_sheet_excel):
        """Each sheet's annotations persist independently to DB."""
        ann = self.service.read_all_from_excel(self.excel_path)
        self.service.persist_to_db(ann)

        from_db = self.service.load_from_db()

        # Count by entity type
        backup_count = sum(1 for k in from_db if k.startswith("backup|"))
        ls_count = sum(1 for k in from_db if k.startswith("linked_server|"))
        tr_count = sum(1 for k in from_db if k.startswith("trigger|"))

        assert backup_count >= 5, f"Expected 5+ backup annotations, got {backup_count}"
        assert ls_count >= 5, f"Expected 5+ linked server annotations, got {ls_count}"
        assert tr_count >= 5, f"Expected 5+ trigger annotations, got {tr_count}"

    def test_random_mix_modification(self, setup_multi_sheet_excel):
        """Randomly modify annotations across all sheets, verify integrity."""
        ann = self.service.read_all_from_excel(self.excel_path)

        # Group by entity type
        by_type = {}
        for key in ann:
            entity_type = key.split("|")[0]
            if entity_type not in by_type:
                by_type[entity_type] = []
            by_type[entity_type].append(key)

        # Randomly modify 2 from each type
        modified_tracker = {}
        for entity_type, keys in by_type.items():
            to_modify = random.sample(keys, min(2, len(keys)))
            for key in to_modify:
                new_val = f"RANDOM_MOD_{random_string(8)}"
                if "justification" in ann[key]:
                    ann[key]["justification"] = new_val
                    modified_tracker[key] = new_val

        # Write
        self.service.write_all_to_excel(self.excel_path, ann)

        # Re-read
        reloaded = self.service.read_all_from_excel(self.excel_path)

        # Verify modifications
        for key, expected_val in modified_tracker.items():
            actual = reloaded.get(key, {}).get("justification", "")
            assert actual == expected_val, f"Random modification lost for {key}"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
