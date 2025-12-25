"""
Layer Tests - Shared Configuration.

Provides fixtures and utilities for the 6-layer testing architecture.
Note: Imports are deferred to fixture functions to ensure root conftest runs first.
"""

from __future__ import annotations

import sqlite3
from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Generator

import pytest
from openpyxl import Workbook, load_workbook


class LayerTestContext:
    """
    Test context for layered tests.

    Provides isolated DB and Excel environment for each test.
    """

    def __init__(self, temp_dir: Path):
        # Import here to ensure root conftest has added src to path
        from autodbaudit.infrastructure.sqlite.history_store import HistoryStore
        from autodbaudit.application.annotation_sync import AnnotationSyncService

        self.temp_dir = temp_dir
        self.db_path = temp_dir / "audit_history.db"
        self.excel_path = temp_dir / "Audit_Report.xlsx"

        # Initialize DB
        self.store = HistoryStore(self.db_path)
        self.store.initialize_schema()

        # Create initial audit run
        self.conn = self.store._get_connection()
        now = datetime.now(timezone.utc).isoformat()
        cursor = self.conn.execute(
            "INSERT INTO audit_runs (started_at, status, run_type) VALUES (?, ?, ?)",
            (now, "completed", "audit"),
        )
        self.conn.commit()
        self.run_id = cursor.lastrowid

        # Create dummy server/instance for findings
        self.conn.execute("INSERT INTO servers (hostname) VALUES (?)", ("TestServer",))
        self.conn.commit()
        self.server_id = self.conn.execute(
            "SELECT id FROM servers WHERE hostname = ?", ("TestServer",)
        ).fetchone()[0]

        self.conn.execute(
            "INSERT INTO instances (server_id, instance_name, port) VALUES (?, ?, ?)",
            (self.server_id, "TestInstance", 1433),
        )
        self.conn.commit()
        self.instance_id = self.conn.execute(
            "SELECT id FROM instances WHERE server_id = ?", (self.server_id,)
        ).fetchone()[0]

        # Link instance to run
        self.conn.execute(
            "INSERT INTO audit_run_instances (audit_run_id, instance_id, checked_at) VALUES (?, ?, ?)",
            (self.run_id, self.instance_id, now),
        )
        self.conn.commit()

        # Annotation sync service
        self.annot_sync = AnnotationSyncService(self.db_path)

    def cleanup(self) -> None:
        """Clean up resources."""
        self.conn.close()
        if hasattr(self.store, "_connection") and self.store._connection:
            self.store._connection.close()

    def create_excel(self, sheets: list[dict[str, Any]] | None = None) -> None:
        """
        Create an Excel file with specified sheets.

        Args:
            sheets: List of sheet configs {name, headers, data}
        """
        wb = Workbook()
        ws = wb.active

        if sheets:
            for i, sheet_config in enumerate(sheets):
                if i == 0:
                    ws.title = sheet_config["name"]
                else:
                    ws = wb.create_sheet(sheet_config["name"])

                # Add headers (row 1)
                for col_idx, header in enumerate(sheet_config.get("headers", []), 1):
                    ws.cell(row=1, column=col_idx, value=header)

                # Add data rows
                for row_idx, row_data in enumerate(sheet_config.get("data", []), 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            ws.title = "Cover"

        wb.save(self.excel_path)

    def read_excel_cell(self, sheet: str, row: int, column: int) -> Any:
        """Read a cell value from Excel."""
        wb = load_workbook(self.excel_path)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            return ws.cell(row=row, column=column).value
        return None

    def write_excel_cell(self, sheet: str, row: int, column: int, value: Any) -> bool:
        """Write a cell value to Excel."""
        try:
            wb = load_workbook(self.excel_path)
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                ws.cell(row=row, column=column, value=value)
                wb.save(self.excel_path)
                return True
            return False
        except Exception:
            return False

    def set_db_annotation(
        self,
        entity_type: str,
        entity_key: str,
        field: str,
        value: str,
    ) -> bool:
        """Set an annotation in the database."""
        from autodbaudit.infrastructure.sqlite.schema import set_annotation

        try:
            set_annotation(
                connection=self.conn,
                entity_type=entity_type,
                entity_key=entity_key,
                field_name=field,
                field_value=value,
            )
            self.conn.commit()
            return True
        except Exception:
            return False

    def get_db_annotation(
        self,
        entity_type: str,
        entity_key: str,
        field: str,
    ) -> str | None:
        """Get an annotation from the database."""
        cursor = self.conn.execute(
            f"""SELECT {field} FROM row_annotations
                WHERE entity_type = ? AND entity_key LIKE ?""",
            (entity_type, f"%{entity_key}%"),
        )
        row = cursor.fetchone()
        return row[0] if row else None


@pytest.fixture
def layer_ctx(tmp_path: Path) -> Generator[LayerTestContext, None, None]:
    """Create a LayerTestContext for each test."""
    ctx = LayerTestContext(tmp_path)
    yield ctx
    ctx.cleanup()
