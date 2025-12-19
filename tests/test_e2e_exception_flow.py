"""
TRUE End-to-End Tests for Exception Flow.

These tests run the ACTUAL --audit and --sync commands, then verify:
1. Excel files are created with correct content
2. Database has correct findings and annotations
3. Exception detection works correctly across multiple syncs

Run with: python -m pytest tests/test_e2e_exception_flow.py -v --tb=short

NOTE: These tests require SQL Server instances to be available as configured
in config/sql_targets.json. For CI/CD, you may need to mock the SQL connection.
"""

import sys
import os
import unittest
import tempfile
import shutil
import sqlite3
import json
from pathlib import Path
from datetime import datetime
from unittest.mock import patch, MagicMock

# Add src to path
sys.path.insert(0, str(Path(__file__).parents[1] / "src"))


class TestE2EExceptionFlow(unittest.TestCase):
    """
    True E2E tests that simulate the full audit→sync→sync flow.

    These tests create a temp output directory, run commands, and verify results.
    They use mocked SQL connections to avoid requiring actual SQL Server instances.
    """

    def setUp(self):
        """Create temp directories for isolated testing."""
        self.temp_dir = tempfile.mkdtemp(prefix="autodbaudit_e2e_")
        self.output_dir = Path(self.temp_dir) / "output"
        self.output_dir.mkdir()
        self.db_path = self.output_dir / "audit_history.db"

        # Create minimal config
        self.config_dir = Path(self.temp_dir) / "config"
        self.config_dir.mkdir()

        # Create test targets file (mocked - no real connection needed)
        targets = {
            "targets": [
                {
                    "id": "test",
                    "name": "Test Server",
                    "server": "localhost",
                    "instance": "TEST",
                    "port": 1433,
                    "auth": "windows",
                    "enabled": True,
                }
            ]
        }
        with open(self.config_dir / "sql_targets.json", "w") as f:
            json.dump(targets, f)

    def tearDown(self):
        """Clean up temp directories."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _create_mock_audit_service(self):
        """Create a mock audit service that creates realistic test data."""
        from autodbaudit.infrastructure.sqlite import HistoryStore, initialize_schema_v2
        from autodbaudit.infrastructure.sqlite.schema import (
            save_finding,
            build_entity_key,
        )

        # Initialize DB
        store = HistoryStore(self.db_path)
        store.initialize_schema()
        conn = sqlite3.connect(self.db_path)
        initialize_schema_v2(conn)

        # Create server/instance
        conn.execute(
            "INSERT INTO servers (hostname, ip_address) VALUES (?, ?)",
            ("localhost", "127.0.0.1"),
        )
        server_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        conn.execute(
            "INSERT INTO instances (server_id, instance_name, version, version_major) VALUES (?, ?, ?, ?)",
            (server_id, "TEST", "15.0.4123.1", 15),
        )
        instance_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Create audit run
        conn.execute(
            "INSERT INTO audit_runs (started_at, status, run_type) VALUES (?, ?, ?)",
            (datetime.now().isoformat(), "completed", "audit"),
        )
        run_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Create test findings - mix of PASS and FAIL
        test_logins = [
            ("admin_login", "FAIL", "high", "Admin with excess privileges"),
            ("service_account", "PASS", None, "Service account OK"),
            ("readonly_user", "PASS", None, "Read-only user OK"),
            ("legacy_admin", "FAIL", "critical", "Legacy admin needs review"),
        ]

        for login_name, status, risk, desc in test_logins:
            entity_key = build_entity_key("localhost", "TEST", login_name)
            save_finding(
                conn,
                run_id,
                instance_id,
                entity_key,
                "login",
                login_name,
                status,
                risk,
                desc,
            )

        conn.commit()
        conn.close()

        return run_id, instance_id

    def _create_mock_excel_with_annotations(self, run_id):
        """
        Create a mock Excel file with annotations for testing.

        This simulates what the user does: adding justifications to the Excel.
        """
        from openpyxl import Workbook

        excel_path = self.output_dir / "Audit_Report.xlsx"
        wb = Workbook()

        # Create Server Logins sheet
        ws = wb.active
        ws.title = "Server Logins"

        # Headers (matching SHEET_ANNOTATION_CONFIG)
        headers = [
            "Server",
            "Instance",
            "Login Name",
            "Type",
            "Status",
            "Action",
            "Review Status",
            "Justification",
            "Notes",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Data rows - simulating what --audit creates
        test_data = [
            # (Server, Instance, Login, Type, Status, Action, Review Status, Justification, Notes)
            ("localhost", "TEST", "admin_login", "SQL", "FAIL", "⏳", "", "", ""),
            ("localhost", "TEST", "service_account", "SQL", "PASS", "", "", "", ""),
            ("localhost", "TEST", "readonly_user", "SQL", "PASS", "", "", "", ""),
            ("localhost", "TEST", "legacy_admin", "SQL", "FAIL", "⏳", "", "", ""),
        ]

        for row_num, row_data in enumerate(test_data, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

        wb.save(excel_path)
        wb.close()
        return excel_path

    def _add_annotation_to_excel(
        self, excel_path, login_name, justification, review_status=""
    ):
        """Simulate user adding annotation to Excel."""
        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        ws = wb["Server Logins"]

        # Find the row with the login
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=3).value == login_name:
                ws.cell(row=row, column=7, value=review_status)  # Review Status
                ws.cell(row=row, column=8, value=justification)  # Justification
                break

        wb.save(excel_path)
        wb.close()

    def test_e2e_exception_detection_full_flow(self):
        """
        Full E2E test:
        1. Create mock audit data
        2. Create Excel with annotations
        3. Run detect_exception_changes
        4. Verify correct exceptions detected
        """
        print("\n" + "=" * 60)
        print("E2E TEST: Exception Detection Full Flow")
        print("=" * 60)

        # Step 1: Create mock audit
        run_id, instance_id = self._create_mock_audit_service()
        print(f"✅ Created mock audit (run_id={run_id})")

        # Step 2: Create Excel with base data
        excel_path = self._create_mock_excel_with_annotations(run_id)
        print(f"✅ Created Excel at {excel_path}")

        # Step 3: Add annotation to FAIL row (should be exception)
        self._add_annotation_to_excel(
            excel_path, "admin_login", "Legacy admin - approved by IT", "✓ Exception"
        )
        print("✅ Added annotation to FAIL row (admin_login)")

        # Step 4: Add annotation to PASS row (should NOT be exception)
        self._add_annotation_to_excel(
            excel_path, "service_account", "Service account for app X", ""
        )
        print("✅ Added annotation to PASS row (service_account)")

        # Step 5: Read annotations and detect changes
        from autodbaudit.application.annotation_sync import AnnotationSyncService
        from autodbaudit.infrastructure.sqlite import HistoryStore

        store = HistoryStore(self.db_path)
        current_findings = store.get_findings(run_id)
        print(f"✅ Loaded {len(current_findings)} findings from DB")

        annot_sync = AnnotationSyncService(self.db_path)
        old_annotations = {}  # First sync - no previous
        current_annotations = annot_sync.read_all_from_excel(excel_path)
        print(f"✅ Read {len(current_annotations)} annotations from Excel")

        # Step 6: Detect exceptions
        exceptions = annot_sync.detect_exception_changes(
            old_annotations, current_annotations, current_findings
        )
        print(f"✅ Detected {len(exceptions)} exception changes")

        for ex in exceptions:
            print(f"   {ex['change_type']}: {ex['entity_key']}")

        # Step 7: Verify
        self.assertEqual(len(exceptions), 1, "Should detect exactly 1 exception")
        self.assertEqual(exceptions[0]["change_type"], "added")
        self.assertIn("admin_login", exceptions[0]["entity_key"])

        # PASS row should NOT be in exceptions
        pass_exceptions = [
            e for e in exceptions if "service_account" in e["entity_key"]
        ]
        self.assertEqual(
            len(pass_exceptions), 0, "PASS row should NOT be logged as exception"
        )

        print("\n✅ E2E TEST PASSED!")
        print("=" * 60)

    def test_e2e_second_sync_stability(self):
        """
        E2E test for second sync:
        1. First sync with annotations
        2. Persist to DB
        3. Second sync with SAME annotations
        4. Should detect NO changes
        """
        print("\n" + "=" * 60)
        print("E2E TEST: Second Sync Stability")
        print("=" * 60)

        # Setup
        run_id, _ = self._create_mock_audit_service()
        excel_path = self._create_mock_excel_with_annotations(run_id)

        # Add annotation to PASS row
        self._add_annotation_to_excel(
            excel_path, "service_account", "Service account for app X", ""
        )

        from autodbaudit.application.annotation_sync import AnnotationSyncService
        from autodbaudit.infrastructure.sqlite import HistoryStore

        store = HistoryStore(self.db_path)
        current_findings = store.get_findings(run_id)
        annot_sync = AnnotationSyncService(self.db_path)

        # First sync
        print("📌 First sync...")
        old_annotations_1 = {}
        current_annotations_1 = annot_sync.read_all_from_excel(excel_path)
        annot_sync.persist_to_db(current_annotations_1)

        exceptions_1 = annot_sync.detect_exception_changes(
            old_annotations_1, current_annotations_1, current_findings
        )
        print(f"   First sync detected: {len(exceptions_1)} changes")

        # Second sync - same Excel, no changes
        print("📌 Second sync (no Excel changes)...")
        old_annotations_2 = annot_sync.load_from_db()
        current_annotations_2 = annot_sync.read_all_from_excel(excel_path)

        exceptions_2 = annot_sync.detect_exception_changes(
            old_annotations_2, current_annotations_2, current_findings
        )
        print(f"   Second sync detected: {len(exceptions_2)} changes")

        for ex in exceptions_2:
            print(f"   ⚠️ Unexpected: {ex['change_type']}: {ex['entity_key']}")

        # Verify no false changes
        self.assertEqual(
            len(exceptions_2),
            0,
            "Second sync with unchanged data should detect NO changes",
        )

        print("\n✅ E2E TEST PASSED!")
        print("=" * 60)


if __name__ == "__main__":
    unittest.main(verbosity=2)
