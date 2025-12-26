"""
Excel Assertions - Sheet/column structure verification.

Provides utilities for verifying Excel report structure without
caring about specific content (that's for the test cases).
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


class ExcelAssertions:
    """
    Excel structure verification utilities.

    These work on any workbook - mock or real.
    """

    @staticmethod
    def assert_sheet_exists(wb: Workbook, name: str) -> None:
        """Verify sheet exists in workbook."""
        assert (
            name in wb.sheetnames
        ), f"Sheet '{name}' not found. Available: {wb.sheetnames}"

    @staticmethod
    def assert_all_sheets_exist(wb: Workbook, expected: list[str]) -> list[str]:
        """
        Verify all expected sheets exist.

        Returns list of missing sheets (empty if all present).
        """
        missing = [s for s in expected if s not in wb.sheetnames]
        assert not missing, f"Missing sheets: {missing}"
        return missing

    @staticmethod
    def assert_column_headers(
        ws: Worksheet, expected: list[str], strict: bool = False
    ) -> list[str]:
        """
        Verify column headers match expected.

        Args:
            ws: Worksheet to check
            expected: List of expected header names (in order)
            strict: If True, fail on any mismatch. If False, return mismatches.

        Returns:
            List of mismatches in format "Col N: expected 'X', got 'Y'"
        """
        actual = [cell.value for cell in ws[1]]
        mismatches = []

        for i, exp in enumerate(expected):
            if i >= len(actual):
                mismatches.append(f"Col {i+1}: expected '{exp}', column missing")
            elif actual[i] != exp:
                mismatches.append(f"Col {i+1}: expected '{exp}', got '{actual[i]}'")

        if strict:
            assert not mismatches, f"Column mismatches: {mismatches}"

        return mismatches

    @staticmethod
    def find_entity_row(ws: Worksheet, pattern: str, column: int = 2) -> int | None:
        """
        Find row number where entity pattern matches.

        Args:
            ws: Worksheet to search
            pattern: Regex pattern to match
            column: Column index (1-based) to search in

        Returns:
            Row number (1-based) or None if not found
        """
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=column).value
            if cell_value and re.search(pattern, str(cell_value), re.IGNORECASE):
                return row_num
        return None

    @staticmethod
    def assert_entity_exists(ws: Worksheet, pattern: str, column: int = 2) -> int:
        """
        Assert entity exists and return its row.

        Raises AssertionError if not found.
        """
        row = ExcelAssertions.find_entity_row(ws, pattern, column)
        assert (
            row is not None
        ), f"Entity matching '{pattern}' not found in column {column}"
        return row

    @staticmethod
    def get_cell_value(ws: Worksheet, row: int, header: str) -> any:
        """
        Get cell value by row and header name.

        Args:
            ws: Worksheet
            row: Row number (1-based)
            header: Header name to find column

        Returns:
            Cell value or None if header not found
        """
        headers = [cell.value for cell in ws[1]]
        try:
            col = headers.index(header) + 1
            return ws.cell(row=row, column=col).value
        except ValueError:
            return None

    @staticmethod
    def assert_cell_value(ws: Worksheet, row: int, header: str, expected: any) -> None:
        """Assert cell value matches expected."""
        actual = ExcelAssertions.get_cell_value(ws, row, header)
        assert (
            actual == expected
        ), f"Row {row} '{header}': expected '{expected}', got '{actual}'"

    @staticmethod
    def assert_dropdowns_exist(ws: Worksheet) -> int:
        """
        Assert data validations (dropdowns) exist on sheet.

        Returns count of validations.
        """
        count = len(ws.data_validations.dataValidation)
        assert count > 0, f"No data validations (dropdowns) found on sheet '{ws.title}'"
        return count

    @staticmethod
    def assert_conditional_formatting_exists(ws: Worksheet) -> int:
        """
        Assert conditional formatting rules exist.

        Returns count of CF rules.
        """
        rules = ws.conditional_formatting._cf_rules
        count = len(rules)
        assert count > 0, f"No conditional formatting on sheet '{ws.title}'"
        return count
