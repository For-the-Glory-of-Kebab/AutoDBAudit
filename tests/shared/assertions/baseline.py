"""
Baseline Manager - Track pre-existing discrepancies for delta assertions.

This module handles:
- Capturing baseline state before tests
- Tracking what test fixtures add
- Computing deltas for assertions
- Protecting privileged entities (King login)
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook


@dataclass
class SheetBaseline:
    """Baseline state for a single sheet."""

    sheet_name: str
    entity_column: str
    entity_keys: set[str] = field(default_factory=set)
    discrepant_keys: set[str] = field(default_factory=set)  # FAIL/WARN
    exception_keys: set[str] = field(default_factory=set)  # Has justification
    row_count: int = 0

    def to_dict(self) -> dict:
        return {
            "sheet_name": self.sheet_name,
            "entity_column": self.entity_column,
            "entity_keys": list(self.entity_keys),
            "discrepant_keys": list(self.discrepant_keys),
            "exception_keys": list(self.exception_keys),
            "row_count": self.row_count,
        }

    @classmethod
    def from_dict(cls, data: dict) -> "SheetBaseline":
        return cls(
            sheet_name=data["sheet_name"],
            entity_column=data["entity_column"],
            entity_keys=set(data.get("entity_keys", [])),
            discrepant_keys=set(data.get("discrepant_keys", [])),
            exception_keys=set(data.get("exception_keys", [])),
            row_count=data.get("row_count", 0),
        )


@dataclass
class AuditBaseline:
    """Complete audit baseline state."""

    audit_id: int
    sheets: dict[str, SheetBaseline] = field(default_factory=dict)
    protected_entities: set[str] = field(default_factory=set)
    timestamp: str = ""

    def add_sheet(self, baseline: SheetBaseline) -> None:
        self.sheets[baseline.sheet_name] = baseline

    def get_sheet(self, name: str) -> SheetBaseline | None:
        return self.sheets.get(name)

    def to_json(self) -> str:
        data = {
            "audit_id": self.audit_id,
            "timestamp": self.timestamp,
            "protected_entities": list(self.protected_entities),
            "sheets": {name: bl.to_dict() for name, bl in self.sheets.items()},
        }
        return json.dumps(data, indent=2)

    @classmethod
    def from_json(cls, json_str: str) -> "AuditBaseline":
        data = json.loads(json_str)
        baseline = cls(
            audit_id=data["audit_id"],
            timestamp=data.get("timestamp", ""),
            protected_entities=set(data.get("protected_entities", [])),
        )
        for name, sheet_data in data.get("sheets", {}).items():
            baseline.add_sheet(SheetBaseline.from_dict(sheet_data))
        return baseline


# Protected entities that should NEVER be modified by tests
PROTECTED_ENTITIES = frozenset(
    [
        "King",  # King SQL login
        "sa",  # Built-in SA (though we test enable/disable)
        "SYSTEM",
        "NT AUTHORITY",
        "NT SERVICE",
    ]
)


# Sheet entity column mapping
SHEET_ENTITY_COLUMNS = {
    "Logins": "Login",
    "SA Account": "SA Name",
    "Configuration": "Setting",
    "Databases": "Database",
    "DB Users": "User",
    "Orphaned Users": "User",
    "Roles": "Role",
    "DB Roles": "Role",
    "Permissions": "Permission",
    "Services": "Service",
    "Linked Servers": "Linked Server",
    "Backups": "Database",
    "Encryption": "Database",
    "Triggers": "Trigger",
    "Audit Settings": "Setting",
}


class BaselineCapture:
    """
    Capture and manage baseline state for delta assertions.

    Usage:
        # Before tests
        baseline = BaselineCapture.capture(workbook, audit_id)

        # After applying fixtures and syncing
        delta = BaselineCapture.compute_delta(baseline, workbook)

        # Assert only on delta
        assert "WeakPolicyAdmin_TEST" in delta.new_discrepancies["Logins"]
    """

    @staticmethod
    def capture(wb: "Workbook", audit_id: int) -> AuditBaseline:
        """
        Capture baseline state from workbook.

        Args:
            wb: Excel workbook
            audit_id: Current audit ID

        Returns:
            AuditBaseline with all sheet states
        """
        from datetime import datetime

        baseline = AuditBaseline(
            audit_id=audit_id,
            timestamp=datetime.now().isoformat(),
            protected_entities=set(PROTECTED_ENTITIES),
        )

        for sheet_name in wb.sheetnames:
            if sheet_name in ("Cover", "Actions", "Action Log"):
                continue

            entity_col = SHEET_ENTITY_COLUMNS.get(sheet_name, "Entity")
            sheet_bl = BaselineCapture._capture_sheet(
                wb[sheet_name], sheet_name, entity_col
            )
            baseline.add_sheet(sheet_bl)

        return baseline

    @staticmethod
    def _capture_sheet(ws, sheet_name: str, entity_col: str) -> SheetBaseline:
        """Capture single sheet baseline."""
        baseline = SheetBaseline(
            sheet_name=sheet_name,
            entity_column=entity_col,
        )

        # Find column indices
        headers = [c.value for c in ws[1]]
        entity_idx = None
        result_idx = None
        just_idx = None

        for i, h in enumerate(headers):
            if h == entity_col:
                entity_idx = i + 1
            elif h == "Result":
                result_idx = i + 1
            elif h == "Justification":
                just_idx = i + 1

        if entity_idx is None:
            return baseline

        # Capture all entities
        for row_num in range(2, ws.max_row + 1):
            entity = ws.cell(row=row_num, column=entity_idx).value
            if not entity:
                continue

            entity_key = str(entity)
            baseline.entity_keys.add(entity_key)
            baseline.row_count += 1

            # Check if discrepant
            if result_idx:
                result = ws.cell(row=row_num, column=result_idx).value
                if result in ("FAIL", "WARN"):
                    baseline.discrepant_keys.add(entity_key)

            # Check if exception
            if just_idx:
                just = ws.cell(row=row_num, column=just_idx).value
                if just and str(just).strip():
                    baseline.exception_keys.add(entity_key)

        return baseline

    @staticmethod
    def save(baseline: AuditBaseline, path: Path) -> None:
        """Save baseline to JSON file."""
        path.write_text(baseline.to_json())

    @staticmethod
    def load(path: Path) -> AuditBaseline:
        """Load baseline from JSON file."""
        return AuditBaseline.from_json(path.read_text())

    @staticmethod
    def compute_delta(
        baseline: AuditBaseline,
        current_wb: "Workbook",
    ) -> "BaselineDelta":
        """
        Compute delta between baseline and current state.

        Returns:
            BaselineDelta with new/removed entities per sheet
        """
        delta = BaselineDelta()

        for sheet_name in current_wb.sheetnames:
            if sheet_name in ("Cover", "Actions", "Action Log"):
                continue

            baseline_sheet = baseline.get_sheet(sheet_name)
            if baseline_sheet is None:
                continue

            current_sheet = BaselineCapture._capture_sheet(
                current_wb[sheet_name],
                sheet_name,
                baseline_sheet.entity_column,
            )

            # New entities
            new_entities = current_sheet.entity_keys - baseline_sheet.entity_keys
            delta.new_entities[sheet_name] = new_entities

            # New discrepancies
            new_disc = current_sheet.discrepant_keys - baseline_sheet.discrepant_keys
            delta.new_discrepancies[sheet_name] = new_disc

            # New exceptions
            new_exc = current_sheet.exception_keys - baseline_sheet.exception_keys
            delta.new_exceptions[sheet_name] = new_exc

            # Removed entities
            removed = baseline_sheet.entity_keys - current_sheet.entity_keys
            delta.removed_entities[sheet_name] = removed

            # Fixed (was discrepant, now not)
            fixed = baseline_sheet.discrepant_keys - current_sheet.discrepant_keys
            delta.fixed_entities[sheet_name] = fixed

        return delta


@dataclass
class BaselineDelta:
    """Delta between baseline and current state."""

    new_entities: dict[str, set[str]] = field(default_factory=dict)
    removed_entities: dict[str, set[str]] = field(default_factory=dict)
    new_discrepancies: dict[str, set[str]] = field(default_factory=dict)
    fixed_entities: dict[str, set[str]] = field(default_factory=dict)
    new_exceptions: dict[str, set[str]] = field(default_factory=dict)

    def get_new_discrepancies(self, sheet: str) -> set[str]:
        return self.new_discrepancies.get(sheet, set())

    def get_fixed(self, sheet: str) -> set[str]:
        return self.fixed_entities.get(sheet, set())

    def assert_new_discrepancy(self, sheet: str, entity: str) -> None:
        """Assert entity is a new discrepancy (not in baseline)."""
        new_disc = self.get_new_discrepancies(sheet)
        assert entity in new_disc, (
            f"'{entity}' should be new discrepancy in {sheet}. "
            f"New discrepancies: {new_disc}"
        )

    def assert_fixed(self, sheet: str, entity: str) -> None:
        """Assert entity was fixed (was discrepant, now not)."""
        fixed = self.get_fixed(sheet)
        assert entity in fixed, f"'{entity}' should be fixed in {sheet}. Fixed: {fixed}"
