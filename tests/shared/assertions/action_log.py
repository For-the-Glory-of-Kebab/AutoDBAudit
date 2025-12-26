"""
Action Log Assertions - Verify action log entries.

Checks that transitions are correctly recorded.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import sqlite3
    from openpyxl.workbook import Workbook


@dataclass
class ActionLogEntry:
    """A single action log entry."""

    entity_key: str
    action_type: str
    notes: str | None = None
    action_date: str | None = None


class ActionLogAssertions:
    """
    Action log verification utilities.
    """

    @staticmethod
    def get_db_entries(
        conn: sqlite3.Connection,
        run_id: int,
        action_type: str | None = None,
    ) -> list[ActionLogEntry]:
        """
        Get action log entries from database.

        Args:
            conn: SQLite connection
            run_id: Sync run ID
            action_type: Filter by action type (optional)
        """
        if action_type:
            cursor = conn.execute(
                """
                SELECT entity_key, action_type, notes, action_date
                FROM action_log
                WHERE sync_run_id = ? AND action_type = ?
            """,
                (run_id, action_type),
            )
        else:
            cursor = conn.execute(
                """
                SELECT entity_key, action_type, notes, action_date
                FROM action_log
                WHERE sync_run_id = ?
            """,
                (run_id,),
            )

        return [
            ActionLogEntry(
                entity_key=row[0],
                action_type=row[1],
                notes=row[2],
                action_date=row[3],
            )
            for row in cursor
        ]

    @staticmethod
    def get_excel_entries(wb: Workbook) -> list[ActionLogEntry]:
        """
        Get entries from Actions sheet in Excel.
        """
        entries = []

        if "Actions" not in wb.sheetnames:
            return entries

        ws = wb["Actions"]
        headers = [cell.value for cell in ws[1]]

        # Find column indices
        try:
            entity_col = headers.index("Entity") + 1
            type_col = headers.index("Description") + 1  # or "Action Type"
        except ValueError:
            return entries

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            entity = row[entity_col - 1].value if entity_col <= len(row) else None
            action_type = row[type_col - 1].value if type_col <= len(row) else None

            if entity:
                entries.append(
                    ActionLogEntry(
                        entity_key=str(entity),
                        action_type=str(action_type) if action_type else "",
                    )
                )

        return entries

    @staticmethod
    def assert_entry_exists(
        entries: list[ActionLogEntry],
        entity_pattern: str,
        action_type: str | None = None,
    ) -> ActionLogEntry:
        """
        Assert an entry exists for entity.

        Args:
            entries: List of entries to search
            entity_pattern: Substring to match in entity_key
            action_type: Expected action type (optional)

        Returns:
            The matching entry
        """
        for entry in entries:
            if entity_pattern.lower() in entry.entity_key.lower():
                if (
                    action_type is None
                    or action_type.lower() in entry.action_type.lower()
                ):
                    return entry

        available = [f"{e.entity_key}: {e.action_type}" for e in entries[:5]]
        raise AssertionError(
            f"No action log entry for '{entity_pattern}' "
            f"(type={action_type}). Available: {available}"
        )

    @staticmethod
    def assert_no_entry(
        entries: list[ActionLogEntry],
        entity_pattern: str,
    ) -> None:
        """Assert no entry exists for entity."""
        for entry in entries:
            if entity_pattern.lower() in entry.entity_key.lower():
                raise AssertionError(
                    f"Unexpected action log entry for '{entity_pattern}': {entry}"
                )

    @staticmethod
    def assert_count(
        entries: list[ActionLogEntry],
        action_type: str,
        expected: int,
    ) -> None:
        """Assert count of entries with action type."""
        actual = sum(1 for e in entries if action_type.lower() in e.action_type.lower())
        assert (
            actual == expected
        ), f"Expected {expected} '{action_type}' entries, got {actual}"
