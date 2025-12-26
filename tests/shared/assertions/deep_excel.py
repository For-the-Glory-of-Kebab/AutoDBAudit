"""
Deep Excel Assertions - Row-level, content-based, style-based verification.

This module provides comprehensive assertions that verify:
- Specific row content (not just counts)
- Cell styling (fonts, colors, fills)
- Conditional formatting correctness
- Annotation field exact values
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell import Cell


@dataclass
class RowData:
    """Complete data from a single worksheet row."""

    row_num: int
    cells: dict[str, any] = field(default_factory=dict)
    styles: dict[str, dict] = field(default_factory=dict)

    def get(self, column: str, default: any = None) -> any:
        """Get cell value by column name."""
        return self.cells.get(column, default)

    def get_style(self, column: str) -> dict | None:
        """Get cell style by column name."""
        return self.styles.get(column)


@dataclass
class CellStyle:
    """Extracted cell styling information."""

    font_name: str | None = None
    font_size: float | None = None
    font_bold: bool = False
    font_color: str | None = None
    fill_color: str | None = None
    fill_pattern: str | None = None
    border_style: str | None = None
    alignment: str | None = None


class DeepExcelAssertions:
    """
    Deep, content-based Excel assertions.

    These go beyond counts to verify actual cell content and styling.
    """

    @staticmethod
    def extract_headers(ws: Worksheet) -> list[str]:
        """Extract column headers from row 1."""
        return [cell.value for cell in ws[1] if cell.value]

    @staticmethod
    def header_to_column_map(ws: Worksheet) -> dict[str, int]:
        """Map header names to column indices (1-based)."""
        return {cell.value: i + 1 for i, cell in enumerate(ws[1]) if cell.value}

    @staticmethod
    def extract_row(ws: Worksheet, row_num: int) -> RowData:
        """
        Extract complete row data including values and styles.

        Args:
            ws: Worksheet
            row_num: Row number (1-based)

        Returns:
            RowData with all cell values and styles
        """
        headers = DeepExcelAssertions.extract_headers(ws)
        row_data = RowData(row_num=row_num)

        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=i)
            row_data.cells[header] = cell.value

            # Extract styling
            row_data.styles[header] = {
                "font_name": cell.font.name if cell.font else None,
                "font_bold": cell.font.bold if cell.font else False,
                "font_color": (
                    DeepExcelAssertions._extract_color(cell.font.color)
                    if cell.font
                    else None
                ),
                "fill_color": (
                    DeepExcelAssertions._extract_fill_color(cell.fill)
                    if cell.fill
                    else None
                ),
                "fill_pattern": cell.fill.patternType if cell.fill else None,
            }

        return row_data

    @staticmethod
    def _extract_color(color) -> str | None:
        """Extract color as hex string."""
        if color is None:
            return None
        if hasattr(color, "rgb") and color.rgb:
            return str(color.rgb)
        if hasattr(color, "theme"):
            return f"theme:{color.theme}"
        return None

    @staticmethod
    def _extract_fill_color(fill) -> str | None:
        """Extract fill foreground color."""
        if fill is None:
            return None
        if hasattr(fill, "fgColor") and fill.fgColor:
            return DeepExcelAssertions._extract_color(fill.fgColor)
        return None

    @staticmethod
    def find_row_by_entity(
        ws: Worksheet,
        pattern: str,
        entity_column: str = "Login",
        start_row: int = 2,
        max_rows: int = 1000,
    ) -> RowData | None:
        """
        Find row where entity column matches pattern.

        Args:
            ws: Worksheet to search
            pattern: Regex pattern to match
            entity_column: Column containing entity name
            start_row: Row to start searching
            max_rows: Maximum rows to search

        Returns:
            RowData if found, None otherwise
        """
        col_map = DeepExcelAssertions.header_to_column_map(ws)

        if entity_column not in col_map:
            # Try common entity columns
            for alt in ["Entity", "Name", "Instance", "Setting", "SA Name"]:
                if alt in col_map:
                    entity_column = alt
                    break
            else:
                return None

        col_idx = col_map[entity_column]

        for row_num in range(start_row, min(start_row + max_rows, ws.max_row + 1)):
            cell_value = ws.cell(row=row_num, column=col_idx).value
            if cell_value and re.search(pattern, str(cell_value), re.IGNORECASE):
                return DeepExcelAssertions.extract_row(ws, row_num)

        return None

    @staticmethod
    def assert_row_exists(
        ws: Worksheet,
        pattern: str,
        entity_column: str = "Login",
    ) -> RowData:
        """
        Assert row exists and return its data.

        Raises AssertionError with details if not found.
        """
        row = DeepExcelAssertions.find_row_by_entity(ws, pattern, entity_column)

        if row is None:
            # Collect sample values for debugging
            col_map = DeepExcelAssertions.header_to_column_map(ws)
            col_idx = col_map.get(entity_column, 2)
            samples = []
            for r in range(2, min(10, ws.max_row + 1)):
                val = ws.cell(row=r, column=col_idx).value
                if val:
                    samples.append(str(val)[:30])

            raise AssertionError(
                f"Row matching '{pattern}' not found in column '{entity_column}'.\n"
                f"Sample values: {samples}"
            )

        return row

    @staticmethod
    def assert_cell_value(
        row: RowData,
        column: str,
        expected: any,
        msg: str = "",
    ) -> None:
        """Assert specific cell value in row."""
        actual = row.get(column)

        if actual != expected:
            raise AssertionError(
                f"{msg}Row {row.row_num} column '{column}': "
                f"expected '{expected}', got '{actual}'"
            )

    @staticmethod
    def assert_cell_not_empty(
        row: RowData,
        column: str,
        msg: str = "",
    ) -> None:
        """Assert cell is not empty."""
        actual = row.get(column)

        if actual is None or actual == "" or str(actual).strip() == "":
            raise AssertionError(f"{msg}Row {row.row_num} column '{column}' is empty")

    @staticmethod
    def assert_cell_style(
        row: RowData,
        column: str,
        bold: bool | None = None,
        fill_color: str | None = None,
        font_color: str | None = None,
    ) -> None:
        """
        Assert cell has expected styling.

        Args:
            row: Row data
            column: Column name
            bold: Expected bold state (None = don't check)
            fill_color: Expected fill color contains this substring
            font_color: Expected font color contains this substring
        """
        style = row.get_style(column)

        if style is None:
            raise AssertionError(f"No style data for column '{column}'")

        if bold is not None and style.get("font_bold") != bold:
            raise AssertionError(
                f"Column '{column}' bold: expected {bold}, got {style.get('font_bold')}"
            )

        if fill_color is not None:
            actual_fill = style.get("fill_color") or ""
            if fill_color.lower() not in actual_fill.lower():
                raise AssertionError(
                    f"Column '{column}' fill color: expected contains '{fill_color}', "
                    f"got '{actual_fill}'"
                )

        if font_color is not None:
            actual_font = style.get("font_color") or ""
            if font_color.lower() not in actual_font.lower():
                raise AssertionError(
                    f"Column '{column}' font color: expected contains '{font_color}', "
                    f"got '{actual_font}'"
                )

    @staticmethod
    def assert_result_is_pass(row: RowData) -> None:
        """Assert Result column is PASS."""
        result = row.get("Result")
        assert result == "PASS", f"Expected PASS, got '{result}'"

    @staticmethod
    def assert_result_is_fail(row: RowData) -> None:
        """Assert Result column is FAIL."""
        result = row.get("Result")
        assert result in ("FAIL", "WARN"), f"Expected FAIL/WARN, got '{result}'"

    @staticmethod
    def assert_has_justification(row: RowData) -> None:
        """Assert row has non-empty justification."""
        just = row.get("Justification")
        assert just and str(just).strip(), f"Row {row.row_num} missing justification"

    @staticmethod
    def assert_exception_indicator(row: RowData, expected: str = "✓") -> None:
        """Assert exception indicator is present."""
        # Try common indicator column names
        for col in ["Indicator", "Exception", "Status Indicator"]:
            val = row.get(col)
            if val:
                assert expected in str(
                    val
                ), f"Expected indicator '{expected}', got '{val}'"
                return
        # If no indicator column, check if justification exists (implies exception)
        just = row.get("Justification")
        if expected == "✓":
            assert (
                just and str(just).strip()
            ), "No indicator column, checking justification"
