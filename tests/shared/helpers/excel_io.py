"""
Excel I/O - Read/write Excel files for testing.

Provides utilities for manipulating Excel files in tests.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


class ExcelIO:
    """
    Excel file I/O utilities for testing.

    Handles opening, reading, writing, and saving Excel files
    without leaving files locked.
    """

    def __init__(self, excel_path: Path) -> None:
        self.excel_path = excel_path
        self._wb: Workbook | None = None

    def open(self) -> Workbook:
        """Open workbook for editing."""
        self._wb = load_workbook(self.excel_path)
        return self._wb

    def save(self) -> None:
        """Save and close workbook."""
        if self._wb:
            self._wb.save(self.excel_path)
            self._wb.close()
            self._wb = None

    def close(self) -> None:
        """Close without saving."""
        if self._wb:
            self._wb.close()
            self._wb = None

    @property
    def wb(self) -> Workbook:
        """Get workbook (opens if needed)."""
        if self._wb is None:
            self.open()
        return self._wb

    def get_sheet(self, name: str) -> Worksheet | None:
        """Get sheet by name."""
        if name in self.wb.sheetnames:
            return self.wb[name]
        return None

    def find_column_index(self, ws: Worksheet, header: str) -> int | None:
        """Find column index (1-based) for header name."""
        for i, cell in enumerate(ws[1], start=1):
            if cell.value == header:
                return i
        return None

    def write_cell(
        self,
        sheet: str,
        row: int,
        column: str | int,
        value: any,
    ) -> bool:
        """
        Write value to cell.

        Args:
            sheet: Sheet name
            row: Row number (1-based)
            column: Column name (header) or index (1-based)
            value: Value to write

        Returns:
            True if successful
        """
        ws = self.get_sheet(sheet)
        if ws is None:
            return False

        if isinstance(column, str):
            col_idx = self.find_column_index(ws, column)
            if col_idx is None:
                return False
        else:
            col_idx = column

        ws.cell(row=row, column=col_idx).value = value
        return True

    def read_cell(
        self,
        sheet: str,
        row: int,
        column: str | int,
    ) -> any:
        """
        Read cell value.

        Args:
            sheet: Sheet name
            row: Row number (1-based)
            column: Column name or index

        Returns:
            Cell value or None
        """
        ws = self.get_sheet(sheet)
        if ws is None:
            return None

        if isinstance(column, str):
            col_idx = self.find_column_index(ws, column)
            if col_idx is None:
                return None
        else:
            col_idx = column

        return ws.cell(row=row, column=col_idx).value

    def write_annotation(
        self,
        sheet: str,
        row: int,
        notes: str | None = None,
        justification: str | None = None,
        review_status: str | None = None,
    ) -> bool:
        """
        Write annotation values to a row.

        Convenience method for writing test annotations.
        """
        success = True

        if notes is not None:
            success &= self.write_cell(sheet, row, "Notes", notes)
        if justification is not None:
            success &= self.write_cell(sheet, row, "Justification", justification)
        if review_status is not None:
            success &= self.write_cell(sheet, row, "Review Status", review_status)

        return success
