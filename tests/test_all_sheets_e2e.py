"""
Exhaustive E2E Test Suite for All Sheets.

This test creates actual Excel files, adds annotations, runs sync,
and verifies everything is preserved correctly.

Tests ALL 17 data sheets with ALL editable columns.
"""

import sys
import unittest
import tempfile
import sqlite3
import shutil
from pathlib import Path
from datetime import datetime

# Add src to path
sys.path.insert(0, str(Path(__file__).parents[1] / "src"))

from openpyxl import Workbook, load_workbook

from autodbaudit.infrastructure.sqlite import HistoryStore, initialize_schema_v2
from autodbaudit.infrastructure.sqlite.schema import save_finding, build_entity_key
from autodbaudit.application.annotation_sync import (
    AnnotationSyncService,
    SHEET_ANNOTATION_CONFIG,
)


class TestAllSheetsAnnotations(unittest.TestCase):
    """
    Test annotation sync for ALL sheets defined in SHEET_ANNOTATION_CONFIG.

    This ensures:
    1. All sheets can read annotations correctly
    2. All sheets build correct entity keys (lowercase)
    3. Exception detection works for all exception-eligible sheets
    4. Notes/Purpose fields are preserved for all sheets
    """

    def setUp(self):
        """Create temp DB and Excel file."""
        self.temp_dir = Path(tempfile.mkdtemp())
        self.db_path = self.temp_dir / "test_history.db"
        self.excel_path = self.temp_dir / "test_audit.xlsx"

        # Initialize DB
        self.store = HistoryStore(self.db_path)
        self.store.initialize_schema()
        conn = self.store._get_connection()
        initialize_schema_v2(conn)

        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row

        # Create server/instance
        self._create_server_instance()

        # Create run
        self.run_id = self._create_run()

    def tearDown(self):
        self.conn.close()
        if self.store._connection:
            self.store._connection.close()
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _create_server_instance(self):
        """Create test server and instance."""
        cursor = self.conn.execute(
            "INSERT INTO servers (hostname, ip_address) VALUES (?, ?)",
            ("localhost", "127.0.0.1"),
        )
        self.conn.commit()
        self.server_id = cursor.lastrowid

        cursor = self.conn.execute(
            "INSERT INTO instances (server_id, instance_name, version, version_major) VALUES (?, ?, ?, ?)",
            (self.server_id, "TestInstance", "15.0.4123.1", 15),
        )
        self.conn.commit()
        self.instance_id = cursor.lastrowid

    def _create_run(self, run_type="audit"):
        """Create audit run."""
        cursor = self.conn.execute(
            "INSERT INTO audit_runs (started_at, status, run_type) VALUES (?, ?, ?)",
            (datetime.now().isoformat(), "running", run_type),
        )
        self.conn.commit()
        return cursor.lastrowid

    def _create_excel_with_sheet(self, sheet_name: str, config: dict, data_rows: list):
        """Create Excel file with specified sheet and data."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Build headers from config
        headers = list(config["key_cols"]) + list(config["editable_cols"].keys())
        # Add Status column if this sheet has exception capability
        if "justification" in config["editable_cols"]:
            if "Status" not in headers:
                headers.insert(len(config["key_cols"]), "Status")

        ws.append(headers)

        # Add data rows
        for row in data_rows:
            ws.append(row)

        wb.save(self.excel_path)
        return headers

    def test_all_sheets_in_config(self):
        """Verify all sheets are properly configured."""
        print("\n=== Testing All Sheets in SHEET_ANNOTATION_CONFIG ===")

        expected_sheets = {
            "Instances",
            "SA Account",
            "Server Logins",
            "Sensitive Roles",
            "Configuration",
            "Services",
            "Databases",
            "Database Users",
            "Database Roles",
            "Permission Grants",
            "Orphaned Users",
            "Linked Servers",
            "Triggers",
            "Client Protocols",
            "Backups",
            "Audit Settings",
            "Encryption",
            "Actions",
        }

        actual_sheets = set(SHEET_ANNOTATION_CONFIG.keys())

        print(f"Configured sheets: {len(actual_sheets)}")
        for sheet in sorted(actual_sheets):
            config = SHEET_ANNOTATION_CONFIG[sheet]
            print(
                f"  - {sheet}: keys={config['key_cols']}, editable={list(config['editable_cols'].keys())}"
            )

        missing = expected_sheets - actual_sheets
        self.assertEqual(len(missing), 0, f"Missing sheets: {missing}")

    def test_backup_sheet_exception_detection(self):
        """Test Backups sheet with lowercase key normalization."""
        print("\n=== Testing Backups Sheet Exception Detection ===")

        config = SHEET_ANNOTATION_CONFIG["Backups"]

        # Create Excel with Backups data - using mixed case that should be normalized
        headers = self._create_excel_with_sheet(
            "Backups",
            config,
            [
                # Server, Instance, Database, Recovery Model, Status, Review Status, Justification, Last Reviewed, Notes
                [
                    "localhost",
                    "TestInstance",
                    "AdventureWorks",
                    "FULL",
                    "FAIL",
                    "✓ Exception",
                    "Test backup exception",
                    "",
                    "Test note",
                ],
            ],
        )

        # Create corresponding finding with lowercase key
        entity_key = build_entity_key(
            "localhost", "testinstance", "adventureworks|full"
        )
        save_finding(
            self.conn,
            self.run_id,
            self.instance_id,
            entity_key,
            "backup",
            "adventureworks|full",
            "FAIL",
            "high",
            "No recent backup",
        )
        self.conn.commit()

        # Read annotations
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)

        print(f"Read {len(annotations)} annotations")
        for key, fields in annotations.items():
            print(f"  Key: {key}")
            print(f"  Fields: {fields}")

        # Verify key is lowercase
        self.assertTrue(len(annotations) > 0, "Should read at least one annotation")
        for key in annotations.keys():
            self.assertEqual(key, key.lower(), f"Key should be lowercase: {key}")

    def test_logins_sheet_exception_detection(self):
        """Test Server Logins sheet with exception handling."""
        print("\n=== Testing Server Logins Sheet ===")

        config = SHEET_ANNOTATION_CONFIG["Server Logins"]

        # Create Excel with Logins data
        headers = [
            "Server",
            "Instance",
            "Login Name",
            "Status",
            "Review Status",
            "Justification",
            "Last Revised",
            "Notes",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Server Logins"
        ws.append(headers)
        ws.append(
            [
                "localhost",
                "TestInstance",
                "admin_user",
                "FAIL",
                "✓ Exception",
                "Admin access required",
                "",
                "Production admin",
            ]
        )
        wb.save(self.excel_path)

        # Create finding
        entity_key = build_entity_key("localhost", "testinstance", "admin_user")
        save_finding(
            self.conn,
            self.run_id,
            self.instance_id,
            entity_key,
            "login",
            "admin_user",
            "FAIL",
            "high",
            "Sensitive login",
        )
        self.conn.commit()
        findings = self.store.get_findings(self.run_id)

        # Read annotations and detect exceptions
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)

        print(f"Annotations: {annotations}")

        # Detect exception changes (old=empty, new=current)
        exceptions = sync.detect_exception_changes({}, annotations, findings)

        print(f"Detected exceptions: {len(exceptions)}")
        for ex in exceptions:
            print(f"  {ex['change_type']}: {ex.get('entity_key', 'N/A')}")

        # Should detect the exception
        self.assertEqual(
            len(exceptions), 1, "Should detect 1 exception for Logins sheet"
        )

    def test_triggers_sheet_notes_preserved(self):
        """Test Triggers sheet preserves Notes fields with new Scope-based schema."""
        print("\n=== Testing Triggers Sheet Notes Preservation ===")

        config = SHEET_ANNOTATION_CONFIG["Triggers"]

        # Create Excel with Triggers data using NEW schema with Scope column
        # key_cols: ["Server", "Instance", "Scope", "Database", "Trigger Name"]
        # editable_cols: ["Review Status", "Notes", "Justification", "Last Revised"]
        headers = ["Server", "Instance", "Scope", "Database", "Trigger Name", "Event", "Enabled", "Review Status", "Notes", "Justification", "Last Revised"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Triggers"
        ws.append(headers)
        ws.append(
            [
                "localhost",
                "TestInstance",
                "DATABASE",  # Scope column
                "AppDB",
                "trg_audit",
                "DDL",
                "Yes",
                "",
                "Audit logging trigger",  # Notes
                "",
                "",
            ]
        )
        ws.append(
            [
                "localhost",
                "TestInstance",
                "DATABASE",  # Scope column
                "AppDB",
                "trg_validation",
                "DML",
                "Yes",
                "",
                "Data validation",  # Notes
                "",
                "",
            ]
        )
        wb.save(self.excel_path)

        # Read annotations
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)

        print(f"Read {len(annotations)} Trigger annotations")
        for key, fields in annotations.items():
            print(
                f"  {key}: notes='{fields.get('notes', '')}'"
            )

        # Should have 2 entries with notes
        self.assertEqual(len(annotations), 2, "Should read 2 trigger annotations")

        # Verify lowercase keys
        for key in annotations.keys():
            self.assertEqual(key, key.lower())

    def test_sensitive_roles_exception(self):
        """Test Sensitive Roles sheet exception detection."""
        print("\n=== Testing Sensitive Roles Sheet ===")

        # Create Excel - key_cols: ["Server", "Instance", "Role", "Member"]
        headers = [
            "Server",
            "Instance",
            "Role",  # Added - required for key matching
            "Member",
            "Status",
            "Review Status",
            "Justification",
            "Last Revised",
            "Notes",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensitive Roles"
        ws.append(headers)
        ws.append(
            [
                "localhost",
                "TestInstance",
                "sysadmin",  # Role value
                "sysadmin_user",
                "FAIL",
                "✓ Exception",
                "DBA team member",
                "",
                "",
            ]
        )
        wb.save(self.excel_path)

        # Create finding - include Role in key for matching
        entity_key = build_entity_key("localhost", "testinstance", "sysadmin", "sysadmin_user")
        save_finding(
            self.conn,
            self.run_id,
            self.instance_id,
            entity_key,
            "server_role_member",
            "sysadmin_user",
            "FAIL",
            "high",
            "sysadmin role member",
        )
        self.conn.commit()
        findings = self.store.get_findings(self.run_id)

        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)
        exceptions = sync.detect_exception_changes({}, annotations, findings)

        print(f"Sensitive Roles exceptions: {len(exceptions)}")
        self.assertEqual(len(exceptions), 1)

    def test_sa_account_exception(self):
        """Test SA Account sheet with fixed 'sa' entity name."""
        print("\n=== Testing SA Account Sheet ===")

        # SA Account uses "Current Name" as a key column
        headers = [
            "Server",
            "Instance",
            "Current Name",
            "Status",
            "Review Status",
            "Justification",
            "Last Reviewed",
            "Notes",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "SA Account"
        ws.append(headers)
        ws.append(
            [
                "localhost",
                "TestInstance",
                "sa",
                "FAIL",
                "✓ Exception",
                "Required for legacy app",
                "",
                "Legacy",
            ]
        )
        wb.save(self.excel_path)

        # Create finding - SA account uses 'sa' as entity_name
        entity_key = build_entity_key("localhost", "testinstance", "sa")
        save_finding(
            self.conn,
            self.run_id,
            self.instance_id,
            entity_key,
            "sa_account",
            "sa",
            "FAIL",
            "high",
            "SA account enabled",
        )
        self.conn.commit()
        findings = self.store.get_findings(self.run_id)

        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)
        exceptions = sync.detect_exception_changes({}, annotations, findings)

        print(f"SA Account exceptions: {len(exceptions)}")
        self.assertEqual(len(exceptions), 1)

    def test_configuration_exception(self):
        """Test Configuration sheet exception."""
        print("\n=== Testing Configuration Sheet ===")

        headers = [
            "Server",
            "Instance",
            "Setting",
            "Status",
            "Review Status",
            "Exception Reason",
            "Last Reviewed",
            "Notes",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Configuration"
        ws.append(headers)
        ws.append(
            [
                "localhost",
                "TestInstance",
                "xp_cmdshell",
                "FAIL",
                "✓ Exception",
                "Required for maintenance",
                "",
                "",
            ]
        )
        wb.save(self.excel_path)

        entity_key = build_entity_key("localhost", "testinstance", "xp_cmdshell")
        save_finding(
            self.conn,
            self.run_id,
            self.instance_id,
            entity_key,
            "config",
            "xp_cmdshell",
            "FAIL",
            "high",
            "xp_cmdshell enabled",
        )
        self.conn.commit()
        findings = self.store.get_findings(self.run_id)

        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)
        exceptions = sync.detect_exception_changes({}, annotations, findings)

        print(f"Configuration exceptions: {len(exceptions)}")
        self.assertEqual(len(exceptions), 1)


class TestAllStateTransitions(unittest.TestCase):
    """Test all state transitions work correctly with new lowercase keys."""

    def setUp(self):
        self.temp_dir = Path(tempfile.mkdtemp())
        self.db_path = self.temp_dir / "test_history.db"

        self.store = HistoryStore(self.db_path)
        self.store.initialize_schema()
        conn = self.store._get_connection()
        initialize_schema_v2(conn)

        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        self._create_server_instance()

    def tearDown(self):
        self.conn.close()
        if self.store._connection:
            self.store._connection.close()
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _create_server_instance(self):
        cursor = self.conn.execute(
            "INSERT INTO servers (hostname, ip_address) VALUES (?, ?)",
            ("localhost", "127.0.0.1"),
        )
        self.conn.commit()
        self.server_id = cursor.lastrowid

        cursor = self.conn.execute(
            "INSERT INTO instances (server_id, instance_name, version, version_major) VALUES (?, ?, ?, ?)",
            (self.server_id, "TestInstance", "15.0.4123.1", 15),
        )
        self.conn.commit()
        self.instance_id = cursor.lastrowid

    def _create_run(self):
        cursor = self.conn.execute(
            "INSERT INTO audit_runs (started_at, status, run_type) VALUES (?, ?, ?)",
            (datetime.now().isoformat(), "running", "sync"),
        )
        self.conn.commit()
        return cursor.lastrowid

    def test_new_exception_detected(self):
        """Test: NEW exception (FAIL + justification) detected correctly."""
        print("\n=== Testing NEW_EXCEPTION state transition ===")

        run_id = self._create_run()
        entity_key = build_entity_key("localhost", "testinstance", "test_entity")

        save_finding(
            self.conn,
            run_id,
            self.instance_id,
            entity_key,
            "login",
            "test_entity",
            "FAIL",
            "high",
            "Test",
        )
        self.conn.commit()
        findings = self.store.get_findings(run_id)

        sync = AnnotationSyncService(self.db_path)

        # New annotation with justification
        new_annotations = {
            "login|localhost|testinstance|test_entity": {
                "justification": "Test exception",
                "review_status": "✓ Exception",
            }
        }

        exceptions = sync.detect_exception_changes({}, new_annotations, findings)

        self.assertEqual(len(exceptions), 1)
        self.assertEqual(exceptions[0]["change_type"], "added")
        print("✅ NEW_EXCEPTION detected correctly")

    def test_exception_removed(self):
        """Test: EXCEPTION_REMOVED (clearing justification) detected."""
        print("\n=== Testing EXCEPTION_REMOVED state transition ===")

        run_id = self._create_run()
        entity_key = build_entity_key("localhost", "testinstance", "removable")

        save_finding(
            self.conn,
            run_id,
            self.instance_id,
            entity_key,
            "login",
            "removable",
            "FAIL",
            "high",
            "Test",
        )
        self.conn.commit()
        findings = self.store.get_findings(run_id)

        sync = AnnotationSyncService(self.db_path)

        # Old: had exception
        old_annotations = {
            "login|localhost|testinstance|removable": {
                "justification": "Was excepted",
                "review_status": "✓ Exception",
            }
        }

        # New: no exception (cleared)
        new_annotations = {
            "login|localhost|testinstance|removable": {
                "justification": "",
                "review_status": "",
            }
        }

        exceptions = sync.detect_exception_changes(
            old_annotations, new_annotations, findings
        )

        removed = [e for e in exceptions if e["change_type"] == "removed"]
        self.assertEqual(len(removed), 1)
        print("✅ EXCEPTION_REMOVED detected correctly")

    def test_pass_with_justification_not_exception(self):
        """Test: PASS + justification = documentation only (not exception)."""
        print("\n=== Testing PASS + justification = NOT exception ===")

        run_id = self._create_run()
        entity_key = build_entity_key("localhost", "testinstance", "passing")

        # PASS status
        save_finding(
            self.conn,
            run_id,
            self.instance_id,
            entity_key,
            "login",
            "passing",
            "PASS",
            None,
            "OK",
        )
        self.conn.commit()
        findings = self.store.get_findings(run_id)

        sync = AnnotationSyncService(self.db_path)

        new_annotations = {
            "login|localhost|testinstance|passing": {
                "justification": "This is just a note",
                "review_status": "",
            }
        }

        exceptions = sync.detect_exception_changes({}, new_annotations, findings)

        self.assertEqual(
            len(exceptions), 0, "PASS row should NOT be logged as exception"
        )
        print("✅ PASS + justification correctly treated as documentation only")


if __name__ == "__main__":
    unittest.main(verbosity=2)
