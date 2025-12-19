"""
Real Excel Roundtrip E2E Test.

This test uses the ACTUAL Excel writer to create sheets,
then reads annotations back using the ACTUAL annotation reader,
to catch any column name or key format mismatches.

This is the DEFINITIVE test for sync reliability.
"""

import sys
import unittest
import tempfile
import sqlite3
import shutil
from pathlib import Path
from datetime import datetime

# Add src to path
sys.path.insert(0, str(Path(__file__).parents[1] / "src"))

from openpyxl import load_workbook

from autodbaudit.infrastructure.sqlite import HistoryStore, initialize_schema_v2
from autodbaudit.infrastructure.sqlite.schema import save_finding, build_entity_key
from autodbaudit.infrastructure.excel.writer import EnhancedReportWriter
from autodbaudit.application.annotation_sync import (
    AnnotationSyncService,
    SHEET_ANNOTATION_CONFIG,
)


class TestRealExcelRoundtrip(unittest.TestCase):
    """
    Test that uses the REAL ExcelWriter to create sheets,
    then reads them back with AnnotationSyncService.

    This catches any mismatch between:
    - Column names in Excel vs SHEET_ANNOTATION_CONFIG
    - Key format in Excel vs findings
    """

    def setUp(self):
        """Create temp DB and Excel file."""
        self.temp_dir = Path(tempfile.mkdtemp())
        self.db_path = self.temp_dir / "test_history.db"
        self.excel_path = self.temp_dir / "test_audit.xlsx"

        # Initialize DB
        self.store = HistoryStore(self.db_path)
        self.store.initialize_schema()
        conn = self.store._get_connection()
        initialize_schema_v2(conn)

        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row

        # Create server/instance
        self._create_server_instance()

        # Create run
        self.run_id = self._create_run()

    def tearDown(self):
        self.conn.close()
        if self.store._connection:
            self.store._connection.close()
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _create_server_instance(self):
        """Create test server and instance."""
        cursor = self.conn.execute(
            "INSERT INTO servers (hostname, ip_address) VALUES (?, ?)",
            ("TestServer", "127.0.0.1"),
        )
        self.conn.commit()
        self.server_id = cursor.lastrowid

        cursor = self.conn.execute(
            "INSERT INTO instances (server_id, instance_name, version, version_major) VALUES (?, ?, ?, ?)",
            (self.server_id, "INST1", "15.0.4123.1", 15),
        )
        self.conn.commit()
        self.instance_id = cursor.lastrowid

    def _create_run(self, run_type="audit"):
        """Create audit run."""
        cursor = self.conn.execute(
            "INSERT INTO audit_runs (started_at, status, run_type) VALUES (?, ?, ?)",
            (datetime.now().isoformat(), "running", run_type),
        )
        self.conn.commit()
        return cursor.lastrowid

    def test_logins_sheet_roundtrip(self):
        """Test Server Logins sheet: write with real writer, read back, verify."""
        print("\n=== Testing Server Logins REAL Roundtrip ===")

        # 1. Create Excel with real writer
        writer = EnhancedReportWriter()

        # Add a login
        writer.add_login(
            server_name="TestServer",
            instance_name="INST1",
            login_name="admin_user",
            login_type="SQL Login",
            is_disabled=False,
            pwd_policy=False,  # FAIL - no password policy
            default_db="master",
        )

        writer.save(self.excel_path)

        # 2. Manually add justification to the Excel file
        wb = load_workbook(self.excel_path)
        ws = wb["Server Logins"]

        # Find the Justification column
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"Headers: {header_row}")

        justification_col = None
        review_status_col = None
        for idx, header in enumerate(header_row):
            if header:
                h = str(header).strip()
                if "Justification" in h:
                    justification_col = idx + 1  # 1-indexed
                if "Review Status" in h:
                    review_status_col = idx + 1

        print(
            f"Justification col: {justification_col}, Review Status col: {review_status_col}"
        )

        # Set justification in row 2
        if justification_col:
            ws.cell(
                row=2,
                column=justification_col,
                value="Admin access needed for DBA team",
            )
        if review_status_col:
            ws.cell(row=2, column=review_status_col, value="✓ Exception")

        wb.save(self.excel_path)
        wb.close()

        # 3. Create corresponding finding
        entity_key = build_entity_key("TestServer", "INST1", "admin_user")
        save_finding(
            self.conn,
            self.run_id,
            self.instance_id,
            entity_key,
            "login",
            "admin_user",
            "FAIL",
            "high",
            "No password policy",
        )
        self.conn.commit()
        findings = self.store.get_findings(self.run_id)

        # 4. Read annotations with real reader
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)

        print(f"\nAnnotations read: {len(annotations)}")
        for key, fields in annotations.items():
            print(f"  Key: {key}")
            print(f"  Fields: {fields}")

        # 5. Verify we got the annotation
        self.assertGreater(len(annotations), 0, "Should read at least one annotation")

        # Check the annotation has justification
        found_justification = False
        for key, fields in annotations.items():
            if "admin_user" in key.lower():
                self.assertIn(
                    "justification",
                    fields,
                    f"Annotation {key} should have justification",
                )
                self.assertEqual(
                    fields["justification"], "Admin access needed for DBA team"
                )
                found_justification = True
                break

        self.assertTrue(found_justification, "Should find the login annotation")

        # 6. Detect exception
        exceptions = sync.detect_exception_changes({}, annotations, findings)

        print(f"\nDetected exceptions: {len(exceptions)}")
        for ex in exceptions:
            print(f"  {ex['change_type']}: entity_key={ex.get('entity_key', 'N/A')}")

        self.assertEqual(len(exceptions), 1, "Should detect 1 exception for admin_user")
        print("✅ Server Logins roundtrip PASSED!")

    def test_backups_sheet_roundtrip(self):
        """Test Backups sheet: write with real writer, read back, verify."""
        print("\n=== Testing Backups REAL Roundtrip ===")

        # 1. Create Excel with real writer
        writer = EnhancedReportWriter()

        # Add a backup info
        writer.add_backup_info(
            server_name="TestServer",
            instance_name="INST1",
            database_name="AppDB",
            recovery_model="FULL",
            last_backup_date=None,  # No backup = FAIL
            days_since=None,
            backup_path="",
            backup_size_mb=None,
        )

        writer.save(self.excel_path)

        # 2. Manually add justification
        wb = load_workbook(self.excel_path)
        ws = wb["Backups"]

        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"Backups Headers: {header_row}")

        justification_col = None
        for idx, header in enumerate(header_row):
            if header and "Justification" in str(header):
                justification_col = idx + 1

        if justification_col:
            ws.cell(row=2, column=justification_col, value="Dev DB - no backup needed")

        wb.save(self.excel_path)
        wb.close()

        # 3. Create corresponding finding - NOTE: entity_name includes recovery model or backup type!
        entity_key = build_entity_key("TestServer", "INST1", "AppDB|full")
        save_finding(
            self.conn,
            self.run_id,
            self.instance_id,
            entity_key,
            "backup",
            "AppDB|full",
            "FAIL",
            "high",
            "No backup",
        )
        self.conn.commit()
        findings = self.store.get_findings(self.run_id)

        # 4. Read annotations
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)

        print(f"\nBackup annotations: {len(annotations)}")
        for key, fields in annotations.items():
            print(f"  Key: {key}")
            print(f"  Fields: {fields}")

        # 5. Verify
        self.assertGreater(
            len(annotations), 0, "Should read at least one backup annotation"
        )

        # The key should be: backup|testserver|inst1|appdb|full
        # (lowercase, and includes Recovery Model)
        found = False
        for key, fields in annotations.items():
            if "appdb" in key.lower() and "backup" in key.lower():
                found = True
                print(f"  Found backup annotation: {key}")

        self.assertTrue(found, "Should find backup annotation for AppDB")
        print("✅ Backups roundtrip PASSED!")

    def test_linked_servers_sheet_roundtrip(self):
        """Test Linked Servers sheet: write with real writer, read back, verify."""
        print("\n=== Testing Linked Servers REAL Roundtrip ===")

        # 1. Create Excel with real writer
        writer = EnhancedReportWriter()

        writer.add_linked_server(
            server_name="TestServer",
            instance_name="INST1",
            linked_server_name="REMOTE_DB",
            product="SQL Server",
            provider="SQLNCLI11",
            data_source="remotehost",
            rpc_out=True,
            local_login="sa",
            remote_login="sa",
            impersonate=False,
            risk_level="HIGH_PRIVILEGE",
        )

        writer.save(self.excel_path)

        # 2. Add Purpose and Justification
        wb = load_workbook(self.excel_path)
        ws = wb["Linked Servers"]

        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"Linked Servers Headers: {header_row}")

        purpose_col = None
        justification_col = None
        for idx, header in enumerate(header_row):
            h = str(header).strip() if header else ""
            if "Purpose" in h:
                purpose_col = idx + 1
            if "Justification" in h:
                justification_col = idx + 1

        print(f"Purpose col: {purpose_col}, Justification col: {justification_col}")

        if purpose_col:
            ws.cell(row=2, column=purpose_col, value="Legacy reporting connection")
        if justification_col:
            ws.cell(
                row=2, column=justification_col, value="Required for legacy reports"
            )

        wb.save(self.excel_path)
        wb.close()

        # 3. Create finding
        entity_key = build_entity_key("TestServer", "INST1", "REMOTE_DB")
        save_finding(
            self.conn,
            self.run_id,
            self.instance_id,
            entity_key,
            "linked_server",
            "REMOTE_DB",
            "FAIL",
            "high",
            "High-privilege linked server",
        )
        self.conn.commit()
        findings = self.store.get_findings(self.run_id)

        # 4. Read annotations
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)

        print(f"\nLinked Server annotations: {len(annotations)}")
        for key, fields in annotations.items():
            print(f"  Key: {key}")
            print(f"  Fields: {fields}")

        # 5. Verify we got both purpose and justification
        found = False
        for key, fields in annotations.items():
            if "remote_db" in key.lower():
                found = True
                self.assertIn("purpose", fields, "Should have purpose field")
                self.assertEqual(fields["purpose"], "Legacy reporting connection")
                self.assertIn(
                    "justification", fields, "Should have justification field"
                )
                self.assertEqual(fields["justification"], "Required for legacy reports")

        self.assertTrue(found, "Should find linked server annotation")
        print("✅ Linked Servers roundtrip PASSED!")

    def test_database_roles_sheet_roundtrip(self):
        """Test Database Roles sheet roundtrip."""
        print("\n=== Testing Database Roles REAL Roundtrip ===")

        # 1. Create Excel with real writer
        writer = EnhancedReportWriter()

        writer.add_db_role_member(
            server_name="TestServer",
            instance_name="INST1",
            database_name="AppDB",
            role_name="db_owner",
            member_name="app_user",
            member_type="SQL User",
        )

        writer.save(self.excel_path)

        # 2. Add justification
        wb = load_workbook(self.excel_path)
        ws = wb["Database Roles"]

        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"Database Roles Headers: {header_row}")

        justification_col = None
        for idx, header in enumerate(header_row):
            if header and "Justification" in str(header):
                justification_col = idx + 1

        if justification_col:
            ws.cell(
                row=2,
                column=justification_col,
                value="App needs db_owner for schema changes",
            )

        wb.save(self.excel_path)
        wb.close()

        # 3. Read annotations
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)

        print(f"\nDatabase Roles annotations: {len(annotations)}")
        for key, fields in annotations.items():
            print(f"  Key: {key}")
            print(f"  Fields: {fields}")

        # 4. Verify
        # Key should be: db_role|testserver|inst1|appdb|db_owner|app_user
        self.assertGreater(
            len(annotations), 0, "Should read at least one db role annotation"
        )

        found = False
        for key, fields in annotations.items():
            if "db_owner" in key.lower() and "app_user" in key.lower():
                found = True
                self.assertIn("justification", fields)

        self.assertTrue(found, "Should find database role annotation")
        print("✅ Database Roles roundtrip PASSED!")

    def test_column_headers_match_config(self):
        """Verify all sheets have expected columns."""
        print("\n=== Testing Column Headers Match Config ===")

        # Create Excel with real writer - add sample data to create sheets
        writer = EnhancedReportWriter()

        # Add minimal data to trigger sheet creation
        writer.add_login("Srv", "Inst", "login1", "SQL", False, True, "master")
        writer.add_linked_server("Srv", "Inst", "linked1", "", "SQLNCLI", "src", False)
        writer.add_backup_info("Srv", "Inst", "DB1", "FULL", None, None, "", None)
        writer.add_db_role_member("Srv", "Inst", "DB1", "role1", "user1", "SQL")

        writer.save(self.excel_path)

        wb = load_workbook(self.excel_path)

        mismatches = []

        for sheet_name, config in SHEET_ANNOTATION_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                continue  # Skip sheets not created

            ws = wb[sheet_name]
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            headers = [str(h).strip() if h else "" for h in header_row]

            print(f"\n{sheet_name}:")
            print(f"  Expected editable: {list(config['editable_cols'].keys())}")
            print(f"  Actual headers: {[h for h in headers if h]}")

            # Check each expected editable column exists
            for col_name in config["editable_cols"].keys():
                # Direct match or partial match
                found = any(col_name.lower() in h.lower() for h in headers if h)
                if not found:
                    mismatches.append(f"{sheet_name}: Missing '{col_name}'")
                    print(f"  ❌ Missing: {col_name}")

        wb.close()

        if mismatches:
            print(f"\n❌ Column mismatches found: {mismatches}")
        else:
            print("\n✅ All column headers match config!")

        # This test should FAIL if there are mismatches - that's the point!
        self.assertEqual(len(mismatches), 0, f"Column mismatches: {mismatches}")


if __name__ == "__main__":
    unittest.main(verbosity=2)
