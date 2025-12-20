"""
Actions Sheet Integration Test

Verifies the Actions sheet functionality:
1. ID-based row matching (unique key per action entry)
2. Notes column read/write round-trip
3. Detected Date user override persistence
"""

import pytest
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook

from autodbaudit.application.annotation_sync import (
    AnnotationSyncService,
    SHEET_ANNOTATION_CONFIG,
)
from autodbaudit.infrastructure.sqlite.store import HistoryStore


class TestActionsSheetIntegration:
    """Test Actions sheet specific functionality."""

    @pytest.fixture
    def setup_actions_excel(self, tmp_path):
        """Create an Excel file with Actions sheet data."""
        excel_path = tmp_path / "test_actions.xlsx"
        db_path = tmp_path / "test.db"

        # Create workbook with Actions sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Actions"

        # Headers per EXCEL_COLUMNS.md:
        # ID, Server, Instance, Category, Finding, Risk Level, Change Description, Change Type, Detected Date, Notes
        headers = [
            "ID",
            "Server",
            "Instance",
            "Category",
            "Finding",
            "Risk Level",
            "Change Description",
            "Change Type",
            "Detected Date",
            "Notes",
        ]
        ws.append(headers)

        # Add test actions
        actions = [
            [
                1,
                "srv1",
                "inst1",
                "sa_account",
                "SA Disabled",
                "High",
                "Fixed issue",
                "✓ Closed",
                "2025-01-01",
                "User note 1",
            ],
            [
                2,
                "srv1",
                "inst1",
                "trigger",
                "Trigger logged",
                "Medium",
                "New Issue",
                "⏳ Open",
                "2025-01-02",
                "",
            ],
            [
                3,
                "srv2",
                "inst1",
                "config",
                "xp_cmdshell",
                "Critical",
                "Regression",
                "⏳ Open",
                "2025-01-03",
                "Needs review",
            ],
        ]
        for action in actions:
            ws.append(action)

        wb.save(excel_path)

        # Initialize database schema first
        store = HistoryStore(db_path)
        store.initialize_schema()

        # Initialize annotation service
        self.service = AnnotationSyncService(db_path)

        self.excel_path = excel_path
        self.db_path = db_path

        return tmp_path

    def test_actions_config_exists(self):
        """Verify Actions sheet configuration is defined."""
        assert "Actions" in SHEET_ANNOTATION_CONFIG
        config = SHEET_ANNOTATION_CONFIG["Actions"]

        assert config["entity_type"] == "action"
        assert config["key_cols"] == ["ID"]
        assert "Detected Date" in config["editable_cols"]
        assert "Notes" in config["editable_cols"]

    def test_read_actions_annotations(self, setup_actions_excel):
        """Test reading annotations from Actions sheet."""
        annotations = self.service.read_all_from_excel(self.excel_path)

        # Actions are keyed by entity_type|ID
        # Check for action with ID=1
        expected_keys = [
            "action|1",
            "action|2",
            "action|3",
        ]

        for key in expected_keys:
            assert key in annotations, f"Missing action key: {key}"

        # Verify content of action 1
        action_1 = annotations.get("action|1", {})
        assert action_1.get("notes") == "User note 1"
        assert "2025-01-01" in str(action_1.get("action_date", ""))

    def test_write_actions_annotations(self, setup_actions_excel):
        """Test writing annotations back to Actions sheet."""
        # Read original
        annotations = self.service.read_all_from_excel(self.excel_path)

        # Modify action 2's notes
        if "action|2" in annotations:
            annotations["action|2"]["notes"] = "Updated note for action 2"
            annotations["action|2"]["action_date"] = "2025-06-15"

        # Write back
        cells_updated = self.service.write_all_to_excel(self.excel_path, annotations)

        assert cells_updated > 0, "No cells were updated"

        # Verify by re-reading
        wb = load_workbook(self.excel_path)
        ws = wb["Actions"]

        # Find action 2 row (ID=2)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == 2:  # ID column
                assert (
                    row[9] == "Updated note for action 2"
                )  # Notes column (J = index 9)
                assert "2025-06-15" in str(row[8])  # Detected Date column (I = index 8)
                break
        else:
            pytest.fail("Action ID 2 not found in sheet")

    def test_actions_persist_to_db(self, setup_actions_excel):
        """Test persisting Actions annotations to database."""
        annotations = self.service.read_all_from_excel(self.excel_path)

        # Persist to DB
        saved_count = self.service.persist_to_db(annotations)

        assert saved_count > 0, "No annotations persisted to DB"

        # Load from DB and verify
        loaded = self.service.load_from_db()

        assert "action|1" in loaded or any(
            "action|1" in k for k in loaded
        ), f"Action 1 not in loaded annotations. Keys: {list(loaded.keys())[:10]}"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
