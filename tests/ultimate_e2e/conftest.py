"""
Shared fixtures and test infrastructure for Ultimate E2E tests.
"""

from __future__ import annotations

import sqlite3
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Generator

# Ensure src is in python path
PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import pytest
from openpyxl import load_workbook

# Import application components
from autodbaudit.infrastructure.excel.writer import EnhancedReportWriter
from autodbaudit.infrastructure.sqlite.store import HistoryStore
from autodbaudit.application.annotation_sync import AnnotationSyncService
from autodbaudit.application.actions.action_recorder import ActionRecorder
from autodbaudit.domain.change_types import ChangeType, DetectedChange
from autodbaudit.infrastructure.sqlite.schema import initialize_schema_v2, save_finding

from .sheet_specs import ALL_SHEET_SPECS, get_data_specs, SheetSpec


class TestContext:
    """
    Test context that holds all test infrastructure.

    This provides a clean, isolated environment for each test.
    """

    def __init__(self):
        self.temp_dir = Path(tempfile.mkdtemp(prefix="ultimate_e2e_"))
        self.db_path = self.temp_dir / "audit_history.db"
        self.excel_path = self.temp_dir / "Audit_Report.xlsx"

        # Initialize DB
        self.store = HistoryStore(self.db_path)
        self.store.initialize_schema()

        # Initialize v2 schema
        with sqlite3.connect(self.db_path) as conn:
            initialize_schema_v2(conn)

        # Create test connection
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row

        # Create server/instance
        cursor = self.conn.execute(
            "INSERT INTO servers (hostname, ip_address) VALUES (?, ?)",
            ("TestServer", "127.0.0.1"),
        )
        self.conn.commit()
        self.server_id = cursor.lastrowid

        cursor = self.conn.execute(
            "INSERT INTO instances (server_id, instance_name, version, version_major) VALUES (?, ?, ?, ?)",
            (self.server_id, "INST1", "15.0.4123.1", 15),
        )
        self.conn.commit()
        self.instance_id = cursor.lastrowid

        # Create audit run
        cursor = self.conn.execute(
            "INSERT INTO audit_runs (started_at, status, run_type) VALUES (?, ?, ?)",
            (datetime.now().isoformat(), "completed", "audit"),
        )
        self.conn.commit()
        self.run_id = cursor.lastrowid

        self.cycle_count = 0

    def cleanup(self) -> None:
        """Clean up resources."""
        self.conn.close()
        if self.store._connection:
            self.store._connection.close()

    def create_excel_with_specs(
        self, specs: list[SheetSpec] | None = None
    ) -> EnhancedReportWriter:
        """
        Create an Excel file with data for the given specs.

        Args:
            specs: List of SheetSpecs to include. Defaults to all data specs.
        """
        if specs is None:
            specs = list(get_data_specs())

        writer = EnhancedReportWriter()

        for spec in specs:
            if spec.writer_method is None:
                continue

            method = getattr(writer, spec.writer_method, None)
            if method is None:
                print(f"  ⚠️ Writer method not found: {spec.writer_method}")
                continue

            try:
                method(**spec.sample_kwargs)
            except Exception as e:
                print(f"  ⚠️ Error calling {spec.writer_method}: {e}")

        writer.save(str(self.excel_path))
        return writer

    def add_annotation_to_excel(
        self,
        sheet_name: str,
        row: int,
        column_name: str,
        value: str,
    ) -> bool:
        """
        Add an annotation to an Excel cell.

        Returns True if successful.
        """
        wb = load_workbook(self.excel_path)

        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️ Sheet not found: {sheet_name}")
            wb.close()
            return False

        ws = wb[sheet_name]

        # Find column
        col_idx = None
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=col).value
            if cell_val and column_name.lower() in str(cell_val).lower():
                col_idx = col
                break

        if col_idx is None:
            print(f"  ⚠️ Column not found: {column_name} in {sheet_name}")
            wb.close()
            return False

        ws.cell(row=row, column=col_idx).value = value
        wb.save(self.excel_path)
        wb.close()
        return True

    def mark_row_as_fail(self, sheet_name: str, row: int = 2) -> bool:
        """
        Mark a row as FAIL/discrepant by setting Status and Action indicator.

        This is needed for exception detection tests since sample data creates PASS rows.
        """
        wb = load_workbook(self.excel_path)

        if sheet_name not in wb.sheetnames:
            wb.close()
            return False

        ws = wb[sheet_name]

        # Find Status column and Action indicator column
        status_col = None
        action_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                header_str = str(header).strip().lower()
                if header_str == "status":
                    status_col = col
                if "⏳" in str(header):
                    action_col = col

        # Set Status to FAIL if column exists
        if status_col:
            ws.cell(row=row, column=status_col).value = "FAIL"

        # Set Action indicator to ⏳ (needs action) if column exists
        if action_col:
            ws.cell(row=row, column=action_col).value = "⏳"

        wb.save(self.excel_path)
        wb.close()
        return True

    def read_annotation_from_excel(
        self,
        sheet_name: str,
        row: int,
        column_name: str,
    ) -> Any:
        """Read an annotation value from Excel."""
        wb = load_workbook(self.excel_path)

        if sheet_name not in wb.sheetnames:
            wb.close()
            return None

        ws = wb[sheet_name]

        # Find column
        col_idx = None
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=col).value
            if cell_val and column_name.lower() in str(cell_val).lower():
                col_idx = col
                break

        if col_idx is None:
            wb.close()
            return None

        value = ws.cell(row=row, column=col_idx).value
        wb.close()
        return value

    def run_sync_cycle(
        self,
        annotations_to_add: dict[str, dict[str, str]] | None = None,
        generate_mock_findings: bool = True,
    ) -> dict[str, Any]:
        """
        Run a complete sync cycle.

        Args:
            annotations_to_add: {sheet_name: {col_name: value}} to add before sync
            generate_mock_findings: If True, mark all items as FAIL to test exceptions.
                                  If False, findings will be empty (PASS).

        Returns:
            Dict with cycle results including detected exceptions per sheet.
        """
        self.cycle_count += 1

        # Add annotations if provided
        if annotations_to_add:
            for sheet_name, cols in annotations_to_add.items():
                for col_name, value in cols.items():
                    self.add_annotation_to_excel(sheet_name, 2, col_name, value)

        # Create annotation sync service
        sync = AnnotationSyncService(self.db_path)

        # Read annotations from Excel
        current_annotations = sync.read_all_from_excel(self.excel_path)

        # Get old annotations from DB
        old_annotations = sync.load_from_db()

        # Persist current to DB
        persist_count = sync.persist_to_db(current_annotations)

        # Build mock findings for exception detection
        mock_findings = []
        if generate_mock_findings:
            mock_findings = self._build_mock_findings()

        # Detect exception changes
        exception_changes = sync.detect_exception_changes(
            old_annotations=old_annotations,
            new_annotations=current_annotations,
            current_findings=mock_findings,
        )

        # Reload and write back to EXISTING Excel (don't recreate, that wipes FAIL status)
        loaded = sync.load_from_db()
        # NOTE: Removed create_excel_with_specs() here - it was overwriting the Excel
        # and erasing FAIL status set by mark_row_as_fail()
        write_count = sync.write_all_to_excel(self.excel_path, loaded)

        # Helper to map string change types to Enum
        def get_change_type_enum(ct_str: str) -> ChangeType:
            mapping = {
                "added": ChangeType.EXCEPTION_ADDED,
                "removed": ChangeType.EXCEPTION_REMOVED,
                "updated": ChangeType.EXCEPTION_UPDATED,
                "new": ChangeType.NEW_ISSUE,
                "fixed": ChangeType.FIXED,
                "regression": ChangeType.REGRESSION,
                "active": ChangeType.STILL_FAILING,
            }
            return mapping.get(ct_str, ChangeType.NEW_ISSUE)

        # Record actions using ActionRecorder
        recorder = ActionRecorder(self.store)
        detected_actions = []

        for chg in exception_changes:
            action = DetectedChange(
                entity_key=chg["entity_key"],
                entity_type=None,
                change_type=get_change_type_enum(chg["change_type"]),
                description=f"Test change: {chg['change_type']}",
                detected_at=datetime.now(),
            )
            detected_actions.append(action)

        recorder.record_actions(
            actions=detected_actions,
            initial_run_id=self.run_id,
            sync_run_id=self.run_id,  # Simplified: use same run_id
        )

        return {
            "cycle": self.cycle_count,
            "current_annotations": current_annotations,
            "old_annotations": old_annotations,
            "persist_count": persist_count,
            "exception_changes": exception_changes,
            "loaded": loaded,
            "write_count": write_count,
        }

    def _build_mock_findings(self) -> list[dict]:
        """Build mock findings that mark all test data as FAIL."""
        findings = []

        for spec in get_data_specs():
            # The expected_key_pattern has format: entity_type|server|instance|...
            # But findings use entity_key WITHOUT entity_type prefix: server|instance|...
            # Strip the entity_type prefix from the key
            if "|" in spec.expected_key_pattern:
                # Format: "entity_type|server|instance|entity_name"
                # Strip first part: "server|instance|entity_name"
                entity_key_parts = spec.expected_key_pattern.split("|", 1)
                if len(entity_key_parts) > 1:
                    finding_key = entity_key_parts[1]  # Just the data portion
                else:
                    finding_key = spec.expected_key_pattern
            else:
                finding_key = spec.expected_key_pattern

            # Create a finding that marks this as discrepant
            finding = {
                "entity_type": spec.entity_type,
                "entity_key": finding_key,  # Without entity_type prefix
                "status": "FAIL",
                "is_discrepant": True,
            }
            findings.append(finding)

            # Insert into DB for ActionRecorder
            save_finding(
                connection=self.conn,
                audit_run_id=self.run_id,
                instance_id=self.instance_id,
                entity_key=finding_key,  # Without entity_type prefix
                finding_type=spec.entity_type,
                entity_name=finding_key.split("|")[-1],
                status="FAIL",
                risk_level="High",
                finding_description="Mock Finding",
            )
            self.conn.commit()

        return findings

    def count_action_log(self) -> int:
        """Count entries in the action_log table."""
        cursor = self.conn.execute("SELECT COUNT(*) FROM action_log")
        return cursor.fetchone()[0]

    def get_action_log_entries(self) -> list[dict]:
        """Get all action log entries."""
        cursor = self.conn.execute("SELECT * FROM action_log ORDER BY id")
        return [dict(row) for row in cursor.fetchall()]


@pytest.fixture
def ctx() -> Generator[TestContext, None, None]:
    """Provide a fresh test context for each test."""
    context = TestContext()
    yield context
    context.cleanup()


@pytest.fixture
def excel_with_all_sheets(ctx: TestContext) -> Path:
    """Create an Excel file with all data sheets populated."""
    ctx.create_excel_with_specs()
    return ctx.excel_path


@pytest.fixture
def annotation_sync(ctx: TestContext) -> AnnotationSyncService:
    """Create an AnnotationSyncService instance."""
    return AnnotationSyncService(ctx.db_path)
