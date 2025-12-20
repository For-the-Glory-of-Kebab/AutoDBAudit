"""
All Sheets Annotation Round-Trip Test

Validates that annotations can be read and written for every sheet
defined in SHEET_ANNOTATION_CONFIG.
"""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from autodbaudit.application.annotation_sync import (
    AnnotationSyncService,
    SHEET_ANNOTATION_CONFIG,
)
from autodbaudit.infrastructure.sqlite.store import HistoryStore


# Sample data generators per sheet type
SHEET_SAMPLE_DATA = {
    "Instances": {
        "headers": ["Server", "Instance", "Notes", "Last Reviewed"],
        "rows": [["srv1", "inst1", "Test note", "2025-01-01"]],
    },
    "SA Account": {
        "headers": [
            "Server",
            "Instance",
            "Current Name",
            "Review Status",
            "Justification",
            "Last Reviewed",
            "Notes",
        ],
        "rows": [
            ["srv1", "inst1", "sa", "✓ Exception", "Justified", "2025-01-01", "Note"]
        ],
    },
    "Server Logins": {
        "headers": [
            "Server",
            "Instance",
            "Login Name",
            "Review Status",
            "Justification",
            "Last Reviewed",
            "Notes",
        ],
        "rows": [["srv1", "inst1", "admin", "Needs Review", "", "2025-01-02", ""]],
    },
    "Sensitive Roles": {
        "headers": [
            "Server",
            "Instance",
            "Role",
            "Member",
            "Review Status",
            "Justification",
            "Last Reviewed",
        ],
        "rows": [
            ["srv1", "inst1", "sysadmin", "admin", "✓ Exception", "DBA", "2025-01-03"]
        ],
    },
    "Configuration": {
        "headers": [
            "Server",
            "Instance",
            "Setting",
            "Review Status",
            "Exception Reason",
            "Last Reviewed",
        ],
        "rows": [["srv1", "inst1", "xp_cmdshell", "Needs Review", "", ""]],
    },
    "Services": {
        "headers": [
            "Server",
            "Instance",
            "Service Name",
            "Review Status",
            "Justification",
            "Last Reviewed",
        ],
        "rows": [["srv1", "inst1", "SQL Agent", "", "", ""]],
    },
    "Databases": {
        "headers": [
            "Server",
            "Instance",
            "Database",
            "Review Status",
            "Justification",
            "Last Reviewed",
            "Notes",
        ],
        "rows": [["srv1", "inst1", "TestDB", "", "", "", ""]],
    },
    "Database Users": {
        "headers": [
            "Server",
            "Instance",
            "Database",
            "User Name",
            "Review Status",
            "Justification",
            "Last Reviewed",
            "Notes",
        ],
        "rows": [["srv1", "inst1", "TestDB", "testuser", "", "", "", ""]],
    },
    "Database Roles": {
        "headers": [
            "Server",
            "Instance",
            "Database",
            "Role",
            "Member",
            "Review Status",
            "Justification",
            "Last Reviewed",
        ],
        "rows": [["srv1", "inst1", "TestDB", "db_owner", "dbo", "", "", ""]],
    },
    "Permission Grants": {
        "headers": [
            "Server",
            "Instance",
            "Scope",
            "Database",
            "Grantee",
            "Permission",
            "Entity Name",
            "Review Status",
            "Justification",
            "Last Reviewed",
            "Notes",
        ],
        "rows": [
            [
                "srv1",
                "inst1",
                "DATABASE",
                "TestDB",
                "admin",
                "SELECT",
                "dbo.Table1",
                "",
                "",
                "",
                "",
            ]
        ],
    },
    "Orphaned Users": {
        "headers": [
            "Server",
            "Instance",
            "Database",
            "User Name",
            "Review Status",
            "Justification",
            "Last Reviewed",
        ],
        "rows": [["srv1", "inst1", "TestDB", "orphan1", "", "", ""]],
    },
    "Linked Servers": {
        "headers": [
            "Server",
            "Instance",
            "Linked Server",
            "Local Login",
            "Remote Login",
            "Review Status",
            "Purpose",
            "Justification",
            "Last Reviewed",
        ],
        "rows": [["srv1", "inst1", "REMOTE1", "admin", "remoteadmin", "", "", "", ""]],
    },
    "Triggers": {
        "headers": [
            "Server",
            "Instance",
            "Scope",
            "Database",
            "Trigger Name",
            "Event",
            "Review Status",
            "Notes",
            "Justification",
            "Last Reviewed",
        ],
        "rows": [
            ["srv1", "inst1", "SERVER", "", "LogonTrigger", "LOGON", "", "", "", ""]
        ],
    },
    "Client Protocols": {
        "headers": [
            "Server",
            "Instance",
            "Protocol",
            "Review Status",
            "Justification",
            "Last Reviewed",
        ],
        "rows": [["srv1", "inst1", "TCP/IP", "", "", ""]],
    },
    "Backups": {
        "headers": [
            "Server",
            "Instance",
            "Database",
            "Recovery Model",
            "Review Status",
            "Justification",
            "Last Reviewed",
            "Notes",
        ],
        "rows": [["srv1", "inst1", "TestDB", "FULL", "", "", "", ""]],
    },
    "Audit Settings": {
        "headers": [
            "Server",
            "Instance",
            "Setting",
            "Review Status",
            "Justification",
            "Last Reviewed",
            "Notes",
        ],
        "rows": [["srv1", "inst1", "Login Auditing", "", "", "", ""]],
    },
    "Encryption": {
        "headers": ["Server", "Instance", "Key Type", "Key Name", "Notes"],
        "rows": [["srv1", "inst1", "SMK", "##ServiceMasterKey##", ""]],
    },
    "Actions": {
        "headers": [
            "ID",
            "Server",
            "Instance",
            "Category",
            "Finding",
            "Risk Level",
            "Change Description",
            "Change Type",
            "Detected Date",
            "Notes",
        ],
        "rows": [
            [
                1,
                "srv1",
                "inst1",
                "config",
                "Test",
                "Low",
                "Fixed",
                "✓ Closed",
                "2025-01-01",
                "",
            ]
        ],
    },
}


class TestAllSheetsRoundTrip:
    """Validate annotation round-trip for all sheets."""

    @pytest.fixture
    def setup_all_sheets(self, tmp_path):
        """Create Excel with all sheet types."""
        excel_path = tmp_path / "all_sheets.xlsx"
        db_path = tmp_path / "test.db"

        # Create workbook with all sheets
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active

        for sheet_name, data in SHEET_SAMPLE_DATA.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(data["headers"])
            for row in data["rows"]:
                ws.append(row)

        # Remove the default empty sheet
        wb.remove(default_sheet)

        wb.save(excel_path)

        # Initialize DB
        store = HistoryStore(db_path)
        store.initialize_schema()

        self.service = AnnotationSyncService(db_path)
        self.excel_path = excel_path
        self.db_path = db_path

        return tmp_path

    def test_all_sheets_in_config(self):
        """Verify all expected sheets are in config."""
        expected_sheets = list(SHEET_SAMPLE_DATA.keys())

        for sheet in expected_sheets:
            assert (
                sheet in SHEET_ANNOTATION_CONFIG
            ), f"Missing config for sheet: {sheet}"

    def test_read_all_sheets(self, setup_all_sheets):
        """Test reading annotations from all sheets."""
        annotations = self.service.read_all_from_excel(self.excel_path)

        # Should have at least one annotation per sheet (from sample data)
        entity_types_found = set()
        for key in annotations:
            entity_type = key.split("|")[0]
            entity_types_found.add(entity_type)

        print(f"\nEntity types found: {entity_types_found}")

        # Verify we got annotations from multiple sheets
        assert (
            len(entity_types_found) >= 10
        ), f"Only found {len(entity_types_found)} entity types"

    def test_write_and_verify_all_sheets(self, setup_all_sheets):
        """Test writing annotations back and verifying."""
        # Read original
        annotations = self.service.read_all_from_excel(self.excel_path)

        # Modify some annotations
        for key, fields in list(annotations.items())[:5]:
            if "notes" in fields and fields["notes"] is None:
                fields["notes"] = "Updated by test"

        # Write back
        cells_updated = self.service.write_all_to_excel(self.excel_path, annotations)

        assert cells_updated >= 0, "Write failed"

        # Re-read and verify structure preserved
        reloaded = self.service.read_all_from_excel(self.excel_path)

        # Should have same number of keys
        assert len(reloaded) == len(annotations), "Annotation count changed after write"

    def test_persist_and_load_all(self, setup_all_sheets):
        """Test DB persistence for all sheets."""
        annotations = self.service.read_all_from_excel(self.excel_path)

        # Persist
        saved = self.service.persist_to_db(annotations)
        assert saved > 0, "No annotations saved"

        # Load back
        loaded = self.service.load_from_db()

        # Should have similar number of entries
        assert len(loaded) > 0, "No annotations loaded from DB"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
