"""
BaselineManager - Handles pre-existing discrepancies.

The test instances have existing discrepancies (like King login)
that we need to ignore when asserting test results.

Strategy:
1. Capture baseline before any test fixtures
2. Track what fixtures we apply
3. Assert only on DELTA (new findings vs baseline)
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class BaselineSnapshot:
    """Snapshot of pre-existing findings."""

    captured_at: str = ""
    sheets: dict[str, list[str]] = field(default_factory=dict)
    counts: dict[str, int] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "captured_at": self.captured_at,
            "sheets": self.sheets,
            "counts": self.counts,
        }

    @classmethod
    def from_dict(cls, data: dict) -> BaselineSnapshot:
        return cls(
            captured_at=data.get("captured_at", ""),
            sheets=data.get("sheets", {}),
            counts=data.get("counts", {}),
        )


class BaselineManager:
    """
    Manages baseline snapshot for pre-existing discrepancies.

    Protected entities (never modified):
    - King login (test connection user)
    - System databases
    """

    PROTECTED_ENTITIES = [
        "King",
        "master",
        "msdb",
        "tempdb",
        "model",
    ]

    def __init__(self, snapshot_path: Path) -> None:
        self.snapshot_path = snapshot_path
        self.baseline: BaselineSnapshot | None = None
        self.test_additions: list[str] = []

        self._load()

    def _load(self) -> None:
        """Load baseline from disk."""
        if self.snapshot_path.exists():
            try:
                data = json.loads(self.snapshot_path.read_text(encoding="utf-8"))
                self.baseline = BaselineSnapshot.from_dict(data)
            except Exception:
                self.baseline = None

    def _save(self) -> None:
        """Save baseline to disk."""
        if self.baseline:
            self.snapshot_path.parent.mkdir(parents=True, exist_ok=True)
            self.snapshot_path.write_text(
                json.dumps(self.baseline.to_dict(), indent=2),
                encoding="utf-8",
            )

    def capture_baseline(self, excel_path: Path) -> None:
        """
        Capture baseline from Excel report.

        Run this once before applying test fixtures.
        """
        from datetime import datetime

        self.baseline = BaselineSnapshot(
            captured_at=datetime.now().isoformat(),
        )

        wb = load_workbook(excel_path, data_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name in ["Cover", "Actions", "Action Log"]:
                continue

            ws = wb[sheet_name]
            entities = self._extract_entities(ws)
            self.baseline.sheets[sheet_name] = entities
            self.baseline.counts[sheet_name] = len(entities)

        wb.close()
        self._save()

    def _extract_entities(self, ws) -> list[str]:
        """Extract entity identifiers from worksheet."""
        entities = []

        # Usually entity is in column 2 (after Server)
        for row_num in range(2, min(ws.max_row + 1, 1000)):
            cell_value = ws.cell(row=row_num, column=2).value
            if cell_value:
                entities.append(str(cell_value))

        return entities

    def record_addition(self, fixture_name: str) -> None:
        """Record a test fixture we're adding."""
        self.test_additions.append(fixture_name)

    def get_delta(self, excel_path: Path) -> dict[str, list[str]]:
        """
        Get findings that are NEW compared to baseline.

        Returns:
            Dict mapping sheet name to list of new entity names
        """
        if not self.baseline:
            return {}

        delta = {}
        wb = load_workbook(excel_path, data_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name in ["Cover", "Actions", "Action Log"]:
                continue

            ws = wb[sheet_name]
            current_entities = self._extract_entities(ws)
            baseline_entities = self.baseline.sheets.get(sheet_name, [])

            # Find new ones
            new_entities = [
                e
                for e in current_entities
                if e not in baseline_entities and e not in self.PROTECTED_ENTITIES
            ]

            if new_entities:
                delta[sheet_name] = new_entities

        wb.close()
        return delta

    def assert_addition_found(
        self,
        excel_path: Path,
        sheet: str,
        entity_pattern: str,
    ) -> bool:
        """
        Assert that a test addition appears in delta.

        Returns True if found in new findings.
        """
        delta = self.get_delta(excel_path)

        if sheet not in delta:
            return False

        for entity in delta[sheet]:
            if entity_pattern.lower() in entity.lower():
                return True

        return False

    def clear_additions(self) -> None:
        """Clear test additions list."""
        self.test_additions.clear()
