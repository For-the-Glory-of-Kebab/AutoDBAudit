"""
Excel Driver for E2E Tests.
Wrapper around OpenPyXL to interact with the generated audit reports.
"""

from pathlib import Path
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


class ExcelDriver:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir

    def _get_latest_file(self) -> Path:
        files = list(self.output_dir.rglob("*.xlsx"))
        if not files:
            raise FileNotFoundError("No Excel report found in output dir")
        return max(files, key=lambda p: p.stat().st_mtime)

    def perform_action(
        self, sheet_name: str, entity_name: str, action_type: str, value: str
    ):
        """
        Perform a user action on the Excel sheet.
        """
        path = self._get_latest_file()
        wb = load_workbook(path)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name} not found")

        ws = wb[sheet_name]

        # Locate row
        # We assume entity name is in column D or E depending on sheet,
        # OR we search loosely. For robustness, let's search specifically.
        # Server Logins -> Col D (Login Name)
        # Sensitive Roles -> Col E (Member) or C/D.

        # Dynamic Header Search
        headers = [c.value for c in ws[1]]

        name_col_idx = -1
        target_col_idx = -1

        # Identify Name Column
        possible_name_headers = ["Login Name", "Member", "Database", "Setting"]
        for i, h in enumerate(headers, 1):
            if str(h) in possible_name_headers:
                name_col_idx = i
                break

        if name_col_idx == -1:
            # Fallback search
            pass

        # Identify Target Column (Justification or Review Status)
        target_header = (
            "Justification" if action_type == "ADD_JUSTIFICATION" else "Review Status"
        )
        if action_type == "CLEAR_ALL":
            target_header = None  # Special case

        just_col_idx = -1
        status_col_idx = -1

        for i, h in enumerate(headers, 1):
            if "Justification" in str(h):
                just_col_idx = i
            if "Review Status" in str(h):
                status_col_idx = i

        # Find Row
        target_row = None
        for row in ws.iter_rows(min_row=2):
            val = row[name_col_idx - 1].value
            if val and entity_name in str(val):
                target_row = row
                break

        if not target_row:
            raise ValueError(f"Entity {entity_name} not found in {sheet_name}")

        # Perform Edit
        if action_type == "ADD_JUSTIFICATION":
            ws.cell(row=target_row[0].row, column=just_col_idx).value = value
        elif action_type == "SET_STATUS":
            ws.cell(row=target_row[0].row, column=status_col_idx).value = value
        elif action_type == "CLEAR_JUSTIFICATION":
            ws.cell(row=target_row[0].row, column=just_col_idx).value = None
        elif action_type == "CLEAR_ALL":
            ws.cell(row=target_row[0].row, column=just_col_idx).value = None
            ws.cell(row=target_row[0].row, column=status_col_idx).value = None

        wb.save(path)
        wb.close()

    def verify_indicator(
        self, sheet_name: str, entity_name: str, expected_indicator: str
    ):
        """Verify the logic indicator in Column A."""
        path = self._get_latest_file()
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name]

        # Find row again
        headers = [c.value for c in ws[1]]
        name_col_idx = -1
        for i, h in enumerate(headers, 1):
            if str(h) in ["Login Name", "Member", "Setting"]:
                name_col_idx = i
                break

        for row in ws.iter_rows(min_row=2):
            val = row[name_col_idx - 1].value
            if val and entity_name in str(val):
                indicator = row[0].value  # Column A
                if expected_indicator not in str(indicator):
                    raise AssertionError(
                        f"Expected indicator {expected_indicator} for {entity_name}, got {indicator}"
                    )
                return

        raise ValueError(f"Entity {entity_name} not found for verification")
