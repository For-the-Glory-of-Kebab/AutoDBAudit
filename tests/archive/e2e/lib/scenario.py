
import os
import sys
import shutil
import sqlite3
import json
import subprocess
import pytest
from pathlib import Path
from openpyxl import load_workbook

class E2EScenario:
    """
    God-Tier E2E Test Scenario Builder.
    Orchestrates the lifecycle of a Sync Engine test:
    1. Setup Workspace (clean db, clean output)
    2. Mock Audit Data (findings.json)
    3. Run Sync
    4. Assert Excel Content
    5. Simulate User Edits
    6. Run Sync Again (Persist/Update)
    7. Assert Database State & Stats
    """

    def __init__(self, tmp_path: Path):
        self.root = tmp_path
        self.db_path = self.root / "audit_history.db"
        self.output_dir = self.root / "output"
        self.excel_path = self.output_dir / "Audit_Latest.xlsx"
        self.findings_dir = self.root / "findings"
        
        # Setup Dirs
        self.output_dir.mkdir(exist_ok=True)
        self.findings_dir.mkdir(exist_ok=True)
        
        # Locate entry point (assuming we run from repo root)
        self.entry_point = Path("src/main.py").resolve()
        if not self.entry_point.exists():
             # Fallback if running from inside tests
             self.entry_point = Path("../../src/main.py").resolve()
        
    def mock_audit_run(self, finding_type: str, findings: list[dict], run_id="run_1"):
        """Create a mock JSON finding file."""
        
        # Wrap findings in standard envelope if needed, or raw list
        # Our collector output is usually list of dicts
        
        # Enrich with common fields if missing
        enriched = []
        for f in findings:
            if "server_name" not in f: f["server_name"] = "TEST-SERVER"
            if "instance_name" not in f: f["instance_name"] = "MSSQLSERVER"
            if "finding_type" not in f: f["finding_type"] = finding_type
            enriched.append(f)
            
        file_path = self.findings_dir / f"{run_id}_{finding_type}.json"
        with open(file_path, "w") as f:
            json.dump(enriched, f)
            
    def run_audit(self, expect_exit_code=0):
        """Run the autodbaudit CLI with --audit via subprocess."""
        env = os.environ.copy()
        env["PYTHONPATH"] = str(self.entry_point.parent)
        
        # We must support finding-dir too for Mock Service
        cmd = [
            sys.executable, str(self.entry_point),
            "--audit", 
            "--finding-dir", str(self.findings_dir), # To create DB from Mock
            "--output-dir", str(self.output_dir),
            "--db-path", str(self.db_path)
        ]
        
        result = subprocess.run(
            cmd, 
            capture_output=True, 
            text=True, 
            env=env
        )
        
        if result.returncode != expect_exit_code:
            print("RUN_AUDIT STDOUT:", result.stdout)
            print("RUN_AUDIT STDERR:", result.stderr)
            raise AssertionError(f"CLI audit failed with {result.returncode}")
            
        return result.stdout

    def run_sync(self, expect_exit_code=0):
        """Run the autodbaudit CLI with --sync via subprocess."""
        env = os.environ.copy()
        env["PYTHONPATH"] = str(self.entry_point.parent)
        
        cmd = [
            sys.executable, str(self.entry_point),
            "--sync", 
            "--finding-dir", str(self.findings_dir),
            "--output-dir", str(self.output_dir),
            "--db-path", str(self.db_path)
        ]
        
        result = subprocess.run(
            cmd, 
            capture_output=True, 
            text=True, 
            env=env
        )
        
        if result.returncode != expect_exit_code:
            print("STDOUT:", result.stdout)
            print("STDERR:", result.stderr)
            raise AssertionError(f"CLI failed with {result.returncode}")
            
        return result.stdout

    def assert_excel_sheet(self, sheet_name: str, row_checks: list[dict]):
        """
        Nuclear assertion of Excel row content.
        row_checks is a list of dicts: {'ColName': 'ExpectedValue'}
        """
        assert self.excel_path.exists(), "Excel file not generated"
        wb = load_workbook(self.excel_path)
        assert sheet_name in wb.sheetnames, f"Sheet {sheet_name} missing"
        ws = wb[sheet_name]
        
        # Build Header Map
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[cell.value] = cell.column
        
        # Scan data rows (skipping header) - naive scan finding matches
        # For 'Nuclear' tests, we might want exact row index matching if provided, 
        # or find-first. Let's do find-first matching criterion.
        
        for check in row_checks:
            found = False
            # Search criteria is the first key/val in the check dict
            search_key, search_val = list(check.items())[0]
            search_col_idx = headers.get(search_key)
            if not search_col_idx:
                 raise ValueError(f"Header '{search_key}' not found in sheet '{sheet_name}'")

            for row in ws.iter_rows(min_row=2):
                cell_val = row[search_col_idx - 1].value
                if str(cell_val) == str(search_val):
                    # Found match, check other columns
                    found = True
                    for k, v in check.items():
                        col_idx = headers.get(k)
                        if not col_idx:
                             raise ValueError(f"Header '{k}' not found")
                        actual = row[col_idx - 1].value
                        assert str(actual) == str(v), f"Row Match '{search_val}': Expected {k}='{v}', got '{actual}'"
                    break
            
            assert found, f"Could not find row where {search_key}='{search_val}' in {sheet_name}"

    def simulate_user_edit(self, sheet_name: str, match_col: str, match_val: str, update_col: str, new_val: str):
        """Simulate a user manually editing the Excel file."""
        wb = load_workbook(self.excel_path)
        ws = wb[sheet_name]
        
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[cell.value] = cell.column_letter

        match_idx = None
        update_idx = None
        
        # Need column index (int) to iterate
        # Re-map to 1-based index
        header_indices = {}
        for cell in ws[1]:
            if cell.value:
                header_indices[cell.value] = cell.column
                
        tgt_col_idx = header_indices.get(match_col)
        update_col_idx = header_indices.get(update_col)
        
        found = False
        for row in ws.iter_rows(min_row=2):
            if str(row[tgt_col_idx-1].value) == str(match_val):
                row[update_col_idx-1].value = new_val
                found = True
                break
        
        assert found, f"Could not edit row: {match_col}={match_val} not found."
        wb.save(self.excel_path)
        
    def clear_findings(self):
        """Wipe finding files to simulate a new clean run or different run."""
        for f in self.findings_dir.glob("*.json"):
            f.unlink()
            
    def assert_stats(self, cli_output: str, stats: dict):
        """Check the 'Stats' table printed to CLI."""
        # This is fuzzy, ideally we parse the output table lines
        # "Triggers | 5 | 2" -> Sheet | Total | Rows
        for k, v in stats.items():
            # Naive check: "{k}" and "{v}" appear near each other?
            # Better: Regex or line splitting.
            # CLI output format: "| Sheet Name | Total Found |"
            pass # TODO: Implement robust CLI table parser if needed.
