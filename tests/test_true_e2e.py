"""
TRUE E2E Test - Simulating Exact App Flow

This test simulates:
1. Excel generation using ACTUAL writers (not mock data)
2. User edits in Excel (adding exceptions, notes)
3. Multiple --sync runs
4. Verification that annotations persist correctly

This is the definitive test for the sync engine.
"""

import unittest
import tempfile
import shutil
import sqlite3
from pathlib import Path
from datetime import datetime

import sys
sys.path.insert(0, str(Path(__file__).parents[2] / "src"))

from openpyxl import load_workbook

from autodbaudit.infrastructure.excel import EnhancedReportWriter
from autodbaudit.infrastructure.sqlite import HistoryStore
from autodbaudit.application.annotation_sync import (
    AnnotationSyncService,
    SHEET_ANNOTATION_CONFIG,
)


class TrueE2ETest(unittest.TestCase):
    """Full E2E simulation of the actual app flow."""

    def setUp(self):
        """Create temp environment with DB and Excel."""
        self.temp_dir = Path(tempfile.mkdtemp())
        self.db_path = self.temp_dir / "audit_history.db"
        self.excel_path = self.temp_dir / "Audit_Latest.xlsx"
        
        # Initialize SQLite store - let it create its own schema
        self.store = HistoryStore(self.db_path)
        self.store.initialize_schema()

    def tearDown(self):
        if self.store._connection:
            self.store._connection.close()
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def _generate_report_with_writer(self):
        """Generate an Excel report using the ACTUAL writer."""
        writer = EnhancedReportWriter()
        writer.set_audit_info(
            run_id=1,
            organization="Test Org",
            audit_name="E2E Test Audit",
            started_at=datetime.now(),
        )
        
        # Add test data for Linked Servers
        writer.add_linked_server(
            server_name="localhost",
            instance_name="TestInstance",
            linked_server_name="REMOTE_SERVER_1",
            product="SQL Server",
            provider="SQLNCLI",
            data_source="remote.server.com",
            rpc_out=True,
            local_login="sa",
            remote_login="admin",
            impersonate=False,
            risk_level="HIGH_PRIVILEGE",
        )
        writer.add_linked_server(
            server_name="localhost",
            instance_name="TestInstance",
            linked_server_name="REMOTE_SERVER_2",
            product="SQL Server",
            provider="SQLNCLI",
            data_source="other.server.com",
            rpc_out=False,
        )
        
        # Add test data for Triggers
        writer.add_trigger(
            server_name="localhost",
            instance_name="TestInstance",
            trigger_name="trg_audit_login",
            event_type="LOGON",
            is_enabled=True,
            level="SERVER",
            database_name=None,  # SERVER-level triggers have no database
        )
        writer.add_trigger(
            server_name="localhost",
            instance_name="TestInstance",
            trigger_name="trg_data_audit",
            event_type="DDL",
            is_enabled=True,
            level="DATABASE",
            database_name="AppDB",
        )
        
        # Add test data for Sensitive Roles
        writer.add_role_member(
            server_name="localhost",
            instance_name="TestInstance",
            role_name="sysadmin",
            member_name="DOMAIN\\AppSvc",
            member_type="Windows",
            is_disabled=False,
        )
        
        writer.create_cover_sheet()
        writer.save(self.excel_path)
        return writer

    def _simulate_user_edit(self, sheet_name: str, row: int, col_name: str, value: str):
        """Simulate a user editing a cell in Excel."""
        wb = load_workbook(self.excel_path)
        ws = wb[sheet_name]
        
        # Find column index by header name
        headers = [str(c.value or "").strip() for c in ws[1]]
        # Remove emoji prefixes
        headers = [h.replace("⏳ ", "").replace("✓ ", "").strip() for h in headers]
        
        if col_name not in headers:
            wb.close()
            self.fail(f"Column '{col_name}' not found in {sheet_name}. Headers: {headers}")
        
        col_idx = headers.index(col_name) + 1  # 1-indexed
        ws.cell(row=row, column=col_idx).value = value
        wb.save(self.excel_path)
        wb.close()

    def _read_cell_value(self, sheet_name: str, row: int, col_name: str) -> str:
        """Read a cell value from Excel."""
        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb[sheet_name]
        
        headers = [str(c.value or "").strip() for c in ws[1]]
        headers = [h.replace("⏳ ", "").replace("✓ ", "").strip() for h in headers]
        
        if col_name not in headers:
            wb.close()
            return None
        
        col_idx = headers.index(col_name) + 1
        value = ws.cell(row=row, column=col_idx).value
        wb.close()
        return str(value) if value else ""

    def test_linked_servers_purpose_persists(self):
        """CRITICAL: Purpose values should persist across syncs without duplication."""
        # Step 1: Generate report
        self._generate_report_with_writer()
        
        # Verify Linked Servers sheet exists with correct headers
        wb = load_workbook(self.excel_path, read_only=True)
        self.assertIn("Linked Servers", wb.sheetnames)
        ws = wb["Linked Servers"]
        headers = [str(c.value or "") for c in ws[1]]
        wb.close()
        
        # Headers should include Purpose
        header_names = [h.replace("⏳ ", "").replace("✓ ", "").strip() for h in headers]
        self.assertIn("Purpose", header_names, f"Purpose column missing. Headers: {header_names}")
        
        # Step 2: User adds Purpose values
        self._simulate_user_edit("Linked Servers", 2, "Purpose", "p1 - Production data sync")
        self._simulate_user_edit("Linked Servers", 3, "Purpose", "p2 - Backup replication")
        
        # Verify values are set
        p1 = self._read_cell_value("Linked Servers", 2, "Purpose")
        p2 = self._read_cell_value("Linked Servers", 3, "Purpose")
        self.assertEqual(p1, "p1 - Production data sync")
        self.assertEqual(p2, "p2 - Backup replication")
        
        # Step 3: Simulate sync (read annotations from Excel)
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)
        
        # Should have 2 linked server annotations
        ls_annotations = {k: v for k, v in annotations.items() if k.startswith("linked_server|")}
        self.assertEqual(len(ls_annotations), 2, f"Expected 2 linked server annotations, got: {ls_annotations}")
        
        # Step 4: Write annotations back (simulating sync output)
        sync.persist_to_db(annotations)
        sync.write_all_to_excel(self.excel_path, annotations)
        
        # Step 5: Read again - values should NOT be duplicated or shifted
        p1_after = self._read_cell_value("Linked Servers", 2, "Purpose")
        p2_after = self._read_cell_value("Linked Servers", 3, "Purpose")
        
        self.assertEqual(p1_after, "p1 - Production data sync", 
            f"Purpose value corrupted! Expected 'p1 - Production data sync', got '{p1_after}'")
        self.assertEqual(p2_after, "p2 - Backup replication",
            f"Purpose value corrupted! Expected 'p2 - Backup replication', got '{p2_after}'")

    def test_triggers_notes_column_exists(self):
        """Triggers sheet should have a Notes column."""
        self._generate_report_with_writer()
        
        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb["Triggers"]
        headers = [str(c.value or "").strip() for c in ws[1]]
        headers = [h.replace("⏳ ", "").replace("✓ ", "").strip() for h in headers]
        wb.close()
        
        self.assertIn("Notes", headers, f"Notes column missing from Triggers. Headers: {headers}")

    def test_triggers_notes_persists(self):
        """Triggers Notes should persist across syncs."""
        self._generate_report_with_writer()
        
        # Add notes to triggers
        self._simulate_user_edit("Triggers", 2, "Notes", "Audit login trigger for security")
        self._simulate_user_edit("Triggers", 3, "Notes", "Schema change tracking")
        
        # Sync
        sync = AnnotationSyncService(self.db_path)
        annotations = sync.read_all_from_excel(self.excel_path)
        
        # Verify trigger annotations have notes
        trigger_annotations = {k: v for k, v in annotations.items() if k.startswith("trigger|")}
        self.assertEqual(len(trigger_annotations), 2)
        
        notes_values = [v.get("notes", "") for v in trigger_annotations.values()]
        self.assertIn("Audit login trigger for security", notes_values)

    def test_multiple_sync_runs_preserve_annotations(self):
        """Annotations should persist through multiple sync cycles."""
        self._generate_report_with_writer()
        
        # Add various annotations
        self._simulate_user_edit("Linked Servers", 2, "Purpose", "Original purpose")
        self._simulate_user_edit("Linked Servers", 2, "Review Status", "✓ Exception")
        self._simulate_user_edit("Linked Servers", 2, "Justification", "Business requirement")
        
        sync = AnnotationSyncService(self.db_path)
        
        # First sync
        annotations_1 = sync.read_all_from_excel(self.excel_path)
        sync.persist_to_db(annotations_1)
        sync.write_all_to_excel(self.excel_path, annotations_1)
        
        # Second sync (simulating re-audit that regenerates the file)
        # In real app, the writer would regenerate the sheet content
        # For this test, we just verify annotations persist
        annotations_2 = sync.read_all_from_excel(self.excel_path)
        
        # Values should remain stable
        ls_keys = [k for k in annotations_2 if k.startswith("linked_server|")]
        self.assertTrue(len(ls_keys) > 0)
        
        first_key = ls_keys[0]
        self.assertEqual(annotations_2[first_key].get("purpose"), "Original purpose")
        self.assertEqual(annotations_2[first_key].get("justification"), "Business requirement")


if __name__ == "__main__":
    unittest.main(verbosity=2)
