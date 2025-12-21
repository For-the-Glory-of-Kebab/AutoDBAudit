"""Debug script to check actual Excel headers."""
import subprocess
from pathlib import Path
import os

PROJECT_ROOT = Path(__file__).parent.parent
VENV_PYTHON = PROJECT_ROOT / "venv" / "Scripts" / "python.exe"

debug_script = '''
import sys
sys.path.insert(0, ".")
sys.path.insert(0, "src")

from openpyxl import load_workbook

EXCEL_PATH = "output/audit_001/Audit_001_Latest.xlsx"

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Linked Servers"]

print("=== Linked Servers Sheet Headers ===")
headers = []
for i, cell in enumerate(ws[1]):
    print(f"  Column {i}: {repr(cell.value)}")
    headers.append(cell.value)

print(f"\\n=== First 3 Data Rows ===")
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), 2):
    print(f"Row {row_num}:")
    for i, val in enumerate(row[:10]):  # First 10 columns
        print(f"  {i} ({headers[i] if i < len(headers) else '?'}): {repr(val)}")
    print()

wb.close()
'''

result = subprocess.run(
    [str(VENV_PYTHON), "-c", debug_script],
    cwd=str(PROJECT_ROOT),
    capture_output=True,
    text=True,
    env={
        **os.environ,
        "PYTHONPATH": str(PROJECT_ROOT / "src"),
    }
)

print("STDOUT:")
print(result.stdout)
if result.stderr:
    print("\nSTDERR (last 500):")
    print(result.stderr[-500:])
