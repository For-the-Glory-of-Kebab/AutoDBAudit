"""
Persian Excel Generator Service.

Creates a Persian (RTL) copy of an English Excel report.
Translates:
- Sheet names
- Column headers
- Dropdown values
- Conditional formatting rules
- Cover page text

Does NOT translate:
- User-entered data (notes, justifications, etc.)
- Server names, database names, technical values
"""

from __future__ import annotations

import logging
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

from autodbaudit.infrastructure.i18n import get_translator, Translator
from autodbaudit.infrastructure.i18n.rtl_excel import (
    apply_rtl_to_worksheet,
    translate_headers_in_place,
    apply_rtl_to_header_row,
)
from autodbaudit.infrastructure.i18n.dropdowns import (
    RISK_LEVELS,
    CHANGE_TYPE_OPTIONS,
    ACTION_CATEGORIES,
)

logger = logging.getLogger(__name__)


class PersianExcelGenerator:
    """
    Generates Persian Excel report from English source.

    Creates a copy with:
    - Translated sheet names
    - Translated headers
    - RTL direction
    - Persian fonts (IRTitr, IRNazanin)
    - Updated CF rules with Persian values
    - Updated dropdown validations with Persian options
    """

    def __init__(self, translator: Translator | None = None) -> None:
        """Initialize with translator."""
        self.translator = translator or get_translator("fa")

    def generate(self, source_path: Path, output_path: Path | None = None) -> Path:
        """
        Generate Persian copy of Excel report.

        Args:
            source_path: Path to English Excel file
            output_path: Optional output path (default: source_fa.xlsx)

        Returns:
            Path to generated Persian file
        """
        if output_path is None:
            stem = source_path.stem
            output_path = source_path.parent / f"{stem}_fa.xlsx"

        logger.info("Generating Persian Excel: %s -> %s", source_path, output_path)

        # Copy source file
        shutil.copy2(source_path, output_path)

        # Load and transform
        wb = load_workbook(output_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._transform_sheet(ws, sheet_name)

        # Rename sheets (must be done after iterating)
        self._rename_sheets(wb)

        wb.save(output_path)
        wb.close()

        logger.info("✓ Generated Persian Excel: %s", output_path)
        return output_path

    def _transform_sheet(self, ws: Worksheet, original_name: str) -> None:
        """Transform a single sheet to Persian."""
        logger.debug("Transforming sheet: %s", original_name)

        # Apply RTL direction
        apply_rtl_to_worksheet(ws, self.translator)

        # Special handling for Cover sheet - translate all text, preserve formatting
        if original_name == "Cover":
            self._translate_cover_sheet(ws)
        else:
            # Regular sheets: translate headers (row 1)
            translate_headers_in_place(ws, 1, self.translator)

            # Apply RTL styling to headers
            apply_rtl_to_header_row(ws, 1, self.translator)

            # Translate dropdown values in cells
            self._translate_dropdown_values(ws)

        # Apply Persian font to all cells (preserves other styling)
        self._apply_persian_fonts(ws)

        # Update dropdown validations with Persian options
        self._update_dropdown_validations(ws, original_name)

        # Update conditional formatting with Persian values
        self._update_conditional_formatting(ws, original_name)

    def _translate_cover_sheet(self, ws: Worksheet) -> None:
        """Translate Cover sheet text while preserving all formatting."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original = cell.value
                    # Try text translation (preserves icons like 📊)
                    translated = self.translator.text(original)

                    if translated == original:
                        # Try partial text translations for compound strings
                        for eng, fa in self._get_cover_translations().items():
                            if eng in original:
                                translated = original.replace(eng, fa)
                                break

                    if translated != original:
                        cell.value = translated

    def _get_cover_translations(self) -> dict[str, str]:
        """Get cover text translations for partial matching."""
        from autodbaudit.infrastructure.i18n.cover import ALL_COVER_TEXT

        return ALL_COVER_TEXT

    def _rename_sheets(self, wb) -> None:
        """Rename all sheets to Persian."""
        renames = {}
        for sheet_name in wb.sheetnames:
            translated = self.translator.sheet(sheet_name)
            if translated != sheet_name:
                renames[sheet_name] = translated

        for old_name, new_name in renames.items():
            ws = wb[old_name]
            ws.title = new_name
            logger.debug("Renamed sheet: %s -> %s", old_name, new_name)

    def _apply_persian_fonts(self, ws: Worksheet) -> None:
        """Apply Persian fonts to all cells while preserving ALL other formatting."""
        content_font_name = self.translator.content_font
        heading_font_name = self.translator.heading_font

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            is_header = row_idx == 1
            font_name = heading_font_name if is_header else content_font_name

            for cell in row:
                if cell.value is not None:
                    current = cell.font
                    # Preserve ALL font properties, only change the name
                    cell.font = Font(
                        name=font_name,
                        size=current.size if current else 11,
                        bold=current.bold if current else (is_header),
                        italic=current.italic if current else False,
                        color=current.color if current else None,
                        underline=current.underline if current else None,
                        strike=current.strike if current else False,
                        vertAlign=current.vertAlign if current else None,
                    )

    def _translate_dropdown_values(self, ws: Worksheet) -> None:
        """Translate known dropdown values in cells."""
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value

                    translated = self.translator.status(value)
                    if translated != value:
                        cell.value = translated
                        continue

                    translated = self.translator.risk(value)
                    if translated != value:
                        cell.value = translated
                        continue

                    translated = self.translator.change_type(value)
                    if translated != value:
                        cell.value = translated
                        continue

    def _update_dropdown_validations(self, ws: Worksheet, sheet_name: str) -> None:
        """Update dropdown validations with Persian options."""
        # Remove existing data validations
        ws.data_validations.dataValidation.clear()

        # Add Persian dropdowns for Actions sheet
        if sheet_name == "Actions":
            max_row = ws.max_row + 500

            # Category dropdown (column D)
            cat_options = list(ACTION_CATEGORIES.values())
            dv_cat = DataValidation(
                type="list",
                formula1=f'"{",".join(cat_options)}"',
                allow_blank=True,
            )
            dv_cat.add(f"D2:D{max_row}")
            ws.add_data_validation(dv_cat)

            # Risk Level dropdown (column F)
            risk_options = list(RISK_LEVELS.values())
            dv_risk = DataValidation(
                type="list",
                formula1=f'"{",".join(risk_options)}"',
                allow_blank=True,
            )
            dv_risk.add(f"F2:F{max_row}")
            ws.add_data_validation(dv_risk)

            # Change Type dropdown (column H)
            change_options = list(CHANGE_TYPE_OPTIONS.values())
            dv_change = DataValidation(
                type="list",
                formula1=f'"{",".join(change_options)}"',
                allow_blank=True,
            )
            dv_change.add(f"H2:H{max_row}")
            ws.add_data_validation(dv_change)

            logger.debug("Added Persian dropdowns for Actions sheet")

    def _update_conditional_formatting(self, ws: Worksheet, sheet_name: str) -> None:
        """Update conditional formatting with Persian values."""
        # Clear existing CF rules
        ws.conditional_formatting._cf_rules.clear()

        if sheet_name == "Actions":
            self._add_actions_persian_cf(ws)

    def _add_actions_persian_cf(self, ws: Worksheet) -> None:
        """Add Persian conditional formatting for Actions sheet."""
        # Define colors
        pass_fill = PatternFill(
            start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"
        )
        pass_font = Font(color="1B5E20", bold=True, name=self.translator.content_font)
        warn_fill = PatternFill(
            start_color="FFE082", end_color="FFE082", fill_type="solid"
        )
        warn_font = Font(color="E65100", bold=True, name=self.translator.content_font)
        fail_fill = PatternFill(
            start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
        )
        fail_font = Font(color="B71C1C", bold=True, name=self.translator.content_font)

        max_row = ws.max_row + 500

        # Risk Level (Column F) - Persian values
        f_range = f"F2:F{max_row}"

        # کم (Low) = Green
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{RISK_LEVELS["Low"]}",F2))'],
                stopIfTrue=True,
                fill=pass_fill,
                font=pass_font,
            ),
        )
        # متوسط (Medium) = Orange
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{RISK_LEVELS["Medium"]}",F2))'],
                stopIfTrue=True,
                fill=warn_fill,
                font=warn_font,
            ),
        )
        # بالا (High) = Red
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{RISK_LEVELS["High"]}",F2))'],
                stopIfTrue=True,
                fill=fail_fill,
                font=fail_font,
            ),
        )

        # Change Type (Column H) - Persian values
        h_range = f"H2:H{max_row}"

        # Fixed/Closed/Exception = Green
        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=[
                    f'OR(ISNUMBER(SEARCH("{CHANGE_TYPE_OPTIONS["✓ Fixed"]}",H2)), '
                    f'ISNUMBER(SEARCH("{CHANGE_TYPE_OPTIONS["✓ Closed"]}",H2)), '
                    f'ISNUMBER(SEARCH("{CHANGE_TYPE_OPTIONS["✓ Exception"]}",H2)))'
                ],
                stopIfTrue=True,
                fill=pass_fill,
                font=pass_font,
            ),
        )
        # Regression = Red
        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=[
                    f'ISNUMBER(SEARCH("{CHANGE_TYPE_OPTIONS["❌ Regression"]}",H2))'
                ],
                stopIfTrue=True,
                fill=fail_fill,
                font=fail_font,
            ),
        )
        # Open/Pending = Orange
        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{CHANGE_TYPE_OPTIONS["⏳ Open"]}",H2))'],
                stopIfTrue=True,
                fill=warn_fill,
                font=warn_font,
            ),
        )

        logger.debug("Added Persian CF rules for Actions sheet")


def generate_persian_report(source_path: Path, output_path: Path | None = None) -> Path:
    """
    Convenience function to generate Persian report.

    Args:
        source_path: Path to English Excel
        output_path: Optional output path

    Returns:
        Path to generated Persian file
    """
    generator = PersianExcelGenerator()
    return generator.generate(source_path, output_path)
