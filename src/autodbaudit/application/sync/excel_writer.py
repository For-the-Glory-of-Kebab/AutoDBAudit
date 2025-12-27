"""
Excel Writer for Annotation Sync.

Writes annotations back to Excel worksheets.
Updates only editable columns, preserves all other data.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING

from autodbaudit.application.sync.key_builder import (
    normalize_key,
    clean_key_value,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def write_all_sheets(
    excel_path: Path | str,
    annotations: dict[str, dict],
    sheet_configs: dict[str, dict],
) -> int:
    """
    Write annotations to all configured sheets.

    Args:
        excel_path: Path to Excel file
        annotations: Dict of {entity_type|entity_key: {field: value}}
        sheet_configs: Dict of {sheet_name: config_dict}

    Returns:
        Number of cells updated
    """
    from openpyxl import load_workbook

    excel_path = Path(excel_path)
    if not excel_path.exists():
        logger.warning("Excel file not found: %s", excel_path)
        return 0

    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        logger.error("Failed to open Excel file for writing: %s", e)
        return 0

    total_updated = 0

    for sheet_name, config in sheet_configs.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        entity_type = config["entity_type"]

        # Filter annotations for this sheet
        sheet_annotations = {
            k.split("|", 1)[1]: v
            for k, v in annotations.items()
            if k.startswith(f"{entity_type}|")
        }

        if sheet_annotations:
            count = write_sheet_annotations(ws, config, sheet_annotations)
            total_updated += count

    try:
        wb.save(excel_path)
        logger.info("Wrote %d annotation cells to %s", total_updated, excel_path)
    except PermissionError:
        logger.error(
            "Cannot write to '%s' - file is open! Close Excel and retry.",
            excel_path.name,
        )
        return 0
    except Exception as e:
        logger.error("Failed to save Excel file: %s", e)
        return 0

    wb.close()
    return total_updated


def write_sheet_annotations(
    ws: "Worksheet",
    config: dict,
    annotations: dict[str, dict],
) -> int:
    """
    Write annotations to a single worksheet.

    Args:
        ws: Worksheet to update
        config: Sheet config with key_cols and editable_cols
        annotations: Dict of {entity_key: {field: value}}

    Returns:
        Number of cells updated
    """
    updated = 0

    # Get header row
    header_row = [cell.value for cell in ws[1]]
    if not header_row:
        return 0

    # Build header -> column index map (1-based for openpyxl)
    header_map = {}
    for idx, header in enumerate(header_row, start=1):
        if header:
            clean = str(header).strip()
            header_map[clean] = idx

    # Find key column indices
    key_indices = []
    for key_col in config["key_cols"]:
        idx = header_map.get(key_col)
        if idx is None:
            for h, i in header_map.items():
                if key_col.lower() == h.lower():
                    idx = i
                    break
        if idx is not None:
            key_indices.append((key_col, idx))

    if len(key_indices) != len(config["key_cols"]):
        return 0

    # Find editable column indices
    editable_indices = {}
    for col_name, field_name in config["editable_cols"].items():
        idx = header_map.get(col_name)
        if idx is None:
            for h, i in header_map.items():
                if col_name.lower() == h.lower():
                    idx = i
                    break
        if idx is not None:
            editable_indices[field_name] = idx

    # Track last values for merged cells
    last_key_values = {}

    # Process data rows
    for row_num in range(2, ws.max_row + 1):
        # Build entity key for this row
        key_parts = []
        for key_col, col_idx in key_indices:
            cell_val = ws.cell(row=row_num, column=col_idx).value
            if cell_val is None:
                val = last_key_values.get(key_col, "")
            else:
                val = str(cell_val)
                if key_col == "Permission":
                    val = clean_key_value(val)
                last_key_values[key_col] = val
            key_parts.append(val)

        entity_key = "|".join(normalize_key(p) for p in key_parts)
        if not entity_key:
            continue

        # Check if we have annotations for this row
        if entity_key not in annotations:
            continue

        row_annotations = annotations[entity_key]

        # Update editable cells
        for field_name, col_idx in editable_indices.items():
            if field_name in row_annotations:
                new_val = row_annotations[field_name]
                current = ws.cell(row=row_num, column=col_idx).value

                if str(new_val or "").strip() != str(current or "").strip():
                    ws.cell(row=row_num, column=col_idx).value = new_val
                    updated += 1

    return updated
