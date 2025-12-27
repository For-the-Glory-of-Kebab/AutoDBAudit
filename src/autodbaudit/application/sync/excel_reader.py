"""
Excel Reader for Annotation Sync.

Reads user annotations from Excel worksheets.
Handles merged cells, key column discovery, and value extraction.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING

from autodbaudit.application.sync.key_builder import (
    normalize_key,
    clean_key_value,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def read_all_sheets(
    excel_path: Path | str,
    sheet_configs: dict[str, dict],
) -> dict[str, dict]:
    """
    Read annotations from all configured sheets.

    Args:
        excel_path: Path to Excel file
        sheet_configs: Dict of {sheet_name: config_dict}

    Returns:
        Dict of {entity_type|entity_key: {field: value}}
    """
    from openpyxl import load_workbook

    excel_path = Path(excel_path)
    all_annotations: dict[str, dict] = {}

    if not excel_path.exists():
        logger.warning("Excel file not found: %s", excel_path)
        return all_annotations

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("Failed to open Excel file: %s", e)
        return all_annotations

    for sheet_name, config in sheet_configs.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        sheet_annotations = read_sheet_annotations(ws, config)

        for entity_key, fields in sheet_annotations.items():
            entity_type = str(config["entity_type"]).lower()
            full_key = f"{entity_type}|{entity_key}"
            all_annotations[full_key] = fields

    wb.close()
    logger.info(
        "Read %d annotations from %d sheets",
        len(all_annotations),
        len(wb.sheetnames),
    )
    return all_annotations


def read_sheet_annotations(
    ws: "Worksheet",
    config: dict,
) -> dict[str, dict]:
    """
    Read annotations from a single worksheet.

    Uses header row to find column positions dynamically.
    Handles merged cells by tracking last non-empty values.

    Args:
        ws: Worksheet to read
        config: Sheet config with key_cols and editable_cols

    Returns:
        Dict of {entity_key: {field: value}}
    """
    annotations: dict[str, dict] = {}

    # Get header row
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    if not header_row:
        return annotations

    # Build header -> index map
    header_map = {}
    for idx, header in enumerate(header_row):
        if header:
            clean = str(header).strip()
            # Remove common emojis/prefixes
            for prefix in ["⏳ ", "✅ ", "⚠️ ", "❌ "]:
                clean = clean.replace(prefix, "")
            header_map[clean] = idx

    # Find key column indices
    key_indices = _find_key_indices(header_map, config["key_cols"])
    if len(key_indices) != len(config["key_cols"]):
        logger.warning(
            "Sheet %s: Missing key columns. Found %d/%d",
            ws.title,
            len(key_indices),
            len(config["key_cols"]),
        )
        return annotations

    # Find editable column indices
    editable_indices = _find_editable_indices(header_map, config["editable_cols"])

    # Track last values for merged cells
    last_key_values = [""] * len(key_indices)

    # Read data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        # Build entity key
        key_parts = []
        for i, idx in enumerate(key_indices):
            val = row[idx] if idx < len(row) else None
            if val is None:
                val = last_key_values[i]
            else:
                val = str(val)
                if config["key_cols"][i] == "Permission":
                    val = clean_key_value(val)
                last_key_values[i] = val
            key_parts.append(val)

        entity_key = "|".join(normalize_key(p) for p in key_parts)
        if not entity_key or entity_key == "||":
            continue

        # Extract editable fields
        fields = {}
        for field_name, idx in editable_indices.items():
            if idx < len(row) and row[idx] is not None:
                val = row[idx]
                if isinstance(val, str):
                    val = val.strip()
                if val:
                    fields[field_name] = val

        if fields:
            annotations[entity_key] = fields

    return annotations


def _find_key_indices(
    header_map: dict[str, int],
    key_cols: list[str],
) -> list[int]:
    """Find indices for key columns with case-insensitive fallback."""
    indices = []
    for key_col in key_cols:
        if key_col in header_map:
            indices.append(header_map[key_col])
        else:
            # Case-insensitive search
            for h, idx in header_map.items():
                if key_col.lower() == h.lower():
                    indices.append(idx)
                    break
    return indices


def _find_editable_indices(
    header_map: dict[str, int],
    editable_cols: dict[str, str],
) -> dict[str, int]:
    """Find indices for editable columns."""
    indices = {}
    for col_name, field_name in editable_cols.items():
        if col_name in header_map:
            indices[field_name] = header_map[col_name]
        else:
            for h, idx in header_map.items():
                if col_name.lower() == h.lower():
                    indices[field_name] = idx
                    break
    return indices
