"""
Exception service for persisting user annotations from Excel to SQLite.

Core principle: Excel is the UI, SQLite is the source of truth.

Flow:
1. User adds Notes/Reason to items in Excel
2. --apply-exceptions reads Excel, matches entity_key
3. Persists to annotations table (idempotent upsert)
4. Next --audit queries annotations and carries forward

Architecture:
- Uses openpyxl to read Excel without modifying
- Entity key matching via server|instance|entity_name pattern
- Annotations table has field_name (notes/reason/status_override)
- History tracked in annotation_history for audit trail
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from autodbaudit.infrastructure.sqlite import HistoryStore

logger = logging.getLogger(__name__)


@dataclass
class ExcelAnnotation:
    """Annotation extracted from Excel."""

    sheet_name: str
    row: int
    server: str
    instance: str
    entity_name: str
    notes: str | None = None
    reason: str | None = None
    status_override: str | None = None

    @property
    def entity_key(self) -> str:
        """Build entity key for SQLite matching."""
        return f"{self.server}|{self.instance or '(Default)'}|{self.entity_name}"


class ExcelAnnotationReader:
    """
    Reads user annotations from Excel audit report.

    Uses central sheet_registry for configuration - NO duplicate definitions.
    """

    def __init__(self, excel_path: str | Path):
        """Initialize reader with Excel path."""
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

    def read_annotations(self) -> Iterator[ExcelAnnotation]:
        """
        Yield annotations from all trackable sheets.

        Uses central sheet_registry for configuration.
        Builds composite entity keys using ALL key_columns.
        """
        from openpyxl import load_workbook
        from autodbaudit.domain.sheet_registry import (
            SHEET_REGISTRY,
            get_all_trackable_sheets,
        )

        wb = load_workbook(self.excel_path, read_only=True, data_only=True)

        for spec in get_all_trackable_sheets():
            sheet_name = spec.name

            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # Build header -> index map from first row
            header_row = list(next(ws.iter_rows(min_row=1, max_row=1)))
            headers = {}
            for idx, cell in enumerate(header_row):
                if cell.value:
                    headers[str(cell.value).strip()] = idx

            # Find indices for key columns
            key_indices = []
            for key_col in spec.key_columns:
                idx = headers.get(key_col)
                if idx is None:
                    # Try case-insensitive match
                    for h, i in headers.items():
                        if h.lower() == key_col.lower():
                            idx = i
                            break
                if idx is not None:
                    key_indices.append((key_col, idx))

            if len(key_indices) != len(spec.key_columns):
                logger.debug(
                    "Sheet %s: Could not find all key columns %s",
                    sheet_name,
                    spec.key_columns,
                )
                continue

            # Find annotation column
            annotation_idx = headers.get(spec.annotation_column)
            if annotation_idx is None:
                # Try alternatives
                for alt in ["Notes", "Justification", "Purpose", "Exception Reason"]:
                    if alt in headers:
                        annotation_idx = headers[alt]
                        break

            # Find status column
            status_idx = headers.get("Review Status") or headers.get("Status Override")

        def get_cell(cells_list: list, idx: int | None) -> str | None:
            if idx is not None and idx < len(cells_list):
                val = cells_list[idx].value
                return str(val) if val is not None else None
            return None

        # Read data rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            cells = list(row)

            # Build composite entity key from ALL key columns
            key_parts = []
            for key_col, idx in key_indices:
                val = get_cell(cells, idx) or ""
                # Handle default instance
                if key_col == "Instance" and not val:
                    val = "(Default)"
                key_parts.append(val)

            entity_key = "|".join(key_parts)

            # Get annotation and status
            annotation = get_cell(cells, annotation_idx)
            status = get_cell(cells, status_idx)

            # Skip if no annotations
            if not annotation and not status:
                continue

                # Skip empty entity keys
                if all(not p or p == "(Default)" for p in key_parts):
                    continue

                # Extract server/instance for backward compatibility
                server = key_parts[0] if len(key_parts) > 0 else ""
                instance = key_parts[1] if len(key_parts) > 1 else ""
                entity_name = (
                    "|".join(key_parts[2:])
                    if len(key_parts) > 2
                    else key_parts[-1] if key_parts else ""
                )

                yield ExcelAnnotation(
                    sheet_name=sheet_name,
                    row=row_num,
                    server=server,
                    instance=instance,
                    entity_name=entity_name,
                    notes=annotation,
                    reason=None,
                    status_override=status,
                )

        wb.close()


class ExceptionService:
    """
    Service for persisting exceptions from Excel to SQLite.
    """

    def __init__(
        self,
        db_path: str | Path = "output/audit_history.db",
        excel_path: str | Path | None = None,
    ):
        """
        Initialize exception service.

        Args:
            db_path: SQLite database path
            excel_path: Excel report to read (auto-detect latest if None)
        """
        self.db_path = Path(db_path)
        self.store = HistoryStore(self.db_path)
        self.excel_path = Path(excel_path) if excel_path else None
        logger.info("ExceptionService initialized")

    def _find_latest_excel(self) -> Path | None:
        """Find most recent audit Excel in output directory."""
        output_dir = self.db_path.parent
        excels = sorted(output_dir.glob("sql_audit_*.xlsx"), reverse=True)
        return excels[0] if excels else None

    def apply_exceptions(self, audit_run_id: int | None = None) -> dict:
        """
        Read annotations from Excel and persist to SQLite.

        Args:
            audit_run_id: Link annotations to this run (latest if None)

        Returns:
            Dict with counts: applied, skipped, errors
        """
        # Find Excel if not specified
        excel_path = self.excel_path or self._find_latest_excel()
        if not excel_path or not excel_path.exists():
            return {"error": "No Excel file found"}

        logger.info("Reading annotations from: %s", excel_path)

        if audit_run_id is None:
            audit_run_id = self.store.get_latest_run_id()

        if audit_run_id is None:
            return {"error": "No audit run found"}

        # Read and persist annotations
        try:
            reader = ExcelAnnotationReader(excel_path)
            counts = {"applied": 0, "skipped": 0, "errors": 0}

            for annotation in reader.read_annotations():
                try:
                    entity_type = annotation.sheet_name.lower().replace(" ", "_")

                    # Persist notes
                    if annotation.notes:
                        self.store.upsert_annotation(
                            entity_type=entity_type,
                            entity_key=annotation.entity_key,
                            field_name="notes",
                            field_value=annotation.notes,
                            modified_by="excel_import",
                            audit_run_id=audit_run_id,
                        )

                    # Persist reason
                    if annotation.reason:
                        self.store.upsert_annotation(
                            entity_type=entity_type,
                            entity_key=annotation.entity_key,
                            field_name="reason",
                            field_value=annotation.reason,
                            modified_by="excel_import",
                            audit_run_id=audit_run_id,
                        )

                    # Persist status override
                    if annotation.status_override:
                        self.store.upsert_annotation(
                            entity_type=entity_type,
                            entity_key=annotation.entity_key,
                            field_name="status_override",
                            field_value=annotation.status_override,
                            status_override=annotation.status_override,
                            modified_by="excel_import",
                            audit_run_id=audit_run_id,
                        )

                    counts["applied"] += 1
                    logger.debug("Applied: %s", annotation.entity_key)

                except Exception as e:
                    counts["errors"] += 1
                    logger.warning(
                        "Failed to apply annotation for %s: %s",
                        annotation.entity_key,
                        e,
                    )

            logger.info(
                "Apply exceptions complete: %d applied, %d errors",
                counts["applied"],
                counts["errors"],
            )

            return counts
        except Exception as e:
            return {"error": f"Failed to read Excel: {e}"}

    def get_exceptions_for_entity(self, entity_key: str) -> dict:
        """
        Get all annotations for an entity.

        Returns dict with notes, reason, status_override fields.
        """
        return self.store.get_annotations_for_entity(entity_key)

    def get_all_exceptions(self) -> list[dict]:
        """Get all exceptions with status_override set."""
        return self.store.get_all_annotations(only_overrides=True)


def main():
    """CLI entry point for exception management."""
    import argparse

    parser = argparse.ArgumentParser(description="Manage exceptions")
    parser.add_argument("--db", default="output/audit_history.db")
    parser.add_argument("--excel", help="Excel file to read (latest if not specified)")
    parser.add_argument("--list", action="store_true", help="List all exceptions")

    args = parser.parse_args()

    service = ExceptionService(db_path=args.db, excel_path=args.excel)

    if args.list:
        exceptions = service.get_all_exceptions()
        print(f"\nExceptions: {len(exceptions)}")
        for exc in exceptions:
            print(f"  {exc['entity_key']}: {exc['status_override']}")
    else:
        result = service.apply_exceptions()
        if "error" in result:
            print(f"❌ {result['error']}")
            return 1
        print(f"\n✅ Applied {result['applied']} annotations")
        if result["errors"]:
            print(f"   ⚠️ {result['errors']} errors")

    return 0


if __name__ == "__main__":
    import sys

    sys.exit(main())
