"""
Sync service for tracking remediation progress.

Core principle: action_log records fixes with REAL timestamps.

Flow:
1. Re-audit current state
2. Diff against INITIAL baseline (not previous sync)
3. For NEW fixes: record with today's timestamp
4. For EXISTING fixes: preserve original timestamp
5. Update action_log with current sync run ID
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from autodbaudit.application.audit_service import AuditService

logger = logging.getLogger(__name__)

# Colors for CLI output
RED = "\033[91m"
RESET = "\033[0m"


class SyncService:
    """
    Service for syncing remediation progress.

    Separate from FinalizeService because:
    - Sync is repeatable (run after each fix)
    - Sync records real timestamps
    - Finalize is one-time (persist everything to DB)
    """

    def __init__(
        self,
        db_path: str | Path = "output/audit_history.db",
    ):
        """
        Initialize sync service.

        Args:
            db_path: Path to SQLite audit database
        """
        self.db_path = Path(db_path)
        logger.info("SyncService initialized")

    def get_initial_run_id(self) -> int | None:
        """Get the first (baseline) audit run ID."""
        import sqlite3

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM audit_runs WHERE run_type = 'audit' ORDER BY started_at ASC LIMIT 1"
        )
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else None

    def get_latest_run_id(self) -> int | None:
        """Get the most recent audit run ID."""
        import sqlite3

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM audit_runs ORDER BY started_at DESC LIMIT 1")
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else None

    def sync(
        self,
        audit_service: "AuditService" = None,
        targets_file: str = "sql_targets.json",
        audit_manager: object = None,
        audit_id: int = None,
    ) -> dict:
        """
        Sync remediation progress.

        1. Re-audit current state (creates new audit_run with type='sync')
        2. Diff against initial baseline
        3. Update action_log with real timestamps
        4. Inject actions into the Excel report and save it

        Args:
            audit_service: AuditService instance (creates new if None)
            targets_file: Targets config for re-audit
            audit_manager: Optional AuditManager for path resolution
            audit_id: Optional Audit ID context
        """
        import sqlite3
        from autodbaudit.infrastructure.excel import EnhancedReportWriter

        # Get initial baseline
        initial_run_id = self.get_initial_run_id()
        if initial_run_id is None:
            logger.error("No baseline audit found. Run --audit first.")
            return {"error": "No baseline audit found"}

        # ---------------------------------------------------------------------
        # Pre-flight Check: Finalized State
        # ---------------------------------------------------------------------
        conn = sqlite3.connect(self.db_path)
        row = conn.execute("SELECT status FROM audit_runs WHERE id = ?", (initial_run_id,)).fetchone()
        conn.close()
        
        if row and row[0] == 'finalized':
            msg = f"Audit Run #{initial_run_id} is FINALIZED. Sync is blocked to preserve state."
            logger.error(msg)
            print(f"\n{RED}⛔ ERROR: {msg}{RESET}")
            print(f"To modify this audit, you must start a NEW baseline run.{RESET}\n")
            return {"error": "Audit is finalized"}

        # ---------------------------------------------------------------------
        # Pre-flight Check: File Locking
        # ---------------------------------------------------------------------
        # Ensure we can write to the output file before doing any work.
        path_latest = None
        if audit_manager and audit_id:
            path_latest = audit_manager.get_latest_excel_path(audit_id)
        else:
            path_latest = self.db_path.parent / "Audit_Latest.xlsx"

        if path_latest and path_latest.exists():
            try:
                with open(path_latest, "r+b"):
                    pass
            except IOError:
                logger.error("Output file is locked: %s", path_latest)
                return {
                    "error": f"FILE EXTANT & LOCKED: Close '{path_latest.name}' and try again."
                }

        logger.info("Sync: initial baseline run ID = %d", initial_run_id)

        # Initialize Writer explicitly with Metadata from Baseline
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        baseline_meta = conn.execute(
            "SELECT started_at, organization FROM audit_runs WHERE id = ?",
            (initial_run_id,),
        ).fetchone()

        # Capture Pre-Scan State (Before re-audit overwrites instances)
        pre_scan_instances = {}
        try:
            rows = conn.execute(
                "SELECT id, version, product_level, edition FROM instances"
            ).fetchall()
            for r in rows:
                pre_scan_instances[r["id"]] = {
                    "v": r["version"],
                    "l": r["product_level"],
                    "e": r["edition"],
                }
        except Exception as e:
            logger.warning("Could not capture pre-scan instances: %s", e)

        writer = EnhancedReportWriter()
        if baseline_meta:
            writer.set_audit_info(
                run_id=initial_run_id,  # Use baseline ID for the report ID
                organization=baseline_meta["organization"] or "Unspecified",
                audit_name="Remediation Sync Report",
                started_at=(
                    datetime.fromisoformat(baseline_meta["started_at"])
                    if baseline_meta["started_at"]
                    else datetime.now()
                ),
            )
        conn.close()

        # Run new audit (Capture Data Only)
        logger.info("Sync: Running re-audit...")
        if audit_service is None:
            from autodbaudit.application.audit_service import AuditService

            audit_service = AuditService(
                config_dir=Path("config"), output_dir=Path("output")
            )

        try:
            # Pass writer and skip saving
            audit_service.run_audit(
                targets_file=targets_file, writer=writer, skip_save=True
            )
            # NOTE: run_audit will scan the updated (live) instances and update the DB 'instances' table
            # with the LATEST version info.
        except Exception as e:
            logger.error("Re-audit failed: %s", e)
            return {"error": f"Re-audit failed: {e}"}

        # Get new run ID and mark as sync type
        current_run_id = self.get_latest_run_id()

        drift_count = 0  # Initialize for stats tracking

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        if current_run_id == initial_run_id:
            # Re-audit might not have created a run if it failed early?
            # Or if it returned early.
            logger.error("Re-audit did not create new run path")
            # Usually run_audit always creates a run.
            pass

        # Mark as sync run
        if current_run_id and current_run_id != initial_run_id:
            conn.execute(
                "UPDATE audit_runs SET run_type = 'sync' WHERE id = ?",
                (current_run_id,),
            )

            # -------------------------------------------------------------------------
            # Detect Version / info Drift
            # -------------------------------------------------------------------------
            # Compare 'instances' (now updated) with pre_scan_instances

            # Fetch current state (post-scan)
            rows = conn.execute(
                "SELECT id, version, product_level, edition FROM instances"
            ).fetchall()

            # Get existing actions context
            existing_actions = {}
            for act_row in conn.execute(
                "SELECT entity_key, action_type, action_date FROM action_log WHERE initial_run_id = ?",
                (initial_run_id,),
            ).fetchall():
                existing_actions[act_row[0]] = {
                    "action_type": act_row[1],
                    "action_date": act_row[2],
                }

            drift_count = 0
            now_iso = datetime.now(timezone.utc).isoformat()

            for r in rows:
                iid = r["id"]
                if iid in pre_scan_instances:
                    old = pre_scan_instances[iid]
                    changes = []
                    if r["version"] != old["v"]:
                        changes.append(f"Version: {old['v']} -> {r['version']}")
                    if r["product_level"] != old["l"]:
                        changes.append(f"Level: {old['l']} -> {r['product_level']}")
                    if r["edition"] != old["e"]:
                        changes.append(f"Edition: {old['e']} -> {r['edition']}")

                    if changes:
                        # Fetch names for key
                        name_row = conn.execute(
                            "SELECT s.hostname, i.instance_name FROM instances i JOIN servers s ON i.server_id = s.id WHERE i.id = ?",
                            (iid,),
                        ).fetchone()

                        if name_row:
                            hostname = name_row[0]
                            inst_name = name_row[1]
                            entity_key = (
                                f"{hostname}\\{inst_name}" if inst_name else hostname
                            )
                            key_suffix = "|System|Version"
                            full_key = f"{entity_key}{key_suffix}"

                            desc = f"System Update Detected: {', '.join(changes)}"
                            logger.info(f"Sync: {desc}")

                            self._upsert_action(
                                conn,
                                initial_run_id,
                                full_key,
                                "System Information",
                                "fixed",
                                now_iso,
                                existing_actions,
                                desc,
                                current_run_id,
                            )
                            drift_count += 1

            if drift_count > 0:
                logger.info(f"Sync: Detected {drift_count} system version changes.")

            conn.commit()
            logger.info("Sync: current run ID = %d", current_run_id)
        else:
            logger.warning("Sync run ID is same as initial? This might be a bug.")

        # Re-connect later (Removed - conn is open)

        # Compare findings and update action log
        # NOTE: This updates the DB using the *diff*, but we also need
        # to capture the list of actions to put into the Excel file.
        # We need to refactor _update_action_log slightly or just pull them from DB.

        # 1. Update DB Action Log
        result = self._update_action_log(conn, initial_run_id, current_run_id)
        
        # CRITICAL: Commit action_log changes before get_actions reads them
        conn.commit()
        logger.debug("Committed %d action_log entries to database", 
                     result["fixed"] + result["still_failing"] + result["regression"] + result["new"])

        # Merge drift count into results
        result["fixed"] += drift_count

        # 2. Preserve ALL User Annotations (Not just Actions sheet)
        # Use AnnotationSyncService to read from ALL sheets
        from autodbaudit.application.annotation_sync import AnnotationSyncService
        annotation_sync = AnnotationSyncService(db_path=self.db_path)
        
        # Read all annotations from existing Excel (all sheets)
        all_annotations = annotation_sync.read_all_from_excel(path_latest)
        logger.info("Preserved %d annotations from existing Excel", len(all_annotations))
        
        # Also read action-specific annotations for backward compatibility
        preserved_annotations = self._read_existing_action_annotations(path_latest)
        
        # 3. Add Actions to Excel Writer
        actions = self.get_actions(initial_run_id)
        valid_actions_count = 0
        
        for action in actions:
            # Skip still_failing - these are not changes, shouldn't appear in changelog
            if action["action_type"] == "still_failing":
                continue
            
            # We list all actual CHANGES in the log (Fixed, Regression, New)
            
            # Parse entity key (format: server|instance|finding_type|entity_name)
            entity = action["entity_key"]
            parts = entity.split("|")
            server = parts[0] if parts else entity
            # Instance is second part, empty means default
            instance = parts[1] if len(parts) > 1 and parts[1] else "(Default)"
            
            # Get finding description for key matching
            finding_desc = action["action_description"] or f"{action['action_type'].title()}: {action['finding_type']}"
            
            # Build lookup key to find preserved annotations
            inst_normalized = "" if instance == "(Default)" else instance
            lookup_key = f"{server}|{inst_normalized}|{finding_desc}"

            # Map action type to display values
            # Actions sheet is a CHANGELOG - show what changed, not recommendations
            action_type = action["action_type"]
            
            if action_type == "fixed":
                risk = "Low"  # Fixed items are good news
                change_desc = f"✅ Fixed: {action['finding_type']}"
                status = "Closed"
            elif action_type == "regression":
                risk = "High"  # Regressions are bad
                change_desc = f"⚠️ Regressed: {action['finding_type']}"
                status = "Open"
            elif action_type == "new":
                risk = "Medium"  # New issues need attention
                change_desc = f"🆕 New Issue: {action['finding_type']}"
                status = "Open"
            else:
                risk = "Medium"
                change_desc = f"{action_type}: {action['finding_type']}"
                status = "Open"

            # Initialize manual fields
            notes = ""

            # Parse timestamp - prefer user's date if they edited it in Excel
            f_date = None
            if lookup_key in preserved_annotations:
                ann = preserved_annotations[lookup_key]
                notes = ann.get("resolution_notes", "")
                
                user_date = ann.get("found_date")
                if user_date:
                    if isinstance(user_date, datetime):
                        f_date = user_date
                    elif isinstance(user_date, str):
                        try:
                            clean_date = user_date.replace("Z", "+00:00")
                            f_date = datetime.fromisoformat(clean_date)
                        except (ValueError, TypeError):
                            pass
            
            # Fall back to DB date (Action Log Date - when change was first detected)
            if not f_date and action["action_date"]:
                try:
                    f_date = datetime.fromisoformat(action["action_date"])
                except (ValueError, TypeError):
                    f_date = datetime.now()

            writer.add_action(
                server_name=server,
                instance_name=instance,
                category=action["finding_type"] or "General",
                finding=finding_desc,
                risk_level=risk,
                recommendation=change_desc,  # Using recommendation column for change description
                status=status,
                found_date=f_date,
                resolution_notes=notes,
            )
            valid_actions_count += 1
            
        if valid_actions_count == 0:
            logger.info("No changes detected for Actions changelog.")
        else:
            logger.info("Added %d change entries to Actions changelog", valid_actions_count)

        # -------------------------------------------------------------------------
        # 3. Calculate Dual Stats (Baseline vs Current AND Previous vs Current)
        # -------------------------------------------------------------------------

        # We already have Baseline stats in 'result' (from _update_action_log)
        # Now we want "Recent Activity" (Recent Sync -> Current)

        # Find previous run ID (ignoring current)
        conn = sqlite3.connect(self.db_path)  # Fresh connection for stats
        conn.row_factory = sqlite3.Row

        prev_run_query = conn.execute(
            "SELECT id FROM audit_runs WHERE id < ? ORDER BY id DESC LIMIT 1",
            (current_run_id,),
        ).fetchone()

        recent_stats = {"fixed": 0, "regression": 0, "new": 0}

        if prev_run_query:
            prev_run_id = prev_run_query[0]
            logger.info(
                "Sync stats: prev_run_id=%d, current_run_id=%d, initial_run_id=%d",
                prev_run_id, current_run_id, initial_run_id
            )
            
            # Use schema's generic comparator
            from autodbaudit.infrastructure.sqlite.schema import compare_findings

            diff_recent = compare_findings(conn, prev_run_id, current_run_id)
            
            logger.debug(
                "Recent diff: fixed=%d, regression=%d, new=%d",
                len(diff_recent["fixed"]),
                len(diff_recent["regression"]),
                len(diff_recent["new"])
            )

            # Count them up
            recent_stats["fixed"] = len(diff_recent["fixed"])
            recent_stats["regression"] = len(diff_recent["regression"])
            recent_stats["new"] = len(diff_recent["new"])

            # Add drift to recent stats too (since it happened in this run)
            recent_stats["fixed"] += drift_count
            
            # If prev_run == initial_run, recent stats == baseline stats (expected)
            if prev_run_id == initial_run_id:
                logger.info("First sync detected: 'Since Last Run' == 'Since Baseline'")
                # Force equality to avoid confusing discrepancies from minor diff logic differences
                recent_stats["fixed"] = result["fixed"]
                recent_stats["regression"] = result["regression"]
                recent_stats["new"] = result["new"]
        else:
            logger.warning("No previous run found for recent stats calculation")

        conn.close()

        # -------------------------------------------------------------------------
        # 4. Save Reports
        # -------------------------------------------------------------------------

        # A. Archived Copy (History)
        if audit_manager and audit_id:
            # Use Hierarchical Structure
            audit_folder = audit_manager._get_audit_folder(audit_id)
            run_folder = audit_folder / "runs" / f"run_{current_run_id:03d}_sync"
            run_folder.mkdir(parents=True, exist_ok=True)

            path_latest = audit_manager.get_latest_excel_path(audit_id)

        else:
            # Fallback to root output (Legacy/No-Audit-Mode)
            run_folder = audit_service.output_dir / f"Run_{current_run_id}_Sync"
            run_folder.mkdir(parents=True, exist_ok=True)
            path_latest = audit_service.output_dir / "Audit_Latest.xlsx"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_archive = f"sql_audit_sync_{timestamp}.xlsx"
        path_archive = run_folder / filename_archive

        writer.save(path_archive)
        logger.info("Saved archive report: %s", path_archive)

        # B. Working Copy ("Latest")
        try:
            writer.save(path_latest)
            logger.info("Updated working copy: %s", path_latest)
        except PermissionError:
            logger.warning("Could not update %s (File open?)", path_latest)
            print(
                f"⚠️  Warning: Could not update {path_latest.name} (File is open). Close it to see updates."
            )

        # C. Restore user annotations to fresh reports
        if all_annotations:
            try:
                cells_updated = annotation_sync.write_all_to_excel(path_latest, all_annotations)
                if cells_updated > 0:
                    logger.info("Restored %d annotation cells to %s", cells_updated, path_latest.name)
                    print(f"📝 Restored {cells_updated} user annotations from previous report")
            except Exception as e:
                logger.warning("Failed to restore annotations: %s", e)
        
        # D. Detect and log exception changes (justification added/changed)
        # Load previous annotations from DB to compare
        db_annotations = annotation_sync.load_from_db()
        exception_changes = annotation_sync.detect_exception_changes(db_annotations, all_annotations)
        
        if exception_changes:
            print(f"📋 {len(exception_changes)} exception(s) documented")
            # Open fresh connection for exception logging (original conn was closed earlier)
            with sqlite3.connect(self.db_path) as exc_conn:
                for exc in exception_changes:
                    # Log each exception to SQLite action_log
                    self._log_exception_action(
                        conn=exc_conn,
                        initial_run_id=initial_run_id,
                        current_run_id=current_run_id,
                        exc=exc,
                    )
                    
                    # ALSO add to Excel Actions sheet (Phase 20B fix)
                    # Parse entity_key for server/instance (format: "Server|Instance|...")
                    parts = exc["entity_key"].split("|")
                    server = parts[0] if len(parts) > 0 else "Unknown"
                    instance = parts[1] if len(parts) > 1 else "(Default)"
                    
                    # Build concise finding description
                    entity_type = exc.get("entity_type", "Item")
                    entity_key_short = exc["entity_key"][:60]
                    justification = exc.get("justification", "")[:100]
                    
                    writer.add_action(
                        server_name=server,
                        instance_name=instance,
                        category=f"Exception: {entity_type.replace('_', ' ').title()}",
                        finding=f"Justified: {entity_key_short}",
                        risk_level="Low",  # Exception = risk documented/accepted
                        recommendation=f"User added justification: {justification}",
                        status="Closed",  # Documented = closed
                        found_date=datetime.now(),
                        resolution_notes=justification,
                    )
                    
                exc_conn.commit()

        result["report_path"] = str(path_archive)
        result["initial_run_id"] = initial_run_id
        result["current_run_id"] = current_run_id
        result["exceptions_documented"] = len(exception_changes) if exception_changes else 0

        # -------------------------------------------------------------------------
        # 5. Console Output
        # -------------------------------------------------------------------------

        print("\n" + "=" * 60)
        print(f"✅ SYNC COMPLETE (Run #{current_run_id})")
        print("=" * 60)
        print(f"{'Metric':<20} | {'Since Last Run':<15} | {'Since Baseline':<15}")
        print("-" * 56)

        # Helper to format numbers
        def fmt(n, is_good=True):
            # Simple indicator? no, just numbers for now to avoid clutter
            return str(n)

        print(f"{'Fixed':<20} | {recent_stats['fixed']:<15} | {result['fixed']:<15}")
        print(
            f"{'Regression':<20} | {recent_stats['regression']:<15} | {result['regression']:<15}"
        )
        print(f"{'New Findings':<20} | {recent_stats['new']:<15} | {result['new']:<15}")
        
        # Add documented exceptions count (counts as positive progress)
        if result["exceptions_documented"] > 0:
            print(f"{'Exceptions Documented':<20} | {result['exceptions_documented']:<15} | {'↑ Audit Trail':<15}")
        
        print("-" * 56)
        print(f"Total Outstanding: {result['still_failing']} (Baseline Deficits)")
        print("-" * 56)
        print(f"\n📂 Archive: {path_archive}")
        print(f"📂 LATEST:  {path_latest}")
        print("=" * 60 + "\n")

        return result

    def _update_action_log(
        self,
        conn,
        initial_run_id: int,
        current_run_id: int,
    ) -> dict:
        """
        Update action_log based on diff.

        Key behavior:
        - NEW fixes get today's timestamp
        - EXISTING fixes keep original timestamp
        - All get updated last_sync_run_id
        """
        from autodbaudit.infrastructure.sqlite.schema import get_findings_for_run

        now = datetime.now(timezone.utc).isoformat()

        # Get findings from both runs
        initial_findings = {
            f["entity_key"]: f for f in get_findings_for_run(conn, initial_run_id)
        }
        current_findings = {
            f["entity_key"]: f for f in get_findings_for_run(conn, current_run_id)
        }

        # Build set of successfully scanned instances (Server|Instance)
        scanned_instances = set()
        for key in current_findings:
            parts = key.split("|")
            if len(parts) >= 2:
                scanned_instances.add(f"{parts[0]}|{parts[1]}")

        unreachable_warnings = set()

        # Get existing action_log entries
        existing_actions = {}
        for row in conn.execute(
            "SELECT entity_key, action_type, action_date FROM action_log WHERE initial_run_id = ?",
            (initial_run_id,),
        ).fetchall():
            existing_actions[row[0]] = {"action_type": row[1], "action_date": row[2]}

        counts = {"fixed": 0, "still_failing": 0, "regression": 0, "new": 0}

        # Process each initial finding
        for key, initial in initial_findings.items():
            if key in current_findings:
                current = current_findings[key]

                if (
                    initial["status"] in ("FAIL", "WARN")
                    and current["status"] == "PASS"
                ):
                    # FIXED
                    self._upsert_action(
                        conn,
                        initial_run_id,
                        key,
                        initial["finding_type"],
                        "fixed",
                        now,
                        existing_actions,
                        f"Fixed: {initial['finding_type']} - {initial['entity_name']}",
                        current_run_id,
                    )
                    counts["fixed"] += 1

                elif initial["status"] in ("FAIL", "WARN") and current["status"] in (
                    "FAIL",
                    "WARN",
                ):
                    # STILL FAILING - NOT A CHANGE, don't log to action sheet
                    # The Actions sheet is a CHANGELOG - only record actual state transitions
                    # still_failing means: was bad, still bad = no change occurred
                    counts["still_failing"] += 1
                    # Do NOT upsert - this is not a change event

                elif initial["status"] == "PASS" and current["status"] in (
                    "FAIL",
                    "WARN",
                ):
                    # REGRESSION
                    self._upsert_action(
                        conn,
                        initial_run_id,
                        key,
                        current["finding_type"],
                        "regression",
                        now,
                        existing_actions,
                        f"Regression: {current['finding_type']} - {current['entity_name']}",
                        current_run_id,
                    )
                    counts["regression"] += 1

            else:
                # Key missing in current run (Entity deleted?)
                parts = key.split("|")
                if len(parts) >= 2:
                    instance_key = f"{parts[0]}|{parts[1]}"
                    
                    if instance_key in scanned_instances:
                        # Instance was scanned, finding is gone -> FIXED
                        # Only mark fixed if it was previously FAIL/WARN
                        if initial["status"] in ("FAIL", "WARN"):
                            self._upsert_action(
                                conn,
                                initial_run_id,
                                key,
                                initial["finding_type"],
                                "fixed",
                                now,
                                existing_actions,
                                f"Fixed (Removed): {initial['finding_type']} - {initial['entity_name']}",
                                current_run_id,
                            )
                            counts["fixed"] += 1
                    else:
                        # Instance wasn't scanned -> Assume Unreachable -> No Change
                        if instance_key not in unreachable_warnings:
                            unreachable_warnings.add(instance_key)

        # --------------------------------------------------------------------
        # PROCESS LATE ARRIVALS: Findings in Current but NOT in Initial
        # --------------------------------------------------------------------
        
        # Build set of instances present in INITIAL run
        initial_instances = set()
        for key in initial_findings:
            parts = key.split("|")
            if len(parts) >= 2:
                initial_instances.add(f"{parts[0]}|{parts[1]}")
        
        for key, current in current_findings.items():
            # If we haven't seen this key in initial findings
            if key not in initial_findings:
                # Check status - only care if it's FAIL/WARN
                if current["status"] in ("FAIL", "WARN"):
                    # Is this a new instance or just a new finding on old instance?
                    parts = key.split("|")
                    instance_key = f"{parts[0]}|{parts[1]}" if len(parts) >= 2 else ""
                    
                    is_late_arrival = instance_key and instance_key not in initial_instances
                    
                    if is_late_arrival:
                        # LATE ARRIVAL: Treat as baseline finding
                        desc = f"Baseline Finding (Late Arrival): {current['finding_type']}"
                        logger.info("Late arrival server detected: %s", instance_key)
                    else:
                        # NEW ISSUE on existing server
                        desc = f"New Issue Detected: {current['finding_type']}"
                    
                    # Upsert as 'new' action so it appears in report
                    self._upsert_action(
                        conn,
                        initial_run_id,
                        key,
                        current["finding_type"],
                        "new",
                        now,
                        existing_actions,
                        desc,
                        current_run_id,
                    )
                    counts["new"] += 1

        for w in unreachable_warnings:
            logger.warning("Target %s appears Unreachable. Skipping fix detection.", w)
            print(f"⚠️  WARNING: Target {w} was unreachable. Assuming NO CHANGE.")

        return counts

    def _upsert_action(
        self,
        conn,
        initial_run_id: int,
        entity_key: str,
        finding_type: str,
        action_type: str,
        now: str,
        existing_actions: dict,
        action_description: str | None,
        current_run_id: int,
    ) -> None:
        """
        Insert or update action_log entry.

        Key: Preserve original action_date for existing entries.
        """
        if entity_key in existing_actions:
            # UPDATE: preserve original action_date
            conn.execute(
                """
                UPDATE action_log 
                SET action_type = ?, action_description = COALESCE(?, action_description),
                    last_sync_run_id = ?
                WHERE initial_run_id = ? AND entity_key = ?
            """,
                (
                    action_type,
                    action_description,
                    current_run_id,
                    initial_run_id,
                    entity_key,
                ),
            )
        else:
            # INSERT: set action_date to NOW
            conn.execute(
                """
                INSERT INTO action_log
                (initial_run_id, entity_key, finding_type, action_type, 
                 action_date, action_description, captured_at, last_sync_run_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
                (
                    initial_run_id,
                    entity_key,
                    finding_type,
                    action_type,
                    now,
                    action_description,
                    now,
                    current_run_id,
                ),
            )

    def _log_exception_action(
        self,
        conn,
        initial_run_id: int,
        current_run_id: int,
        exc: dict,
    ) -> None:
        """
        Log a documented exception as an action in the action_log.
        
        This creates an append-only record when a user adds justification
        to a FAIL/WARN item, marking it as a documented exception.
        
        Args:
            conn: Database connection
            initial_run_id: The baseline run ID
            current_run_id: Current sync run ID
            exc: Exception dict from detect_exception_changes
        """
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Build action description with justification preview
        justification_preview = exc["justification"][:60]
        if len(exc["justification"]) > 60:
            justification_preview += "..."
        
        # Create entity key in format: entity_type|entity_key|exception
        entity_key = f"{exc['entity_type']}|{exc['entity_key']}|exception"
        
        # Determine finding type from entity type
        entity_type_display = exc["entity_type"].replace("_", " ").title()
        
        action_type = "exception_documented"
        action_description = f"Exception Documented: {entity_type_display} - {justification_preview}"
        
        # Always INSERT (append-only for exceptions)
        # Don't update existing - each justification change is a new log entry
        conn.execute(
            """
            INSERT INTO action_log
            (initial_run_id, entity_key, finding_type, action_type, 
             action_date, action_description, captured_at, last_sync_run_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                initial_run_id,
                entity_key,
                f"Exception: {entity_type_display}",
                action_type,
                now,
                action_description,
                now,
                current_run_id,
            ),
        )
        
        logger.info("Logged exception action: %s", entity_key[:50])

    def _read_existing_action_annotations(
        self, excel_path: Path
    ) -> dict:
        """
        Read user-entered data from existing Actions sheet.
        
        Preserves:
        - Found Date (if user edited it)
        - Assigned To, Due Date, Resolution Date, Resolution Notes
        
        Returns:
            Dict keyed by (server, instance, finding) -> user annotations
        """
        from openpyxl import load_workbook
        
        preserved = {}
        
        if not excel_path.exists():
            return preserved
        
        try:
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            if "Actions" not in wb.sheetnames:
                wb.close()
                return preserved
            
            ws = wb["Actions"]
            
            # Column mapping (1-indexed):
            # A=ID, B=Server, C=Instance, D=Category, E=Finding, F=Risk,
            # G=Recommendation, H=Status, I=Found Date, J=Assigned To,
            # K=Due Date, L=Resolution Date, M=Resolution Notes
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:  # Skip empty rows
                    continue
                    
                # Create key from Server|Instance|Finding
                server = row[1] or ""
                instance = row[2] or ""
                finding = row[4] or ""
                
                # Normalize instance for key matching
                inst_normalized = "" if instance == "(Default)" else instance
                key = f"{server}|{inst_normalized}|{finding}"
                
                preserved[key] = {
                    "found_date": row[8] if len(row) > 8 else None,
                    "assigned_to": row[9] if len(row) > 9 else None,
                    "due_date": row[10] if len(row) > 10 else None,
                    "resolution_date": row[11] if len(row) > 11 else None,
                    "resolution_notes": row[12] if len(row) > 12 else None,
                }
            
            wb.close()
            logger.info("Read %d existing action annotations from Excel", len(preserved))
            
        except Exception as e:
            logger.warning("Could not read existing actions: %s", e)
        
        return preserved

    def get_action_summary(self, initial_run_id: int | None = None) -> dict:
        """Get summary of all actions for an audit cycle."""
        import sqlite3

        if initial_run_id is None:
            initial_run_id = self.get_initial_run_id()

        if initial_run_id is None:
            return {"error": "No baseline found"}

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        rows = conn.execute(
            """
            SELECT action_type, COUNT(*) as count
            FROM action_log
            WHERE initial_run_id = ?
            GROUP BY action_type
        """,
            (initial_run_id,),
        ).fetchall()

        conn.close()

        return {row["action_type"]: row["count"] for row in rows}

    def get_actions(self, initial_run_id: int | None = None) -> list[dict]:
        """Get all actions for an audit cycle."""
        import sqlite3

        if initial_run_id is None:
            initial_run_id = self.get_initial_run_id()

        if initial_run_id is None:
            return []

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        rows = conn.execute(
            """
            SELECT * FROM action_log
            WHERE initial_run_id = ?
            ORDER BY action_date, entity_key
        """,
            (initial_run_id,),
        ).fetchall()

        conn.close()

        return [dict(row) for row in rows]


def main():
    """CLI entry point for sync."""
    import argparse

    parser = argparse.ArgumentParser(description="Sync remediation progress")
    parser.add_argument("--db", default="output/audit_history.db")
    parser.add_argument("--targets", default="sql_targets.json")
    parser.add_argument(
        "--summary", action="store_true", help="Show action summary only"
    )

    args = parser.parse_args()

    service = SyncService(db_path=args.db)

    if args.summary:
        summary = service.get_action_summary()
        if "error" in summary:
            print(f"❌ {summary['error']}")
            return 1
        print(f"\n📊 Action Summary:")
        for action_type, count in summary.items():
            print(f"   {action_type}: {count}")
        return 0

    result = service.sync(targets_file=args.targets)

    if "error" in result:
        print(f"❌ {result['error']}")
        return 1

    print(f"\n✅ Sync complete!")
    print(f"   Baseline: Run #{result['initial_run_id']}")
    print(f"   Current:  Run #{result['current_run_id']}")
    print(f"")
    print(f"   ✅ Fixed:         {result['fixed']}")
    print(f"   ⚠️  Still Failing: {result['still_failing']}")
    print(f"   🔴 Regression:    {result['regression']}")
    print(f"   🆕 New:           {result['new']}")

    return 0


if __name__ == "__main__":
    exit(main())
