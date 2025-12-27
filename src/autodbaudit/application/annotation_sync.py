"""
Annotation Sync Service.

Bidirectional synchronization of user annotations between Excel and SQLite.

This module provides:
- Read annotations from Excel worksheets
- Write annotations back to Excel
- Persist to/from SQLite database
- Detect exception and documentation changes

NOTE: A modular refactor exists at autodbaudit.application.sync/
but it is INCOMPLETE. This module remains the working implementation.
"""

from __future__ import annotations


def _clean_key_value(val: str) -> str:
    """
    Clean key value by stripping non-ASCII characters (icons/emojis).
    Used to normalize keys that contain decoration icons in Excel.
    """
    if not val:
        return ""
    # Encode to ASCII (ignore errors drops non-ascii) then decode back
    return str(val).encode("ascii", "ignore").decode("ascii").strip()


import logging
from datetime import datetime
from pathlib import Path

from autodbaudit.domain.entity_key import (
    normalize_key_string,
    annotation_key_to_finding_key,
)

logger = logging.getLogger(__name__)


# Sheet configuration: entity type, key columns, editable columns
# Key columns are column header names (matched case-insensitively)
# Editable columns map Excel header name to DB field name
# Justification columns mark FAIL items as documented exceptions when filled
#
# ⚠️ DEPRECATION NOTICE:
# This configuration is DUPLICATED and should be migrated to:
#   autodbaudit.domain.sheet_registry.SHEET_REGISTRY
#
# The sheet_registry is the SINGLE SOURCE OF TRUTH.
# This local config exists for backward compatibility and will be removed.
# See: domain/sheet_registry.py
#
SHEET_ANNOTATION_CONFIG = {
    "SA Account": {
        "entity_type": "sa_account",
        "key_cols": ["Server", "Instance", "Current Name"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",  # Excel has 'Last Reviewed'
            "Notes": "notes",
        },
    },
    "Server Logins": {
        "entity_type": "login",
        "key_cols": ["Server", "Instance", "Login Name"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
            "Notes": "notes",
        },
    },
    "Sensitive Roles": {
        "entity_type": "server_role_member",
        # Key includes Role to disambiguate same member in multiple roles
        "key_cols": ["Server", "Instance", "Role", "Member"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
        },
    },
    "Configuration": {
        "entity_type": "config",
        "key_cols": ["Server", "Instance", "Setting"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",  # Excel has 'Last Reviewed'
            # No Notes column in Excel for this sheet
        },
    },
    "Instances": {
        "entity_type": "instance",
        "key_cols": ["Server", "Instance"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
            "Notes": "notes",
        },
    },
    "Services": {
        "entity_type": "service",
        "key_cols": ["Server", "Instance", "Service Name"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
            # No Notes column in Excel for this sheet
        },
    },
    "Databases": {
        "entity_type": "database",
        "key_cols": ["Server", "Instance", "Database"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",  # Excel has 'Last Reviewed'
            "Notes": "notes",
        },
    },
    "Database Users": {
        "entity_type": "db_user",
        "key_cols": ["Server", "Instance", "Database", "User Name"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",  # Excel has 'Last Reviewed'
            "Notes": "notes",
        },
    },
    "Database Roles": {
        "entity_type": "db_role",
        "key_cols": ["Server", "Instance", "Database", "Role", "Member"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",  # Excel has 'Last Reviewed'
            # No Notes column in Excel for this sheet
        },
    },
    "Permission Grants": {
        "entity_type": "permission",
        # Key includes Database + Entity Name to disambiguate same permission on different objects
        "key_cols": [
            "Server",
            "Instance",
            "Scope",
            "Database",
            "Grantee",
            "Permission",
            "Entity Name",
        ],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
            "Notes": "notes",
        },
    },
    "Orphaned Users": {
        "entity_type": "orphaned_user",
        "key_cols": ["Server", "Instance", "Database", "User Name"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
            # No Notes column in Excel for this sheet
        },
    },
    "Linked Servers": {
        "entity_type": "linked_server",
        # Just Server/Instance/Linked Server name - logins are optional and cause trailing pipes
        "key_cols": ["Server", "Instance", "Linked Server"],
        "editable_cols": {
            "Review Status": "review_status",
            "Purpose": "purpose",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
            # No Notes column in Excel for this sheet
        },
    },
    "Triggers": {
        "entity_type": "trigger",
        # Key includes Scope to distinguish SERVER vs DATABASE level triggers
        # Added Event to distinguish multiple events for same trigger
        "key_cols": [
            "Server",
            "Instance",
            "Scope",
            "Database",
            "Trigger Name",
            "Event",
        ],
        "editable_cols": {
            "Review Status": "review_status",
            "Notes": "notes",  # Purpose/notes for this trigger
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
        },
    },
    "Client Protocols": {
        "entity_type": "protocol",
        "key_cols": ["Server", "Instance", "Protocol"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",
            "Status": "status",
            "Enabled": "enabled",
        },
    },
    "Backups": {
        "entity_type": "backup",
        "key_cols": ["Server", "Instance", "Database"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",  # Excel has 'Last Reviewed'
            "Notes": "notes",
        },
    },
    "Audit Settings": {
        "entity_type": "audit_settings",
        "key_cols": ["Server", "Instance", "Setting"],
        "editable_cols": {
            "Review Status": "review_status",
            "Justification": "justification",
            "Last Reviewed": "last_reviewed",  # Excel has 'Last Reviewed'
            "Notes": "notes",
        },
    },
    "Encryption": {
        "entity_type": "encryption",
        # Column names match writer: "Key Type" and "Key Name"
        "key_cols": ["Server", "Instance", "Key Type", "Key Name"],
        "editable_cols": {
            "Notes": "notes",
        },
    },
    "Actions": {
        "entity_type": "action",
        "key_cols": ["ID"],  # Use DB ID for unique, reliable row matching
        "editable_cols": {
            "Detected Date": "action_date",
            "Notes": "notes",
            "Server": "server",
            "Instance": "instance",
            "Category": "category",
            "Finding": "finding",
            "Risk Level": "risk_level",
            "Change Description": "change_description",
            "Change Type": "change_type",
        },
    },
}


class AnnotationSyncService:
    """
    Bidirectional annotation sync between Excel and SQLite.

    Usage:
        sync = AnnotationSyncService("output/audit_history.db")

        # Before regenerating Excel:
        annotations = sync.read_all_from_excel("output/Audit_Latest.xlsx")

        # After regenerating Excel:
        sync.write_all_to_excel("output/Audit_Latest.xlsx", annotations)

        # Persist to DB (for finalize):
        sync.persist_to_db(annotations)
    """

    def __init__(self, db_path: str | Path = "output/audit_history.db"):
        """Initialize annotation sync service."""
        self.db_path = Path(db_path)
        logger.info("AnnotationSyncService initialized")

    def read_all_from_excel(self, excel_path: Path | str) -> dict[str, dict]:
        """
        Read all annotations from all configured sheets in Excel.

        Args:
            excel_path: Path to Excel file

        Returns:
            Dict of {entity_key: {field_name: value}}
        """
        from openpyxl import load_workbook

        excel_path = Path(excel_path)
        all_annotations: dict[str, dict] = {}

        if not excel_path.exists():
            logger.warning("Excel file not found: %s", excel_path)
            return all_annotations

        try:
            wb = load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error("Failed to open Excel file: %s", e)
            return all_annotations

        for sheet_name, config in SHEET_ANNOTATION_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            sheet_annotations = self._read_sheet_annotations(ws, config)

            for entity_key, fields in sheet_annotations.items():
                # Prefix with entity type for uniqueness (lowercase for consistency)
                entity_type = str(config["entity_type"]).lower()
                full_key = f"{entity_type}|{entity_key}"
                all_annotations[full_key] = fields

        wb.close()
        logger.info(
            "Read %d annotations from %d sheets",
            len(all_annotations),
            len(wb.sheetnames),
        )
        return all_annotations

    def _read_sheet_annotations(self, ws, config: dict) -> dict[str, dict]:
        """
        Read annotations from a single worksheet.

        Uses header row to find column positions dynamically.
        Handles merged cells by tracking last non-empty values for key columns.
        """
        annotations: dict[str, dict] = {}

        # Get header row to find column indices
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        if not header_row:
            return annotations

        # Map header names to column indices (0-indexed)
        header_map = {}
        for idx, header in enumerate(header_row):
            if header:
                # Normalize header (strip whitespace, handle emojis)
                clean_header = str(header).strip()
                # Remove common prefixes/emojis
                for prefix in ["⏳ ", "✅ ", "⚠️ ", "❌ "]:
                    clean_header = clean_header.replace(prefix, "")
                header_map[clean_header] = idx

        # Find key column indices
        key_indices = []
        found_keys = 0

        # Try primary key config
        for key_col in config["key_cols"]:
            found = False
            if key_col in header_map:
                key_indices.append(header_map[key_col])
                found = True
            else:
                # Try EXACT match first (case-insensitive)
                exact_found = False
                for h, idx in header_map.items():
                    if key_col.lower() == h.lower():
                        key_indices.append(idx)
                        found = True
                        exact_found = True
                        break
                # Only use partial match if exact match failed
                if not exact_found:
                    for h, idx in header_map.items():
                        # Partial match but avoid substring collisions
                        if key_col.lower() in h.lower() and len(key_col) >= len(h) // 2:
                            key_indices.append(idx)
                            found = True
                            break
            if found:
                found_keys += 1

        # Fallback for Backups (backward compatibility for existing reports)
        if found_keys != len(config["key_cols"]) and config["entity_type"] == "backup":
            logger.warning("Backups sheet missing 'Recovery Model'. Trying legacy key.")
            legacy_keys = ["Server", "Instance", "Database"]
            key_indices = []
            found_keys = 0
            for key_col in legacy_keys:
                if key_col in header_map:
                    key_indices.append(header_map[key_col])
                    found_keys += 1

            # If we found the legacy keys, valid.
            # Note: This means entity_key will be "Server|Instance|Database"
            # Logic later needs to handle this or migration.
            if found_keys == 3:
                logger.info("Using legacy key for Backups sheet.")

        if found_keys != len(key_indices) and found_keys == 0:
            # Logic above: key_indices length might differ if we fallback?
            # Actually, key_indices contains the indices. found_keys tracks count.
            # If we used fallback, key_indices has 3 items. found_keys is 3.
            pass

        # Strict check: Did we find enough keys for strict match OR fallback match?
        # If we are using standard config, we need match.
        required_keys = (
            3
            if (config["entity_type"] == "backup" and len(key_indices) == 3)
            else len(config["key_cols"])
        )

        if len(key_indices) != required_keys:
            logger.warning(
                "Could not find all key columns for %s sheet (Found %d/%d)",
                ws.title,
                len(key_indices),
                required_keys,
            )
            return annotations

        # Find editable column indices
        editable_indices = {}
        missing_cols = []
        for col_name, field_name in config["editable_cols"].items():
            found = False
            # Try EXACT match first (case-sensitive)
            if col_name in header_map:
                editable_indices[field_name] = header_map[col_name]
                found = True
            else:
                # Try EXACT match (case-insensitive)
                for h, idx in header_map.items():
                    if col_name.lower() == h.lower():
                        editable_indices[field_name] = idx
                        found = True
                        break
                # Only use partial match if exact match failed
                if not found:
                    for h, idx in header_map.items():
                        if col_name.lower() in h.lower():
                            editable_indices[field_name] = idx
                            found = True
                            break
            if not found:
                missing_cols.append(col_name)

        # Log warning if columns are missing
        if missing_cols:
            logger.warning(
                "Sheet %s: Could not find editable columns: %s (Available: %s)",
                ws.title,
                missing_cols,
                list(header_map.keys()),
            )

        # Find action indicator column (column with header containing ⏳)
        # MUST use raw header_row because header_map keys are stripped of emojis
        action_col_idx = None
        status_col_idx = None  # Track Status column for discrepancy detection
        for idx, val in enumerate(header_row):
            if val:
                val_str = str(val)
                if "⏳" in val_str:
                    action_col_idx = idx
                # Find Status column (may contain PASS/FAIL/WARN)
                if val_str.strip().lower() == "status" or val_str.strip() == "Status":
                    status_col_idx = idx

        logger.warning(
            f"DEBUG: Sheet {ws.title} ActionCol={action_col_idx} StatusCol={status_col_idx}"
        )

        # Track last non-empty values for key columns (handles merged cells)
        # When cells are merged, only first row has value, rest are None
        last_key_values = [""] * len(key_indices)

        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue

            # === UUID-BASED MATCHING (v3) ===
            # Column A contains stable UUID - use this as primary key
            # This ensures annotations persist even if data in key columns changes
            row_uuid = None
            if len(row) > 0 and row[0]:
                uuid_val = str(row[0]).strip().lower()
                # Validate: 8 hex chars
                if len(uuid_val) == 8:
                    try:
                        int(uuid_val, 16)
                        row_uuid = uuid_val
                    except ValueError:
                        pass

            # Build legacy entity key from key columns (for backwards compatibility)
            # Use last non-empty value if current is None (merged cell)
            key_parts = []
            for i, idx in enumerate(key_indices):
                val = row[idx] if idx < len(row) else None

                if val is None:
                    val = last_key_values[i]
                else:
                    val = str(val)

                    # Clean permission strings (remove icons)
                    if config["key_cols"][i] == "Permission":
                        val = _clean_key_value(val)

                    # Update last known value
                    last_key_values[i] = val

                key_parts.append(val)

            # Build legacy entity key (normalized to lowercase)
            legacy_entity_key = "|".join(str(p).lower() for p in key_parts)

            # === PRIMARY KEY SELECTION ===
            # Prefer UUID if available, fall back to legacy entity_key
            # UUID format: uppercase 8-char hex (e.g., "A7F3B2C1")
            # Legacy format: entity_type prefixed key (e.g., "server|instance|name")
            if row_uuid:
                entity_key = row_uuid
            else:
                entity_key = legacy_entity_key
                if any(key_parts):
                    logger.debug(
                        "No UUID found for row in %s, using legacy key: %s",
                        ws.title,
                        entity_key[:50],
                    )

            # Skip if we couldn't build a valid key
            if not row_uuid and not any(key_parts):
                continue

            # Extract editable fields
            fields = {}
            has_any_value = False
            for field_name, col_idx in editable_indices.items():
                if col_idx < len(row):
                    val = row[col_idx]

                    # Robust Parsing & Casting
                    # Robust Parsing & Casting
                    val_str = str(val).strip() if val is not None else ""

                    if val_str:
                        # A) Date Fields: Use robust parser
                        if (
                            "date" in field_name.lower()
                            or "revised" in field_name.lower()
                            or "reviewed" in field_name.lower()
                        ):
                            from autodbaudit.infrastructure.excel.base import (
                                parse_datetime_flexible,
                            )

                            parsed_date = parse_datetime_flexible(
                                val, log_errors=True, context=f"{ws.title}|{entity_key}"
                            )
                            if parsed_date:
                                fields[field_name] = parsed_date
                                has_any_value = True

                        # B) Text Fields: Strict String Cast
                        else:
                            fields[field_name] = val_str
                            has_any_value = True
                    else:
                        # Capture empty strings to allow clearing annotations in DB
                        fields[field_name] = ""
                        has_any_value = True

            # AUTO-STATUS: Check Justification ONLY (notes/purpose are documentation, NOT exception triggers)
            raw_just = fields.get("justification")

            # Check review status for "Exception"
            raw_status = fields.get("review_status")
            is_explicit_exception = raw_status and "Exception" in str(raw_status)

            if (raw_just and str(raw_just).strip()) or is_explicit_exception:
                from autodbaudit.infrastructure.excel.base import STATUS_VALUES

                # We enforce Exception status so it syncs back to Excel next time
                fields["review_status"] = STATUS_VALUES.EXCEPTION
                # NOTE: Do NOT set action_needed here - that's only for actual discrepant rows
                # Justification on a PASS row is just documentation
                has_any_value = True
                logger.debug(
                    "Set review_status to Exception for %s (justification/status present)",
                    entity_key,
                )

            # If explicit "Exception" status but no justification, we still treat as documented
            # (Logic handled in detect_exception_changes)

            # Check if this row has action indicator (⏳ means needs action = FAIL/WARN)
            # This is used by detect_exception_changes to only log exceptions for FAIL items
            if action_col_idx is not None and action_col_idx < len(row):
                action_val = row[action_col_idx]
                if "Sensitive-Role" in entity_key:
                    logger.warning(
                        f"DEBUG: {entity_key} ActionVal='{action_val}' ActionCol={action_col_idx}"
                    )
                # ⏳ = needs action (FAIL), ✅ = documented exception
                if action_val and "⏳" in str(action_val):
                    fields["action_needed"] = True
                elif action_val and "✅" in str(action_val):
                    # Already documented, but was previously a FAIL
                    fields["action_needed"] = True

            # Read actual Status column value for discrepancy detection
            if status_col_idx is not None and status_col_idx < len(row):
                status_val = row[status_col_idx]
                if status_val:
                    fields["status"] = str(status_val).strip()

            # Only store if there are actual annotations
            if has_any_value:
                # Store metadata for exception detection matching
                # When annotations are keyed by UUID, we need legacy_entity_key for findings lookup
                fields["_legacy_entity_key"] = legacy_entity_key
                fields["_row_uuid"] = row_uuid if row_uuid else None

                # IMPORTANT: Store with entity_key (UUID or legacy) as dict key
                # The write_all_to_excel expects these keys to match Excel rows
                # DB persistence in persist_to_db() expects full keys (type|key)
                # So we MUST prepend entity_type here
                full_key = f"{config['entity_type']}|{entity_key}"
                annotations[full_key] = fields

        logger.debug("Read %d annotations from %s sheet", len(annotations), ws.title)
        return annotations

    def write_all_to_excel(
        self, excel_path: Path | str, annotations: dict[str, dict]
    ) -> int:
        """
        Write annotations back to Excel file.

        Args:
            excel_path: Path to Excel file
            annotations: Dict from read_all_from_excel

        Returns:
            Number of cells updated
        """
        from openpyxl import load_workbook

        excel_path = Path(excel_path)
        if not excel_path.exists():
            logger.warning("Excel file not found: %s", excel_path)
            return 0

        try:
            wb = load_workbook(excel_path)
        except Exception as e:
            logger.error("Failed to open Excel file for writing: %s", e)
            return 0

        total_updated = 0

        for sheet_name, config in SHEET_ANNOTATION_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            entity_type = config["entity_type"]

            # Filter annotations for this sheet
            sheet_annotations = {
                k.split("|", 1)[1]: v
                for k, v in annotations.items()
                if k.startswith(f"{entity_type}|")
            }

            if sheet_annotations:
                count = self._write_sheet_annotations(ws, config, sheet_annotations)
                total_updated += count

        try:
            wb.save(excel_path)
            logger.info("Wrote %d annotation cells to %s", total_updated, excel_path)
        except PermissionError:
            logger.error(
                f"❌ Cannot write to '{excel_path.name}' - file is open!\n"
                f"   Please close the file in Excel and try again."
            )
            return 0
        except Exception as e:
            logger.error("Failed to save Excel file: %s", e)
            return 0

        wb.close()
        return total_updated

    def _write_sheet_annotations(
        self, ws, config: dict, annotations: dict[str, dict]
    ) -> int:
        """Write annotations to a single worksheet."""
        updated = 0
        logger.warning(
            f"DEBUG: _write_sheet_annotations for {ws.title} with {len(annotations)} items"
        )

        # Get header row to find column indices
        header_row = [cell.value for cell in ws[1]]
        if not header_row:
            logger.warning(f"DEBUG: No header row found in {ws.title}")
            return 0

        # Map header names to column indices (1-indexed for openpyxl write)
        header_map = {}
        for idx, header in enumerate(header_row):
            if header:
                clean_header = str(header).strip()
                for prefix in ["⏳ ", "✅ ", "⚠️ ", "❌ "]:
                    clean_header = clean_header.replace(prefix, "")
                header_map[clean_header] = idx + 1  # 1-indexed

        # Find key column indices
        key_indices = []
        for key_col in config["key_cols"]:
            # Try EXACT match first
            exact_found = False
            for h, idx in header_map.items():
                if key_col.lower() == h.lower():
                    key_indices.append(idx)
                    exact_found = True
                    break
            # Only use partial match if exact match failed
            if not exact_found:
                for h, idx in header_map.items():
                    # Partial match but avoid substring collisions
                    if key_col.lower() in h.lower() and len(key_col) >= len(h) // 2:
                        key_indices.append(idx)
                        break

        if len(key_indices) != len(config["key_cols"]):
            logger.warning(f"DEBUG: Header mismatch in {ws.title}. Start Search.")

        # Find editable column indices (1-indexed)
        editable_indices = {}
        for col_name, field_name in config["editable_cols"].items():
            # Try EXACT match first
            exact_found = False
            for h, idx in header_map.items():
                if col_name.lower() == h.lower():
                    editable_indices[field_name] = idx
                    exact_found = True
                    break
            # Only use partial match if exact match failed
            if not exact_found:
                for h, idx in header_map.items():
                    if col_name.lower() in h.lower():
                        editable_indices[field_name] = idx
                        break

        # Find action indicator column (usually column 1 with header "⏳")
        action_col_idx = None
        status_col_idx = None  # Column with PASS/FAIL/WARN values
        for h, idx in header_map.items():
            if "⏳" in str(h) or h == "":  # Action column header is usually just ⏳
                action_col_idx = idx
            if h.strip().lower() == "status":
                status_col_idx = idx

        # Check if this sheet has a justification field
        has_justification = "justification" in config["editable_cols"].values()

        # Track last non-empty values for key columns (handles merged cells)
        last_key_values = [""] * len(key_indices)

        # Build legacy key map for fallback matching (recovers annotations if UUIDs changed)
        legacy_key_map = {}
        for key, fields in annotations.items():
            legacy_key = fields.get("_legacy_entity_key")
            if legacy_key:
                legacy_key_map[legacy_key.lower()] = fields

        # Iterate through data rows and update matching cells
        rows_processed = 0
        for row_num in range(2, ws.max_row + 1):
            rows_processed += 1

            # === UUID-BASED MATCHING (v3) ===
            # Read UUID from Column A (column 1)
            row_uuid = None
            uuid_cell = ws.cell(row=row_num, column=1).value
            if uuid_cell:
                uuid_val = str(uuid_cell).strip().lower()
                # Validate: 8 hex chars
                if len(uuid_val) == 8:
                    try:
                        int(uuid_val, 16)
                        row_uuid = uuid_val
                    except ValueError:
                        pass

            # Build legacy entity key from key columns
            # Handle merged cells by using last non-empty value
            key_parts = []
            for i, col_idx in enumerate(key_indices):
                cell_val = ws.cell(row=row_num, column=col_idx).value

                if cell_val is None:
                    # Merged cell - use last known value
                    key_parts.append(last_key_values[i])
                elif cell_val == "(Default)":
                    key_parts.append("")
                    last_key_values[i] = ""
                else:
                    val = str(cell_val)

                    # Clean special columns to match read logic
                    if config["key_cols"][i] == "Permission":
                        val = _clean_key_value(val)

                    key_parts.append(val)
                    last_key_values[i] = val

            legacy_entity_key = "|".join(key_parts)
            # Normalize to lowercase for consistent matching with DB keys
            entity_key_lower = legacy_entity_key.lower()

            # Skip if we couldn't build a valid key
            if not row_uuid and not any(key_parts):
                continue

            # === ANNOTATION MATCHING ===
            # Try UUID first (preferred), then legacy entity_key (direct match or map lookup)
            row_annotations = None
            display_key = row_uuid if row_uuid else legacy_entity_key  # For logging

            if row_uuid and row_uuid in annotations:
                row_annotations = annotations[row_uuid]
            elif entity_key_lower in annotations:
                row_annotations = annotations[entity_key_lower]
            elif entity_key_lower in legacy_key_map:
                # Fallback: Match by legacy key (e.g., if UUID changed due to regen)
                row_annotations = legacy_key_map[entity_key_lower]
                if rows_processed <= 5:  # Log first few only to avoid spam
                    logger.debug("Matched by legacy fallback: %s", entity_key_lower)

            # Check if we have annotations for this row
            if row_annotations:
                for field_name, col_idx in editable_indices.items():
                    if field_name in row_annotations:
                        val = row_annotations[field_name]

                        # Convert ISO date strings back to datetime objects for Excel
                        if (
                            val
                            and isinstance(val, str)
                            and (
                                "date" in field_name.lower()
                                or "revised" in field_name.lower()
                                or "reviewed" in field_name.lower()
                            )
                        ):
                            from autodbaudit.infrastructure.excel.base import (
                                parse_datetime_flexible,
                            )

                            dt = parse_datetime_flexible(val, log_errors=False)
                            if dt:
                                val = dt

                        ws.cell(row=row_num, column=col_idx).value = val
                        updated += 1

                # Update action indicator (⏳→✅) ONLY if:
                # - Row is FAIL/WARN (discrepant), AND
                # - justification is filled OR review_status is "Exception"
                # PASS rows: keep justification as text but NO indicator, CLEAR Exception status
                if has_justification and action_col_idx:
                    justification = row_annotations.get("justification", "")
                    review_status = row_annotations.get("review_status", "")
                    has_just = justification and str(justification).strip()
                    has_exception = review_status and "Exception" in str(review_status)

                    # Check row status - must be discrepant to apply indicator
                    is_discrepant = False
                    status_val = None
                    if status_col_idx:
                        status_val = ws.cell(row=row_num, column=status_col_idx).value
                        if status_val:
                            status_str = str(status_val).upper()
                            is_discrepant = status_str in ("FAIL", "WARN", "⏳", "⚠")

                    # Fallback: Check Action Indicator Column (for sheets like Logins)
                    if not is_discrepant and action_col_idx:
                        action_val = ws.cell(row=row_num, column=action_col_idx).value
                        if action_val:
                            s_val = str(action_val)
                            if "⏳" in s_val or "✓" in s_val or "✅" in s_val:
                                is_discrepant = True

                    if (has_just or has_exception) and is_discrepant:
                        # DISCREPANT + (justification OR Exception status) = Valid Exception
                        from autodbaudit.infrastructure.excel.base import (
                            apply_exception_documented_styling,
                        )

                        apply_exception_documented_styling(
                            ws.cell(row=row_num, column=action_col_idx)
                        )
                        updated += 1
                        logger.warning(
                            "DEBUG: Updated Excel Indicator for %s (Just=%s, Exc=%s, Stat=%s, Row=%d)",
                            display_key,
                            has_just,
                            has_exception,
                            status_val,
                            row_num,
                        )
                    elif is_discrepant and not (has_just or has_exception):
                        # Was an exception (or marked discrepant), but justification removed.
                        # Revert to Needs Action (⏳)
                        from autodbaudit.infrastructure.excel.base import (
                            apply_action_needed_styling,
                        )

                        apply_action_needed_styling(
                            ws.cell(row=row_num, column=action_col_idx),
                            needs_action=True,
                        )
                        updated += 1
                        logger.warning(
                            "DEBUG: Reverted Excel Indicator for %s (Just=%s, Exc=%s, Row=%d)",
                            display_key,
                            has_just,
                            has_exception,
                            row_num,
                        )
                    elif has_exception and not is_discrepant:
                        # PASS row with Exception dropdown: CLEAR the exception status
                        # Per requirements: "Non-discrepant + Exception dropdown → Ignored, cleared"
                        # Keep justification as documentation (don't touch it)
                        review_status_col = editable_indices.get("review_status")
                        if review_status_col:
                            ws.cell(row=row_num, column=review_status_col).value = ""
                            logger.debug(
                                "Cleared Exception status for PASS row: %s", display_key
                            )
                    else:
                        logger.warning(
                            "DEBUG: Skipped Indicator Update for %s (Just=%s, Exc=%s, Disc=%s, Stat=%s, Row=%d)",
                            display_key,
                            has_just,
                            has_exception,
                            is_discrepant,
                            status_val,
                            row_num,
                        )
                    # Note: has_just and not is_discrepant → justification kept as documentation (no action needed)

        return updated

    def persist_to_db(self, annotations: dict[str, dict]) -> int:
        """
        Persist all annotations to SQLite database.

        Args:
            annotations: Dict from read_all_from_excel

        Returns:
            Number of annotations saved
        """
        import sqlite3
        from autodbaudit.infrastructure.sqlite.schema import set_annotation

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        count = 0
        for full_key, fields in annotations.items():
            parts = full_key.split("|", 1)
            if len(parts) != 2:
                continue

            entity_type, entity_key = parts

            for field_name, value in fields.items():
                if value is not None:
                    # Convert datetime to string if needed
                    if isinstance(value, datetime):
                        value = value.isoformat()

                    set_annotation(
                        connection=conn,
                        entity_type=entity_type,
                        entity_key=entity_key,
                        field_name=field_name,
                        field_value=str(value),
                    )
                    count += 1

        conn.close()
        logger.info("Persisted %d annotations to database", count)
        return count

    def load_from_db(self) -> dict[str, dict]:
        """
        Load all annotations from SQLite database.

        Returns:
            Dict of {entity_type|entity_key: {field_name: value}}
        """
        import sqlite3

        annotations: dict[str, dict] = {}

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        rows = conn.execute(
            "SELECT entity_type, entity_key, field_name, field_value FROM annotations"
        ).fetchall()

        for row in rows:
            # Normalize to lowercase for consistent matching
            entity_type = row["entity_type"].lower() if row["entity_type"] else ""
            entity_key = row["entity_key"].lower() if row["entity_key"] else ""
            full_key = f"{entity_type}|{entity_key}"
            if full_key not in annotations:
                annotations[full_key] = {}
            annotations[full_key][row["field_name"]] = row["field_value"]

        conn.close()
        logger.info("Loaded %d annotation entries from database", len(annotations))
        return annotations

    def get_all_annotations(self) -> dict[str, dict]:
        """
        Get all annotations keyed by entity_key.

        This is an alias for load_from_db() to satisfy the AnnotationsProvider protocol
        expected by StatsService.

        Returns:
            Dict of {entity_type|entity_key: {field_name: value}}
        """
        return self.load_from_db()

    def detect_exception_changes(
        self,
        old_annotations: dict[str, dict],
        new_annotations: dict[str, dict],
        current_findings: list[dict] | None = None,
    ) -> list[dict]:
        """
        Compare annotations to detect new/changed/removed justifications for FAIL items.

        Only logs as "exception" when an item that had a pending action (⏳)
        now has justification. Items that were already PASS are just notes,
        not exceptions.

        Also detects REMOVED exceptions: when user clears BOTH Justification
        AND Review Status (reverts from Exception).

        Args:
            old_annotations: Previous annotations from DB
            new_annotations: Current annotations from Excel
            current_findings: Current findings from DB (for accurate PASS/FAIL status)

        Returns:
            List of {entity_key, entity_type, justification, is_new, change_type}
            where change_type is 'added', 'updated', or 'removed'
        """
        exceptions = []

        # Build findings map for status AND server/instance lookup with LOWERCASE keys
        # Key format in findings: entity_key like "type|SERVER|INSTANCE|entity_name"
        # Key format in annotations: "entity_type|UUID" with _legacy_entity_key storing "server|instance|entity"
        # We normalize ALL keys to lowercase for case-insensitive matching
        findings_status_map: dict[str, str] = {}
        findings_full_map: dict[str, dict] = {}  # Full finding data for server/instance
        if current_findings:
            for f in current_findings:
                raw_key = f.get("entity_key", "")
                # Normalize to lowercase for matching
                normalized_key = normalize_key_string(raw_key)
                findings_status_map[normalized_key] = f.get("status", "PASS")
                findings_full_map[normalized_key] = f

                # Also index by non-type key portion for legacy matching
                parts = normalized_key.split("|", 1)
                if len(parts) > 1:
                    non_type_key = parts[1]
                    # Don't overwrite if exists (first match wins)
                    if non_type_key not in findings_full_map:
                        findings_full_map[non_type_key] = f

        # Helper to check if DB annotation was already an exception
        # DB annotations don't have status field - just check justification/review_status
        def was_previously_exception(fields: dict) -> bool:
            if not fields:
                return False
            raw_just = fields.get("justification")
            justification = str(raw_just).strip() if raw_just else ""
            raw_status = fields.get("review_status", "")
            has_exception_status = "Exception" in str(raw_status)
            return bool(justification) or has_exception_status

        # Helper to check if Excel annotation qualifies as NEW exception
        # Must be DISCREPANT (FAIL/WARN) to count as exception
        def is_new_exception(fields: dict) -> bool:
            if not fields:
                return False

            # First check status - must be FAIL/WARN to be exception
            row_status = str(fields.get("status", "")).upper()
            is_discrepant = row_status in ("FAIL", "WARN", "⏳", "⚠")

            # Also treat as discrepant if action_needed was explicitly set
            if fields.get("action_needed", False):
                is_discrepant = True

            # If row is passing, it's NOT an exception (just documentation)
            if not is_discrepant:
                return False

            # Now check if it has justification or Exception status
            raw_just = fields.get("justification")
            justification = str(raw_just).strip() if raw_just else ""
            raw_status = fields.get("review_status", "")
            has_exception_status = "Exception" in str(raw_status)
            return bool(justification) or has_exception_status

        # 1. Detect NEW and UPDATED exceptions
        for full_key, fields in new_annotations.items():
            # Check if this has justification OR review_status set to Exception
            # Notes/Purpose are DOCUMENTATION ONLY, not exception triggers
            raw_just = fields.get("justification")
            justification = str(raw_just).strip() if raw_just else ""

            raw_review_status = fields.get("review_status", "")
            review_status = str(raw_review_status).strip() if raw_review_status else ""
            has_exception_status = "Exception" in review_status

            # Need at least one of: justification text OR exception status
            if not justification and not has_exception_status:
                continue

            # CRITICAL: Check if this row is actually DISCREPANT (FAIL/WARN)
            # PASS rows with justification are documentation only, NOT exceptions
            # Use findings_status_map for accurate lookup (not Excel status column)

            # === UUID-AWARE KEY MATCHING ===
            # Extract entity_type from full_key first (always valid as type|uuid)
            parts = full_key.split("|", 1)
            # Default entity_type from key structure
            current_entity_type = parts[0] if len(parts) == 2 else ""

            # When annotations are keyed by UUID, use _legacy_entity_key for findings lookup
            if "_legacy_entity_key" in fields and fields["_legacy_entity_key"]:
                # Use stored legacy key (UUID-based annotations)
                finding_key = fields["_legacy_entity_key"]
                # entity_type = ""  <-- REMOVED clearing of entity_type, we use current_entity_type variable
            else:
                # Fallback: Extract entity_key from full_key (legacy format)
                # Note: This overwrites local variables, but main loop uses full_key
                et, finding_key = annotation_key_to_finding_key(full_key)
                if not current_entity_type:
                    current_entity_type = et

            normalized_finding_key = normalize_key_string(finding_key)

            # Look up actual status from findings (using normalized lowercase key)
            finding_status = findings_status_map.get(normalized_finding_key, "")

            # Log key matching for troubleshooting at debug level
            if justification or has_exception_status:
                logger.debug(
                    "Exception candidate: %s -> finding_key=%s, status=%s",
                    full_key,
                    normalized_finding_key,
                    finding_status or "NOT_FOUND",
                )

            # Fallback to annotation's status field if no finding match
            if not finding_status:
                row_status = str(fields.get("status", "")).upper()
                # Check Status column - must be FAIL/WARN to be discrepant
                is_discrepant = row_status in ("FAIL", "WARN", "EXCEPTION")
                # Also check for emoji indicators in status
                if "⏳" in row_status or "⚠" in row_status or "✗" in row_status:
                    is_discrepant = True
                # CRITICAL: For sheets without Status column (Sensitive Roles, Services),
                # the ONLY way to detect discrepancy is via action_needed field
                # which was set from the ⏳ icon in Action column during Excel read
                if fields.get("action_needed", False):
                    is_discrepant = True
                logger.info(
                    "Using Excel status fallback for %s: status=%s, action_needed=%s, discrepant=%s",
                    full_key,
                    row_status or "(empty)",
                    fields.get("action_needed", False),
                    is_discrepant,
                )
            else:
                is_discrepant = finding_status.upper() in ("FAIL", "WARN")

            # If row is passing and no explicit exception status, skip it
            # Keep the justification in DB as documentation for future reference
            if not is_discrepant:
                logger.debug(
                    "Skipping non-discrepant row with justification: %s (status=%s)",
                    full_key,
                    finding_status or "unknown",
                )
                continue

            # Check if this is NEW or CHANGED
            # FALLBACK MATCHING: If direct key lookup fails (UUID mismatch), try legacy key
            old_fields = old_annotations.get(full_key)
            if not old_fields:
                # Try finding by legacy key
                legacy_key_val = fields.get("_legacy_entity_key")
                if legacy_key_val:
                    # We need to find the old annotation that has this legacy key
                    # This is O(N) unless we build an index. Given N is usually small (<1000), iteration is fine.
                    # Or build index at start of method. Let's build index for performance.
                    pass

            # Optimization: Build legacy map for old_annotations once at start of method?
            # Doing it locally here for minimal diff, but let's do it right.
            # Actually, let's use a helper function or build it at top of loop if needed?
            # No, let's just iterate if missing.

            if not old_fields and "_legacy_entity_key" in fields:
                target_legacy = fields["_legacy_entity_key"]
                # Search old_annotations for matching legacy key (filtered by entity type)
                for old_k, old_v in old_annotations.items():
                    if (
                        old_k.startswith(f"{current_entity_type}|")
                        and old_v.get("_legacy_entity_key") == target_legacy
                    ):
                        old_fields = old_v
                        logger.debug(
                            "Matched old annotation by legacy fallback: %s",
                            target_legacy,
                        )
                        break

            old_fields = old_fields or {}

            old_raw = old_fields.get("justification")
            old_justification = str(old_raw).strip() if old_raw else ""

            old_status = old_fields.get("review_status", "")
            was_exception = was_previously_exception(old_fields)

            # Detect change: new/updated justification or new Exception status
            # OR status changed (e.g. Exception -> Needs Review)
            status_changed = review_status != old_status
            justification_changed = justification != old_justification

            if justification_changed or status_changed:
                parts = full_key.split("|", 1)
                if len(parts) == 2:
                    entity_type, entity_key = parts

                    if (
                        not was_exception
                        and not has_exception_status
                        and not justification
                    ):
                        # Nothing interesting (empty -> empty or irrelevant change)
                        continue

                    if not was_exception:
                        change_type = "added"
                    elif justification_changed:
                        change_type = "updated"
                    elif status_changed:
                        change_type = "updated"  # Status change counts as update
                    else:
                        change_type = "updated"

                    # Add detailed note about what changed
                    change_note = ""
                    if status_changed:
                        change_note += f"Status: '{old_status}' -> '{review_status}'. "
                    if justification_changed:
                        change_note += "Justification updated."

                    # CRITICAL FIX: Get semantic entity_key, server, instance from finding
                    # NOT from the UUID-based annotation key
                    semantic_entity_key = entity_key  # Fallback
                    server_name = ""
                    instance_name = ""

                    # Use _legacy_entity_key for semantic key (it's server|instance|entity)
                    legacy_key = fields.get("_legacy_entity_key", "")
                    if legacy_key:
                        semantic_entity_key = legacy_key
                        # Parse server/instance from legacy key format: server|instance|entity
                        legacy_parts = legacy_key.split("|")
                        if len(legacy_parts) >= 2:
                            server_name = legacy_parts[0]
                            instance_name = legacy_parts[1]

                    # Try to get server/instance from the actual finding
                    normalized_legacy = (
                        normalize_key_string(legacy_key) if legacy_key else ""
                    )
                    finding_data = None

                    # First try full key lookup
                    full_finding_key = (
                        f"{entity_type}|{normalized_legacy}"
                        if normalized_legacy
                        else ""
                    )
                    if full_finding_key and full_finding_key in findings_full_map:
                        finding_data = findings_full_map[full_finding_key]
                    # Then try non-type key lookup
                    elif normalized_legacy and normalized_legacy in findings_full_map:
                        finding_data = findings_full_map[normalized_legacy]

                    if finding_data:
                        # Extract proper semantic entity_key and server/instance
                        semantic_entity_key = finding_data.get(
                            "entity_key", semantic_entity_key
                        )
                        server_name = (
                            finding_data.get("server_name")
                            or finding_data.get("server")
                            or server_name
                        )
                        instance_name = (
                            finding_data.get("instance_name")
                            or finding_data.get("instance")
                            or instance_name
                        )

                        # Parse from finding entity_key if still missing
                        if not server_name or not instance_name:
                            fk_parts = semantic_entity_key.split("|")
                            if len(fk_parts) >= 3:
                                # Format: type|server|instance|entity
                                server_name = server_name or fk_parts[1]
                                instance_name = instance_name or fk_parts[2]

                    exceptions.append(
                        {
                            "full_key": full_key,
                            "entity_type": entity_type,
                            "entity_key": semantic_entity_key,  # Use semantic key, not UUID
                            "justification": justification,
                            "is_new": not was_exception,
                            "change_type": change_type,
                            "note": change_note.strip(),
                            "server": server_name,
                            "instance": instance_name,
                        }
                    )
                    logger.info(
                        "Exception %s: %s - %s (%s)",
                        change_type,
                        entity_type,
                        entity_key[:50],
                        change_note,
                    )

        # 2. Detect REMOVED exceptions
        # A "removed exception" is when:
        #   - Row was discrepant (FAIL/WARN) with justification (was exception)
        #   - Row is STILL discrepant but user CLEARED the justification
        # NOT a "removed exception" if:
        #   - Row became PASS (that's a FIX, not removal)
        #   - Row was never discrepant (was just documentation)
        for full_key, old_fields in old_annotations.items():
            if not was_previously_exception(old_fields):
                continue  # Was not marked as exception before

            # Look up current finding status
            parts = full_key.split("|", 1)
            entity_key_for_lookup = parts[1] if len(parts) == 2 else full_key

            # CRITICAL: Also try _legacy_entity_key for UUID-based annotations
            if "_legacy_entity_key" in old_fields and old_fields["_legacy_entity_key"]:
                entity_key_for_lookup = old_fields["_legacy_entity_key"]

            # CRITICAL: Normalize to lowercase for case-insensitive matching
            normalized_lookup_key = normalize_key_string(entity_key_for_lookup)
            current_status = findings_status_map.get(normalized_lookup_key, "")

            # If row is now PASS, this is a FIX - not a "removed exception"
            # If status is unknown/missing (not in current findings), assume FIX or GONE
            # The exception documentation becomes historical note
            if not current_status or current_status.upper() == "PASS":
                logger.debug(
                    "Row became PASS/Gone (%s), exception docs now historical: %s",
                    current_status,
                    entity_key_for_lookup[:50],
                )
                continue

            # Row is still discrepant - check if exception was removed
            # Exception removal happens when:
            # 1. User clears justification AND status
            # 2. User changes Review Status FROM "Exception" to something else (Needs Review, etc.)

            # CRITICAL: Try to find matching new annotation with fallback key matching
            new_fields = new_annotations.get(full_key)
            if not new_fields:
                # Try matching by _legacy_entity_key
                old_legacy_key = old_fields.get("_legacy_entity_key")
                if old_legacy_key:
                    for new_key, nf in new_annotations.items():
                        if nf.get("_legacy_entity_key") == old_legacy_key:
                            new_fields = nf
                            break

            # If still no matching new annotation found, this entity may have been removed
            # from Excel entirely - that's NOT an "exception removed", it's just gone
            if not new_fields:
                logger.debug(
                    "No matching new annotation for %s - entity may have been removed",
                    full_key[:60],
                )
                continue

            new_just = new_fields.get("justification", "")
            new_status = new_fields.get("review_status", "")
            has_new_just = new_just and str(new_just).strip()
            has_new_exception = "Exception" in str(new_status)

            # Check if OLD status was Exception
            old_status = old_fields.get("review_status", "")
            was_exception_status = "Exception" in str(old_status)

            # Case 1: Status changed FROM Exception to something else
            status_reverted = was_exception_status and not has_new_exception

            # Case 2: Both justification and exception status cleared
            both_cleared = not has_new_just and not has_new_exception

            if status_reverted or both_cleared:
                if len(parts) == 2:
                    entity_type, entity_key = parts
                    old_just = old_fields.get("justification", "")

                    exceptions.append(
                        {
                            "full_key": full_key,
                            "entity_type": entity_type,
                            "entity_key": entity_key,
                            "justification": str(new_just).strip() if new_just else "",
                            "old_justification": (
                                str(old_just).strip() if old_just else ""
                            ),
                            "is_new": False,
                            "change_type": "removed",
                            "note": (
                                f"Status changed: '{old_status}' -> '{new_status}'"
                                if status_reverted
                                else "Cleared"
                            ),
                        }
                    )
                    logger.info(
                        "Exception removed: %s - %s (was: %s, now: %s)",
                        entity_type,
                        entity_key[:50],
                        old_status,
                        new_status or "(empty)",
                    )

        return exceptions
