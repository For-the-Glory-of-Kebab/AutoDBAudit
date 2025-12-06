"""
Excel report generation using openpyxl.

Phase 1: Minimal instance inventory report.
Future phases will add requirement results, security matrix, charts, etc.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from autodbaudit.domain.models import AuditRun, Server, Instance

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def write_instance_inventory(
    instances: list[tuple[Server, Instance]],
    audit_run: AuditRun,
    output_path: Path
) -> Path:
    """
    Write instance inventory to Excel file.
    
    Creates a simple report showing all audited SQL Server instances
    with their version information.
    
    Args:
        instances: List of (Server, Instance) tuples from HistoryStore
        audit_run: The audit run metadata
        output_path: Path to write Excel file
        
    Returns:
        Path to the created Excel file
    """
    logger.info("Generating instance inventory report: %s", output_path)
    
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Create InstanceInventory sheet
    ws = wb.create_sheet("InstanceInventory")
    
    # Define columns
    columns = [
        ("Server", 20),
        ("Instance", 20),
        ("Version", 18),
        ("Major Ver", 10),
        ("Edition", 35),
        ("Patch Level", 12),
        ("IP Address", 15),
    ]
    
    # Write header row
    for col_idx, (col_name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Write data rows
    for row_idx, (server, instance) in enumerate(instances, start=2):
        instance_display = instance.instance_name or "(default)"
        version_display = _format_version_display(instance.version_major)
        
        row_data = [
            server.hostname,
            instance_display,
            instance.version,
            version_display,
            instance.edition,
            instance.product_level,
            server.ip_address or "",
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
    
    # Add metadata sheet
    _add_metadata_sheet(wb, audit_run, len(instances))
    
    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Save workbook
    wb.save(output_path)
    logger.info("Instance inventory report saved: %s (%d instances)", output_path, len(instances))
    
    return output_path


def _format_version_display(version_major: int) -> str:
    """Convert major version number to readable SQL Server version."""
    version_map = {
        10: "2008/2008R2",
        11: "2012",
        12: "2014",
        13: "2016",
        14: "2017",
        15: "2019",
        16: "2022",
    }
    return version_map.get(version_major, f"Unknown ({version_major})")


def _add_metadata_sheet(wb: Workbook, audit_run: AuditRun, instance_count: int) -> None:
    """Add a metadata sheet with audit run information."""
    ws = wb.create_sheet("AuditInfo")
    
    # Metadata rows
    metadata = [
        ("Audit Run ID", audit_run.id),
        ("Organization", audit_run.organization or "(not specified)"),
        ("Started At", audit_run.started_at.isoformat() if audit_run.started_at else ""),
        ("Ended At", audit_run.ended_at.isoformat() if audit_run.ended_at else ""),
        ("Status", audit_run.status),
        ("Instances Audited", instance_count),
        ("Report Generated", datetime.now().isoformat()),
    ]
    
    for row_idx, (label, value) in enumerate(metadata, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
