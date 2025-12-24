"""
Excel styling configuration and utilities.

Provides consistent styling across all audit report sheets:
- Color palette
- Icons (Unicode with text fallbacks)
- Font definitions
- Cell style presets
- Formatting helpers
"""

from __future__ import annotations

from dataclasses import dataclass
from enum import Enum

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


# ============================================================================
# Color Palette
# ============================================================================


class Colors:
    """Audit report color palette (hex codes without #)."""

    # Headers - Modern Deep Slate Blue
    HEADER_BG = "203764"  # Darker, more premium Navy
    HEADER_TEXT = "FFFFFF"  # White
    SUBHEADER_BG = "4472C4"  # Steel blue (kept for hierarchy)

    # Status colors - Vibrant but professional
    PASS_BG = "C6EFCE"  # Vibrant Green
    PASS_TEXT = "006100"  # Dark Green
    FAIL_BG = "FFC7CE"  # Vibrant Red
    FAIL_TEXT = "9C0006"  # Dark Red
    WARN_BG = "FFEB9C"  # Vibrant Yellow
    WARN_TEXT = "9C5700"  # Dark Orange/Brown
    EXCEPTION_BG = "FFEB9C"
    EXCEPTION_TEXT = "9C5700"
    INFO_BG = "BDD7EE"  # Blue
    INFO_TEXT = "1F4E79"  # Dark Blue

    # Special
    MANUAL_BG = "F2F2F2"  # Light gray
    DISABLED_BG = "D9D9D9"  # Medium gray
    CRITICAL_BG = "FF0000"  # Bright Red
    CRITICAL_TEXT = "FFFFFF"
    NEW_BG = "FFFF00"  # Bright Yellow
    NEW_TEXT = "000000"
    CHANGED_BG = "FFC000"  # Orange
    CHANGED_TEXT = "000000"

    # Action needed indicator
    ACTION_BG = "FFC000"  # Orange

    # Row alternation
    SERVER_ALT_BG = "DEEBF7"  # Slightly more saturated blue tint
    ROW_ALT_BG = "FAFAFA"  # Almost white

    # Server group color rotation - Premium Macaron Palette
    SERVER_1_MAIN = "A9D18E"  # Fresh Green
    SERVER_1_LIGHT = "E2F0D9"

    SERVER_2_MAIN = "F4B084"  # Soft Orange
    SERVER_2_LIGHT = "FCE4D6"

    SERVER_3_MAIN = "BF9000"  # Gold Accent
    SERVER_3_LIGHT = "FFF2CC"

    SERVER_4_MAIN = "9DC3E6"  # Sky Blue
    SERVER_4_LIGHT = "DDEBF7"


# ... (Server Group Colors Tuple remains same logic, just using new constants)
SERVER_GROUP_COLORS = [
    (Colors.SERVER_1_MAIN, Colors.SERVER_1_LIGHT),
    (Colors.SERVER_2_MAIN, Colors.SERVER_2_LIGHT),
    (Colors.SERVER_3_MAIN, Colors.SERVER_3_LIGHT),
    (Colors.SERVER_4_MAIN, Colors.SERVER_4_LIGHT),
]


# ============================================================================
# Icons (Unchanged)
# ============================================================================
class Icons:
    """
    Unicode icons with text fallbacks.
    """

    PASS = "✓"
    PASS_FALLBACK = "[PASS]"
    FAIL = "✗"
    FAIL_FALLBACK = "[FAIL]"
    WARN = "⚠"
    WARN_FALLBACK = "[WARN]"
    EXCEPTION = "⚡"  # Updated to bolt
    EXCEPTION_FALLBACK = "[EXCPT]"
    INFO = "ℹ"
    INFO_FALLBACK = "[INFO]"
    RUNNING = "▶"  # Updated to play arrow
    RUNNING_FALLBACK = "[RUN]"
    STOPPED = "⏹"  # Updated to stop square
    STOPPED_FALLBACK = "[STOP]"
    PENDING = "⏳"
    PENDING_FALLBACK = "[PEND]"
    LOCKED = "🔒"
    LOCKED_FALLBACK = "[LOCK]"
    UNLOCKED = "🔓"
    UNLOCKED_FALLBACK = "[RISK]"
    NEW = "★"
    NEW_FALLBACK = "[NEW]"
    CHANGED = "Δ"
    CHANGED_FALLBACK = "[CHG]"

    @classmethod
    def get(cls, icon_name: str, use_fallback: bool = False) -> str:
        icon = getattr(cls, icon_name.upper(), "?")
        if use_fallback:
            return getattr(cls, f"{icon_name.upper()}_FALLBACK", icon)
        return icon


# ============================================================================
# Fonts - Premium "Segoe UI"
# ============================================================================


class Fonts:
    """Font definitions for the report."""

    # Main fonts
    TITLE = Font(
        name="Segoe UI", size=18, bold=True, color=Colors.HEADER_BG
    )  # Larger Title
    HEADER = Font(name="Segoe UI", size=11, bold=True, color=Colors.HEADER_TEXT)
    SUBHEADER = Font(name="Segoe UI", size=11, bold=True, color=Colors.HEADER_TEXT)
    DATA = Font(name="Segoe UI", size=10)
    DATA_BOLD = Font(name="Segoe UI", size=10, bold=True)
    NOTES = Font(name="Segoe UI", size=9, italic=True, color="666666")
    MONOSPACE = Font(name="Consolas", size=10)

    # Status fonts
    PASS = Font(name="Segoe UI", size=10, bold=True, color=Colors.PASS_TEXT)
    FAIL = Font(name="Segoe UI", size=10, bold=True, color=Colors.FAIL_TEXT)
    WARN = Font(name="Segoe UI", size=10, bold=True, color=Colors.WARN_TEXT)
    CRITICAL = Font(name="Segoe UI", size=10, bold=True, color=Colors.CRITICAL_TEXT)
    INFO = Font(name="Segoe UI", size=10, bold=True, color=Colors.INFO_TEXT)
    NEW = Font(name="Segoe UI", size=10, bold=True, color=Colors.NEW_TEXT)


# ============================================================================
# Fills (Backgrounds)
# ============================================================================


class Fills:
    """Background fill patterns."""

    HEADER = PatternFill(
        start_color=Colors.HEADER_BG, end_color=Colors.HEADER_BG, fill_type="solid"
    )
    SUBHEADER = PatternFill(
        start_color=Colors.SUBHEADER_BG,
        end_color=Colors.SUBHEADER_BG,
        fill_type="solid",
    )

    PASS = PatternFill(
        start_color=Colors.PASS_BG, end_color=Colors.PASS_BG, fill_type="solid"
    )
    FAIL = PatternFill(
        start_color=Colors.FAIL_BG, end_color=Colors.FAIL_BG, fill_type="solid"
    )
    WARN = PatternFill(
        start_color=Colors.WARN_BG, end_color=Colors.WARN_BG, fill_type="solid"
    )
    EXCEPTION = PatternFill(
        start_color=Colors.EXCEPTION_BG,
        end_color=Colors.EXCEPTION_BG,
        fill_type="solid",
    )
    INFO = PatternFill(
        start_color=Colors.INFO_BG, end_color=Colors.INFO_BG, fill_type="solid"
    )

    MANUAL = PatternFill(
        start_color=Colors.MANUAL_BG, end_color=Colors.MANUAL_BG, fill_type="solid"
    )
    DISABLED = PatternFill(
        start_color=Colors.DISABLED_BG, end_color=Colors.DISABLED_BG, fill_type="solid"
    )
    CRITICAL = PatternFill(
        start_color=Colors.CRITICAL_BG, end_color=Colors.CRITICAL_BG, fill_type="solid"
    )

    NEW = PatternFill(
        start_color=Colors.NEW_BG, end_color=Colors.NEW_BG, fill_type="solid"
    )
    CHANGED = PatternFill(
        start_color=Colors.CHANGED_BG, end_color=Colors.CHANGED_BG, fill_type="solid"
    )
    ACTION = PatternFill(
        start_color=Colors.ACTION_BG, end_color=Colors.ACTION_BG, fill_type="solid"
    )

    SERVER_ALT = PatternFill(
        start_color=Colors.SERVER_ALT_BG,
        end_color=Colors.SERVER_ALT_BG,
        fill_type="solid",
    )
    ROW_ALT = PatternFill(
        start_color=Colors.ROW_ALT_BG, end_color=Colors.ROW_ALT_BG, fill_type="solid"
    )


# ============================================================================
# Borders
# ============================================================================


class Borders:
    """Border styles."""

    THIN = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    HEADER = Border(
        left=Side(style="thin", color="1F4E79"),
        right=Side(style="thin", color="1F4E79"),
        top=Side(style="thin", color="1F4E79"),
        bottom=Side(style="medium", color="1F4E79"),
    )

    SERVER_BOTTOM = Border(bottom=Side(style="medium", color="1F4E79"))

    NONE = Border()


# ============================================================================
# Alignments
# ============================================================================


class Alignments:
    """Text alignment definitions."""

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
    CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
    LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)


# ============================================================================
# Status Enum
# ============================================================================


class Status(Enum):
    """Assessment status values."""

    PASS = "pass"
    FAIL = "fail"
    WARN = "warn"
    EXCEPTION = "exception"
    INFO = "info"
    NEW = "new"
    CHANGED = "changed"
    CRITICAL = "critical"


# ============================================================================
# Column Definition
# ============================================================================


@dataclass
class ColumnDef:
    """
    Column definition for a report sheet.

    Attributes:
        name: Column header text
        width: Column width in characters
        alignment: Text alignment
        is_manual: If True, this is a user-input column (gray background)
        is_monospace: If True, use monospace font
        is_status: If True, apply status-based formatting
    """

    name: str
    width: int = 12
    alignment: Alignment = Alignments.LEFT
    is_manual: bool = False
    is_monospace: bool = False
    is_status: bool = False


# ============================================================================
# Helper Functions
# ============================================================================


def apply_header_row(ws: Worksheet, columns: list[ColumnDef], row: int = 1) -> None:
    """
    Apply header styling to a row.

    Args:
        ws: Worksheet
        columns: List of column definitions
        row: Row number (1-indexed)
    """
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_def.name
        cell.font = Fonts.HEADER
        cell.fill = Fills.HEADER
        cell.alignment = Alignments.CENTER_WRAP
        cell.border = Borders.HEADER

        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def.width


def apply_status_styling(cell, status: str | Status) -> None:
    """
    Apply status-based styling to a cell.

    Args:
        cell: openpyxl cell
        status: Status value (string or Status enum)
    """
    if isinstance(status, Status):
        status = status.value

    status = status.lower() if status else ""

    if status == "pass":
        cell.font = Fonts.PASS
        cell.fill = Fills.PASS
        cell.value = f"{Icons.PASS} Pass"
    elif status == "fail":
        cell.font = Fonts.FAIL
        cell.fill = Fills.FAIL
        cell.value = f"{Icons.FAIL} Fail"
    elif status == "warn":
        cell.font = Fonts.WARN
        cell.fill = Fills.WARN
        cell.value = f"{Icons.WARN} Warning"
    elif status == "exception":
        cell.font = Fonts.WARN
        cell.fill = Fills.EXCEPTION
        cell.value = f"{Icons.EXCEPTION} Exception"
    elif status == "critical":
        cell.font = Fonts.CRITICAL
        cell.fill = Fills.CRITICAL
        cell.value = f"{Icons.FAIL} Critical"
    elif status == "new":
        cell.font = Fonts.NEW
        cell.fill = Fills.NEW
        cell.value = f"{Icons.NEW} New"
    elif status == "changed":
        cell.font = Fonts.DATA
        cell.fill = Fills.CHANGED
        cell.value = f"{Icons.CHANGED} Changed"


def apply_boolean_styling(cell, value: bool | int | None, invert: bool = False) -> None:
    """
    Apply styling for boolean values (checkmarks/crosses).

    Args:
        cell: openpyxl cell
        value: Boolean value (True/False/1/0/None)
        invert: If True, True=bad (red), False=good (green)
    """
    if value is None:
        cell.value = ""
        return

    is_true = bool(value)

    if invert:
        # True = bad, False = good
        if is_true:
            cell.value = Icons.FAIL
            cell.font = Fonts.FAIL
            cell.fill = Fills.FAIL
        else:
            cell.value = Icons.PASS
            cell.font = Fonts.PASS
            cell.fill = Fills.PASS
    else:
        # True = good, False = bad
        if is_true:
            cell.value = Icons.PASS
            cell.font = Fonts.PASS
            cell.fill = Fills.PASS
        else:
            cell.value = Icons.FAIL
            cell.font = Fonts.FAIL
            cell.fill = Fills.FAIL

    cell.alignment = Alignments.CENTER


def apply_service_status_styling(cell, status: str) -> None:
    """Apply styling for service status (Running/Stopped)."""
    status_lower = (status or "").lower()

    if status_lower == "running":
        cell.value = f"{Icons.RUNNING} Running"
        cell.font = Fonts.PASS
    elif status_lower == "stopped":
        cell.value = f"{Icons.STOPPED} Stopped"
        cell.font = Font(
            name="Calibri", size=10, color=Colors.DISABLED_BG.replace("BF", "00")
        )
    else:
        cell.value = status or ""


def freeze_panes(ws: Worksheet, row: int = 2, col: int = 1) -> None:
    """
    Freeze panes in a worksheet.

    Args:
        ws: Worksheet
        row: First unfrozen row (freeze rows above)
        col: First unfrozen column (freeze columns to the left)
    """
    ws.freeze_panes = ws.cell(row=row, column=col)


def add_autofilter(
    ws: Worksheet, columns: list[ColumnDef], header_row: int = 1
) -> None:
    """
    Add autofilter to header row.

    Args:
        ws: Worksheet
        columns: Column definitions
        header_row: Header row number
    """
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"


def merge_server_cells(
    ws: Worksheet,
    server_col: int,
    start_row: int,
    end_row: int,
    server_name: str,
    is_alt: bool = False,
) -> None:
    """
    Merge cells vertically for server grouping.

    Args:
        ws: Worksheet
        server_col: Column index for server name
        start_row: First row to merge
        end_row: Last row to merge
        server_name: Server name to display
        is_alt: If True, use alternate background color
    """
    if end_row > start_row:
        ws.merge_cells(
            start_row=start_row,
            start_column=server_col,
            end_row=end_row,
            end_column=server_col,
        )

    cell = ws.cell(row=start_row, column=server_col)
    cell.value = server_name
    cell.font = Fonts.DATA_BOLD
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Borders.THIN

    if is_alt:
        cell.fill = Fills.SERVER_ALT
