"""
Instances Sheet Module.

Handles the Instances worksheet for SQL Server instance properties.
Visual grouping with rotating colors for different servers.


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    apply_boolean_styling,
    merge_server_cells,
    SERVER_GROUP_COLORS,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    get_sql_year,
    LAST_REVIEWED_COLUMN,
    STATUS_COLUMN,
    ACTION_COLUMN,
    apply_action_needed_styling,
)


__all__ = ["InstanceSheetMixin", "INSTANCE_CONFIG"]


INSTANCE_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Config Name", 22, Alignments.LEFT),  # Column C
    ColumnDef("Server", 18, Alignments.LEFT),  # Column C - merged
    ColumnDef("Instance", 12, Alignments.LEFT),  # Column D - merged
    ColumnDef("Machine Name", 16, Alignments.LEFT),  # From SERVERPROPERTY
    ColumnDef("IP Address", 16, Alignments.LEFT),  # From dm_exec_connections
    ColumnDef("Version", 12, Alignments.LEFT),
    ColumnDef("Build", 14, Alignments.LEFT),
    ColumnDef("Version Status", 12, Alignments.CENTER),  # NEW: PASS/WARN
    ColumnDef("SQL Year", 8, Alignments.CENTER),
    ColumnDef("Edition", 20, Alignments.LEFT),
    ColumnDef("Clustered", 8, Alignments.CENTER),
    ColumnDef("HADR", 6, Alignments.CENTER),
    ColumnDef("OS", 18, Alignments.LEFT),
    ColumnDef("CPU", 4, Alignments.CENTER),
    ColumnDef("RAM", 5, Alignments.CENTER),
    STATUS_COLUMN,
    ColumnDef("Justification", 30, Alignments.LEFT, is_manual=True),
    ColumnDef("Notes", 24, Alignments.LEFT_WRAP, is_manual=True),
    LAST_REVIEWED_COLUMN,
)

INSTANCE_CONFIG = SheetConfig(name="Instances", columns=INSTANCE_COLUMNS)


class InstanceSheetMixin(BaseSheetMixin):
    """Mixin for Instances sheet functionality."""

    _instance_sheet = None
    _instance_last_server: str = ""
    _instance_last_instance: str = ""  # Value to display in merged cell
    _instance_group_mixed: bool = (
        False  # Track if multiple instances in same server group
    )
    _instance_server_start_row: int = 2
    _instance_server_idx: int = 0

    def add_instance(
        self,
        config_name: str,  # NEW: display_name from config
        server_name: str,  # Connection target (localhost, localhost,1434)
        instance_name: str,
        machine_name: str,  # NEW: from SERVERPROPERTY('MachineName')
        ip_address: str,  # From dm_exec_connections
        tcp_port: int | None,
        version: str,
        version_major: int,
        edition: str,
        product_level: str,
        is_clustered: bool = False,
        is_hadr: bool = False,
        os_info: str = "",
        cpu_count: int | None = None,
        memory_gb: int | None = None,
        cu_level: str = "",
        build_number: int | None = None,
        version_status: str = "PASS",  # NEW: PASS, WARN, or FAIL
        version_status_note: str = "",  # NEW: tooltip/description
    ) -> None:
        """Add an instance row to the Instances sheet."""
        from autodbaudit.infrastructure.excel_styles import apply_status_styling

        if self._instance_sheet is None:
            self._instance_sheet = self._ensure_sheet_with_uuid(INSTANCE_CONFIG)
            self._instance_last_server = ""
            self._instance_last_instance = ""
            self._instance_group_mixed = False
            self._instance_server_start_row = 2
            self._instance_server_idx = 0
            self._add_instance_dropdowns()

        ws = self._instance_sheet
        current_row = self._row_counters[INSTANCE_CONFIG.name]

        display_instance = instance_name or "(Default)"

        # Check if server changed
        if server_name != self._instance_last_server:
            if self._instance_last_server:
                self._merge_instance_server(ws)
                self._instance_server_idx += 1

            self._instance_server_start_row = current_row
            self._instance_last_server = server_name
            self._instance_last_instance = display_instance
            self._instance_group_mixed = False
        else:
            # Same server, check if instance is different
            if display_instance != self._instance_last_instance:
                self._instance_group_mixed = True

        _, color_light = SERVER_GROUP_COLORS[
            self._instance_server_idx % len(SERVER_GROUP_COLORS)
        ]

        # Build info string
        build_parts = [product_level]
        if cu_level:
            build_parts.append(cu_level)
        if build_number:
            build_parts.append(f"({build_number})")
        build_info = " ".join(build_parts)

        # Format IP with port
        ip_display = ip_address or ""
        if ip_display and tcp_port:
            ip_display = f"{ip_address}:{tcp_port}"
        elif tcp_port and not ip_display:
            ip_display = f":{tcp_port}"  # Just port if no IP

        data = [
            None,  # Col 2: Action
            config_name,  # Col 3: Config Name
            server_name,  # Col 4: Server
            display_instance,  # Col 5: Instance
            machine_name or "",  # Col 6: Machine Name
            ip_display,  # Col 7: IP Address
            version,  # Col 8: Version
            build_info,  # Col 9: Build
            None,  # Col 10: Version Status (styled separately)
            get_sql_year(version_major),  # Col 11: SQL Year
            edition,  # Col 12: Edition
            None,  # Col 13: Clustered (styled separately)
            None,  # Col 14: HADR (styled separately)
            os_info or "",  # Col 15: OS
            str(cpu_count) if cpu_count else "",  # Col 16: CPU
            str(memory_gb) if memory_gb else "",  # Col 17: RAM
            "",  # Col 18: Notes
            "",  # Col 19: Last Reviewed
        ]

        row, row_uuid = self._write_row_with_uuid(ws, INSTANCE_CONFIG, data)

        # Apply action indicator (column 2)
        # Check version status (FAIL/WARN needs action)
        needs_action = version_status in ("FAIL", "WARN")
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_action)

        fill = PatternFill(
            start_color=color_light, end_color=color_light, fill_type="solid"
        )
        # All data columns start from 3 (Config Name)
        # Col indices shifted +1 due to Action column
        # Config=3, Server=4, Inst=5, Mach=6, IP=7, Ver=8, Build=9, Stat=10, Year=11, Ed=12, Clust=13, HADR=14, OS=15, CPU=16, RAM=17
        for col in [3, 4, 5, 6, 7, 8, 9, 11, 12, 15, 16, 17]:
            ws.cell(row=row, column=col).fill = fill

        # Apply version status styling (column 10)
        status_cell = ws.cell(row=row, column=10)
        apply_status_styling(status_cell, version_status)
        if version_status_note:
            from openpyxl.comments import Comment

            status_cell.comment = Comment(version_status_note, "AutoDBAudit")

        # Apply boolean styling for Clustered (col 13) and HADR (col 14)
        apply_boolean_styling(ws.cell(row=row, column=13), is_clustered)
        apply_boolean_styling(ws.cell(row=row, column=14), is_hadr)

    def _merge_instance_server(self, ws) -> None:
        """Merge Server and Instance cells for current server group."""
        current_row = self._row_counters[INSTANCE_CONFIG.name]
        if current_row > self._instance_server_start_row:
            color_main, _ = SERVER_GROUP_COLORS[
                self._instance_server_idx % len(SERVER_GROUP_COLORS)
            ]

            # Merge 1: Server Name (Always merge for the group)
            # Server column is now 4 (A=UUID, B=Action, C=Config, D=Server)
            merge_server_cells(
                ws,
                server_col=4,
                start_row=self._instance_server_start_row,
                end_row=current_row - 1,
                server_name=self._instance_last_server,
                is_alt=True,
            )
            merged_cell = ws.cell(row=self._instance_server_start_row, column=4)
            merged_cell.fill = PatternFill(
                start_color=color_main, end_color=color_main, fill_type="solid"
            )

            # Merge 2: Instance Name (Only if not mixed)
            if not self._instance_group_mixed:
                # Instance column is now 5
                merge_server_cells(
                    ws,
                    server_col=5,
                    start_row=self._instance_server_start_row,
                    end_row=current_row - 1,
                    server_name=self._instance_last_instance,
                    is_alt=True,  # Use same style
                )
                merged_cell_inst = ws.cell(
                    row=self._instance_server_start_row, column=5
                )
                merged_cell_inst.fill = PatternFill(
                    start_color=color_main, end_color=color_main, fill_type="solid"
                )

    def _finalize_instances(self) -> None:
        """Finalize instances sheet - merge remaining server group."""
        if self._instance_sheet and self._instance_last_server:
            self._merge_instance_server(self._instance_sheet)
            self._finalize_sheet_with_uuid(self._instance_sheet)

    def _add_instance_dropdowns(self) -> None:
        """Add dropdown validations for boolean columns."""
        from autodbaudit.infrastructure.excel.base import add_dropdown_validation

        ws = self._instance_sheet
        # Clustered column (M) - column 13 (shifted +1 from L)
        add_dropdown_validation(ws, "M", ["✓", "✗"])
        # HADR column (N) - column 14 (shifted +1 from M)
        add_dropdown_validation(ws, "N", ["✓", "✗"])
