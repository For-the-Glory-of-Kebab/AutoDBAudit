"""
Server Logins Sheet Module.

Visual Features:
    - Server column merged with main color
    - Instance column merged with alternating shade within server
    - Enabled/Policy columns keep status colors (green/red/yellow)


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fonts,
    Fills,
    Icons,
    merge_server_cells,
    SERVER_GROUP_COLORS,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    LAST_REVIEWED_COLUMN,
    STATUS_COLUMN,
    ACTION_COLUMN,
    apply_action_needed_styling,
)


__all__ = ["LoginSheetMixin", "LOGIN_CONFIG"]


LOGIN_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Server", 16, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 14, Alignments.LEFT),
    ColumnDef("Login Name", 28, Alignments.LEFT),
    ColumnDef("Login Type", 18, Alignments.LEFT),
    ColumnDef("Enabled", 10, Alignments.CENTER),
    ColumnDef("Password Policy", 14, Alignments.CENTER),
    ColumnDef("Default Database", 18, Alignments.LEFT),
    STATUS_COLUMN,  # Review Status dropdown
    ColumnDef("Justification", 35, Alignments.CENTER_WRAP, is_manual=True),
    LAST_REVIEWED_COLUMN,  # Last Reviewed (was Last Revised)
    ColumnDef("Notes", 25, Alignments.CENTER_WRAP, is_manual=True),
)

LOGIN_CONFIG = SheetConfig(name="Server Logins", columns=LOGIN_COLUMNS)


class LoginSheetMixin(BaseSheetMixin):
    """Mixin for Server Logins sheet with server/instance grouping."""

    _login_sheet = None
    _login_last_server: str = ""
    _login_last_instance: str = ""
    _login_server_start_row: int = 2
    _login_instance_start_row: int = 2
    _login_server_idx: int = 0  # Color rotation for servers
    _login_instance_alt: bool = False  # Alternate shade for instances

    def add_login(
        self,
        server_name: str,
        instance_name: str,
        login_name: str,
        login_type: str,
        is_disabled: bool,
        pwd_policy: bool | None = None,
        default_db: str = "",
    ) -> None:
        """Add a login row with server/instance grouping."""
        # Lazy-initialize
        if self._login_sheet is None:
            self._login_sheet = self._ensure_sheet_with_uuid(LOGIN_CONFIG)
            self._login_last_server = ""
            self._login_last_instance = ""
            self._login_server_start_row = 2
            self._login_instance_start_row = 2
            self._login_server_idx = 0
            self._login_instance_alt = False
            self._add_login_dropdowns()

        ws = self._login_sheet
        current_row = self._row_counters[LOGIN_CONFIG.name]
        inst_display = instance_name or "(Default)"

        # Check if SERVER changed
        if server_name != self._login_last_server:
            # Finalize previous server group (merges both server & instance)
            if self._login_last_server:
                self._merge_login_groups(ws)
                self._login_server_idx += 1

            # Start new server
            self._login_server_start_row = current_row
            self._login_instance_start_row = current_row
            self._login_last_server = server_name
            self._login_last_instance = inst_display
            self._login_instance_alt = False

        # Check if INSTANCE changed (within same server)
        elif inst_display != self._login_last_instance:
            # Merge previous instance group
            self._merge_login_instance(ws)
            self._login_instance_start_row = current_row
            self._login_last_instance = inst_display
            self._login_instance_alt = not self._login_instance_alt  # Toggle shade

        # Get colors
        color_main, color_light = SERVER_GROUP_COLORS[
            self._login_server_idx % len(SERVER_GROUP_COLORS)
        ]
        # Use main color for even instances, light for odd
        row_color = color_main if self._login_instance_alt else color_light

        # Prepare row data
        is_enabled = not is_disabled
        data = [
            None,  # Action indicator (column A)
            server_name,
            inst_display,
            login_name,
            login_type,
            None,  # Enabled
            None,  # Password Policy
            default_db or "",
            "",  # Notes
            "",  # Last Revised
        ]

        row, row_uuid = self._write_row_with_uuid(ws, LOGIN_CONFIG, data)

        # Determine if this is a system/internal login (Q1 decision)
        # ##...## pattern logins are SQL Server internal accounts - exclude from discrepancy
        is_system_login = (
            login_name.startswith("##") and login_name.endswith("##")
        ) or login_name.lower() in {
            "sa",
            "nt authority\\system",
            "nt service\\mssqlserver",
            "nt service\\sqlserveragent",
        }

        # Determine if action is needed:
        # - System logins: never discrepant (Q1 decision)
        # - Disabled logins: COMPLIANT (disabled = not a risk)
        # - SQL logins without password policy: discrepant (pwd_policy=False)
        # - Windows logins: pwd_policy is None (N/A), not discrepant
        if is_system_login:
            needs_action = False
        else:
            # Only SQL logins with password policy disabled need action
            # Disabled logins WITH NO POLICY are still findings (defense in depth)
            needs_action = pwd_policy is not None and not pwd_policy
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_action)

        # Apply shade to informational columns (NOT Action column which stays white)
        # A=UUID, B=Action(2), C=Server(3), D=Instance(4), E=LoginName(5), F=LoginType(6), G=Enabled(7), H=Policy(8), I=DefaultDB(9)
        fill = PatternFill(
            start_color=row_color, end_color=row_color, fill_type="solid"
        )
        for col in [3, 4, 5, 6, 9]:  # Server=3, Instance=4, Name=5, Type=6, DefaultDB=9
            ws.cell(row=row, column=col).fill = fill

        # Style Enabled column with icon + color (column G = 7)
        enabled_cell = ws.cell(row=row, column=7)
        if is_enabled:
            enabled_cell.value = f"{Icons.PASS} Yes"
            enabled_cell.font = Fonts.PASS
            enabled_cell.fill = Fills.PASS
        else:
            enabled_cell.value = f"{Icons.FAIL} No"
            enabled_cell.font = Fonts.FAIL
            enabled_cell.fill = Fills.FAIL

        # Style Password Policy (column H = 8)
        policy_cell = ws.cell(row=row, column=8)
        if pwd_policy is None:
            policy_cell.value = "N/A"
            policy_cell.fill = fill
        elif pwd_policy:
            policy_cell.value = f"{Icons.PASS} Yes"
            policy_cell.font = Fonts.PASS
            policy_cell.fill = Fills.PASS
        else:
            policy_cell.value = f"{Icons.FAIL} No"
            policy_cell.font = Fonts.WARN
            policy_cell.fill = Fills.WARN

    def _merge_login_instance(self, ws) -> None:
        """Merge Instance cells for current instance group."""
        current_row = self._row_counters[LOGIN_CONFIG.name]
        if current_row > self._login_instance_start_row:
            merge_server_cells(
                ws,
                server_col=4,  # Instance column (A=UUID, B=Action, C=Server, D=Instance)
                start_row=self._login_instance_start_row,
                end_row=current_row - 1,
                server_name=self._login_last_instance,
                is_alt=self._login_instance_alt,
            )

    def _merge_login_groups(self, ws) -> None:
        """Merge both Server and Instance cells for current server group."""
        # First merge the last instance group
        self._merge_login_instance(ws)

        # Then merge the server group
        current_row = self._row_counters[LOGIN_CONFIG.name]
        if current_row > self._login_server_start_row:
            color_main, _ = SERVER_GROUP_COLORS[
                self._login_server_idx % len(SERVER_GROUP_COLORS)
            ]
            merge_server_cells(
                ws,
                server_col=3,  # Server column (A=UUID, B=Action, C=Server)
                start_row=self._login_server_start_row,
                end_row=current_row - 1,
                server_name=self._login_last_server,
                is_alt=True,
            )
            # Apply main color to merged server cell
            merged = ws.cell(row=self._login_server_start_row, column=3)
            merged.fill = PatternFill(
                start_color=color_main, end_color=color_main, fill_type="solid"
            )

    def _add_login_dropdowns(self) -> None:
        """Add dropdown validations for status columns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation,
            add_review_status_conditional_formatting,
            STATUS_VALUES,
        )

        ws = self._login_sheet
        # With UUID: A=UUID, B=Action, C=Server, D=Instance, E=Login, F=Type, G=Enabled, H=Policy, I=DefaultDB, J=Status
        add_dropdown_validation(ws, "G", ["✓ Yes", "✗ No"])  # Enabled
        add_dropdown_validation(ws, "H", ["✓ Yes", "✗ No", "N/A"])  # Password Policy
        add_dropdown_validation(ws, "J", STATUS_VALUES.all())  # Review Status
        add_review_status_conditional_formatting(ws, "J")

    def _finalize_logins(self) -> None:
        """Finalize login sheet - merge remaining groups."""
        if self._login_sheet and self._login_last_server:
            self._merge_login_groups(self._login_sheet)
            self._finalize_sheet_with_uuid(self._login_sheet)
