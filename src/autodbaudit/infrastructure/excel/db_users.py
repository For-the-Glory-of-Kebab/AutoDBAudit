"""
Database Users Sheet Module.

Handles the Database Users worksheet for per-database user security matrix.
Uses ServerGroupMixin for server/instance grouping.

Security logic:
- Orphaned users (no mapped login) = review needed, but system users are OK
- GUEST user with CONNECT permission = FAIL (per requirement 13)
- System users (dbo, guest, INFORMATION_SCHEMA, sys) = expected, not failures


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fills,
    Fonts,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    ACTION_COLUMN,
    STATUS_COLUMN,
    LAST_REVIEWED_COLUMN,
    apply_action_needed_styling,
)
from autodbaudit.infrastructure.excel.server_group import ServerGroupMixin


__all__ = ["DBUserSheetMixin", "DB_USER_CONFIG"]


# System/built-in users that are expected to exist
SYSTEM_USERS = frozenset(
    {
        "dbo",
        "guest",
        "INFORMATION_SCHEMA",
        "sys",
        "##MS_PolicyEventProcessingLogin##",
        "##MS_AgentSigningCertificate##",
    }
)


DB_USER_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Server", 18, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 14, Alignments.LEFT),  # Column D
    ColumnDef("Database", 18, Alignments.LEFT),
    ColumnDef("User Name", 22, Alignments.LEFT),
    ColumnDef("Type", 16, Alignments.LEFT),
    ColumnDef("Type", 16, Alignments.LEFT),
    ColumnDef("Login Status", 14, Alignments.CENTER),
    ColumnDef("Mapped Login", 22, Alignments.LEFT),
    ColumnDef("Compliant", 10, Alignments.CENTER),
    STATUS_COLUMN,  # Review Status dropdown
    ColumnDef("Justification", 35, Alignments.CENTER_WRAP, is_manual=True),
    LAST_REVIEWED_COLUMN,
    ColumnDef("Notes", 25, Alignments.CENTER_WRAP, is_manual=True),
)

DB_USER_CONFIG = SheetConfig(name="Database Users", columns=DB_USER_COLUMNS)


class DBUserSheetMixin(ServerGroupMixin, BaseSheetMixin):
    """Mixin for Database Users sheet with server/instance grouping."""

    _db_user_sheet = None

    def add_db_user(
        self,
        server_name: str,
        instance_name: str,
        database_name: str,
        user_name: str,
        user_type: str,
        mapped_login: str | None,
        is_orphaned: bool,
        has_connect: bool = True,
    ) -> None:
        """Add a database user row with security assessment.

        Args:
            server_name: Server name
            instance_name: Instance name
            database_name: Database containing the user
            user_name: Database user name
            user_type: User type (SQL_USER, WINDOWS_USER, etc.)
            mapped_login: Server login this user is mapped to (None if orphaned)
            is_orphaned: True if user has no matching server login
            has_connect: True if user has CONNECT permission (for GUEST check)
        """
        if self._db_user_sheet is None:
            self._db_user_sheet = self._ensure_sheet_with_uuid(DB_USER_CONFIG)
            self._init_grouping(self._db_user_sheet, DB_USER_CONFIG)
            self._add_db_user_dropdowns()

        ws = self._db_user_sheet

        # Track grouping and get row color
        row_color = self._track_group(
            server_name, instance_name, DB_USER_CONFIG.name, database_name
        )

        # Determine login status and compliance
        is_system_user = user_name in SYSTEM_USERS
        is_guest = user_name.lower() == "guest"

        # Login status determination
        if mapped_login:
            login_status = "✓ Mapped"
            login_color = Fills.PASS
        elif is_system_user:
            login_status = "🔧 System"
            login_color = PatternFill(
                start_color="E8EAF6", end_color="E8EAF6", fill_type="solid"
            )
        else:
            login_status = "⚠️ Orphaned"
            login_color = Fills.WARN

        # Compliance check:
        # - GUEST with CONNECT = FAIL (req 13: "Guest user should be disabled")
        # - Orphaned non-system user = WARN (needs review)
        # - Everything else = PASS
        if is_guest and has_connect:
            self._increment_issue()  # GUEST with access = critical issue
        elif is_orphaned and not is_system_user:
            self._increment_warn()  # Orphaned user needs review
        else:
            self._increment_pass()

        # Determine if this row needs action (for ⏳ indicator)
        needs_action = (is_guest and has_connect) or (
            is_orphaned and not is_system_user
        )

        data = [
            None,  # Action indicator (column B)
            server_name,  # Column C
            instance_name or "(Default)",
            database_name,
            user_name,
            user_type,
            None,  # Login Status - styled separately
            mapped_login or ("(system)" if is_system_user else "(none)"),
            None,  # Compliant - styled separately
            "",  # Justification
            "",  # Notes
        ]

        row, row_uuid = self._write_row_with_uuid(ws, DB_USER_CONFIG, data)

        # Apply action indicator (column 1)
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_action)

        # Apply row color to data columns (shifted +1 for action column)
        self._apply_row_color(row, row_color, data_cols=[3, 4, 5, 6, 7, 9], ws=ws)

        # Style User Type (Column 6)
        self._apply_user_type_styling(ws.cell(row=row, column=6), user_type)

        # Style Login Status column (A=UUID, B=Action, C=Server, D=Instance, E=Database, F=User, G=Type, H=LoginStatus, I=MappedLogin)
        # Login Status is now Column H = 8
        status_cell = ws.cell(row=row, column=8)
        status_cell.value = login_status
        status_cell.fill = login_color
        if "Orphaned" in login_status:
            status_cell.font = Fonts.WARN
        elif "Mapped" in login_status:
            status_cell.font = Fonts.PASS

        # Style Compliant column (column J = 10)
        compliant_cell = ws.cell(row=row, column=10)
        if is_guest and has_connect:
            compliant_cell.value = "❌ GUEST"
            compliant_cell.fill = Fills.FAIL
            compliant_cell.font = Fonts.FAIL
        elif is_orphaned and not is_system_user:
            compliant_cell.value = "⚠️ Review"
            compliant_cell.fill = Fills.WARN
            compliant_cell.font = Fonts.WARN
        else:
            compliant_cell.value = "✓"
            compliant_cell.fill = Fills.PASS
            compliant_cell.font = Fonts.PASS

    def _finalize_db_users(self) -> None:
        """Finalize db users sheet - merge remaining groups."""
        if self._db_user_sheet:
            self._finalize_grouping(DB_USER_CONFIG.name)
            self._finalize_sheet_with_uuid(self._db_user_sheet)

    def _add_db_user_dropdowns(self) -> None:
        """Add dropdown validations for status columns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation,
            add_review_status_conditional_formatting,
            STATUS_VALUES,
        )

        ws = self._db_user_sheet
        if ws is None:
            return

        # Login Status column (H) - column 8
        add_dropdown_validation(ws, "H", ["✓ Mapped", "🔧 System", "⚠️ Orphaned"])
        # Compliant column (J) - column 10
        add_dropdown_validation(ws, "J", ["✓", "⚠️ Review", "❌ GUEST"])
        # Review Status column (K) - column 11
        add_dropdown_validation(ws, "K", STATUS_VALUES.all())
        add_review_status_conditional_formatting(ws, "K")

        # Type column (F) - column 6
        add_dropdown_validation(
            ws,
            "F",
            ["SQL_USER", "WINDOWS_USER", "WINDOWS_GROUP", "CERTIFICATE_MAPPED_USER"],
        )

    def _apply_user_type_styling(self, cell, user_type: str) -> None:
        """Apply styling to user type cell."""
        val = (user_type or "").upper()
        if "SQL" in val:
            cell.font = Fonts.WARN  # Orange-ish for SQL auth
        elif "WINDOWS" in val:
            cell.font = Fonts.PASS  # Green-ish for Windows auth
