"""
Services Sheet Module.

Handles the Services worksheet for SQL Server services audit.
Groups by Server with color rotation.

Discrepancy Logic:
- Essential services (Database Engine, SQL Agent): Only check account compliance
- Non-essential services (Browser, Integration, Analysis, etc.): Need justification


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fonts,
    Fills,
    apply_boolean_styling,
    merge_server_cells,
    SERVER_GROUP_COLORS,
)
from autodbaudit.infrastructure.excel.server_group import ServerGroupMixin
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    ACTION_COLUMN,
    STATUS_COLUMN,
    LAST_REVIEWED_COLUMN,
    apply_action_needed_styling,
)


__all__ = ["ServiceSheetMixin", "SERVICE_CONFIG"]


SERVICE_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Server", 16, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 14, Alignments.LEFT),  # Column D
    ColumnDef("Service Name", 40, Alignments.LEFT),
    ColumnDef("Type", 18, Alignments.LEFT),
    ColumnDef("Status", 12, Alignments.CENTER),
    ColumnDef("Startup", 12, Alignments.CENTER),
    ColumnDef("Service Account", 35, Alignments.LEFT),
    ColumnDef("Compliant", 10, Alignments.CENTER),
    STATUS_COLUMN,  # Review Status dropdown
    ColumnDef("Justification", 40, Alignments.CENTER_WRAP, is_manual=True),
    LAST_REVIEWED_COLUMN,
)

SERVICE_CONFIG = SheetConfig(name="Services", columns=SERVICE_COLUMNS)

# Essential services that are expected on a SQL Server (no justification needed)
ESSENTIAL_SERVICE_TYPES = frozenset(
    {
        "database engine",
        "sql agent",
        "sql server",
        "agent",
    }
)

# Non-essential services that need justification if enabled
# These are not strictly required for core SQL Server functionality
NON_ESSENTIAL_SERVICE_TYPES = frozenset(
    {
        "sql browser",
        "full-text search",
        "analysis services",
        "reporting services",
        "integration services",
        "launchpad",
        "launchpad (ml)",
        "vss writer",
        "ceip telemetry",
        "polybase",
        "ad helper",  # Active Directory helper
        "other",
        "other sql service",
    }
)

# Virtual/built-in accounts that are non-compliant per requirements
NON_COMPLIANT_ACCOUNTS = frozenset(
    {
        "localservice",
        "local service",
        "nt authority\\localservice",
        "networkservice",
        "network service",
        "nt authority\\networkservice",
        "localsystem",
        "local system",
        "nt authority\\system",
        "nt authority\\local service",
        "nt authority\\network service",
    }
)


class ServiceSheetMixin(ServerGroupMixin, BaseSheetMixin):
    """Mixin for Services sheet with server grouping."""

    _service_sheet = None
    _svc_last_server: str = ""
    _svc_last_instance: str = ""
    _svc_server_start_row: int = 2
    _svc_instance_start_row: int = 2
    _svc_server_idx: int = 0
    _svc_instance_alt: bool = False

    def add_service(
        self,
        server_name: str,
        instance_name: str,
        service_name: str,
        service_type: str,
        status: str,
        startup_type: str,
        service_account: str,
    ) -> None:
        """Add a SQL service row."""
        # Track grouping and get row color
        # USE MIXIN for grouping logic instead of custom duplication
        # Pass service_name as discriminator for 3rd level merging (so we don't merge all services into one block)
        if self._service_sheet is None:
            self._service_sheet = self._ensure_sheet_with_uuid(SERVICE_CONFIG)
            # Initialize mixin state
            self._init_grouping(
                self._service_sheet, SERVICE_CONFIG, database_col_idx=5
            )  # Service Name is col 5
            self._add_service_dropdowns()

        ws = self._service_sheet

        # Track group - Service Name is the "database" equivalent here for 3rd level differentiation
        # This prevents merging all services under one instance
        row_color = self._track_group(
            server_name, instance_name, SERVICE_CONFIG.name, service_name
        )

        # Check account compliance - service account should not be virtual
        account_compliant = True
        if service_account:
            acct_lower = service_account.lower().strip()
            if acct_lower in NON_COMPLIANT_ACCOUNTS:
                account_compliant = False

        # Check if this is a non-essential service that needs justification
        service_type_lower = (service_type or "").lower().strip()
        is_essential = any(s in service_type_lower for s in ESSENTIAL_SERVICE_TYPES)
        is_non_essential = (
            any(s in service_type_lower for s in NON_ESSENTIAL_SERVICE_TYPES)
            or not is_essential
        )

        # Check if this is specifically SQL Agent
        is_sql_agent = (
            "agent" in service_type_lower or "sql agent" in service_type_lower
        )

        # Service is running and enabled? Non-essential services need justification
        is_running = "running" in (status or "").lower()
        is_stopped = "stopped" in (status or "").lower()
        is_auto = "auto" in (startup_type or "").lower()
        is_disabled = "disabled" in (startup_type or "").lower()

        # Needs action if:
        # 1. Account is non-compliant (for any service), OR
        # 2. Non-essential service is running/enabled (needs justification), OR
        # 3. SQL Agent is stopped or disabled (Q2: WARNING, may need justification)
        needs_action = (
            not account_compliant
            or (is_non_essential and is_running and is_auto)
            or (
                is_sql_agent and (is_stopped or is_disabled)
            )  # Q2: Agent down = discrepancy
        )

        # Compliant column should match the action indicator
        # If needs_action = True (⏳), then compliant = False (✗)
        # This prevents confusion where ⏳ shows but Compliant says ✓
        is_compliant = not needs_action

        if needs_action:
            self._increment_warn()
        else:
            self._increment_pass()

        inst_display = instance_name or "(Default)"

        data = [
            None,  # Action indicator (column A)
            server_name,
            inst_display,
            service_name,
            service_type,
            None,  # Status - styled separately
            None,  # Startup - styled separately
            service_account,
            None,  # Compliant
            "",  # Justification
        ]

        row, _ = self._write_row_with_uuid(ws, SERVICE_CONFIG, data)

        # Apply action indicator (column 1)
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_action)

        # Apply color to data columns (NOT Action column which stays white)
        # A=UUID, B=Action(2), C=Server(3), D=Instance(4), E=ServiceName(5), F=Type(6), G=Status(7), H=Startup(8), I=Account(9), J=Compliant(10)
        # Apply color to data columns
        # Use mixin helper
        # Columns: Server(3), Instance(4), Service(5), Type(6), Status(7), Startup(8), Account(9), Compliant(10)
        # Note: Status/Startup/Compliant have their own coloring logic which overrides this background
        self._apply_row_color(row, row_color, data_cols=[3, 4, 5, 6, 9], ws=ws)

        # Style status column (column G = 7)
        status_cell = ws.cell(row=row, column=7)
        status_lower = (status or "").lower()
        if "running" in status_lower:
            status_cell.value = "✓ Running"
            status_cell.fill = Fills.PASS
            status_cell.font = Fonts.PASS
        elif "stopped" in status_lower:
            status_cell.value = "✗ Stopped"
            status_cell.fill = Fills.WARN
            status_cell.font = Fonts.WARN
        else:
            status_cell.value = status or "Unknown"

        # Style startup type column (column H = 8)
        startup_cell = ws.cell(row=row, column=8)
        startup_lower = (startup_type or "").lower()
        if "auto" in startup_lower:
            startup_cell.value = "⚡ Auto"
            startup_cell.fill = Fills.PASS
        elif "manual" in startup_lower:
            startup_cell.value = "🔧 Manual"
            startup_cell.fill = PatternFill(
                start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
            )
        elif "disabled" in startup_lower:
            startup_cell.value = "⛔ Disabled"
            startup_cell.fill = Fills.WARN
        else:
            startup_cell.value = startup_type or ""

        # Style Compliant column (column J = 10)
        apply_boolean_styling(ws.cell(row=row, column=10), is_compliant)

    def _merge_svc_instance(self, ws) -> None:
        """Merge Instance cells for current instance group."""
        current_row = self._row_counters[SERVICE_CONFIG.name]
        if current_row > self._svc_instance_start_row:
            merge_server_cells(
                ws,
                server_col=4,  # Instance column (A=UUID, B=Action, C=Server, D=Instance)
                start_row=self._svc_instance_start_row,
                end_row=current_row - 1,
                server_name=self._svc_last_instance,
                is_alt=self._svc_instance_alt,
            )

    def _merge_svc_groups(self, ws) -> None:
        """Merge both Server and Instance cells."""
        self._merge_svc_instance(ws)
        current_row = self._row_counters[SERVICE_CONFIG.name]
        if current_row > self._svc_server_start_row:
            color_main, _ = SERVER_GROUP_COLORS[
                self._svc_server_idx % len(SERVER_GROUP_COLORS)
            ]
            merge_server_cells(
                ws,
                server_col=3,  # Server column (A=UUID, B=Action, C=Server)
                start_row=self._svc_server_start_row,
                end_row=current_row - 1,
                server_name=self._svc_last_server,
                is_alt=True,
            )
            merged = ws.cell(row=self._svc_server_start_row, column=3)
            merged.fill = PatternFill(
                start_color=color_main, end_color=color_main, fill_type="solid"
            )

    def _finalize_services(self) -> None:
        """Finalize services sheet."""
        if self._service_sheet:
            self._finalize_grouping(SERVICE_CONFIG.name)
            self._finalize_sheet_with_uuid(self._service_sheet)

    def _add_service_dropdowns(self) -> None:
        """Add dropdown validations for status columns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation,
            add_review_status_conditional_formatting,
            STATUS_VALUES,
        )

        ws = self._service_sheet
        # With UUID: A=UUID, B=Action, C=Server, D=Instance, E=ServiceName, F=Type, G=Status, H=Startup, I=Account, J=Compliant, K=ReviewStatus
        add_dropdown_validation(
            ws, "G", ["✓ Running", "✗ Stopped", "Unknown"]
        )  # Status
        add_dropdown_validation(
            ws, "H", ["⚡ Auto", "🔧 Manual", "⛔ Disabled"]
        )  # Startup
        add_dropdown_validation(ws, "J", ["✓", "✗"])  # Compliant
        add_dropdown_validation(ws, "K", STATUS_VALUES.all())  # Review Status
        add_review_status_conditional_formatting(ws, "K")
