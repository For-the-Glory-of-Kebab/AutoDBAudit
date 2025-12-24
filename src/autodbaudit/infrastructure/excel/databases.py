"""
Databases Sheet Module.

Handles the Databases worksheet for database properties audit.
Uses ServerGroupMixin for server/instance grouping.
Enhanced with visual icons for State and Recovery Model.


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fills,
    Fonts,
    apply_boolean_styling,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    format_size_mb,
    ACTION_COLUMN,
    STATUS_COLUMN,
    LAST_REVIEWED_COLUMN,
    apply_action_needed_styling,
)
from autodbaudit.infrastructure.excel.server_group import ServerGroupMixin


__all__ = ["DatabaseSheetMixin", "DATABASE_CONFIG"]


DATABASE_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Server", 18, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 15, Alignments.LEFT),  # Column D
    ColumnDef("Database", 25, Alignments.LEFT),  # Column E
    ColumnDef("Owner", 20, Alignments.LEFT),  # Column F
    ColumnDef("Recovery", 14, Alignments.CENTER),  # Column G
    ColumnDef("State", 14, Alignments.CENTER),  # Column H
    ColumnDef("Data (MB)", 12, Alignments.RIGHT),
    ColumnDef("Log (MB)", 12, Alignments.RIGHT),
    ColumnDef("Trustworthy", 12, Alignments.CENTER),
    STATUS_COLUMN,  # Review Status dropdown
    ColumnDef("Justification", 35, Alignments.LEFT, is_manual=True),
    LAST_REVIEWED_COLUMN,
    ColumnDef("Notes", 30, Alignments.LEFT_WRAP, is_manual=True),
)

DATABASE_CONFIG = SheetConfig(name="Databases", columns=DATABASE_COLUMNS)


class DatabaseSheetMixin(ServerGroupMixin, BaseSheetMixin):
    """Mixin for Databases sheet with server/instance grouping."""

    _database_sheet = None

    def add_database(
        self,
        server_name: str,
        instance_name: str,
        database_name: str,
        owner: str,
        recovery_model: str,
        state: str,
        data_size_mb: Any,
        log_size_mb: Any,
        is_trustworthy: bool,
    ) -> None:
        """Add a database row."""
        if self._database_sheet is None:
            self._database_sheet = self._ensure_sheet_with_uuid(DATABASE_CONFIG)
            self._init_grouping(self._database_sheet, DATABASE_CONFIG)
            self._add_database_dropdowns()

        ws = self._database_sheet

        # Track grouping and get row color
        row_color = self._track_group(server_name, instance_name, DATABASE_CONFIG.name)

        # Determine if action needed (TRUSTWORTHY ON for user DBs)
        system_dbs = {"master", "msdb", "model", "tempdb"}
        is_system_db = database_name.lower() in system_dbs
        needs_action = is_trustworthy and not is_system_db

        data = [
            None,  # Action indicator (column B)
            server_name,  # Column C
            instance_name or "(Default)",
            database_name,
            owner or "",
            None,  # Recovery - styled separately
            None,  # State - styled separately
            format_size_mb(data_size_mb),
            format_size_mb(log_size_mb),
            None,  # Trustworthy
            "",  # Justification
            "",  # Notes
        ]

        row, row_uuid = self._write_row_with_uuid(ws, DATABASE_CONFIG, data)

        # Apply action indicator (column 1)
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_action)

        # Apply row color to data columns (A=UUID, B=Action, C=Server, D=Instance, E=Database, F=Owner, ...)
        self._apply_row_color(row, row_color, data_cols=[3, 4, 5, 6, 9, 10], ws=ws)

        # Style Recovery Model column (column G = 7)
        recovery_cell = ws.cell(row=row, column=7)
        recovery_lower = (recovery_model or "").lower()
        if "full" in recovery_lower:
            recovery_cell.value = "🛡️ Full"
            recovery_cell.fill = Fills.PASS
        elif "bulk" in recovery_lower:
            recovery_cell.value = "📦 Bulk-Logged"
            recovery_cell.fill = PatternFill(
                start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
            )
        elif "simple" in recovery_lower:
            recovery_cell.value = "⚡ Simple"
            recovery_cell.fill = Fills.WARN
        else:
            recovery_cell.value = recovery_model or ""

        # Style State column (column H = 8)
        state_cell = ws.cell(row=row, column=8)
        state_lower = (state or "").lower()
        if "online" in state_lower:
            state_cell.value = "✓ Online"
            state_cell.fill = Fills.PASS
            state_cell.font = Fonts.PASS
        elif "offline" in state_lower:
            state_cell.value = "⛔ Offline"
            state_cell.fill = Fills.WARN
            state_cell.font = Fonts.WARN
        elif "restoring" in state_lower:
            state_cell.value = "🔄 Restoring"
            state_cell.fill = PatternFill(
                start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
            )
        elif "recovering" in state_lower:
            state_cell.value = "⏳ Recovering"
            state_cell.fill = PatternFill(
                start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
            )
        elif "suspect" in state_lower:
            state_cell.value = "⚠️ Suspect"
            state_cell.fill = Fills.FAIL
            state_cell.font = Fonts.FAIL
        elif "emergency" in state_lower:
            state_cell.value = "🚨 Emergency"
            state_cell.fill = Fills.FAIL
            state_cell.font = Fonts.FAIL
        else:
            state_cell.value = state or ""

        # Trustworthy column (column K = 11)
        # is_system_db already calculated above for needs_action
        trustworthy_cell = ws.cell(row=row, column=11)
        if is_system_db:
            # System DB - show value but don't mark as pass/fail
            trustworthy_cell.value = "✓ ON" if is_trustworthy else "✗ OFF"
            trustworthy_cell.fill = PatternFill(
                start_color="E8EAF6", end_color="E8EAF6", fill_type="solid"
            )
        else:
            # User DB - TRUSTWORTHY ON is a security concern
            apply_boolean_styling(trustworthy_cell, is_trustworthy, invert=True)

    def _finalize_databases(self) -> None:
        """Finalize databases sheet - merge remaining groups."""
        if self._database_sheet:
            self._finalize_grouping(DATABASE_CONFIG.name)
            self._finalize_sheet_with_uuid(self._database_sheet)

    def _add_database_dropdowns(self) -> None:
        """Add dropdown validations for status columns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation,
            add_review_status_conditional_formatting,
            STATUS_VALUES,
        )

        ws = self._database_sheet
        # Recovery Model column (G) - column 7 (A=UUID, B=Action, C=Server, D=Instance, E=DB, F=Owner, G=Recovery)
        add_dropdown_validation(ws, "G", ["🛡️ Full", "📦 Bulk-Logged", "⚡ Simple"])
        # State column (H) - column 8
        add_dropdown_validation(
            ws,
            "H",
            [
                "✓ Online",
                "⛔ Offline",
                "🔄 Restoring",
                "⏳ Recovering",
                "⚠️ Suspect",
                "🚨 Emergency",
            ],
        )
        # Trustworthy column (K) - column 11
        add_dropdown_validation(ws, "K", ["✓ ON", "✗ OFF", "✓", "✗"])
        # Review Status column (L) - column 12
        add_dropdown_validation(ws, "L", STATUS_VALUES.all())
        add_review_status_conditional_formatting(ws, "L")

        # --- Dynamic CF for Trustworthy ---
        from openpyxl.formatting.rule import FormulaRule
        from autodbaudit.infrastructure.excel_styles import Fills, Fonts

        # Trustworthy is Column K
        # If "ON" or "Checked", and NOT System DB -> Warning/Fail
        # System DBs logic is harder in straight CF without helper cols.
        # But for general rule: TRUSTWORTHY ON = BAD (Red/Orange).
        # We assume standard text "✓ ON" / "✗ OFF" or boolean.

        k_range = f"K2:K{ws.max_row+100}"

        # 1. Trustworthy ON = FAIL (Red) - Security Risk
        # Match "ON", "True", "Yes", "Checked"
        ws.conditional_formatting.add(
            k_range,
            FormulaRule(
                formula=[
                    'OR(ISNUMBER(SEARCH("ON",K2)), ISNUMBER(SEARCH("True",K2)), ISNUMBER(SEARCH("Yes",K2)))'
                ],
                stopIfTrue=True,
                fill=Fills.FAIL,
                font=Fonts.FAIL,
            ),
        )

        # 2. Trustworthy OFF = PASS (Green)
        ws.conditional_formatting.add(
            k_range,
            FormulaRule(
                formula=[
                    'OR(ISNUMBER(SEARCH("OFF",K2)), ISNUMBER(SEARCH("False",K2)), ISNUMBER(SEARCH("No",K2)))'
                ],
                stopIfTrue=True,
                fill=Fills.PASS,
                font=Fonts.PASS,
            ),
        )
