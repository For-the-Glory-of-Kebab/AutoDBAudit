"""
Row UUID Utilities for Excel ↔ SQLite Synchronization.

This module provides THE authoritative implementation for Row UUIDs:
- UUID generation and validation
- Hidden/locked UUID column definition
- Sheet protection helpers
- UUID detection and integrity checking

All sheet modules MUST use these utilities for consistency.

Architecture Note:
    - UUID is stored in hidden Column A of every sheet
    - UUID is immutable once created (never changes across syncs)
    - UUID uniquely identifies a row for annotation matching
"""

from __future__ import annotations

import logging
import uuid
from typing import TYPE_CHECKING, Any

from openpyxl.styles import Protection
from openpyxl.worksheet.worksheet import Worksheet

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fonts,
)

if TYPE_CHECKING:
    pass

logger = logging.getLogger(__name__)

__all__ = [
    "UUID_COLUMN",
    "generate_row_uuid",
    "is_valid_uuid",
    "read_row_uuid",
    "write_row_uuid",
    "apply_uuid_column_protection",
    "hide_uuid_column",
    "detect_uuid_issues",
]


# ============================================================================
# UUID Column Definition
# ============================================================================

# Standard UUID column - always Column A, hidden and locked
UUID_COLUMN = ColumnDef(
    name="_UUID",  # Underscore prefix indicates internal/hidden
    width=0,  # Hidden (width=0)
    alignment=Alignments.CENTER,
    is_manual=False,  # System-managed, not user-editable
)


# ============================================================================
# UUID Generation
# ============================================================================


def generate_row_uuid() -> str:
    """
    Generate a new Row UUID.
    
    Uses UUID v4 (random) for guaranteed uniqueness.
    No clashes possible (1 in 2^122 probability).
    
    Returns:
        8-character uppercase hex string for readability
        (e.g., "A7F3B2C1")
    """
    return uuid.uuid4().hex[:8].upper()


def is_valid_uuid(value: Any) -> bool:
    """
    Check if a value is a valid Row UUID.
    
    Valid = 8 character hex string (uppercase or lowercase).
    
    Args:
        value: Any value to check
        
    Returns:
        True if valid UUID format
    """
    if not value:
        return False
    
    if not isinstance(value, str):
        try:
            value = str(value).strip()
        except Exception:
            return False
    
    value = value.strip()
    if len(value) != 8:
        return False
    
    try:
        int(value, 16)
        return True
    except ValueError:
        return False


# ============================================================================
# UUID Cell Operations
# ============================================================================


def read_row_uuid(ws: Worksheet, row: int, uuid_column: int = 1) -> str | None:
    """
    Read UUID from a row.
    
    Args:
        ws: Worksheet to read from
        row: Row number (1-indexed)
        uuid_column: Column index for UUID (default 1 = Column A)
        
    Returns:
        UUID string if valid, None if empty or invalid
    """
    cell = ws.cell(row=row, column=uuid_column)
    value = cell.value
    
    if not value:
        return None
    
    value_str = str(value).strip().upper()
    if is_valid_uuid(value_str):
        return value_str
    
    logger.warning("Invalid UUID in row %d: %s", row, repr(value)[:20])
    return None


def write_row_uuid(
    ws: Worksheet,
    row: int,
    row_uuid: str,
    uuid_column: int = 1,
    apply_protection: bool = True,
) -> None:
    """
    Write UUID to a row.
    
    Args:
        ws: Worksheet to write to
        row: Row number (1-indexed)
        row_uuid: UUID string to write
        uuid_column: Column index for UUID (default 1 = Column A)
        apply_protection: Whether to lock the cell
    """
    cell = ws.cell(row=row, column=uuid_column)
    cell.value = row_uuid.upper()
    cell.font = Fonts.DATA
    cell.alignment = Alignments.CENTER
    
    if apply_protection:
        cell.protection = Protection(locked=True)


# ============================================================================
# Sheet Protection
# ============================================================================


def apply_uuid_column_protection(ws: Worksheet, max_row: int = 1000) -> None:
    """
    Apply protection to UUID column (Column A) ONLY.
    
    Makes UUID cells locked while allowing edits to ALL other columns.
    
    CRITICAL: In Excel, all cells are locked by default. We must explicitly
    UNLOCK all non-UUID columns before enabling sheet protection.
    
    Args:
        ws: Worksheet to protect
        max_row: Maximum row to protect (default 1000)
    """
    # STEP 1: Unlock ALL cells first (columns B onwards)
    # This is necessary because Excel cells default to locked=True
    max_col = ws.max_column or 50  # Use worksheet max or reasonable default
    for row in range(1, max_row + 1):
        for col in range(2, max_col + 1):  # Start at column 2 (B)
            cell = ws.cell(row=row, column=col)
            cell.protection = Protection(locked=False)
    
    # STEP 2: Lock UUID column cells (column A)
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.protection = Protection(locked=True)
    
    # STEP 3: Enable sheet protection (NO password = advisory only)
    # User can disable via Review > Unprotect Sheet (no password needed)
    ws.protection.sheet = True
    # Note: Do NOT set password=None, it causes TypeError in openpyxl
    ws.protection.objects = False
    ws.protection.scenarios = False
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.insertHyperlinks = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    
    logger.debug("Applied UUID column protection to sheet %s (unlocked %d columns)", ws.title, max_col - 1)


def hide_uuid_column(ws: Worksheet) -> None:
    """
    Hide the UUID column (Column A).
    
    Sets column width to 0 and marks as hidden.
    
    Args:
        ws: Worksheet to modify
    """
    ws.column_dimensions["A"].width = 0
    ws.column_dimensions["A"].hidden = True
    
    logger.debug("Hidden UUID column in sheet %s", ws.title)


def unlock_editable_columns(
    ws: Worksheet,
    editable_columns: list[int],
    start_row: int = 2,
    end_row: int = 1000,
) -> None:
    """
    Unlock specific columns for user editing.
    
    Use this after applying sheet protection to allow editing
    of annotation columns (Notes, Justification, Review Status, Last Reviewed).
    
    Args:
        ws: Worksheet to modify
        editable_columns: List of column indices (1-indexed) to unlock
        start_row: First data row (default 2, after header)
        end_row: Last row to unlock (default 1000)
    """
    for col in editable_columns:
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.protection = Protection(locked=False)
    
    logger.debug(
        "Unlocked columns %s in sheet %s", editable_columns, ws.title
    )


# ============================================================================
# UUID Integrity Checks
# ============================================================================


def detect_uuid_issues(ws: Worksheet, max_row: int = 1000) -> dict[str, list]:
    """
    Detect UUID integrity issues in a worksheet.
    
    Checks for:
    - Missing UUIDs (rows with data but no UUID)
    - Invalid UUIDs (wrong format)
    - Duplicate UUIDs (same UUID on multiple rows)
    
    Args:
        ws: Worksheet to check
        max_row: Maximum row to check
        
    Returns:
        Dict with keys: 'missing', 'invalid', 'duplicates'
        Each value is a list of row numbers with that issue
    """
    issues = {
        "missing": [],
        "invalid": [],
        "duplicates": [],
    }
    
    seen_uuids: dict[str, int] = {}  # uuid -> first row seen
    
    for row in range(2, max_row + 1):  # Skip header
        # Check if row has data (any non-empty cell after UUID column)
        has_data = False
        for col in range(2, min(10, ws.max_column + 1)):  # Check first few columns
            if ws.cell(row=row, column=col).value:
                has_data = True
                break
        
        if not has_data:
            continue  # Empty row, skip
        
        uuid_cell = ws.cell(row=row, column=1)
        uuid_value = uuid_cell.value
        
        if not uuid_value:
            issues["missing"].append(row)
            continue
        
        uuid_str = str(uuid_value).strip().upper()
        
        if not is_valid_uuid(uuid_str):
            issues["invalid"].append(row)
            continue
        
        if uuid_str in seen_uuids:
            issues["duplicates"].append(row)
            # Also mark first occurrence if not already
            first_row = seen_uuids[uuid_str]
            if first_row not in issues["duplicates"]:
                issues["duplicates"].append(first_row)
        else:
            seen_uuids[uuid_str] = row
    
    # Sort duplicates for consistency
    issues["duplicates"].sort()
    
    return issues


def repair_uuid_issues(ws: Worksheet, issues: dict[str, list]) -> int:
    """
    Repair UUID issues by generating new UUIDs.
    
    For missing and invalid: generates new UUID
    For duplicates: regenerates all but the first occurrence
    
    Args:
        ws: Worksheet to repair
        issues: Issues dict from detect_uuid_issues
        
    Returns:
        Number of rows repaired
    """
    repaired = 0
    
    # Fix missing UUIDs
    for row in issues.get("missing", []):
        new_uuid = generate_row_uuid()
        write_row_uuid(ws, row, new_uuid)
        logger.info("Generated UUID for row %d: %s", row, new_uuid)
        repaired += 1
    
    # Fix invalid UUIDs
    for row in issues.get("invalid", []):
        new_uuid = generate_row_uuid()
        write_row_uuid(ws, row, new_uuid)
        logger.info("Regenerated invalid UUID for row %d: %s", row, new_uuid)
        repaired += 1
    
    # Fix duplicates (skip first occurrence in duplicates list)
    duplicates = issues.get("duplicates", [])
    if len(duplicates) > 1:
        # Keep first, regenerate rest
        seen: set[str] = set()
        for row in duplicates:
            current = ws.cell(row=row, column=1).value
            current_str = str(current).strip().upper() if current else ""
            
            if current_str in seen:
                # This is a duplicate, regenerate
                new_uuid = generate_row_uuid()
                write_row_uuid(ws, row, new_uuid)
                logger.info("Regenerated duplicate UUID for row %d: %s", row, new_uuid)
                repaired += 1
            else:
                seen.add(current_str)
    
    return repaired
