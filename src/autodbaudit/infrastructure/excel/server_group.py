"""
Server Group Mixin Module.

Provides reusable server/instance grouping functionality with:
- Cell merging for Server and Instance columns
- Color rotation for server groups (Teal→Coral→Gold→Lavender)
- Shade alternation for instances within a server

Each sheet that uses this mixin gets its own isolated state via a dict keyed by sheet name.

UUID Support (v3):
    - Column A is now UUID (hidden)
    - Column B is Action indicator (OPTIONAL)
    - Server/Instance columns shift based on whether Action column exists
"""

from __future__ import annotations

from dataclasses import dataclass
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from autodbaudit.infrastructure.excel_styles import (
    merge_server_cells,
    SERVER_GROUP_COLORS,
)


__all__ = ["ServerGroupMixin"]


@dataclass
class GroupState:
    """State for tracking server/instance grouping per sheet."""

    ws: Worksheet = None
    config: object = None
    last_server: str = ""
    last_instance: str = ""
    last_database: str = ""  # New: Track database
    server_start_row: int = 2
    instance_start_row: int = 2
    database_start_row: int = 2  # New: Track database start
    server_idx: int = 0
    instance_alt: bool = False
    # Column indices (default to standard layout with Action column)
    # With Action: UUID=1, Action=2, Server=3, Instance=4, Database=5
    server_col_idx: int = 3
    instance_col_idx: int = 4
    database_col_idx: int = 5  # New: Database column index


class ServerGroupMixin:
    """
    Mixin providing server/instance/database grouping with merging and colors.

    Uses a dict to store per-sheet grouping state, avoiding conflicts
    when multiple sheets use this mixin.
    """

    def _init_grouping(
        self,
        ws: Worksheet,
        config,
        has_action_col: bool = True,
        database_col_idx: int | None = None,
    ) -> None:
        """
        Initialize grouping state for a sheet.

        Args:
            ws: Worksheet
            config: Sheet config
            has_action_col: Whether sheet has Action/Indicator column (Col 2)
        """
        if not hasattr(self, "_grp_states"):
            self._grp_states = {}

        # Determine column indices
        # If has_action_col: Server=3, Instance=4, Database=5
        # If no action col: Server=2, Instance=3, Database=4
        offset = 1 if has_action_col else 0
        server_col = 2 + offset
        instance_col = 3 + offset
        database_col = 4 + offset

        # Allow override for sheets with extra columns (e.g. Scope)
        if database_col_idx is not None:
            database_col = database_col_idx

        self._grp_states[config.name] = GroupState(
            ws=ws,
            config=config,
            last_server="",
            last_instance="",
            last_database="",
            server_start_row=2,
            instance_start_row=2,
            database_start_row=2,
            server_idx=0,
            instance_alt=False,
            server_col_idx=server_col,
            instance_col_idx=instance_col,
            database_col_idx=database_col,
        )

    def _get_state(self, config_name: str) -> GroupState:
        """Get grouping state for a sheet."""
        return self._grp_states.get(config_name)

    def _track_group(
        self,
        server_name: str,
        instance_name: str,
        config_name: str,
        database_name: str | None = None,  # New optional arg
    ) -> str:
        """
        Track server/instance/database grouping and return row background color.

        Args:
            server_name: Server hostname
            instance_name: Instance name
            config_name: Sheet config name
            database_name: Database name (optional, enables 3rd level grouping)

        Returns:
            Hex color code for row background
        """
        state = self._grp_states[config_name]
        current_row = self._row_counters[config_name]
        inst_display = instance_name or "(Default)"
        db_display = database_name or "(Instance)"

        # Check if SERVER changed
        if server_name != state.last_server:
            if state.last_server:
                self._merge_groups(config_name)
                state.server_idx += 1

            state.server_start_row = current_row
            state.instance_start_row = current_row
            state.database_start_row = current_row
            state.last_server = server_name
            state.last_instance = inst_display
            state.last_database = db_display
            state.instance_alt = False

        # Check if INSTANCE changed (within same server)
        elif inst_display != state.last_instance:
            self._merge_instance(config_name)  # Will also merge DB
            state.instance_start_row = current_row
            state.database_start_row = current_row
            state.last_instance = inst_display
            state.last_database = db_display
            state.instance_alt = not state.instance_alt

        # Check if DATABASE changed (within same instance)
        elif database_name is not None and db_display != state.last_database:
            self._merge_database(config_name)
            state.database_start_row = current_row
            state.last_database = db_display

        # Get color for this row
        color_main, color_light = SERVER_GROUP_COLORS[
            state.server_idx % len(SERVER_GROUP_COLORS)
        ]
        return color_main if state.instance_alt else color_light

    def _apply_row_color(
        self,
        row: int,
        color: str,
        data_cols: list[int],
        ws: Worksheet,
    ) -> None:
        """
        Apply background color to specified columns.
        Note: Callers must include the Instance/Database column in data_cols if they want it colored.
        """
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in data_cols:
            ws.cell(row=row, column=col).fill = fill

    def _merge_database(self, config_name: str) -> None:
        """Merge Database cells for current database group."""
        state = self._grp_states[config_name]
        current_row = self._row_counters[config_name]

        if current_row > state.database_start_row:
            # Only merge if we have a database column tracked (simple check: implies logic is used)
            # Use merge_server_cells generic helper
            merge_server_cells(
                state.ws,
                server_col=state.database_col_idx,
                start_row=state.database_start_row,
                end_row=current_row - 1,
                server_name=state.last_database,
                is_alt=state.instance_alt,
            )

            # Apply color
            color_main, color_light = SERVER_GROUP_COLORS[
                state.server_idx % len(SERVER_GROUP_COLORS)
            ]
            fill_color = color_main if state.instance_alt else color_light

            merged = state.ws.cell(
                row=state.database_start_row, column=state.database_col_idx
            )
            merged.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

    def _merge_instance(self, config_name: str) -> None:
        """Merge Instance cells for current instance group."""
        state = self._grp_states[config_name]

        # First ensure database group is closed/merged
        self._merge_database(config_name)

        current_row = self._row_counters[config_name]
        if current_row > state.instance_start_row:
            merge_server_cells(
                state.ws,
                server_col=state.instance_col_idx,
                start_row=state.instance_start_row,
                end_row=current_row - 1,
                server_name=state.last_instance,
                is_alt=state.instance_alt,
            )
            # Always apply fill to merged instance cell
            color_main, color_light = SERVER_GROUP_COLORS[
                state.server_idx % len(SERVER_GROUP_COLORS)
            ]
            fill_color = color_main if state.instance_alt else color_light
            merged = state.ws.cell(
                row=state.instance_start_row, column=state.instance_col_idx
            )
            merged.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

    def _merge_groups(self, config_name: str) -> None:
        """Merge Server, Instance, and Database cells for current server group."""
        state = self._grp_states[config_name]

        # First merge the last instance group (which merges the last db group)
        self._merge_instance(config_name)

        # Then merge the server group
        current_row = self._row_counters[config_name]
        if current_row > state.server_start_row:
            color_main, _ = SERVER_GROUP_COLORS[
                state.server_idx % len(SERVER_GROUP_COLORS)
            ]
            merge_server_cells(
                state.ws,
                server_col=state.server_col_idx,
                start_row=state.server_start_row,
                end_row=current_row - 1,
                server_name=state.last_server,
                is_alt=True,
            )
            # Apply main color to merged server cell
            merged = state.ws.cell(
                row=state.server_start_row, column=state.server_col_idx
            )
            merged.fill = PatternFill(
                start_color=color_main, end_color=color_main, fill_type="solid"
            )

    def _finalize_grouping(self, config_name: str) -> None:
        """Finalize by merging any remaining groups."""
        if hasattr(self, "_grp_states") and config_name in self._grp_states:
            state = self._grp_states[config_name]
            if state.last_server:
                self._merge_groups(config_name)
