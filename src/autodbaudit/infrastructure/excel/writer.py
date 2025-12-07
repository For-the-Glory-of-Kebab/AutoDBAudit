"""
Enhanced Excel Report Writer.

Main writer class that composes all sheet mixins into a single,
fully-featured audit report generator using the Mixin Pattern.

Architecture Overview:
    The report writer uses Python's multiple inheritance (mixin pattern)
    to compose functionality from 16 specialized sheet modules. Each
    module provides:
    - Column definitions (SheetConfig)
    - Data entry method (add_*)
    - Sheet-specific styling logic

Mixin Pattern Benefits:
    1. Single Responsibility: Each sheet's logic is isolated
    2. Open/Closed: New sheets are added without modifying existing code
    3. Testability: Each mixin can be tested independently
    4. Maintainability: ~100 lines per file vs 1000+ in monolith

Sheet Order (16 total):
    1.  Cover           - Audit summary and statistics
    2.  Instances       - SQL Server instance inventory
    3.  SA Account      - SA account security status
    4.  Server Logins   - Server-level login audit
    5.  Sensitive Roles - Sysadmin role membership
    6.  Configuration   - sp_configure security settings
    7.  Services        - SQL Server services audit
    8.  Databases       - Database properties and security
    9.  Database Users  - Per-database user security matrix
    10. Database Roles  - Database role membership
    11. Orphaned Users  - Users without server logins
    12. Linked Servers  - Linked server configuration
    13. Triggers        - Server and database triggers
    14. Backups         - Backup history and compliance
    15. Audit Settings  - Login audit configuration
    16. Actions         - Remediation action items

Usage Example:
    from autodbaudit.infrastructure.excel import EnhancedReportWriter
    
    # Create writer instance
    writer = EnhancedReportWriter()
    
    # Set audit metadata
    writer.set_audit_info(
        run_id=1,
        organization="Acme Corporation",
        started_at=datetime.now(),
    )
    
    # Add data using type-safe methods
    writer.add_instance(server_name="SQLPROD01", ...)
    writer.add_login(server_name="SQLPROD01", ...)
    writer.add_database(...)
    
    # Save generates all sheets (including empty ones)
    writer.save("audit_report.xlsx")

Design Decisions:
    - All sheets created with headers even if no data added
    - Alternating row colors for multi-server visual grouping
    - Gray background for manual/user-editable columns
    - Status columns use pass/fail/warn icons
    - Sensitive items (sysadmin, db_owner) highlighted

Dependencies:
    - openpyxl: Excel file generation
    - All sheet mixins from this package
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook

# ============================================================================
# Sheet Mixin Imports
# Each mixin provides one sheet's functionality
# ============================================================================

from autodbaudit.infrastructure.excel.base import SheetConfig
from autodbaudit.infrastructure.excel.cover import CoverSheetMixin
from autodbaudit.infrastructure.excel.instances import (
    InstanceSheetMixin, INSTANCE_CONFIG
)
from autodbaudit.infrastructure.excel.sa_account import (
    SAAccountSheetMixin, SA_ACCOUNT_CONFIG
)
from autodbaudit.infrastructure.excel.logins import (
    LoginSheetMixin, LOGIN_CONFIG
)
from autodbaudit.infrastructure.excel.roles import (
    RoleSheetMixin, ROLE_CONFIG
)
from autodbaudit.infrastructure.excel.config import (
    ConfigSheetMixin, CONFIG_CONFIG
)
from autodbaudit.infrastructure.excel.services import (
    ServiceSheetMixin, SERVICE_CONFIG
)
from autodbaudit.infrastructure.excel.databases import (
    DatabaseSheetMixin, DATABASE_CONFIG
)
from autodbaudit.infrastructure.excel.db_users import (
    DBUserSheetMixin, DB_USER_CONFIG
)
from autodbaudit.infrastructure.excel.db_roles import (
    DBRoleSheetMixin, DB_ROLE_CONFIG
)
from autodbaudit.infrastructure.excel.orphaned_users import (
    OrphanedUserSheetMixin, ORPHANED_USER_CONFIG
)
from autodbaudit.infrastructure.excel.linked_servers import (
    LinkedServerSheetMixin, LINKED_SERVER_CONFIG
)
from autodbaudit.infrastructure.excel.triggers import (
    TriggerSheetMixin, TRIGGER_CONFIG
)
from autodbaudit.infrastructure.excel.backups import (
    BackupSheetMixin, BACKUP_CONFIG
)
from autodbaudit.infrastructure.excel.audit_settings import (
    AuditSettingSheetMixin, AUDIT_SETTING_CONFIG
)
from autodbaudit.infrastructure.excel.actions import (
    ActionSheetMixin, ACTION_CONFIG
)


__all__ = ["EnhancedReportWriter"]

logger = logging.getLogger(__name__)


# ============================================================================
# Sheet Order Configuration
# Defines the order sheets appear in the final workbook
# ============================================================================

SHEET_ORDER: tuple[SheetConfig, ...] = (
    INSTANCE_CONFIG,        # 2. Instances
    SA_ACCOUNT_CONFIG,      # 3. SA Account
    LOGIN_CONFIG,           # 4. Server Logins
    ROLE_CONFIG,            # 5. Sensitive Roles
    CONFIG_CONFIG,          # 6. Configuration
    SERVICE_CONFIG,         # 7. Services
    DATABASE_CONFIG,        # 8. Databases
    DB_USER_CONFIG,         # 9. Database Users
    DB_ROLE_CONFIG,         # 10. Database Roles
    ORPHANED_USER_CONFIG,   # 11. Orphaned Users
    LINKED_SERVER_CONFIG,   # 12. Linked Servers
    TRIGGER_CONFIG,         # 13. Triggers
    BACKUP_CONFIG,          # 14. Backups
    AUDIT_SETTING_CONFIG,   # 15. Audit Settings
    ACTION_CONFIG,          # 16. Actions
)


# ============================================================================
# Main Writer Class
# ============================================================================

class EnhancedReportWriter(
    # Cover sheet (provides set_audit_info and create_cover_sheet)
    CoverSheetMixin,
    
    # Instance-level sheets
    InstanceSheetMixin,     # add_instance
    SAAccountSheetMixin,    # add_sa_account
    
    # Login and role sheets
    LoginSheetMixin,        # add_login
    RoleSheetMixin,         # add_role_member
    
    # Configuration sheets
    ConfigSheetMixin,       # add_config_setting
    ServiceSheetMixin,      # add_service
    
    # Database-level sheets
    DatabaseSheetMixin,     # add_database
    DBUserSheetMixin,       # add_db_user
    DBRoleSheetMixin,       # add_db_role_member
    OrphanedUserSheetMixin, # add_orphaned_user
    
    # Auxiliary sheets
    LinkedServerSheetMixin, # add_linked_server
    TriggerSheetMixin,      # add_trigger
    BackupSheetMixin,       # add_backup_info
    AuditSettingSheetMixin, # add_audit_setting
    
    # Action tracking
    ActionSheetMixin,       # add_action
):
    """
    Enhanced Excel Report Writer.
    
    Generates comprehensive, styled audit reports with 16 worksheets.
    Uses multiple inheritance (mixin pattern) to compose sheet
    functionality from specialized modules.
    
    Each mixin provides:
        - Column definitions (via SheetConfig)
        - add_* method to populate the sheet
        - Sheet-specific styling and validation
    
    Inherited Methods (from mixins):
        set_audit_info()    - Set audit run metadata
        add_instance()      - Add SQL Server instance
        add_sa_account()    - Add SA account status
        add_login()         - Add server login
        add_role_member()   - Add server role membership
        add_config_setting()- Add configuration setting
        add_service()       - Add SQL service
        add_database()      - Add database
        add_db_user()       - Add database user
        add_db_role_member()- Add database role membership
        add_orphaned_user() - Add orphaned user
        add_linked_server() - Add linked server
        add_trigger()       - Add trigger
        add_backup_info()   - Add backup status
        add_audit_setting() - Add audit setting
        add_action()        - Add remediation action
    
    Instance Methods:
        save()              - Generate and save the workbook
    
    Attributes:
        wb: The openpyxl Workbook instance
        _issue_count: Number of FAIL findings
        _pass_count: Number of PASS findings
        _warn_count: Number of WARN findings
        _row_counters: Dict tracking current row per sheet
    """
    
    def __init__(self) -> None:
        """
        Initialize the report writer with an empty workbook.
        
        Creates a fresh workbook and initializes all counters.
        The default "Sheet" worksheet is removed when the first
        real sheet is created.
        """
        # Core workbook instance
        self.wb = Workbook()
        
        # Summary statistics counters
        # These are displayed on the Cover sheet
        self._issue_count: int = 0  # Critical/fail findings
        self._pass_count: int = 0   # Passing checks
        self._warn_count: int = 0   # Warnings
        
        # Row counters for each sheet
        # All sheets start at row 2 (row 1 is the header)
        self._row_counters: dict[str, int] = {
            config.name: 2 for config in SHEET_ORDER
        }
        
        logger.debug("EnhancedReportWriter initialized with empty workbook")
    
    def _ensure_all_sheets(self) -> None:
        """
        Create all sheets with headers, even if no data was added.
        
        This ensures a consistent report structure regardless of
        which data was collected. Empty sheets still have:
            - Header row with column names
            - Column widths set correctly
            - Freeze panes enabled
            - AutoFilter applied
        
        Called automatically by save() before writing to disk.
        """
        for config in SHEET_ORDER:
            # _ensure_sheet is inherited from BaseSheetMixin
            self._ensure_sheet(config)
    
    def _reorder_sheets(self) -> None:
        """
        Reorder sheets to the standard logical order.
        
        The order is:
            1. Cover (always first)
            2-16. Data sheets in SHEET_ORDER
        
        This ensures consistent sheet ordering regardless of
        the order data was added to the writer.
        """
        # Build desired order: Cover first, then all data sheets
        desired_order = ["Cover"] + [config.name for config in SHEET_ORDER]
        
        # Move sheets to correct positions
        for target_idx, name in enumerate(desired_order):
            if name in self.wb.sheetnames:
                current_idx = self.wb.sheetnames.index(name)
                if current_idx != target_idx:
                    offset = target_idx - current_idx
                    self.wb.move_sheet(name, offset=offset)
    
    def save(self, path: Path | str) -> Path:
        """
        Save the report to an Excel file.
        
        This method performs the following steps:
        
        1. Creates any sheets that weren't populated (headers only)
        2. Creates the Cover sheet with summary statistics
        3. Reorders all sheets to the standard order
        4. Saves the workbook to the specified path
        
        The output directory is created if it doesn't exist.
        
        Args:
            path: Output file path (str or Path)
                  Parent directories are created automatically.
        
        Returns:
            Path object pointing to the saved file.
        
        Raises:
            PermissionError: If the file is open in another program
            OSError: If the path is invalid or disk is full
        
        Example:
            path = writer.save("output/audit_2024.xlsx")
            print(f"Saved: {path} ({path.stat().st_size} bytes)")
        """
        path = Path(path)
        
        # Create output directory if needed
        path.parent.mkdir(parents=True, exist_ok=True)
        
        # Step 1: Ensure all sheets exist (even empty ones)
        self._ensure_all_sheets()
        
        # Step 2: Create cover sheet with summary statistics
        # This uses counters populated by the add_* methods
        self.create_cover_sheet()
        
        # Step 3: Reorder sheets to logical flow
        self._reorder_sheets()
        
        # Step 4: Save to disk
        self.wb.save(path)
        
        # Log summary
        logger.info(
            "Report saved: %s (%d sheets, %d issues, %d passes, %d warnings)",
            path,
            len(self.wb.sheetnames),
            self._issue_count,
            self._pass_count,
            self._warn_count,
        )
        
        return path
