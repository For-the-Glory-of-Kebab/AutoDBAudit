"""
SA Account Sheet Module.

Handles the SA Account worksheet for SA account security audit.
Merges Server column when multiple instances on same server.

UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - Column B: Action indicator
    - Column C onwards: Data columns
    - All indices +1 from pre-UUID layout
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    apply_boolean_styling,
    apply_status_styling,
    merge_server_cells,
    SERVER_GROUP_COLORS,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    ACTION_COLUMN,
    STATUS_COLUMN,
    LAST_REVIEWED_COLUMN,
    apply_action_needed_styling,
)


__all__ = ["SAAccountSheetMixin", "SA_ACCOUNT_CONFIG"]


# Column definitions (WITHOUT UUID - added automatically)
SA_ACCOUNT_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (⏳/✅)
    ColumnDef("Server", 18, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 15, Alignments.LEFT),  # Column D
    ColumnDef("Status", 12, Alignments.CENTER, is_status=True),  # Column E
    ColumnDef("Is Disabled", 12, Alignments.CENTER),  # Column F
    ColumnDef("Is Renamed", 12, Alignments.CENTER),  # Column G
    ColumnDef("Current Name", 20, Alignments.LEFT),  # Column H
    ColumnDef("Default DB", 15, Alignments.LEFT),  # Column I
    STATUS_COLUMN,  # Column J: Review Status dropdown
    ColumnDef("Justification", 40, Alignments.LEFT, is_manual=True),  # Column K
    LAST_REVIEWED_COLUMN,  # Column L
    ColumnDef("Notes", 35, Alignments.LEFT, is_manual=True),  # Column M
)

SA_ACCOUNT_CONFIG = SheetConfig(name="SA Account", columns=SA_ACCOUNT_COLUMNS)

# Column indices WITH UUID offset (A=UUID, B=Action, C=Server, etc.)
COL_UUID = 1
COL_ACTION = 2
COL_SERVER = 3
COL_INSTANCE = 4
COL_STATUS = 5
COL_IS_DISABLED = 6
COL_IS_RENAMED = 7
COL_CURRENT_NAME = 8
COL_DEFAULT_DB = 9
COL_REVIEW_STATUS = 10
COL_JUSTIFICATION = 11
COL_LAST_REVIEWED = 12
COL_NOTES = 13


class SAAccountSheetMixin(BaseSheetMixin):
    """Mixin for SA Account sheet with server grouping."""
    
    _sa_account_sheet = None
    _sa_last_server: str = ""
    _sa_server_start_row: int = 2
    _sa_server_idx: int = 0
    
    def add_sa_account(
        self,
        server_name: str,
        instance_name: str,
        is_disabled: bool,
        is_renamed: bool,
        current_name: str,
        default_db: str,
    ) -> tuple[int, str]:
        """Add SA account audit row. Returns (row_number, row_uuid)."""
        if self._sa_account_sheet is None:
            self._sa_account_sheet = self._ensure_sheet_with_uuid(SA_ACCOUNT_CONFIG)
            self._sa_last_server = ""
            self._sa_server_start_row = 2
            self._sa_server_idx = 0
            self._add_sa_dropdowns()
        
        ws = self._sa_account_sheet
        current_row = self._row_counters[SA_ACCOUNT_CONFIG.name]
        
        # Check if server changed
        if server_name != self._sa_last_server:
            if self._sa_last_server:
                self._merge_sa_server(ws)
                self._sa_server_idx += 1
            
            self._sa_server_start_row = current_row
            self._sa_last_server = server_name
        
        # Get color
        color_main, color_light = SERVER_GROUP_COLORS[
            self._sa_server_idx % len(SERVER_GROUP_COLORS)
        ]
        
        # Determine compliance status
        if is_disabled and is_renamed:
            status = "pass"
            self._increment_pass()
        elif is_disabled or is_renamed:
            status = "warn"
            self._increment_warn()
        else:
            status = "fail"
            self._increment_issue()
        
        # Data for columns B onwards (UUID is handled automatically)
        data = [
            None,  # Action indicator (column B)
            server_name,  # Column C
            instance_name or "(Default)",  # Column D
            None,  # Status (column E)
            None,  # Is Disabled (column F)
            None,  # Is Renamed (column G)
            current_name,  # Column H
            default_db or "",  # Column I
            "",  # Review Status (column J)
            "",  # Justification (column K)
            "",  # Last Reviewed (column L)
            "",  # Notes (column M)
        ]
        
        row, row_uuid = self._write_row_with_uuid(ws, SA_ACCOUNT_CONFIG, data)
        
        # Apply action indicator - show ⏳ for non-pass items
        needs_action = status != "pass"
        apply_action_needed_styling(ws.cell(row=row, column=COL_ACTION), needs_action)
        
        # Apply light color to Server, Instance, Current Name, Default DB
        fill = PatternFill(start_color=color_light, end_color=color_light, fill_type="solid")
        for col in [COL_SERVER, COL_INSTANCE, COL_CURRENT_NAME, COL_DEFAULT_DB]:
            ws.cell(row=row, column=col).fill = fill
        
        # Apply status styling
        apply_status_styling(ws.cell(row=row, column=COL_STATUS), status)
        apply_boolean_styling(ws.cell(row=row, column=COL_IS_DISABLED), is_disabled)
        apply_boolean_styling(ws.cell(row=row, column=COL_IS_RENAMED), is_renamed)
        
        return row, row_uuid
    
    def _merge_sa_server(self, ws) -> None:
        """Merge Server cells for current server group."""
        current_row = self._row_counters[SA_ACCOUNT_CONFIG.name]
        if current_row > self._sa_server_start_row:
            color_main, _ = SERVER_GROUP_COLORS[
                self._sa_server_idx % len(SERVER_GROUP_COLORS)
            ]
            merge_server_cells(
                ws,
                server_col=COL_SERVER,  # Server is column C
                start_row=self._sa_server_start_row,
                end_row=current_row - 1,
                server_name=self._sa_last_server,
                is_alt=True,
            )
            merged = ws.cell(row=self._sa_server_start_row, column=COL_SERVER)
            merged.fill = PatternFill(
                start_color=color_main, end_color=color_main, fill_type="solid"
            )
    
    def _finalize_sa_accounts(self) -> None:
        """Finalize SA Account sheet - merge remaining server group and apply UUID protection."""
        if self._sa_account_sheet and self._sa_last_server:
            self._merge_sa_server(self._sa_account_sheet)
            self._finalize_sheet_with_uuid(self._sa_account_sheet)
    
    def _add_sa_dropdowns(self) -> None:
        """Add dropdown validations for status columns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation, add_review_status_conditional_formatting, STATUS_VALUES
        )
        
        ws = self._sa_account_sheet
        # Column letters with UUID: A=UUID, B=Action, C=Server, ..., E=Status, F=IsDisabled, G=IsRenamed, J=ReviewStatus
        add_dropdown_validation(ws, "E", ["PASS", "FAIL", "WARN"])  # Status
        add_dropdown_validation(ws, "F", ["✓", "✗"])  # Is Disabled
        add_dropdown_validation(ws, "G", ["✓", "✗"])  # Is Renamed
        add_dropdown_validation(ws, "J", STATUS_VALUES.all())  # Review Status
        add_review_status_conditional_formatting(ws, "J")

