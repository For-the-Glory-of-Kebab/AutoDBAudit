"""
Sensitive Roles Sheet Module.

Handles the Sensitive Roles worksheet for server role membership audit.
Server and Instance columns are merged with color rotation.


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fonts,
    Fills,
    Icons,
    merge_server_cells,
    SERVER_GROUP_COLORS,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    LAST_REVIEWED_COLUMN,
    STATUS_COLUMN,
    ACTION_COLUMN,
    apply_action_needed_styling,
)


__all__ = ["RoleSheetMixin", "ROLE_CONFIG"]


ROLE_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Server", 18, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 15, Alignments.LEFT),  # Column D
    ColumnDef("Role", 20, Alignments.LEFT),
    ColumnDef("Member", 28, Alignments.LEFT),
    ColumnDef("Member Type", 18, Alignments.LEFT),
    ColumnDef("Enabled", 10, Alignments.CENTER),
    STATUS_COLUMN,  # Review Status dropdown
    ColumnDef("Justification", 40, Alignments.CENTER_WRAP, is_manual=True),
    LAST_REVIEWED_COLUMN,
)

ROLE_CONFIG = SheetConfig(name="Sensitive Roles", columns=ROLE_COLUMNS)

# Accounts to exclude from highlighting (Strict Policy)
SAFE_PREFIXES = ("NT SERVICE\\",)
SAFE_NAMES = set()


class RoleSheetMixin(BaseSheetMixin):
    """Mixin for Sensitive Roles sheet with server/instance grouping."""

    _role_sheet = None
    _role_last_server: str = ""
    _role_last_instance: str = ""
    _role_server_start_row: int = 2
    _role_instance_start_row: int = 2
    _role_server_idx: int = 0
    _role_instance_alt: bool = False

    def add_role_member(
        self,
        server_name: str,
        instance_name: str,
        role_name: str,
        member_name: str,
        member_type: str,
        is_disabled: bool,
    ) -> None:
        """Add a server role membership row."""
        if self._role_sheet is None:
            self._role_sheet = self._ensure_sheet_with_uuid(ROLE_CONFIG)
            self._role_last_server = ""
            self._role_last_instance = ""
            self._role_server_start_row = 2
            self._role_instance_start_row = 2
            self._role_server_idx = 0
            self._role_instance_alt = False
            self._add_role_dropdowns()

        ws = self._role_sheet
        current_row = self._row_counters[ROLE_CONFIG.name]
        inst_display = instance_name or "(Default)"

        # Check if SERVER changed
        if server_name != self._role_last_server:
            if self._role_last_server:
                self._merge_role_groups(ws)
                self._role_server_idx += 1

            self._role_server_start_row = current_row
            self._role_instance_start_row = current_row
            self._role_last_server = server_name
            self._role_last_instance = inst_display
            self._role_instance_alt = False

        # Check if INSTANCE changed (within same server)
        elif inst_display != self._role_last_instance:
            self._merge_role_instance(ws)
            self._role_instance_start_row = current_row
            self._role_last_instance = inst_display
            self._role_instance_alt = not self._role_instance_alt

        # Get colors
        color_main, color_light = SERVER_GROUP_COLORS[
            self._role_server_idx % len(SERVER_GROUP_COLORS)
        ]
        row_color = color_main if self._role_instance_alt else color_light

        # Convert is_disabled to is_enabled (inverted logic)
        is_enabled = not is_disabled

        # Check if this is a privileged role needing justification
        is_safe = member_name.lower() in SAFE_NAMES or any(
            member_name.upper().startswith(p) for p in SAFE_PREFIXES
        )
        needs_justification = (
            role_name.lower() == "sysadmin" and is_enabled and not is_safe
        )

        data = [
            None,  # Action indicator (column A)
            server_name,
            inst_display,
            role_name,
            member_name,
            member_type,
            None,  # Enabled
            "",  # Justification
            "",  # Last Revised
        ]

        row, row_uuid = self._write_row_with_uuid(ws, ROLE_CONFIG, data)

        # Apply action indicator - show ⏳ for sysadmin members needing justification
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_justification)

        # Apply shade to informational columns (NOT Action column which stays white)
        # A=UUID, B=Action(2), C=Server(3), D=Instance(4), E=Role(5), F=Member(6), G=MemberType(7), H=Enabled(8)
        fill = PatternFill(
            start_color=row_color, end_color=row_color, fill_type="solid"
        )
        for col in [3, 4, 5, 6, 7]:  # Server=3, Instance=4, Role=5, Member=6, Type=7
            ws.cell(row=row, column=col).fill = fill

        # Style Enabled column with icon + color (column H = 8)
        enabled_cell = ws.cell(row=row, column=8)
        if is_enabled:
            enabled_cell.value = f"{Icons.PASS} Yes"
            enabled_cell.font = Fonts.PASS
            enabled_cell.fill = Fills.PASS
        else:
            enabled_cell.value = f"{Icons.FAIL} No"
            enabled_cell.font = Fonts.FAIL
            enabled_cell.fill = Fills.FAIL

        # Highlight sysadmin with warning if NOT disabled and NOT whitelisted
        if needs_justification:
            for col in [5, 6, 7]:  # Role=E(5), Member=F(6), Type=G(7)
                ws.cell(row=row, column=col).fill = Fills.WARN

    def _merge_role_instance(self, ws) -> None:
        """Merge Instance cells for current instance group."""
        current_row = self._row_counters[ROLE_CONFIG.name]
        if current_row > self._role_instance_start_row:
            merge_server_cells(
                ws,
                server_col=4,  # Instance column (A=UUID, B=Action, C=Server, D=Instance)
                start_row=self._role_instance_start_row,
                end_row=current_row - 1,
                server_name=self._role_last_instance,
                is_alt=self._role_instance_alt,
            )

    def _merge_role_groups(self, ws) -> None:
        """Merge both Server and Instance cells."""
        self._merge_role_instance(ws)

        current_row = self._row_counters[ROLE_CONFIG.name]
        if current_row > self._role_server_start_row:
            color_main, _ = SERVER_GROUP_COLORS[
                self._role_server_idx % len(SERVER_GROUP_COLORS)
            ]
            merge_server_cells(
                ws,
                server_col=3,  # Server column (A=UUID, B=Action, C=Server)
                start_row=self._role_server_start_row,
                end_row=current_row - 1,
                server_name=self._role_last_server,
                is_alt=True,
            )
            merged = ws.cell(row=self._role_server_start_row, column=3)
            merged.fill = PatternFill(
                start_color=color_main, end_color=color_main, fill_type="solid"
            )

    def _finalize_roles(self) -> None:
        """Finalize roles sheet - merge remaining groups."""
        if self._role_sheet and self._role_last_server:
            self._merge_role_groups(self._role_sheet)
            self._finalize_sheet_with_uuid(self._role_sheet)

    def _add_role_dropdowns(self) -> None:
        """Add dropdown validations for status columns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation,
            add_review_status_conditional_formatting,
            STATUS_VALUES,
        )

        ws = self._role_sheet
        # Column letters with UUID: A=UUID, B=Action, C=Server, D=Instance, E=Role, F=Member, G=Member Type, H=Enabled, I=Review Status

        # E: Role Dropdown
        server_roles = [
            "👑 sysadmin",
            "securityadmin",
            "serveradmin",
            "setupadmin",
            "processadmin",
            "diskadmin",
            "dbcreator",
            "bulkadmin",
        ]
        add_dropdown_validation(ws, "E", server_roles)

        # G: Member Type Dropdown
        add_dropdown_validation(ws, "G", ["🪟 Windows", "🔑 SQL"])

        # H: Enabled
        add_dropdown_validation(ws, "H", ["✓ Yes", "✗ No"])

        # I: Review Status
        add_dropdown_validation(ws, "I", STATUS_VALUES.all())
        add_review_status_conditional_formatting(ws, "I")
