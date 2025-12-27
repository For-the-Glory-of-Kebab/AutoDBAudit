"""
Central Conditional Formatting Styles.

Provides reusable conditional formatting rules for Excel sheets.
Separates styling concerns from data population logic.

Usage:
    from autodbaudit.infrastructure.excel.conditional_styles import (
        apply_status_formatting,
        apply_review_status_formatting,
        apply_risk_level_formatting,
    )

    # Apply to a column range
    apply_status_formatting(worksheet, "G")  # Status column
"""

from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles import PatternFill, Font


# ============================================================================
# Color Constants (matching excel_styles.py)
# ============================================================================

# Pass/Green
PASS_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
PASS_FONT = Font(color="1B5E20", bold=True)

# Warn/Yellow-Orange
WARN_FILL = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
WARN_FONT = Font(color="E65100", bold=True)

# Fail/Red
FAIL_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
FAIL_FONT = Font(color="B71C1C", bold=True)

# Info/Blue
INFO_FILL = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
INFO_FONT = Font(color="0D47A1", bold=True)


# ============================================================================
# Formatting Functions
# ============================================================================


def apply_status_formatting(ws, column: str, max_row: int = 500) -> None:
    """
    Apply conditional formatting for PASS/FAIL/WARN status columns.

    Matches: Running, Stopped, Enabled, Disabled, Yes, No, ✓, ✗
    """
    cell_range = f"{column}2:{column}{max_row}"

    # Pass states (green)
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[
                f'OR(ISNUMBER(SEARCH("Running",{column}2)),ISNUMBER(SEARCH("✓",{column}2)),ISNUMBER(SEARCH("Yes",{column}2)),ISNUMBER(SEARCH("Enabled",{column}2)),ISNUMBER(SEARCH("PASS",{column}2)))'
            ],
            stopIfTrue=True,
            fill=PASS_FILL,
            font=PASS_FONT,
        ),
    )

    # Warn states (yellow)
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[
                f'OR(ISNUMBER(SEARCH("Manual",{column}2)),ISNUMBER(SEARCH("WARN",{column}2)))'
            ],
            stopIfTrue=True,
            fill=WARN_FILL,
            font=WARN_FONT,
        ),
    )

    # Fail states (red)
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[
                f'OR(ISNUMBER(SEARCH("Stopped",{column}2)),ISNUMBER(SEARCH("✗",{column}2)),ISNUMBER(SEARCH("No",{column}2)),ISNUMBER(SEARCH("Disabled",{column}2)),ISNUMBER(SEARCH("FAIL",{column}2)))'
            ],
            stopIfTrue=True,
            fill=FAIL_FILL,
            font=FAIL_FONT,
        ),
    )


def apply_review_status_formatting(ws, column: str, max_row: int = 500) -> None:
    """
    Apply conditional formatting for Review Status dropdown columns.

    - Exception = Blue (documented and approved)
    - Pending Review = Yellow (needs attention)
    - Reviewed = Green (reviewed and cleared)
    """
    cell_range = f"{column}2:{column}{max_row}"

    # Exception = Blue
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Exception",{column}2))'],
            stopIfTrue=True,
            fill=INFO_FILL,
            font=INFO_FONT,
        ),
    )

    # Reviewed = Green
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[
                f'OR(ISNUMBER(SEARCH("Reviewed",{column}2)),ISNUMBER(SEARCH("Approved",{column}2)))'
            ],
            stopIfTrue=True,
            fill=PASS_FILL,
            font=PASS_FONT,
        ),
    )

    # Pending = Yellow
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Pending",{column}2))'],
            stopIfTrue=True,
            fill=WARN_FILL,
            font=WARN_FONT,
        ),
    )


def apply_risk_level_formatting(ws, column: str, max_row: int = 500) -> None:
    """
    Apply conditional formatting for Risk Level columns.

    - High/Critical = Red
    - Medium = Yellow
    - Low = Green
    """
    cell_range = f"{column}2:{column}{max_row}"

    # Low = Green
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Low",{column}2))'],
            stopIfTrue=True,
            fill=PASS_FILL,
            font=PASS_FONT,
        ),
    )

    # Medium = Yellow
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Medium",{column}2))'],
            stopIfTrue=True,
            fill=WARN_FILL,
            font=WARN_FONT,
        ),
    )

    # High/Critical = Red
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[
                f'OR(ISNUMBER(SEARCH("High",{column}2)),ISNUMBER(SEARCH("Critical",{column}2)))'
            ],
            stopIfTrue=True,
            fill=FAIL_FILL,
            font=FAIL_FONT,
        ),
    )


def apply_boolean_formatting(ws, column: str, max_row: int = 500) -> None:
    """
    Apply conditional formatting for boolean Yes/No or ✓/✗ columns.
    """
    cell_range = f"{column}2:{column}{max_row}"

    # Yes/True/✓ = Green
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'OR({column}2="✓",{column}2="Yes",{column}2=TRUE)'],
            stopIfTrue=True,
            fill=PASS_FILL,
            font=PASS_FONT,
        ),
    )

    # No/False/✗ = Red
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'OR({column}2="✗",{column}2="No",{column}2=FALSE)'],
            stopIfTrue=True,
            fill=FAIL_FILL,
            font=FAIL_FONT,
        ),
    )


def apply_days_since_formatting(ws, column: str, max_row: int = 500) -> None:
    """
    Apply conditional formatting for "Days Since" columns.

    - > 365 days = Red (stale)
    - > 90 days = Yellow (aging)
    - <= 90 days = Green (recent)
    """
    cell_range = f"{column}2:{column}{max_row}"

    # > 365 = Red
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan",
            formula=["365"],
            stopIfTrue=True,
            fill=FAIL_FILL,
            font=FAIL_FONT,
        ),
    )

    # > 90 = Yellow
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan",
            formula=["90"],
            stopIfTrue=True,
            fill=WARN_FILL,
            font=WARN_FONT,
        ),
    )

    # <= 90 = Green
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThanOrEqual",
            formula=["90"],
            stopIfTrue=True,
            fill=PASS_FILL,
            font=PASS_FONT,
        ),
    )


def apply_change_type_formatting(ws, column: str, max_row: int = 500) -> None:
    """
    Apply conditional formatting for Change Type columns (Actions sheet).

    - Fixed/Closed/Exception = Green
    - Regression = Red
    - Open/Pending = Yellow
    """
    cell_range = f"{column}2:{column}{max_row}"

    # Fixed/Closed/Exception = Green
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[
                f'OR(ISNUMBER(SEARCH("Fixed",{column}2)),ISNUMBER(SEARCH("Closed",{column}2)),ISNUMBER(SEARCH("Exception",{column}2)))'
            ],
            stopIfTrue=True,
            fill=PASS_FILL,
            font=PASS_FONT,
        ),
    )

    # Regression = Red
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Regression",{column}2))'],
            stopIfTrue=True,
            fill=FAIL_FILL,
            font=FAIL_FONT,
        ),
    )

    # Open/Pending = Yellow
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[
                f'OR(ISNUMBER(SEARCH("Open",{column}2)),ISNUMBER(SEARCH("Pending",{column}2)))'
            ],
            stopIfTrue=True,
            fill=WARN_FILL,
            font=WARN_FONT,
        ),
    )
