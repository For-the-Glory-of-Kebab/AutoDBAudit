"""
Actions Sheet Module.

Handles the Actions worksheet for audit changelog tracking.
This sheet provides a dated log of all detected CHANGES during
sync operations - what was fixed, what regressed, what's new.

Sheet Purpose:
    - Track all state transitions detected during --sync
    - Record when changes were first detected
    - Allow user notes to document context
    - NOT a TODO list - it's an audit trail of changes

Columns:
    - ID: Unique action item number
    - Server/Instance: Location of finding
    - Category: Type (SA Account, Configuration, Backup, etc.)
    - Finding: Description of the change
    - Risk Level: Severity (Low for fixes, High for regressions)
    - Change Description: What happened (Fixed/Regressed/New)
    - Change Type: Type icon (Closed for fixes, Open for issues)
    - Detected Date: When the change was first discovered (editable)
    - Notes: User commentary (synced each run)

Visual Features:
    - Risk level color coding (Low=green, High=red)
    - Change type icons (Closed=checkmark, Open=pending)
    - Gray background for manual input columns
"""

from __future__ import annotations

from datetime import datetime

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fills,
    Icons,
    Fonts,
    Icons,
    Fonts,
    Borders,
)
from openpyxl.styles import PatternFill, Font
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    format_date,
)


__all__ = ["ActionSheetMixin", "ACTION_CONFIG"]


# Changelog columns - simplified for audit trail purpose
ACTION_COLUMNS = (
    ColumnDef("ID", 6, Alignments.CENTER),
    ColumnDef("Server", 16, Alignments.CENTER),
    ColumnDef("Instance", 14, Alignments.CENTER),
    ColumnDef("Category", 20, Alignments.CENTER),  # Wider category
    ColumnDef("Finding", 50, Alignments.CENTER_WRAP),  # Wider finding description
    ColumnDef("Risk Level", 12, Alignments.CENTER),
    ColumnDef("Change Description", 55, Alignments.CENTER_WRAP),  # Wider change desc
    ColumnDef(
        "Change Type", 15, Alignments.CENTER, is_status=True
    ),  # Fixed/Regressed/New
    ColumnDef(
        "Detected Date", 15, Alignments.CENTER_WRAP
    ),  # When change detected (editable)
    ColumnDef("Notes", 60, Alignments.CENTER_WRAP, is_manual=True),  # Much wider notes
)

ACTION_CONFIG = SheetConfig(name="Actions", columns=ACTION_COLUMNS)


class ActionSheetMixin(BaseSheetMixin):
    """
    Mixin for Actions sheet functionality.
    """

    _action_sheet = None
    _action_count: int = 0

    def add_action(
        self,
        server_name: str,
        instance_name: str,
        category: str,
        finding: str,
        risk_level: str,
        recommendation: str,  # Change Description
        status: str = "Open",  # Change Type
        found_date: datetime | None = None,
        notes: str | None = None,
        action_id: int | None = None,
    ) -> None:
        """Add a changelog entry."""
        import logging

        logger = logging.getLogger(__name__)

        # CRITICAL: Assign to self._action_sheet so _finalize_actions can find it
        if self._action_sheet is None:
            self._action_sheet = self._ensure_sheet(ACTION_CONFIG)
        ws = self._action_sheet

        # Prioritize DB ID (Rigorous Requirement)
        if action_id is not None:
            display_id = str(action_id)
            # Update internal counter only if we see a higher ID (to prevent collisions on new rows)
            if isinstance(action_id, int) and action_id > self._action_count:
                self._action_count = action_id
        else:
            self._action_count += 1
            display_id = str(self._action_count)

        if found_date is None:
            found_date = datetime.now()

        # Handle "Unknown" server elegantly (though upstream should fix it)
        srv = (
            server_name
            if server_name and server_name.lower() != "unknown"
            else (server_name or "Unknown")
        )
        inst = instance_name or "(Default)"

        data = [
            display_id,
            srv,
            inst,
            category,
            finding,
            risk_level.title(),
            recommendation,  # Change Description
            None,  # Change Type - styled separately
            format_date(found_date),  # Detected Date
            notes or "",  # Notes
        ]

        row = self._write_row(ws, ACTION_CONFIG, data)

        # Style Change Type cell (column 8)
        status_cell = ws.cell(row=row, column=8)
        status_lower = (
            status.lower().replace("✓", "").replace("⚠", "").replace("⏳", "").strip()
        )

        # Map to Spec Values
        if status_lower == "fixed":
            status_cell.value = "✅ Fixed"
            status_cell.fill = Fills.PASS
            status_cell.font = Fonts.PASS
        elif status_lower == "exception":
            status_cell.value = "🛡️ Exception Documented"
            status_cell.fill = PatternFill(
                start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"
            )  # Blue
        elif status_lower == "regression":
            status_cell.value = "❌ Regression"
            status_cell.fill = Fills.FAIL
            status_cell.font = Fonts.FAIL
        elif status_lower == "closed":  # Exception Removed
            status_cell.value = "🗑️ Exception Removed"
            status_cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )  # Gray
        elif status_lower == "configuration change":
            status_cell.value = "📝 Configuration Change"
            status_cell.fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )  # White
        else:  # Open / New
            status_cell.value = "🆕 New Issue"
            status_cell.fill = Fills.WARN
            status_cell.font = Fonts.WARN

        # Style risk level cell (column 6)
        risk_cell = ws.cell(row=row, column=6)
        risk_lower = risk_level.lower()
        if risk_lower == "low":
            risk_cell.fill = Fills.PASS
            risk_cell.font = Fonts.PASS
        elif risk_lower == "high":
            risk_cell.fill = Fills.FAIL
            risk_cell.font = Fonts.FAIL
        elif risk_lower == "medium":
            risk_cell.fill = Fills.WARN
            risk_cell.font = Fonts.WARN
        elif risk_lower == "critical":
            risk_cell.fill = PatternFill(
                start_color="B71C1C", end_color="B71C1C", fill_type="solid"
            )  # Dark Red
            risk_cell.font = Font(color="FFFFFF", bold=True)
        elif risk_lower == "info":
            risk_cell.fill = PatternFill(
                start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
            )  # Light Blue
            risk_cell.font = Font(color="0D47A1", bold=True)

    def _finalize_actions(self) -> None:
        """
        Finalize Actions sheet - hide/lock ID column and add dropdowns.

        ALWAYS creates the sheet with dropdowns/CF/protection, even if empty.
        This allows users to manually add entries.
        """
        # CRITICAL: Ensure Actions sheet exists even if no actions were recorded
        if self._action_sheet is None:
            self._action_sheet = self._ensure_sheet(ACTION_CONFIG)

        ws = self._action_sheet

        # Rigorous Requirement: Lock & Hide ID column (A)
        try:
            ws.column_dimensions["A"].hidden = True
        except Exception:
            pass

        # ALWAYS add dropdowns and CF - even for empty sheet for manual entries
        self._add_action_dropdowns()

        # Rigorous Requirement: Protect Sheet
        # Unlock user-editable columns before protecting
        self._protect_actions_sheet()

    def _protect_actions_sheet(self) -> None:
        """
        Configure Action sheet for user editing.

        CRITICAL: We do NOT protect the sheet at all.
        Protection was causing the entire sheet to be locked.
        The only thing we do is HIDE column A (ID) for cleanliness.
        Users can edit everything freely.
        """
        ws = self._action_sheet
        if ws is None:
            return

        # Just hide the ID column (column A), do NOT protect the sheet
        # User can still unhide it if they want, but it's hidden by default
        try:
            ws.column_dimensions["A"].hidden = True
        except Exception:
            pass

        # Note: NO sheet protection is applied.
        # Protection was causing the sheet to be locked even with "unlock" settings.
        # This is simpler and more user-friendly.

    def _add_action_dropdowns(self) -> None:
        """Add rigorous dropdown validations for action columns."""
        from autodbaudit.infrastructure.excel.base import add_dropdown_validation

        ws = self._action_sheet
        if not ws:
            return

        # Category column (D)
        add_dropdown_validation(
            ws,
            "D",
            [
                "SA Account",
                "Server Logins",
                "Sensitive Roles",
                "Configuration",
                "Services",
                "Databases",
                "DB Users",
                "DB Roles",
                "Permissions",
                "Audit Settings",
                "Backups",
                "Encryption",
                "Linked Servers",
                "Network",
                "Other",
            ],
        )
        # Risk Level column (F)
        add_dropdown_validation(ws, "F", ["Critical", "High", "Medium", "Low", "Info"])

        # Change Type column (H) - Rigorous values
        add_dropdown_validation(
            ws,
            "H",
            [
                "✅ Fixed",
                "❌ Regression",
                "🆕 New Issue",
                "🛡️ Exception Documented",
                "🗑️ Exception Removed",
                "📝 Configuration Change",
            ],
        )

        # --- Add Conditional Formatting for dynamic styling ---
        self._add_action_conditional_formatting()

    def _add_action_conditional_formatting(self) -> None:
        """Add Conditional Formatting rules for Risk Level and Change Type columns."""
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import PatternFill, Font

        ws = self._action_sheet
        if ws is None:
            return

        # Define fills and fonts for CF
        pass_fill = PatternFill(
            start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"
        )  # Green
        pass_font = Font(color="1B5E20", bold=True)
        warn_fill = PatternFill(
            start_color="FFE082", end_color="FFE082", fill_type="solid"
        )  # Yellow/Orange
        warn_font = Font(color="E65100", bold=True)
        fail_fill = PatternFill(
            start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
        )  # Red
        fail_font = Font(color="B71C1C", bold=True)

        # New Colors
        info_fill = PatternFill(
            start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
        )  # Light Blue for Info
        info_font = Font(color="0D47A1", bold=True)

        crit_fill = PatternFill(
            start_color="B71C1C", end_color="B71C1C", fill_type="solid"
        )  # Dark Red for Critical
        crit_font = Font(color="FFFFFF", bold=True)

        blue_fill = PatternFill(
            start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"
        )  # Blue for Exceptions

        gray_fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )  # Gray for Removed

        # Risk Level (Column F) - Dynamic based on value
        f_range = f"F2:F{ws.max_row + 500}"

        # Low = Green
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Low",F2))'],
                stopIfTrue=True,
                fill=pass_fill,
                font=pass_font,
            ),
        )
        # Medium = Orange/Yellow
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Medium",F2))'],
                stopIfTrue=True,
                fill=warn_fill,
                font=warn_font,
            ),
        )
        # High = Red
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("High",F2))'],
                stopIfTrue=True,
                fill=fail_fill,
                font=fail_font,
            ),
        )
        # Critical = Dark Red
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Critical",F2))'],
                stopIfTrue=True,
                fill=crit_fill,
                font=crit_font,
            ),
        )
        # Info = Blue
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Info",F2))'],
                stopIfTrue=True,
                fill=info_fill,
                font=info_font,
            ),
        )

        # Change Type (Column H) - Dynamic based on value
        h_range = f"H2:H{ws.max_row + 500}"

        # Fixed = Green
        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Fixed",H2))'],
                stopIfTrue=True,
                fill=pass_fill,
                font=pass_font,
            ),
        )
        # Exception Documented = Blue
        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Exception Documented",H2))'],
                stopIfTrue=True,
                fill=blue_fill,
                # font not specified in old Spec but seems logical to keep black or standard
            ),
        )
        # Exception Removed = Gray
        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Exception Removed",H2))'],
                stopIfTrue=True,
                fill=gray_fill,
            ),
        )
        # Regression = Red
        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Regression",H2))'],
                stopIfTrue=True,
                fill=fail_fill,
                font=fail_font,
            ),
        )
        # New Issue = Yellow (Warn)
        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("New Issue",H2))'],
                stopIfTrue=True,
                fill=warn_fill,
                font=warn_font,
            ),
        )
