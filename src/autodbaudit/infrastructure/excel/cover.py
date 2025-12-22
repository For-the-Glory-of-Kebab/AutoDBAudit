"""
Cover Sheet Module.

Handles the Cover worksheet with audit summary and statistics.
Includes SQL Server branding and professional presentation.
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from autodbaudit.infrastructure.excel_styles import (
    Fonts,
    Fills,
    Alignments,
    Icons,
    Colors,
)
from autodbaudit.infrastructure.excel.base import BaseSheetMixin
from autodbaudit.utils.resources import get_asset_path

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


__all__ = ["CoverSheetMixin"]

logger = logging.getLogger(__name__)

# Local colors for banding
LIGHT_BLUE = "E6F2FF"


class CoverSheetMixin(BaseSheetMixin):
    """Mixin for Cover sheet functionality."""

    _cover_sheet: Worksheet | None = None
    audit_run_id: int | None = None
    organization: str = ""
    audit_name: str = ""
    started_at: datetime | None = None
    ended_at: datetime | None = None

    # Granular change stats (default 0)
    _stats_fixed: int = 0
    _stats_regressed: int = 0
    _stats_new: int = 0
    _stats_docs: int = 0
    _stats_exceptions_changed: int = 0

    def set_audit_info(
        self,
        run_id: int,
        organization: str,
        audit_name: str = "",
        started_at: datetime | None = None,
        ended_at: datetime | None = None,
    ) -> None:
        """
        Set audit metadata for the cover sheet.

        Args:
            run_id: Audit run identifier
            organization: Organization name
            audit_name: Name of the audit (from CLI)
            started_at: Audit start time
            ended_at: Audit end time
        """
        self.audit_run_id = run_id
        self.organization = organization
        self.audit_name = audit_name
        self.started_at = started_at
        self.ended_at = ended_at

    def set_stats_from_service(
        self,
        active_issues: int,
        documented_exceptions: int,
        compliant_items: int,
        # Optional granular stats (kwargs to be safe, or explicit)
        fixed: int = 0,
        regressed: int = 0,
        new_issues: int = 0,
        docs_changed: int = 0,
        exceptions_changed: int = 0,
    ) -> None:
        """
        Set Cover sheet stats from StatsService (unified source of truth).

        This overrides the internal counters with values from StatsService
        to ensure CLI and Cover sheet match exactly.

        Args:
            active_issues: FAIL/WARN without exception
            documented_exceptions: FAIL/WARN with exception
            compliant_items: PASS items
            fixed: Items fixed since baseline/last sync
            regressed: Items regressed
            new_issues: New violations
            docs_changed: Count of doc additions/updates
            exceptions_changed: Count of exception additions/removals/updates
        """
        # Override internal counters
        self._issue_count = active_issues
        self._warn_count = documented_exceptions
        self._pass_count = compliant_items

        # Set granular stats
        self._stats_fixed = fixed
        self._stats_regressed = regressed
        self._stats_new = new_issues
        self._stats_docs = docs_changed
        self._stats_exceptions_changed = exceptions_changed

        logger.info(
            "Cover stats set: issues=%d, exc=%d, pass=%d, fixed=%d, reg=%d",
            active_issues,
            documented_exceptions,
            compliant_items,
            fixed,
            regressed,
        )

    def create_cover_sheet(self) -> None:
        """Create the cover sheet with audit summary and branding."""
        if "Cover" in self.wb.sheetnames:
            return

        # Remove default "Sheet" if present
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        ws = self.wb.create_sheet("Cover", 0)
        self._cover_sheet = ws
        ws.sheet_view.showGridLines = False

        # Set column widths
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 15  # Icon column
        ws.column_dimensions["C"].width = 25  # Labels
        ws.column_dimensions["D"].width = 40  # Values

        # Add SQL Server Audit Icon (Moved to end for Z-Index safety)
        # self._add_cover_image(ws)

        # ═══════════════════════════════════════════════════════════════
        # TITLE SECTION
        # ═══════════════════════════════════════════════════════════════

        row = 3
        # Main Title (The Audit Name)
        ws.merge_cells(f"C{row}:D{row}")
        title_cell = ws[f"C{row}"]
        title_cell.value = (self.audit_name or "SQL Server Security Audit").upper()
        title_cell.font = Font(
            name="Segoe UI", size=24, bold=True, color=Colors.HEADER_BG
        )
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )  # CENTERED
        ws.row_dimensions[row].height = 40
        row += 1

        # Subtitle (Organization + Context)
        ws.merge_cells(f"C{row}:D{row}")
        sub_cell = ws[f"C{row}"]
        sub_cell.value = f"Security Assessment Report | {self.organization}"
        sub_cell.font = Font(name="Segoe UI", size=14, color="555555")
        sub_cell.alignment = Alignment(horizontal="center", vertical="top")  # CENTERED
        ws.row_dimensions[row].height = 25
        row += 2

        # ═══════════════════════════════════════════════════════════════
        # EXECUTIVE SUMMARY BOX
        # ═══════════════════════════════════════════════════════════════

        # Helper to draw a box
        def draw_box_row(
            r,
            c1,
            c2,
            color,
            text,
            font_size=11,
            bold=True,
            align="center",
            text_color="FFFFFF",
        ):
            cell = ws.cell(row=r, column=c1)
            cell.value = text
            cell.font = Font(
                name="Segoe UI", size=font_size, bold=bold, color=text_color
            )
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal=align, vertical="center")

            # If spanning
            if c2 > c1:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
                # Apply fill to merged range
                for c in range(c1, c2 + 1):
                    ws.cell(row=r, column=c).fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

        # Header Bar - VIBRANT DARK BLUE
        draw_box_row(
            row,
            2,
            4,
            Colors.HEADER_BG,
            "  📊 EXECUTIVE SUMMARY",
            font_size=14,
            bold=True,
            text_color="FFFFFF",
            align="center",
        )
        ws.row_dimensions[row].height = 30
        row += 1

        # Stats Row (Current State)
        stats_start_row = row
        summary_items = [
            (
                f"  {Icons.FAIL} Critical Issues",
                self._issue_count,
                "DC3545",
                "FFFFFF",
            ),  # Red
            (
                f"  {Icons.WARN} Warnings Found",
                self._warn_count,
                "FFC107",
                "000000",
            ),  # Yellow/Orange
            (
                f"  {Icons.PASS} Passed Checks",
                self._pass_count,
                "28A745",
                "FFFFFF",
            ),  # Green
        ]

        for label, count, bg_color, text_color in summary_items:
            # Label Column
            draw_box_row(
                row,
                2,
                3,
                bg_color,
                label,
                font_size=12,
                bold=True,
                text_color=text_color,
                align="left",
            )
            # Count Column
            draw_box_row(
                row,
                4,
                4,
                bg_color,
                str(count),
                font_size=12,
                bold=True,
                align="center",
                text_color=text_color,
            )
            ws.row_dimensions[row].height = 25
            row += 1

        # -------------------------------------------------------------------
        # CHANGES SECTION (Dynamic)
        # -------------------------------------------------------------------
        # Calculate if we have any changes to show
        has_changes = (
            self._stats_fixed > 0
            or self._stats_regressed > 0
            or self._stats_new > 0
            or self._stats_docs > 0
            or self._stats_exceptions_changed > 0
        )

        if has_changes:
            row += 1
            # Header
            draw_box_row(
                row,
                2,
                4,
                Colors.SUBHEADER_BG,  # Steel Blue
                "  📈 RECENT ACTIVITY",
                font_size=12,
                bold=True,
                text_color="FFFFFF",
                align="center",
            )
            ws.row_dimensions[row].height = 25
            row += 1

            # Helper for change rows
            def add_change_row(label, count, color="F2F2F2", text_color="000000"):
                if count <= 0:
                    return 0
                nonlocal row
                draw_box_row(
                    row,
                    2,
                    3,
                    color,
                    label,
                    font_size=11,
                    bold=False,
                    text_color=text_color,
                    align="left",
                )
                draw_box_row(
                    row,
                    4,
                    4,
                    color,
                    str(count),
                    font_size=11,
                    bold=True,
                    text_color=text_color,
                    align="center",
                )
                ws.row_dimensions[row].height = 22
                row += 1
                return 1

            # Add rows
            add_change_row(
                f"  {Icons.PASS} Issues Fixed",
                self._stats_fixed,
                Colors.PASS_BG,
                Colors.PASS_TEXT,
            )
            add_change_row(
                f"  {Icons.FAIL} Regressions",
                self._stats_regressed,
                Colors.FAIL_BG,
                Colors.FAIL_TEXT,
            )
            add_change_row(
                f"  {Icons.WARN} New Issues",
                self._stats_new,
                Colors.WARN_BG,
                Colors.WARN_TEXT,
            )

            # Exceptions and Docs as separate rows
            add_change_row(
                f"  🔖 Exception Changes",
                self._stats_exceptions_changed,
                Colors.INFO_BG,
                Colors.INFO_TEXT,
            )
            add_change_row(
                f"  📝 Note Updates",
                self._stats_docs,
                Colors.INFO_BG,
                Colors.INFO_TEXT,
            )

        row += 1

        # ═══════════════════════════════════════════════════════════════
        # AUDIT METADATA
        # ═══════════════════════════════════════════════════════════════

        # Header - VIBRANT DARK BLUE
        draw_box_row(
            row,
            2,
            4,
            Colors.HEADER_BG,
            "  📋 AUDIT METADATA",
            font_size=14,
            bold=True,
            text_color="FFFFFF",
            align="center",
        )
        ws.row_dimensions[row].height = 30
        row += 1

        metadata = [
            (
                "Audit Job ID",
                f"RUN-{self.audit_run_id:03d}" if self.audit_run_id else "N/A",
            ),
            ("Organization", self.organization or "Unspecified"),
            (
                "Start Time",
                (
                    self.started_at.strftime("%Y-%m-%d %H:%M:%S")
                    if self.started_at
                    else "N/A"
                ),
            ),
            (
                "Completion",
                (
                    self.ended_at.strftime("%Y-%m-%d %H:%M:%S")
                    if self.ended_at
                    else "In Progress"
                ),
            ),
            ("Engine Version", "AutoDBAudit v2.0 (E2E Edition)"),
        ]

        for i, (key, value) in enumerate(metadata):
            bg = "E6F2FF" if i % 2 == 0 else "FFFFFF"  # Light Blue for banding

            # Key
            cell_k = ws.cell(row=row, column=2)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            cell_k.value = f"  {key}"
            cell_k.font = Font(name="Segoe UI", size=11, bold=True, color="333333")
            cell_k.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            ws.cell(row=row, column=3).fill = PatternFill(
                start_color=bg, end_color=bg, fill_type="solid"
            )

            # Value
            cell_v = ws.cell(row=row, column=4)
            cell_v.value = value
            cell_v.font = Font(name="Segoe UI", size=11, color="000000")
            cell_v.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell_v.alignment = Alignment(horizontal="right")

            ws.row_dimensions[row].height = 22
            row += 1

        # Footer
        row += 2
        ws.merge_cells(f"B{row}:D{row}")
        footer_cell = ws.cell(row=row, column=2)
        footer_cell.value = (
            f"Generated automatically on {datetime.now().strftime('%Y-%m-%d')}"
        )
        footer_cell.font = Font(name="Segoe UI", size=9, italic=True, color="999999")
        footer_cell.alignment = Alignment(horizontal="center")

        # Add SQL Server Audit Icon (At the end to ensure it sits on top)
        self._add_cover_image(ws)

    def _add_cover_image(self, ws: Worksheet) -> None:
        """Try to add the SQL Server icon to the cover sheet."""
        try:
            # Use robust resource locator
            icon_path = get_asset_path("sql_audit_icon.png")

            img = Image(str(icon_path))

            # Scale to fit nicely in top left corner (approx 2 rows high)
            img.width = 100
            img.height = 100
            ws.add_image(img, "B2")
            logger.info("Successfully added cover icon from %s", icon_path)

        except Exception as e:
            # Log warning if image fails, but don't crash the report
            logger.warning("Failed to load cover image: %s", e)
            pass
