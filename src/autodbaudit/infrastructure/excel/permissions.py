"""
Permission Grants Sheet Module.

Handles the Permission Grants worksheet for explicit GRANT/DENY audits.
Uses ServerGroupMixin for server/instance grouping.
Highlights risky permissions and DENY states.


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fills,
    Fonts,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    ACTION_COLUMN,
    STATUS_COLUMN,
    LAST_REVIEWED_COLUMN,
    apply_action_needed_styling,
)
from autodbaudit.infrastructure.excel.server_group import ServerGroupMixin


__all__ = ["PermissionSheetMixin", "PERMISSION_CONFIG"]


PERMISSION_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Server", 18, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 15, Alignments.LEFT),  # Column D
    ColumnDef("Scope", 10, Alignments.CENTER),  # Column E
    ColumnDef("Database", 20, Alignments.LEFT),  # Column F
    ColumnDef("Grantee", 25, Alignments.CENTER),  # Column G
    ColumnDef("Permission", 25, Alignments.CENTER),  # Column H
    ColumnDef("State", 15, Alignments.CENTER),  # Column I
    ColumnDef("Entity Type", 15, Alignments.CENTER),  # Column J
    ColumnDef("Entity Name", 35, Alignments.CENTER),  # Column K
    ColumnDef("Risk", 12, Alignments.CENTER),  # Column L
    STATUS_COLUMN,  # Review Status dropdown
    ColumnDef("Justification", 35, Alignments.CENTER_WRAP, is_manual=True),
    LAST_REVIEWED_COLUMN,
    ColumnDef("Notes", 30, Alignments.CENTER_WRAP, is_manual=True),
)

PERMISSION_CONFIG = SheetConfig(name="Permission Grants", columns=PERMISSION_COLUMNS)

# Risky permissions that warrant attention
RISKY_PERMISSIONS = frozenset(
    {
        "CONTROL SERVER",
        "ALTER ANY LOGIN",
        "ALTER ANY LINKED SERVER",
        "CONTROL",
        "ALTER",
        "Take Ownership",
    }
)


class PermissionSheetMixin(ServerGroupMixin, BaseSheetMixin):
    """Mixin for Permission Grants sheet with server/instance grouping."""

    _permission_sheet = None

    def add_permission(
        self,
        server_name: str,
        instance_name: str,
        scope: str,
        database_name: str,
        grantee_name: str,
        permission_name: str,
        state: str,
        entity_name: str,
        class_desc: str | None = None,
    ) -> None:
        """Add a permission grant/deny row."""
        if self._permission_sheet is None:
            self._permission_sheet = self._ensure_sheet_with_uuid(PERMISSION_CONFIG)
            # Permission sheet has Scope column at E (5), Database at F (6)
            # Default logic puts DB at 5. We override to 6.
            self._init_grouping(
                self._permission_sheet, PERMISSION_CONFIG, database_col_idx=6
            )
            self._add_permission_dropdowns()

        ws = self._permission_sheet

        # Track grouping and get row color
        row_color = self._track_group(
            server_name,
            instance_name,
            PERMISSION_CONFIG.name,
            database_name=database_name,
        )

        # Analyze Risk
        risk_level = "normal"
        state_upper = state.upper()
        perm_upper = permission_name.upper()

        if state_upper == "DENY":
            risk_level = "deny"  # Not necessarily bad, but important to see
        elif "GRANT" in state_upper and "GRANT" in state_upper.replace("GRANT", "", 1):
            # GRANT_WITH_GRANT_OPTION check (primitive but effective if string contains GRANT twice/WITH GRANT)
            # Actually input is usually 'GRANT_WITH_GRANT_OPTION' from SQL
            if "WITH_GRANT" in state_upper or "WITH GRANT" in state_upper:
                risk_level = "high"

        if perm_upper in RISKY_PERMISSIONS and state_upper != "DENY":
            risk_level = "high"

        # Format State with Icon
        state_display = state
        if state_upper == "DENY":
            state_display = "⛔ DENY"
        elif "WITH" in state_upper:
            state_display = "⚠️ GRANT w/ OPT"
        elif state_upper == "GRANT":
            state_display = "✅ GRANT"

        # Format Permission with Icon
        perm_display = permission_name

        # Standardize common permissions
        perm_key = perm_display.upper().strip()

        # Connection / Session
        if perm_key == "CONNECT SQL":
            perm_display = "🔌 Connect SQL"
        elif perm_key == "CONNECT":
            perm_display = "🔌 Connect"
        elif perm_key == "AUTHENTICATE SERVER":
            perm_display = "🛡️ Authenticate Server"

        # High Privilege (Control/Owner)
        elif perm_key == "CONTROL SERVER":
            perm_display = "👑 Control Server"
        elif perm_key == "CONTROL":
            perm_display = "👑 Control"
        elif perm_key == "TAKE OWNERSHIP":
            perm_display = "👑 Take Ownership"
        elif perm_key == "IMPERSONATE":
            perm_display = "🎭 Impersonate"

        # Data Definition (DDL)
        elif "ALTER" in perm_key:
            perm_display = f"🛠️ {perm_display}"
        elif "CREATE" in perm_key:
            perm_display = f"✨ {perm_display}"
        elif "DROP" in perm_key:
            perm_display = f"🔥 {perm_display}"
        elif "DELETE" in perm_key:
            perm_display = f"❌ {perm_display}"

        # Data Access (DML)
        elif "SELECT" in perm_key:
            perm_display = f"👁️ {perm_display}"
        elif "INSERT" in perm_key:
            perm_display = f"➕ {perm_display}"
        elif "UPDATE" in perm_key:
            perm_display = f"✏️ {perm_display}"
        elif "EXECUTE" in perm_key:
            perm_display = f"⚡ {perm_display}"
        elif "VIEW DEFINITION" in perm_key:
            perm_display = f"🔍 {perm_display}"

        # Format Scope
        scope_display = scope.upper()

        # Determine if action needed (high risk or deny)
        needs_action = risk_level == "high"

        data = [
            None,  # Action indicator (column B)
            server_name,  # Column C
            instance_name or "(Default)",
            scope_display,
            database_name,  # Can be empty for SERVER scope
            grantee_name,
            perm_display,
            state_display,
            class_desc or "",
            entity_name,
            None,  # Risk - styled below
            "",  # Justification
            "",  # Notes
        ]

        row, _ = self._write_row_with_uuid(ws, PERMISSION_CONFIG, data)

        # Apply action indicator (column 1)
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_action)

        # Apply row color to data columns (A=UUID, B=Action, C=Server, D=Instance, E=Scope, F=Database, G=Grantee, H=Permission)
        self._apply_row_color(row, row_color, data_cols=[3, 4, 5, 6, 7, 10, 11], ws=ws)

        # Style State Column (Column I = 9)
        state_cell = ws.cell(row=row, column=9)
        if state_upper == "DENY":
            state_cell.fill = Fills.FAIL
            state_cell.font = Fonts.FAIL
        elif risk_level == "high":
            state_cell.fill = Fills.WARN
            state_cell.font = Fonts.WARN

        # Style Risk Column (Column L = 12)
        risk_cell = ws.cell(row=row, column=12)
        if risk_level == "high":
            risk_cell.value = "🔴 Risk"
            risk_cell.fill = Fills.FAIL
            risk_cell.font = Fonts.FAIL
        elif risk_level == "deny":
            risk_cell.value = "⛔ Blocked"
            risk_cell.fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )  # Light red
        else:
            risk_cell.value = "—"
            risk_cell.fill = PatternFill(
                start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
            )

    def _finalize_permissions(self) -> None:
        """Finalize permissions sheet."""
        if self._permission_sheet:
            self._finalize_grouping(PERMISSION_CONFIG.name)
            self._finalize_sheet_with_uuid(self._permission_sheet)

    def _add_permission_dropdowns(self) -> None:
        """Add validation dropdowns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation,
            add_review_status_conditional_formatting,
            STATUS_VALUES,
        )

        ws = self._permission_sheet
        # Column layout: A=UUID, B=Action, C=Server, D=Instance, E=Scope, F=Database,
        #                G=Grantee, H=Permission, I=State, J=EntityType, K=EntityName,
        #                L=Risk, M=ReviewStatus, N=Justification, O=LastReviewed, P=Notes

        # Scope (E) - column 5
        add_dropdown_validation(ws, "E", ["SERVER", "DATABASE"])
        # State (I) - column 9
        add_dropdown_validation(ws, "I", ["✅ GRANT", "⛔ DENY", "⚠️ GRANT w/ OPT"])
        # Review Status column (M) - column 13
        add_dropdown_validation(ws, "M", STATUS_VALUES.all())
        add_review_status_conditional_formatting(ws, "M")
