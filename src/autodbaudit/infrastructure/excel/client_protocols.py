"""
Client Protocols Sheet Module.

Handles the Client Protocols worksheet for network protocol security audit.
Uses ServerGroupMixin for server/instance grouping.

Discrepancy Logic:
- Acceptable: Shared Memory + TCP/IP (enabled by default for network connectivity)
- Discrepant: Named Pipes, VIA, or other protocols (need justification if enabled)


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fills,
    Fonts,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    ACTION_COLUMN,
    STATUS_COLUMN,
    LAST_REVIEWED_COLUMN,
    apply_action_needed_styling,
)
from autodbaudit.infrastructure.excel.server_group import ServerGroupMixin


__all__ = ["ClientProtocolSheetMixin", "CLIENT_PROTOCOL_CONFIG"]


# Acceptable protocols that don't need justification
ACCEPTABLE_PROTOCOLS = frozenset(
    {
        "shared memory",
        "tcp/ip",
    }
)

# Protocols that need justification if enabled
DISCREPANT_PROTOCOLS = frozenset(
    {
        "named pipes",
        "via",
    }
)


CLIENT_PROTOCOL_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Server", 18, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 15, Alignments.LEFT),  # Column D
    ColumnDef("Protocol", 18, Alignments.LEFT),  # Column E
    ColumnDef("Enabled", 10, Alignments.CENTER),  # Column F
    ColumnDef("Port", 10, Alignments.CENTER),  # Column G
    ColumnDef("Status", 14, Alignments.CENTER),  # Column H
    ColumnDef("Notes", 30, Alignments.CENTER_WRAP),  # Column I
    STATUS_COLUMN,  # Review Status dropdown
    ColumnDef("Justification", 40, Alignments.CENTER_WRAP, is_manual=True),
    LAST_REVIEWED_COLUMN,
)

CLIENT_PROTOCOL_CONFIG = SheetConfig(
    name="Client Protocols", columns=CLIENT_PROTOCOL_COLUMNS
)


class ClientProtocolSheetMixin(ServerGroupMixin, BaseSheetMixin):
    """Mixin for Client Protocols sheet with server/instance grouping."""

    _client_protocol_sheet = None

    def add_client_protocol(
        self,
        server_name: str,
        instance_name: str,
        protocol_name: str,
        is_enabled: bool,
        port: int | None = None,
        notes: str = "",
    ) -> None:
        """Add a client protocol row.

        Args:
            server_name: Server hostname
            instance_name: Instance name
            protocol_name: Protocol name (Shared Memory, TCP/IP, Named Pipes, VIA)
            is_enabled: Whether protocol is enabled
            port: Port number (for TCP/IP)
            notes: Additional notes about configuration
        """
        if self._client_protocol_sheet is None:
            self._client_protocol_sheet = self._ensure_sheet_with_uuid(
                CLIENT_PROTOCOL_CONFIG
            )
            # Disable database merging (database_col_idx=0) to prevent Protocol column from being merged
            self._init_grouping(
                self._client_protocol_sheet, CLIENT_PROTOCOL_CONFIG, database_col_idx=0
            )
            self._add_protocol_dropdowns()

        ws = self._client_protocol_sheet

        # Track grouping and get row color
        row_color = self._track_group(
            server_name, instance_name, CLIENT_PROTOCOL_CONFIG.name
        )

        # Determine discrepancy status
        protocol_lower = protocol_name.lower().strip()
        is_acceptable = protocol_lower in ACCEPTABLE_PROTOCOLS
        is_discrepant = protocol_lower in DISCREPANT_PROTOCOLS

        # Needs action if: enabled AND not an acceptable protocol
        needs_action = is_enabled and not is_acceptable

        # Determine status
        if is_acceptable and is_enabled:
            status = "✅ Compliant"
        elif not is_enabled and is_discrepant:
            status = "✅ Disabled"  # Good - discrepant protocol is off
        elif is_enabled and is_discrepant:
            status = "⚠️ Needs Review"  # Enabled discrepant protocol
        elif not is_enabled and is_acceptable:
            status = "ℹ️ Disabled"  # Acceptable but turned off
        else:
            status = "—"

        data = [
            None,  # Action indicator (column B)
            server_name,  # Column C
            instance_name or "(Default)",
            protocol_name,
            None,  # Enabled - styled separately
            str(port) if port else "—",
            None,  # Status - styled separately
            notes or "",
            "",  # Justification
            "",  # Last Revised
        ]

        row, row_uuid = self._write_row_with_uuid(ws, CLIENT_PROTOCOL_CONFIG, data)

        # Apply action indicator (column 1)
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_action)

        # Apply row color to data columns (A=UUID, B=Action, C=Server, D=Instance, E=Protocol, F=Enabled, G=Port, H=Status, I=Notes)
        # Fix: Ensure Port (7) is colored.
        self._apply_row_color(row, row_color, data_cols=[3, 4, 5, 7, 9], ws=ws)

        # Style Enabled column (column F = 6) - Initial static style (CF will override if value changes)
        enabled_cell = ws.cell(row=row, column=6)
        if is_enabled:
            enabled_cell.value = "✓ Yes"
            if is_acceptable:
                enabled_cell.fill = Fills.PASS
                enabled_cell.font = Fonts.PASS
            else:
                enabled_cell.fill = Fills.WARN
                enabled_cell.font = Fonts.WARN
        else:
            enabled_cell.value = "✗ No"
            # DISABLED style
            enabled_cell.fill = PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            )
            enabled_cell.font = Fonts.DATA

        # Style Status column (column H = 8)
        status_cell = ws.cell(row=row, column=8)
        status_cell.value = status
        if "Compliant" in status or (status == "✅ Disabled"):
            status_cell.fill = Fills.PASS
            status_cell.font = Fonts.PASS
        elif "Needs Review" in status:
            status_cell.fill = Fills.WARN
            status_cell.font = Fonts.WARN
            self._increment_warn()
        elif "Disabled" in status:  # Info Disabled
            status_cell.fill = PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            )

    def _finalize_client_protocols(self) -> None:
        """Finalize client protocols sheet - merge remaining groups."""
        if self._client_protocol_sheet:
            self._finalize_grouping(CLIENT_PROTOCOL_CONFIG.name)
            self._finalize_sheet_with_uuid(self._client_protocol_sheet)

    def _add_protocol_dropdowns(self) -> None:
        """Add dropdown validations for choice columns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation,
            add_review_status_conditional_formatting,
            STATUS_VALUES,
        )

        ws = self._client_protocol_sheet
        # Enabled column (F) - column 6 (Action=B, Server=C, Instance=D, Protocol=E)
        add_dropdown_validation(ws, "F", ["✓ Yes", "✗ No"])
        # Status column (H) - column 8
        add_dropdown_validation(
            ws,
            "H",
            ["✅ Compliant", "✅ Disabled", "⚠️ Needs Review", "ℹ️ Disabled", "—"],
        )
        # Review Status column (J) - column 10 (Notes=I)
        add_dropdown_validation(ws, "J", STATUS_VALUES.all())
        add_review_status_conditional_formatting(ws, "J")

        # --- Dynamic CF for Enabled/Status ---
        from openpyxl.formatting.rule import FormulaRule
        from autodbaudit.infrastructure.excel_styles import Fills, Fonts

        # Enabled (Col F)
        # Yes -> Check if acceptable?
        # Requires logic. Actually, "Yes" isn't always good (Named Pipes).
        # We can't easily do complex logic in Excel CF without helper columns.
        # But we can do basic: "✓ Yes" -> Green? No, Named Pipes Yes is Bad.
        # Protocol name is in Col E.
        # Rule: =AND($F2="✓ Yes", OR($E2="Shared Memory", $E2="TCP/IP")) -> Green
        # Rule: =AND($F2="✓ Yes", NOT(OR($E2="Shared Memory", $E2="TCP/IP"))) -> Warn (Orange)
        # Rule: =$F2="✗ No" -> Gray

        f_range = f"F2:F{ws.max_row+100}"

        # 1. Enabled + Safe = Green
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=[
                    'AND(ISNUMBER(SEARCH("Yes",F2)), OR(ISNUMBER(SEARCH("Shared Memory",E2)), ISNUMBER(SEARCH("TCP/IP",E2))))'
                ],
                stopIfTrue=True,
                fill=Fills.PASS,
                font=Fonts.PASS,
            ),
        )

        # 2. Enabled + Unsafe = Warn
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=[
                    'AND(ISNUMBER(SEARCH("Yes",F2)), NOT(OR(ISNUMBER(SEARCH("Shared Memory",E2)), ISNUMBER(SEARCH("TCP/IP",E2)))))'
                ],
                stopIfTrue=True,
                fill=Fills.WARN,
                font=Fonts.WARN,
            ),
        )

        # 3. Disabled = Gray
        from openpyxl.styles import PatternFill

        gray_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        ws.conditional_formatting.add(
            f_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("No",F2))'], stopIfTrue=True, fill=gray_fill
            ),
        )

        # Status (Col H)
        # Compliant/Disabled(Good) -> Green
        # Needs Review -> Orange
        h_range = f"H2:H{ws.max_row+100}"

        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=['OR(ISNUMBER(SEARCH("Compliant",H2)), H2="✅ Disabled")'],
                stopIfTrue=True,
                fill=Fills.PASS,
                font=Fonts.PASS,
            ),
        )

        ws.conditional_formatting.add(
            h_range,
            FormulaRule(
                formula=['ISNUMBER(SEARCH("Needs Review",H2))'],
                stopIfTrue=True,
                fill=Fills.WARN,
                font=Fonts.WARN,
            ),
        )
