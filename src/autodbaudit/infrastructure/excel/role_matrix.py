"""
Role Matrix Sheet Module.

Handles the Role Matrix worksheet for a pivoted view of database role memberships.
Requirement #27: Visual Matrix of Login-to-Role Mapping.


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fills,
    Fonts,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
)
from autodbaudit.infrastructure.excel.server_group import ServerGroupMixin


__all__ = ["RoleMatrixSheetMixin", "ROLE_MATRIX_CONFIG"]


# Fixed roles to show as individual columns
FIXED_ROLES = [
    "db_owner",
    "db_securityadmin",
    "db_accessadmin",
    "db_backupoperator",
    "db_ddladmin",
    "db_datareader",
    "db_datawriter",
    "db_denydatareader",
    "db_denydatawriter",
]

# Column Definitions
# Server, Instance, Database, Principal, Type, [Fixed Roles...], Other Roles, Risk
# NOTE: Role Matrix is info-only (Q3 decision). Justifications go in Database Roles sheet.
# NO ACTION_COLUMN - this is an info-only sheet (no ⏳ indicator)
ROLE_MATRIX_COLUMNS = [
    ColumnDef("Server", 18, Alignments.LEFT),
    ColumnDef("Instance", 15, Alignments.LEFT),
    ColumnDef("Database", 20, Alignments.LEFT),
    ColumnDef("Principal Name", 25, Alignments.LEFT),
    ColumnDef("Principal Type", 18, Alignments.CENTER),
]

# Add columns for each fixed role (narrow, centered)
for role in FIXED_ROLES:
    # Use shorter headers if needed, or rotate text in final styler if we were fancy
    # For now, just the name. 12 width fits most.
    ROLE_MATRIX_COLUMNS.append(ColumnDef(role, 12, Alignments.CENTER))

# Catch-all for custom roles
ROLE_MATRIX_COLUMNS.append(ColumnDef("Other Roles", 40, Alignments.LEFT))
ROLE_MATRIX_COLUMNS.append(ColumnDef("Risk", 12, Alignments.CENTER))

ROLE_MATRIX_CONFIG = SheetConfig(name="Role Matrix", columns=tuple(ROLE_MATRIX_COLUMNS))


class RoleMatrixSheetMixin(ServerGroupMixin, BaseSheetMixin):
    """Mixin for Role Matrix sheet."""

    _role_matrix_sheet = None

    def add_role_matrix_row(
        self,
        server_name: str,
        instance_name: str,
        database_name: str,
        principal_name: str,
        principal_type: str,
        roles: list[str],
    ) -> None:
        """
        Add a matrix row for a principal.

        Args:
            server_name: Server hostname
            instance_name: Instance name
            database_name: Database name
            principal_name: User/Principal name
            principal_type: SQL_USER, WINDOWS_USER, etc.
            roles: List of role names this user belongs to
        """
        if self._role_matrix_sheet is None:
            self._role_matrix_sheet = self._ensure_sheet_with_uuid(ROLE_MATRIX_CONFIG)
            # Role Matrix has NO ACTION COLUMN, so pass has_action_col=False
            self._init_grouping(
                self._role_matrix_sheet, ROLE_MATRIX_CONFIG, has_action_col=False
            )
            self._add_role_matrix_dropdowns()

        ws = self._role_matrix_sheet

        # Track grouping and get row color
        row_color = self._track_group(
            server_name, instance_name, ROLE_MATRIX_CONFIG.name, database_name
        )

        # Prepare Role Flags
        roles_lower = {r.lower() for r in roles}

        # Format Principal Type
        type_upper = (principal_type or "").upper()
        if "WINDOWS" in type_upper:
            type_display = "🪟 Windows"
        elif "SQL" in type_upper:
            type_display = "👤 SQL"
        elif "CERTIFICATE" in type_upper:
            type_display = "📜 Cert"
        elif "ASYMMETRIC" in type_upper:
            type_display = "🔑 Key"
        elif "ROLE" in type_upper:
            type_display = "📦 Role"
        else:
            type_display = principal_type

        row_data = [
            server_name,
            instance_name or "(Default)",
            database_name,
            principal_name,
            type_display,
        ]

        # Add column for each fixed role
        has_high_risk = False
        for fixed_role in FIXED_ROLES:
            if fixed_role in roles_lower:
                if fixed_role == "db_owner" and principal_name.lower() != "dbo":
                    row_data.append("👑 YES")
                    has_high_risk = True
                else:
                    row_data.append("✓")
            else:
                row_data.append("")  # Empty if not member

        # Other Roles (non-fixed)
        other_roles = [r for r in roles if r.lower() not in FIXED_ROLES]
        row_data.append(", ".join(sorted(other_roles)))

        # Risk
        if has_high_risk:
            row_data.append("🔴 High")
        else:
            row_data.append("—")

        row, _ = self._write_row_with_uuid(ws, ROLE_MATRIX_CONFIG, row_data)

        # Column Layout (1-based, with UUID, NO ACTION column):
        # 1=UUID (hidden), 2=Server, 3=Instance, 4=DB, 5=Principal, 6=Type, 7+=Roles, Last=Risk
        # Apply row color to meta columns (excluding UUID=1)
        meta_cols = [2, 3, 4, 5, 6]  # Server, Instance, DB, Principal, Type
        self._apply_row_color(row, row_color, data_cols=meta_cols, ws=ws)

        # Role columns start at 7 (after Type=6)
        start_col = 7

        # Style role cells based on membership
        for i, fixed_role in enumerate(FIXED_ROLES):
            col_idx = start_col + i

            if fixed_role in roles_lower:
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = Alignments.CENTER
                if fixed_role == "db_owner" and principal_name.lower() != "dbo":
                    cell.fill = Fills.FAIL
                    cell.font = Fonts.FAIL
                else:
                    cell.font = Fonts.PASS  # Checkmark in green

        # Style Principal Name (Col 5, was 6 before ACTION_COLUMN removal)
        # Highlighting risky principals like 'sa' or 'Guest'
        if principal_name.lower() in ("sa", "guest", "public"):
            ws.cell(row=row, column=5).font = Fonts.WARN

        # Style Risk column (last column, accounting for UUID offset)
        # len(ROLE_MATRIX_COLUMNS) gives config count, +1 for UUID = actual Excel column
        risk_col_idx = len(ROLE_MATRIX_COLUMNS) + 1
        risk_cell = ws.cell(row=row, column=risk_col_idx)
        if has_high_risk:
            risk_cell.fill = Fills.FAIL
            risk_cell.font = Fonts.FAIL
        else:
            risk_cell.fill = PatternFill(
                start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
            )

    def _finalize_role_matrix(self) -> None:
        """Finalize Role Matrix sheet - merge remaining groups."""
        if self._role_matrix_sheet:
            self._finalize_grouping(ROLE_MATRIX_CONFIG.name)
            self._finalize_sheet_with_uuid(self._role_matrix_sheet)

    def _add_role_matrix_dropdowns(self) -> None:
        """Add dropdown validations for role columns."""
        from autodbaudit.infrastructure.excel.base import add_dropdown_validation
        from openpyxl.utils import get_column_letter

        ws = self._role_matrix_sheet
        # Column layout: 1=UUID, 2=Server, 3=Inst, 4=DB, 5=Principal, 6=Type, 7+=Roles
        # NO ACTION column (info-only sheet)
        start_col = 7
        for i in range(len(FIXED_ROLES)):
            col_letter = get_column_letter(start_col + i)
            add_dropdown_validation(ws, col_letter, ["✓", "👑 YES"])
