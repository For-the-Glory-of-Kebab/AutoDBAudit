"""
Database Roles Sheet Module.

Handles the Database Roles worksheet for per-database role membership audit.
Uses ServerGroupMixin for server/instance grouping.
Enhanced with visual icons for sensitive roles.


UUID Support (v3):
    - Column A: Hidden UUID for stable row identification
    - All other columns shifted +1 from original positions
"""

from __future__ import annotations

from openpyxl.styles import PatternFill

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fills,
    Fonts,
)
from autodbaudit.infrastructure.excel.base import (
    BaseSheetMixin,
    SheetConfig,
    ACTION_COLUMN,
    STATUS_COLUMN,
    LAST_REVIEWED_COLUMN,
    apply_action_needed_styling,
)
from autodbaudit.infrastructure.excel.server_group import ServerGroupMixin


__all__ = ["DBRoleSheetMixin", "DB_ROLE_CONFIG"]


DB_ROLE_COLUMNS = (
    ACTION_COLUMN,  # Column B: Action indicator (A=UUID hidden)
    ColumnDef("Server", 18, Alignments.LEFT),  # Column C
    ColumnDef("Instance", 15, Alignments.LEFT),  # Column D
    ColumnDef("Database", 20, Alignments.LEFT),  # Column E
    ColumnDef("Role", 22, Alignments.CENTER),
    ColumnDef("Member", 25, Alignments.CENTER),
    ColumnDef("Member Type", 18, Alignments.CENTER),
    ColumnDef("Risk", 12, Alignments.CENTER),
    STATUS_COLUMN,  # Review Status dropdown
    ColumnDef("Justification", 45, Alignments.CENTER_WRAP, is_manual=True),
    LAST_REVIEWED_COLUMN,
)

DB_ROLE_CONFIG = SheetConfig(name="Database Roles", columns=DB_ROLE_COLUMNS)

# High-risk roles that grant significant database access
HIGH_RISK_ROLES = frozenset(
    {
        "db_owner",
    }
)

# Medium-risk roles that grant administrative capabilities
MEDIUM_RISK_ROLES = frozenset(
    {
        "db_securityadmin",
        "db_accessadmin",
        "db_backupoperator",
        "db_ddladmin",
    }
)

# Low-risk but notable roles
LOW_RISK_ROLES = frozenset(
    {
        "db_datareader",
        "db_datawriter",
    }
)


class DBRoleSheetMixin(ServerGroupMixin, BaseSheetMixin):
    """Mixin for Database Roles sheet with server/instance grouping."""

    _db_role_sheet = None

    def add_db_role_member(
        self,
        server_name: str,
        instance_name: str,
        database_name: str,
        role_name: str,
        member_name: str,
        member_type: str,
    ) -> None:
        """Add a database role membership row with risk assessment."""
        if self._db_role_sheet is None:
            self._db_role_sheet = self._ensure_sheet_with_uuid(DB_ROLE_CONFIG)
            self._init_grouping(self._db_role_sheet, DB_ROLE_CONFIG)
            self._add_db_role_dropdowns()

        ws = self._db_role_sheet

        # Track grouping and get row color
        row_color = self._track_group(
            server_name, instance_name, DB_ROLE_CONFIG.name, database_name=database_name
        )

        # dbo is always db_owner by design - not a discrepancy
        member_lower = member_name.lower()
        is_dbo = member_lower == "dbo"

        # Determine risk level
        role_lower = role_name.lower()
        if role_lower in HIGH_RISK_ROLES and not is_dbo:
            # db_owner is high risk EXCEPT for dbo (which is expected)
            risk_level = "high"
        elif role_lower in MEDIUM_RISK_ROLES:
            risk_level = "medium"
        elif role_lower in LOW_RISK_ROLES:
            risk_level = "low"
        else:
            risk_level = "normal"  # Custom roles, public, or dbo as db_owner

        # Format role name with icon
        role_display = role_name
        if role_lower == "db_owner":
            role_display = "👑 db_owner"
        elif role_lower in MEDIUM_RISK_ROLES:
            role_display = f"⚙️ {role_name}"

        # Format member type with icon
        type_lower = (member_type or "").lower()
        type_display = member_type
        if "windows" in type_lower:
            type_display = "🪟 Windows"
        elif "sql" in type_lower:
            type_display = "🔑 SQL"
        elif "role" in type_lower:
            type_display = "📦 Role"

        # Needs action if high risk (db_owner for non-dbo)
        needs_action = risk_level == "high"

        data = [
            None,  # Action indicator (column B)
            server_name,  # Column C
            instance_name or "(Default)",
            database_name,
            role_display,
            member_name,
            type_display,
            None,  # Risk - styled separately
            "",  # Justification
        ]

        row, _ = self._write_row_with_uuid(ws, DB_ROLE_CONFIG, data)

        # Apply action indicator (column 1)
        apply_action_needed_styling(ws.cell(row=row, column=2), needs_action)

        # Apply row color to data columns (A=UUID, B=Action, C=Server, D=Instance, E=Database)
        self._apply_row_color(row, row_color, data_cols=[3, 4, 5, 7, 8], ws=ws)

        # Style Role column (column F = 6)
        role_cell = ws.cell(row=row, column=6)
        if risk_level == "high":
            role_cell.fill = Fills.FAIL
            role_cell.font = Fonts.FAIL
            self._increment_warn()  # db_owner needs justification
        elif risk_level == "medium":
            role_cell.fill = Fills.WARN
            role_cell.font = Fonts.WARN

        # Style Risk column (column I = 9)
        risk_cell = ws.cell(row=row, column=9)
        if risk_level == "high":
            risk_cell.value = "🔴 High"
            risk_cell.fill = Fills.FAIL
            risk_cell.font = Fonts.FAIL
        elif risk_level == "medium":
            risk_cell.value = "🟡 Medium"
            risk_cell.fill = Fills.WARN
            risk_cell.font = Fonts.WARN
        elif risk_level == "low":
            risk_cell.value = "🟢 Low"
            risk_cell.fill = Fills.PASS
        else:
            risk_cell.value = "—"
            risk_cell.fill = PatternFill(
                start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
            )

    def _finalize_db_roles(self) -> None:
        """Finalize db roles sheet - merge remaining groups."""
        if self._db_role_sheet:
            self._finalize_grouping(DB_ROLE_CONFIG.name)
            self._finalize_sheet_with_uuid(self._db_role_sheet)

    def _add_db_role_dropdowns(self) -> None:
        """Add dropdown validations for role/status columns."""
        from autodbaudit.infrastructure.excel.base import (
            add_dropdown_validation,
            add_review_status_conditional_formatting,
            STATUS_VALUES,
        )

        ws = self._db_role_sheet
        # Role column (E) - column 5 (Action=A, Server=B, Instance=C, DB=D)
        add_dropdown_validation(
            ws,
            "F",
            [
                "👑 db_owner",
                "⚙️ db_securityadmin",
                "⚙️ db_accessadmin",
                "⚙️ db_backupoperator",
                "⚙️ db_ddladmin",
                "📖 db_datareader",
                "✏️ db_datawriter",
                "db_denydatareader",
                "db_denydatawriter",
                "public",
                "(Custom)",
            ],
        )
        # Member Type column (G) - column 7 (after Role=E, Member=F)
        add_dropdown_validation(ws, "H", ["🪟 Windows", "🔑 SQL", "📦 Role"])
        # Risk column (H) - column 8
        add_dropdown_validation(ws, "I", ["🔴 High", "🟡 Medium", "🟢 Low", "—"])
        # Review Status column (I) - column 9
        add_dropdown_validation(ws, "J", STATUS_VALUES.all())
        add_review_status_conditional_formatting(ws, "J")
