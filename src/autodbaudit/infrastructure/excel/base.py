"""
Excel Report Base Module.

Contains shared utilities, column definitions, and base classes
for all sheet modules.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import datetime
from typing import TYPE_CHECKING, Any

from openpyxl.worksheet.worksheet import Worksheet

from autodbaudit.infrastructure.excel_styles import (
    ColumnDef,
    Alignments,
    Fonts,
    Fills,
    Borders,
    Icons,
    apply_header_row,
    freeze_panes,
    add_autofilter,
)

if TYPE_CHECKING:
    from openpyxl import Workbook


logger = logging.getLogger(__name__)


__all__ = [
    "ColumnDef",
    "Alignments",
    "SheetConfig",
    "BaseSheetMixin",
    "format_date",
    "format_size_mb",
    "get_sql_year",
    "add_dropdown_validation",
    "add_review_status_conditional_formatting",
    "LAST_REVISED_COLUMN",
    "LAST_REVIEWED_COLUMN",
    "STATUS_COLUMN",
    "ACTION_COLUMN",
    "apply_action_needed_styling",
    "apply_exception_documented_styling",
    "apply_review_status_styling",
    "parse_datetime_flexible",
    "STATUS_VALUES",
]


# ============================================================================
# Status Column Values (Dropdown Options)
# ============================================================================


class StatusValues:
    """Valid values for the Status dropdown column.

    Two values only:
    - NEEDS_REVIEW: Item requires auditor attention (default for discrepancies)
    - EXCEPTION: Auditor has documented a justification/exception
    """

    NEEDS_REVIEW = "⏳ Needs Review"
    EXCEPTION = "✓ Exception"

    @classmethod
    def all(cls) -> list[str]:
        """Return all valid status values for dropdown."""
        return [cls.NEEDS_REVIEW, cls.EXCEPTION]

    @classmethod
    def is_documented(cls, value: str | None) -> bool:
        """Check if a status value indicates documented exception."""
        if not value:
            return False
        value_clean = str(value).strip()
        return cls.EXCEPTION in value_clean or "Exception" in value_clean


STATUS_VALUES = StatusValues


# ============================================================================
# Standard Column Definitions
# ============================================================================

# Reusable "Last Revised" column (legacy name kept for compatibility)
LAST_REVISED_COLUMN = ColumnDef(
    name="Last Revised",
    width=12,
    alignment=Alignments.CENTER,
    is_manual=True,
)

# Reusable "Last Reviewed" column - when auditor last reviewed this row
LAST_REVIEWED_COLUMN = ColumnDef(
    name="Last Reviewed",
    width=14,
    alignment=Alignments.CENTER,
    is_manual=True,
)

# Reusable "Review Status" column - dropdown for exception/review status
STATUS_COLUMN = ColumnDef(
    name="Review Status",
    width=14,  # Wide enough for dropdown content
    alignment=Alignments.CENTER,
    is_manual=True,
)

# Reusable "Action" indicator column - shows ⏳ for rows needing attention
ACTION_COLUMN = ColumnDef(
    name="⏳",
    width=4,
    alignment=Alignments.CENTER,
)


def apply_action_needed_styling(cell, needs_action: bool) -> None:
    """
    Apply action-needed indicator to a cell.

    Shows ⏳ icon with orange background for rows needing user attention
    (i.e., FAIL/WARN status with no Notes/justification).
    """
    if needs_action:
        cell.value = Icons.PENDING
        cell.fill = Fills.ACTION
        cell.font = Fonts.WARN
        cell.alignment = Alignments.CENTER
    else:
        cell.value = ""


def apply_exception_documented_styling(cell) -> None:
    """
    Apply documented-exception indicator to a cell.

    Shows ✅ icon with blue/info background for rows where a FAIL/WARN
    has been documented with a justification. Indicates "addressed but
    not fixed" - an acceptable deviation from policy.
    """
    cell.value = Icons.PASS  # ✅
    cell.fill = Fills.INFO  # Blue background
    cell.font = Fonts.INFO
    cell.alignment = Alignments.CENTER


def apply_review_status_styling(cell, value: str | None = None) -> None:
    """
    Apply conditional formatting to Review Status cell based on value.

    Colors:
    - ⏳ Needs Review: Orange/warn (needs attention)
    - ✓ Exception: Green (documented deviation)
    - ✓ Reviewed: Green (approved)
    - Empty: No styling
    """
    if value is None:
        value = cell.value

    if not value:
        return

    value_str = str(value).strip()

    if STATUS_VALUES.NEEDS_REVIEW in value_str or "Needs Review" in value_str:
        cell.fill = Fills.WARN
        cell.font = Fonts.WARN
    elif STATUS_VALUES.EXCEPTION in value_str or "Exception" in value_str:
        cell.fill = Fills.PASS
        cell.font = Fonts.PASS

    cell.alignment = Alignments.CENTER


# ============================================================================
# Utility Functions
# ============================================================================


def format_date(value: Any) -> str:
    """Format a date value for display."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value)
    return s[:10] if len(s) >= 10 else s


def parse_datetime_flexible(
    value: Any,
    log_errors: bool = True,
    context: str = "",
) -> datetime | None:
    """
    Parse a datetime value from Excel with robust handling of multiple formats.

    Handles:
    - Excel numeric dates (float/int representing days since 1900)
    - datetime objects (passthrough)
    - String formats: MM/DD/YYYY, YYYY-MM-DD, DD/MM/YYYY, with optional time
    - Extra whitespace (trimmed)
    - Various separators (/, -, .)

    Args:
        value: The cell value to parse (can be datetime, float, int, str, None)
        log_errors: If True, log parse failures (don't fail silently)
        context: Optional context string for logging (e.g., "Server Logins row 5")

    Returns:
        datetime object if parsed successfully, None otherwise

    Examples:
        >>> parse_datetime_flexible("12/16/2025")
        datetime(2025, 12, 16, 0, 0)
        >>> parse_datetime_flexible("2025-12-16 10:30")
        datetime(2025, 12, 16, 10, 30)
        >>> parse_datetime_flexible(45678.5)  # Excel date number
        datetime(2025, 1, 15, 12, 0)
    """
    if value is None or value == "":
        return None

    # Already a datetime
    if isinstance(value, datetime):
        return value

    # Excel numeric date (days since 1899-12-30, with fractional days as time)
    if isinstance(value, (int, float)):
        try:
            # Excel epoch is 1899-12-30 (with the famous 1900 leap year bug)
            from datetime import timedelta

            excel_epoch = datetime(1899, 12, 30)
            return excel_epoch + timedelta(days=float(value))
        except (ValueError, OverflowError) as e:
            if log_errors:
                logger.debug(
                    "DateTime parse: Excel numeric %s failed: %s %s",
                    value,
                    e,
                    f"[{context}]" if context else "",
                )
            return None

    # String parsing
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None

        # Normalize separators and whitespace
        # Replace multiple spaces with single space
        text = re.sub(r"\s+", " ", text)

        # Try various formats (most common first for performance)
        formats = [
            # ISO formats (standard & with T)
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            # US formats (MM/DD/YYYY)
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y",
            # US formats with AM/PM (Excel keyboard shortcuts: Ctrl+; and Ctrl+Shift+;)
            "%m/%d/%Y %I:%M:%S %p",
            "%m/%d/%Y %I:%M %p",
            # European formats (DD/MM/YYYY) - try after US
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y",
            # With dots (common in some locales)
            "%Y.%m.%d",
            "%d.%m.%Y",
            # With dashes in different order
            "%d-%m-%Y %H:%M:%S",
            "%d-%m-%Y %H:%M",
            "%d-%m-%Y",
            # With forward slashes YYYY/MM/DD (rare but possible and standard in some systems)
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d",
            # Dot separated YYYY.MM.DD
            "%Y.%m.%d %H:%M:%S",
            "%Y.%m.%d",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue

        # If all formats fail, log it
        if log_errors:
            logger.warning(
                "DateTime parse failed for value '%s' %s - tried %d formats",
                text[:50],
                f"[{context}]" if context else "",
                len(formats),
            )
        return None

    # Unknown type
    if log_errors:
        logger.debug(
            "DateTime parse: unknown type %s for value %s %s",
            type(value).__name__,
            repr(value)[:50],
            f"[{context}]" if context else "",
        )
    return None


def format_size_mb(value: Any) -> str:
    """Format a size in MB for display."""
    if value is None:
        return ""
    try:
        return f"{float(value):,.1f}"
    except (ValueError, TypeError):
        return str(value)


def get_sql_year(version_major: int) -> str:
    """Convert major version number to SQL Server year."""
    mapping = {
        10: "2008",
        11: "2012",
        12: "2014",
        13: "2016",
        14: "2017",
        15: "2019",
        16: "2022",
        17: "2025",
    }
    return mapping.get(version_major, f"v{version_major}")


def add_dropdown_validation(
    ws: Worksheet,
    column_letter: str,
    options: list[str],
    start_row: int = 2,
    end_row: int = 1000,
) -> None:
    """Add dropdown data validation to a column.

    Args:
        ws: The worksheet to add validation to
        column_letter: Column letter (e.g., "E", "F")
        options: List of valid options for the dropdown
        start_row: Starting row (default 2 to skip header)
        end_row: Ending row (default 1000)
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    # Create comma-separated list for formula
    formula = '"' + ",".join(options) + '"'

    dv = DataValidation(
        type="list",
        formula1=formula,
        showDropDown=False,  # False = show dropdown arrow (counterintuitive)
        allow_blank=True,
    )
    dv.error = f"Please select from: {', '.join(options)}"
    dv.errorTitle = "Invalid Value"
    ws.add_data_validation(dv)
    dv.add(f"{column_letter}{start_row}:{column_letter}{end_row}")


def add_review_status_conditional_formatting(
    ws: Worksheet,
    column_letter: str,
    start_row: int = 2,
    end_row: int = 1000,
) -> None:
    """Add Excel conditional formatting rules for Review Status column.

    Creates persistent CF rules that appear in Excel's Conditional Formatting manager:
    - "✓ Exception" -> Green fill (documented/approved)
    - "⏳ Needs Review" -> Orange/yellow fill (needs attention)

    These rules are applied to the entire column range and persist when the file is saved.

    Args:
        ws: The worksheet to add formatting to
        column_letter: Column letter for Review Status (e.g., "I", "J")
        start_row: Starting row (default 2 to skip header)
        end_row: Ending row (default 1000)
    """
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill, Font

    cell_range = f"{column_letter}{start_row}:{column_letter}{end_row}"

    # Rule 1: "✓ Exception" -> Green fill
    # Using SEARCH to find "Exception" anywhere in the cell value
    exception_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    exception_font = Font(color="006100", bold=True)
    exception_rule = FormulaRule(
        formula=[f'SEARCH("Exception",{column_letter}{start_row})'],
        fill=exception_fill,
        font=exception_font,
        stopIfTrue=True,
    )
    ws.conditional_formatting.add(cell_range, exception_rule)

    # Rule 2: "⏳ Needs Review" -> Orange/Yellow fill
    needs_review_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    needs_review_font = Font(color="9C5700", bold=True)
    needs_review_rule = FormulaRule(
        formula=[f'SEARCH("Needs Review",{column_letter}{start_row})'],
        fill=needs_review_fill,
        font=needs_review_font,
        stopIfTrue=True,
    )
    ws.conditional_formatting.add(cell_range, needs_review_rule)


# ============================================================================
# Sheet Configuration
# ============================================================================


@dataclass(frozen=True)
class SheetConfig:
    """Configuration for a worksheet."""

    name: str
    columns: tuple[ColumnDef, ...]

    @property
    def column_count(self) -> int:
        return len(self.columns)


# ============================================================================
# Base Sheet Mixin
# ============================================================================


class BaseSheetMixin:
    """
    Base class for sheet mixins.

    Provides common functionality for sheet creation and data writing.
    Each sheet mixin inherits from this and implements sheet-specific
    add_* methods.
    """

    # Subclasses must define these
    wb: Workbook
    _row_counters: dict[str, int]
    _issue_count: int
    _pass_count: int
    _warn_count: int

    def _ensure_sheet(self, config: SheetConfig) -> Worksheet:
        """
        Get or create a worksheet with the given configuration.

        Creates the sheet if it doesn't exist, applies headers,
        freezes panes, and adds autofilter.
        """
        if config.name in self.wb.sheetnames:
            return self.wb[config.name]

        # Remove default "Sheet" if present
        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) == 1:
            del self.wb["Sheet"]

        ws = self.wb.create_sheet(config.name)
        apply_header_row(ws, list(config.columns), row=1)
        freeze_panes(ws, row=2, col=1)
        add_autofilter(ws, list(config.columns), header_row=1)

        return ws

    def _write_row(
        self,
        ws: Worksheet,
        config: SheetConfig,
        data: list[Any],
    ) -> int:
        """
        Write a row of data to a worksheet.

        Args:
            ws: Target worksheet
            config: Sheet configuration with column definitions
            data: List of values to write

        Returns:
            Row number that was written
        """
        row = self._row_counters[config.name]

        for col, (value, col_def) in enumerate(zip(data, config.columns), start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = Fonts.DATA
            cell.border = Borders.THIN
            cell.alignment = col_def.alignment

            if col_def.is_manual:
                cell.fill = Fills.MANUAL

        self._row_counters[config.name] += 1
        return row

    def _increment_pass(self) -> None:
        """Increment pass count."""
        self._pass_count += 1

    def _increment_warn(self) -> None:
        """Increment warning count."""
        self._warn_count += 1

    def _increment_issue(self) -> None:
        """Increment issue count."""
        self._issue_count += 1
