"""
Quick test script to verify Phase 1 implementation.

Tests:
1. HistoryStore schema initialization
2. CRUD operations for audit runs, servers, instances
3. Excel report generation

Run: python test_phase1.py
"""

import sys
import tempfile
from pathlib import Path
from datetime import datetime

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from autodbaudit.domain.models import AuditRun, Server, Instance
from autodbaudit.infrastructure.history_store import HistoryStore
from autodbaudit.infrastructure.excel_report import write_instance_inventory


def test_history_store():
    """Test HistoryStore with in-memory database."""
    print("\n" + "=" * 60)
    print("Testing HistoryStore")
    print("=" * 60)
    
    # Use temp file (sqlite3 :memory: doesn't persist between calls)
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as f:
        db_path = Path(f.name)
    
    try:
        store = HistoryStore(db_path)
        
        # Test schema initialization
        print("\n1. Initializing schema...")
        store.initialize_schema()
        print("   ✅ Schema created")
        
        # Test begin_audit_run
        print("\n2. Creating audit run...")
        run = store.begin_audit_run(organization="Test Corp", config_hash="abc123")
        assert run.id is not None, "Run should have ID"
        assert run.status == "running", "Run should be running"
        print(f"   ✅ Audit run created: id={run.id}")
        
        # Test upsert_server
        print("\n3. Upserting servers...")
        server1 = store.upsert_server("PROD-SQL01", "192.168.1.10")
        server2 = store.upsert_server("PROD-SQL02", "192.168.1.11")
        assert server1.id is not None, "Server1 should have ID"
        print(f"   ✅ Server1: {server1.hostname} (id={server1.id})")
        print(f"   ✅ Server2: {server2.hostname} (id={server2.id})")
        
        # Test upsert_instance
        print("\n4. Upserting instances...")
        instance1 = store.upsert_instance(
            server=server1,
            instance_name="",
            version="15.0.4298.1",
            version_major=15,
            edition="Standard Edition (64-bit)",
            product_level="CU22"
        )
        instance2 = store.upsert_instance(
            server=server1,
            instance_name="NAMED1",
            version="16.0.1000.6",
            version_major=16,
            edition="Developer Edition (64-bit)",
            product_level="RTM"
        )
        instance3 = store.upsert_instance(
            server=server2,
            instance_name="",
            version="14.0.3451.2",
            version_major=14,
            edition="Enterprise Edition (64-bit)",
            product_level="CU28"
        )
        print(f"   ✅ Instance1: {server1.hostname}\\(default) - SQL 2019")
        print(f"   ✅ Instance2: {server1.hostname}\\NAMED1 - SQL 2022")
        print(f"   ✅ Instance3: {server2.hostname}\\(default) - SQL 2017")
        
        # Test link_instance_to_run
        print("\n5. Linking instances to audit run...")
        store.link_instance_to_run(run.id, instance1.id)
        store.link_instance_to_run(run.id, instance2.id)
        store.link_instance_to_run(run.id, instance3.id)
        print("   ✅ All instances linked")
        
        # Test complete_audit_run
        print("\n6. Completing audit run...")
        store.complete_audit_run(run.id, "completed")
        updated_run = store.get_audit_run(run.id)
        assert updated_run.status == "completed", "Run should be completed"
        assert updated_run.ended_at is not None, "Run should have ended_at"
        print(f"   ✅ Audit run completed: status={updated_run.status}")
        
        # Test get_instances_for_run
        print("\n7. Querying instances for run...")
        instances = store.get_instances_for_run(run.id)
        assert len(instances) == 3, f"Should have 3 instances, got {len(instances)}"
        print(f"   ✅ Retrieved {len(instances)} instances")
        for server, instance in instances:
            print(f"      - {server.hostname}\\{instance.instance_name or '(default)'}: v{instance.version}")
        
        store.close()
        print("\n✅ HistoryStore tests passed!")
        return instances, updated_run
        
    finally:
        # Cleanup
        if db_path.exists():
            db_path.unlink()


def test_excel_report(instances, audit_run):
    """Test Excel report generation."""
    print("\n" + "=" * 60)
    print("Testing Excel Report")
    print("=" * 60)
    
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "test_inventory.xlsx"
        
        print("\n1. Generating Excel report...")
        result_path = write_instance_inventory(instances, audit_run, output_path)
        
        assert result_path.exists(), "Excel file should exist"
        file_size = result_path.stat().st_size
        print(f"   ✅ Excel file created: {result_path.name} ({file_size} bytes)")
        
        # Verify with openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        
        print(f"\n2. Verifying workbook structure...")
        print(f"   Sheets: {wb.sheetnames}")
        assert "InstanceInventory" in wb.sheetnames, "Should have InstanceInventory sheet"
        assert "AuditInfo" in wb.sheetnames, "Should have AuditInfo sheet"
        
        ws = wb["InstanceInventory"]
        row_count = ws.max_row
        print(f"   InstanceInventory: {row_count} rows (1 header + {row_count-1} data)")
        assert row_count == 4, f"Should have 4 rows (1 header + 3 instances), got {row_count}"
        
        print("\n✅ Excel report tests passed!")


def main():
    """Run all tests."""
    print("\n" + "=" * 60)
    print("Phase 1 Implementation Tests")
    print("=" * 60)
    
    try:
        # Test HistoryStore
        instances, audit_run = test_history_store()
        
        # Test Excel report
        test_excel_report(instances, audit_run)
        
        print("\n" + "=" * 60)
        print("🎉 ALL TESTS PASSED!")
        print("=" * 60)
        return 0
        
    except AssertionError as e:
        print(f"\n❌ Test failed: {e}")
        return 1
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
