"""
Excel Annotation Writer - Programmatically writes test annotations.

Since we can't simulate human Excel editing, this module uses openpyxl
to write known test values to specific cells, then saves the file so
sync can pick up the changes.

This tests:
- Notes/Justification persistence
- Exception marking
- Review status dropdowns
- State transitions after sync
"""

from __future__ import annotations

import random
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
from typing import Generator

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class TestAnnotation:
    """A test annotation to write."""

    sheet: str
    row: int  # 1-indexed
    notes_value: str | None = None
    justification_value: str | None = None
    review_status: str | None = None
    purpose: str | None = None  # For Linked Servers


@dataclass
class StateTestCase:
    """A state transition test case."""

    name: str
    initial_result: str  # PASS/FAIL/WARN
    has_justification: bool
    has_exception_status: bool
    expected_after_sync: str  # What should appear in Actions
    description: str


class AnnotationWriter:
    """
    Writes test annotations to Excel for sync verification.

    Workflow:
    1. Open Excel
    2. Write known values to specific cells
    3. Save
    4. Run sync
    5. Verify values persisted and triggered expected behavior
    """

    def __init__(self, excel_path: Path) -> None:
        self.excel_path = excel_path
        self.wb = None
        self.written_annotations: list[TestAnnotation] = []

    def open(self) -> bool:
        """Open workbook for editing."""
        try:
            self.wb = load_workbook(self.excel_path)
            return True
        except Exception as e:
            print(f"Failed to open {self.excel_path}: {e}")
            return False

    def save(self) -> bool:
        """Save and close workbook."""
        try:
            self.wb.save(self.excel_path)
            self.wb.close()
            self.wb = None
            return True
        except Exception as e:
            print(f"Failed to save: {e}")
            return False

    def find_column_index(self, ws: Worksheet, header_name: str) -> int | None:
        """Find column index (1-based) for header name."""
        for i, cell in enumerate(ws[1], start=1):
            if cell.value == header_name:
                return i
        return None

    def write_annotation(
        self,
        sheet: str,
        row: int,
        notes: str | None = None,
        justification: str | None = None,
        review_status: str | None = None,
        purpose: str | None = None,
    ) -> bool:
        """Write annotation values to a specific row."""
        if sheet not in self.wb.sheetnames:
            print(f"Sheet {sheet} not found")
            return False

        ws = self.wb[sheet]

        # Find column indices
        notes_col = self.find_column_index(ws, "Notes")
        just_col = self.find_column_index(ws, "Justification")
        status_col = self.find_column_index(ws, "Review Status")
        purpose_col = self.find_column_index(ws, "Purpose")

        if notes and notes_col:
            ws.cell(row=row, column=notes_col).value = notes

        if justification and just_col:
            ws.cell(row=row, column=just_col).value = justification

        if review_status and status_col:
            ws.cell(row=row, column=status_col).value = review_status

        if purpose and purpose_col:
            ws.cell(row=row, column=purpose_col).value = purpose

        # Track what we wrote
        self.written_annotations.append(
            TestAnnotation(
                sheet=sheet,
                row=row,
                notes_value=notes,
                justification_value=justification,
                review_status=review_status,
                purpose=purpose,
            )
        )

        return True

    def write_test_annotations(self) -> int:
        """
        Write a comprehensive set of test annotations.

        Covers multiple sheets and annotation types.
        Returns count of annotations written.
        """
        tag = datetime.now().strftime("%H%M%S")

        test_data = [
            # Sheet, row, notes, justification, status, purpose
            ("Logins", 2, f"Test note {tag}", None, None, None),
            ("Logins", 3, None, f"Test justification {tag}", "Exception ✓", None),
            ("Databases", 2, f"DB note {tag}", None, None, None),
            ("Configuration", 2, None, f"Config exception {tag}", "Exception ✓", None),
            ("Linked Servers", 2, f"LS note {tag}", None, None, f"Test purpose {tag}"),
            ("SA Account", 2, None, f"SA exception {tag}", "Exception ✓", None),
        ]

        count = 0
        for sheet, row, notes, just, status, purpose in test_data:
            if self.write_annotation(sheet, row, notes, just, status, purpose):
                count += 1

        return count

    def verify_annotations_persisted(self) -> tuple[int, int, list[str]]:
        """
        Verify previously written annotations are still present.

        Returns: (found, missing, error_messages)
        """
        if not self.open():
            return 0, 0, ["Could not open workbook"]

        found = 0
        missing = 0
        errors = []

        for ann in self.written_annotations:
            if ann.sheet not in self.wb.sheetnames:
                missing += 1
                errors.append(f"{ann.sheet} missing")
                continue

            ws = self.wb[ann.sheet]

            # Check each value we wrote
            if ann.notes_value:
                col = self.find_column_index(ws, "Notes")
                if col:
                    actual = ws.cell(row=ann.row, column=col).value
                    if actual == ann.notes_value:
                        found += 1
                    else:
                        missing += 1
                        errors.append(
                            f"{ann.sheet}[{ann.row}] Notes: expected '{ann.notes_value}', got '{actual}'"
                        )

            if ann.justification_value:
                col = self.find_column_index(ws, "Justification")
                if col:
                    actual = ws.cell(row=ann.row, column=col).value
                    if actual == ann.justification_value:
                        found += 1
                    else:
                        missing += 1
                        errors.append(
                            f"{ann.sheet}[{ann.row}] Justification: expected '{ann.justification_value[:30]}...', got '{actual}'"
                        )

        self.wb.close()
        return found, missing, errors


def generate_state_test_cases() -> list[StateTestCase]:
    """
    Generate all state transition test cases.

    Tests the state machine:
    - FAIL + no exception = Active Issue
    - FAIL + exception = Documented Exception
    - PASS after FAIL = FIXED
    - FAIL after PASS = REGRESSION
    """
    return [
        StateTestCase(
            name="FAIL_no_exception",
            initial_result="FAIL",
            has_justification=False,
            has_exception_status=False,
            expected_after_sync="Active Issue (no action)",
            description="FAIL without exception stays as active issue",
        ),
        StateTestCase(
            name="FAIL_with_exception",
            initial_result="FAIL",
            has_justification=True,
            has_exception_status=True,
            expected_after_sync="Documented Exception",
            description="FAIL with exception is documented",
        ),
        StateTestCase(
            name="PASS_after_FAIL",
            initial_result="PASS",  # After being FAIL in baseline
            has_justification=False,
            has_exception_status=False,
            expected_after_sync="FIXED",
            description="PASS after FAIL shows as FIXED",
        ),
        StateTestCase(
            name="FAIL_after_PASS",
            initial_result="FAIL",  # After being PASS in baseline
            has_justification=False,
            has_exception_status=False,
            expected_after_sync="REGRESSION",
            description="FAIL after PASS shows as REGRESSION",
        ),
        StateTestCase(
            name="Exception_on_PASS",
            initial_result="PASS",
            has_justification=True,
            has_exception_status=True,
            expected_after_sync="Status cleared (not discrepant)",
            description="Exception on PASS item gets status cleared",
        ),
    ]
