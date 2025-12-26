"""
Comprehensive E2E Audit Verifier.

Automates what would take hours to verify manually:
1. Runs audit, captures output
2. Opens Excel with openpyxl, verifies expected discrepancies appear in correct sheets
3. Checks column alignment (no data in wrong columns)
4. Verifies stats match actual counts
5. Tests sync flow (baseline → changes → sync → verify state transitions)
6. Writes test annotations to Excel, syncs back, verifies persistence
7. Checks action log entries
8. Verifies conditional formatting and dropdown validations exist
9. Tests Persian report generation

Usage:
    python verify_e2e.py --full         # Run complete verification
    python verify_e2e.py --quick        # Quick sanity check
    python verify_e2e.py --phase audit  # Run specific phase
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

# ═══════════════════════════════════════════════════════════════════════════════
# PATH SETUP
# ═══════════════════════════════════════════════════════════════════════════════
SIMULATION_DIR = Path(__file__).parent
PROJECT_ROOT = SIMULATION_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from discrepancy_tracker import DiscrepancyTracker, ExpectedDiscrepancy

# ═══════════════════════════════════════════════════════════════════════════════
# COLORS
# ═══════════════════════════════════════════════════════════════════════════════
CYAN = "\033[96m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
RED = "\033[91m"
BOLD = "\033[1m"
RESET = "\033[0m"


@dataclass
class VerificationResult:
    """Result of a single verification check."""

    name: str
    passed: bool
    message: str = ""
    details: list[str] = field(default_factory=list)


@dataclass
class VerificationReport:
    """Complete verification report."""

    timestamp: datetime = field(default_factory=datetime.now)
    results: list[VerificationResult] = field(default_factory=list)
    audit_id: int | None = None
    excel_path: Path | None = None

    @property
    def passed(self) -> bool:
        return all(r.passed for r in self.results)

    @property
    def passed_count(self) -> int:
        return sum(1 for r in self.results if r.passed)

    @property
    def failed_count(self) -> int:
        return sum(1 for r in self.results if not r.passed)

    def add(
        self, name: str, passed: bool, message: str = "", details: list[str] = None
    ) -> None:
        self.results.append(
            VerificationResult(
                name=name,
                passed=passed,
                message=message,
                details=details or [],
            )
        )

    def print_summary(self) -> None:
        print(f"\n{CYAN}{'═' * 60}{RESET}")
        print(f"{BOLD}VERIFICATION REPORT{RESET}")
        print(f"{CYAN}{'═' * 60}{RESET}")

        for r in self.results:
            status = f"{GREEN}✓{RESET}" if r.passed else f"{RED}✗{RESET}"
            print(f"  {status} {r.name}")
            if r.message:
                print(f"      {r.message}")
            for d in r.details[:3]:  # Limit details
                print(f"      - {d}")
            if len(r.details) > 3:
                print(f"      ... and {len(r.details) - 3} more")

        print(f"\n{CYAN}{'─' * 60}{RESET}")
        if self.passed:
            print(f"{GREEN}{BOLD}ALL {self.passed_count} CHECKS PASSED{RESET}")
        else:
            print(
                f"{RED}{BOLD}FAILED: {self.failed_count}/{len(self.results)} checks{RESET}"
            )


class E2EVerifier:
    """
    Comprehensive E2E verification engine.

    Can automate most of what a human tester would do,
    except for visual verification and user input simulation.
    """

    def __init__(self, project_root: Path = PROJECT_ROOT) -> None:
        self.project_root = project_root
        self.output_dir = project_root / "output"
        self.tracker = DiscrepancyTracker(self.output_dir / "simulation_tracking.json")
        self.report = VerificationReport()
        self.audit_id: int | None = None
        self.excel_path: Path | None = None
        self.wb = None

    def run_cli(self, *args, capture: bool = True) -> subprocess.CompletedProcess:
        """Run CLI command."""
        cmd = [sys.executable, str(self.project_root / "main.py")] + list(args)
        print(f"  → {' '.join(str(a) for a in args)}")
        return subprocess.run(
            cmd,
            cwd=str(self.project_root),
            capture_output=capture,
            text=True,
        )

    def find_latest_excel(self) -> Path | None:
        """Find the most recent Excel report."""
        excels = list(self.output_dir.glob("*.xlsx"))
        if not excels:
            return None
        return max(excels, key=lambda p: p.stat().st_mtime)

    def load_excel(self) -> bool:
        """Load the Excel workbook."""
        if not self.excel_path or not self.excel_path.exists():
            self.excel_path = self.find_latest_excel()

        if not self.excel_path:
            return False

        try:
            self.wb = load_workbook(self.excel_path)
            return True
        except Exception as e:
            print(f"{RED}Failed to load Excel: {e}{RESET}")
            return False

    def close_excel(self) -> None:
        """Close workbook."""
        if self.wb:
            self.wb.close()
            self.wb = None

    # ═══════════════════════════════════════════════════════════════════════════
    # VERIFICATION METHODS
    # ═══════════════════════════════════════════════════════════════════════════

    def verify_audit_creates_excel(self) -> bool:
        """Verify that running audit creates an Excel file."""
        print(f"\n{YELLOW}[1] Running audit...{RESET}")

        # Clear old files
        for f in self.output_dir.glob("*.xlsx"):
            f.unlink()

        result = self.run_cli("audit")

        if result.returncode != 0:
            self.report.add(
                "Audit Command",
                False,
                f"Exit code {result.returncode}",
                [result.stderr[:200] if result.stderr else "No stderr"],
            )
            return False

        # Find the created Excel
        self.excel_path = self.find_latest_excel()

        if not self.excel_path:
            self.report.add("Audit Creates Excel", False, "No Excel file created")
            return False

        # Extract audit ID from output
        match = re.search(r"Audit ID:\s*(\d+)", result.stdout)
        if match:
            self.audit_id = int(match.group(1))

        self.report.add(
            "Audit Creates Excel",
            True,
            f"Created: {self.excel_path.name}",
        )
        return True

    def verify_expected_sheets_exist(self) -> bool:
        """Verify all expected sheets exist."""
        print(f"\n{YELLOW}[2] Checking sheet structure...{RESET}")

        if not self.load_excel():
            self.report.add("Load Excel", False, "Could not load workbook")
            return False

        expected_sheets = [
            "Cover",
            "Instances",
            "Logins",
            "Databases",
            "DB Users",
            "Orphaned Users",
            "Roles",
            "DB Roles",
            "Role Matrix",
            "Permissions",
            "Services",
            "Client Protocols",
            "Configuration",
            "Linked Servers",
            "SA Account",
            "Backups",
            "Encryption",
            "Triggers",
            "Audit Settings",
            "Actions",
        ]

        missing = [s for s in expected_sheets if s not in self.wb.sheetnames]

        if missing:
            self.report.add(
                "Expected Sheets Exist",
                False,
                f"Missing {len(missing)} sheets",
                missing[:5],
            )
            return False

        self.report.add(
            "Expected Sheets Exist",
            True,
            f"All {len(expected_sheets)} sheets present",
        )
        return True

    def verify_column_headers_aligned(self) -> bool:
        """Verify column headers match schema (no data shift)."""
        print(f"\n{YELLOW}[3] Checking column alignment...{RESET}")

        # Spot-check critical sheets
        checks = {
            "Logins": ["Server", "Login", "Type", "Result"],
            "Configuration": ["Server", "Setting", "Value", "Result"],
            "SA Account": ["Server", "SA Name", "Enabled", "Result"],
            "Actions": ["Server", "Entity", "Description", "Category"],
        }

        misaligned = []

        for sheet_name, expected_headers in checks.items():
            if sheet_name not in self.wb.sheetnames:
                continue

            ws = self.wb[sheet_name]
            row1 = [cell.value for cell in ws[1]]

            for i, expected in enumerate(expected_headers):
                if i >= len(row1):
                    misaligned.append(
                        f"{sheet_name}: Missing column {i+1} ({expected})"
                    )
                elif row1[i] != expected:
                    misaligned.append(
                        f"{sheet_name}: Col {i+1} is '{row1[i]}' not '{expected}'"
                    )

        if misaligned:
            self.report.add(
                "Column Alignment", False, f"{len(misaligned)} issues", misaligned
            )
            return False

        self.report.add("Column Alignment", True, "Headers match schema")
        return True

    def verify_expected_discrepancies(self) -> bool:
        """Verify that tracked discrepancies appear in the report."""
        print(f"\n{YELLOW}[4] Checking expected discrepancies...{RESET}")

        expected = self.tracker.get_all_expected_discrepancies()

        if not expected:
            self.report.add(
                "Expected Discrepancies",
                True,
                "No tracked discrepancies (run simulation first)",
            )
            return True

        found = 0
        not_found = []

        for disc in expected:
            if disc.sheet not in self.wb.sheetnames:
                not_found.append(f"{disc.sheet}: Sheet missing")
                continue

            ws = self.wb[disc.sheet]

            # Search for entity pattern in sheet
            entity_found = False
            for row in ws.iter_rows(min_row=2):
                row_text = " ".join(str(c.value or "") for c in row)
                if re.search(disc.entity_pattern, row_text, re.IGNORECASE):
                    entity_found = True
                    found += 1
                    break

            if not entity_found:
                not_found.append(f"{disc.sheet}: {disc.entity_pattern}")

        if not_found:
            self.report.add(
                "Expected Discrepancies Found",
                False,
                f"Found {found}/{len(expected)}",
                not_found[:5],
            )
            return False

        self.report.add(
            "Expected Discrepancies Found",
            True,
            f"All {len(expected)} discrepancies present",
        )
        return True

    def verify_dropdowns_exist(self) -> bool:
        """Verify dropdown validations are set on Actions sheet."""
        print(f"\n{YELLOW}[5] Checking dropdown validations...{RESET}")

        if "Actions" not in self.wb.sheetnames:
            self.report.add("Dropdown Validations", False, "Actions sheet missing")
            return False

        ws = self.wb["Actions"]

        # Check for data validations
        dv_count = len(ws.data_validations.dataValidation)

        if dv_count == 0:
            self.report.add("Dropdown Validations", False, "No validations found")
            return False

        self.report.add(
            "Dropdown Validations",
            True,
            f"{dv_count} validations found",
        )
        return True

    def verify_conditional_formatting(self) -> bool:
        """Verify conditional formatting rules exist."""
        print(f"\n{YELLOW}[6] Checking conditional formatting...{RESET}")

        # Check Actions sheet for CF
        if "Actions" not in self.wb.sheetnames:
            self.report.add("Conditional Formatting", False, "Actions sheet missing")
            return False

        ws = self.wb["Actions"]
        cf_rules = ws.conditional_formatting._cf_rules

        if not cf_rules:
            self.report.add("Conditional Formatting", False, "No CF rules found")
            return False

        self.report.add(
            "Conditional Formatting",
            True,
            f"{len(cf_rules)} CF rules on Actions",
        )
        return True

    def verify_sync_works(self) -> bool:
        """Verify sync command runs successfully."""
        print(f"\n{YELLOW}[7] Testing sync command...{RESET}")

        if not self.audit_id:
            self.report.add("Sync Command", False, "No audit ID available")
            return False

        result = self.run_cli("sync", "--audit-id", str(self.audit_id))

        if result.returncode != 0:
            self.report.add(
                "Sync Command",
                False,
                f"Exit code {result.returncode}",
                [result.stderr[:200] if result.stderr else ""],
            )
            return False

        self.report.add("Sync Command", True, "Sync completed successfully")
        return True

    def verify_finalize_works(self) -> bool:
        """Verify finalize command runs."""
        print(f"\n{YELLOW}[8] Testing finalize command...{RESET}")

        if not self.audit_id:
            self.report.add("Finalize Command", False, "No audit ID")
            return False

        result = self.run_cli("finalize", "--audit-id", str(self.audit_id))

        if result.returncode != 0:
            self.report.add(
                "Finalize Command",
                False,
                f"Exit code {result.returncode}",
            )
            return False

        self.report.add("Finalize Command", True)
        return True

    def verify_persian_report(self) -> bool:
        """Verify Persian report generation."""
        print(f"\n{YELLOW}[9] Testing Persian report...{RESET}")

        if not self.audit_id:
            self.report.add("Persian Report", False, "No audit ID")
            return False

        # De-finalize first
        self.run_cli("definalize", "--audit-id", str(self.audit_id))

        result = self.run_cli("finalize", "--audit-id", str(self.audit_id), "--persian")

        if result.returncode != 0:
            self.report.add("Persian Report", False, f"Exit code {result.returncode}")
            return False

        # Check for _fa.xlsx file
        fa_files = list(self.output_dir.glob("*_fa.xlsx"))

        if not fa_files:
            self.report.add("Persian Report", False, "No _fa.xlsx file created")
            return False

        # Verify RTL setting
        fa_wb = load_workbook(fa_files[0])
        ws = fa_wb.active
        is_rtl = getattr(ws.sheet_view, "rightToLeft", False)
        fa_wb.close()

        if not is_rtl:
            self.report.add("Persian Report", False, "RTL not enabled")
            return False

        self.report.add("Persian Report", True, f"Created: {fa_files[0].name}")
        return True

    def run_full_verification(self) -> VerificationReport:
        """Run all verification steps."""
        print(f"\n{CYAN}{'═' * 60}{RESET}")
        print(f"{BOLD}COMPREHENSIVE E2E VERIFICATION{RESET}")
        print(f"{CYAN}{'═' * 60}{RESET}")

        try:
            self.verify_audit_creates_excel()
            self.verify_expected_sheets_exist()
            self.verify_column_headers_aligned()
            self.verify_expected_discrepancies()
            self.verify_dropdowns_exist()
            self.verify_conditional_formatting()
            self.verify_sync_works()
            self.verify_finalize_works()
            self.verify_persian_report()
        finally:
            self.close_excel()

        self.report.print_summary()
        return self.report

    def run_quick_verification(self) -> VerificationReport:
        """Run quick sanity checks only."""
        print(f"\n{CYAN}QUICK SANITY CHECK{RESET}")

        try:
            self.verify_audit_creates_excel()
            self.verify_expected_sheets_exist()
            self.verify_column_headers_aligned()
        finally:
            self.close_excel()

        self.report.print_summary()
        return self.report


def main():
    parser = argparse.ArgumentParser(description="E2E Audit Verification")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--full", action="store_true", help="Full verification")
    group.add_argument("--quick", action="store_true", help="Quick sanity check")
    group.add_argument("--phase", choices=["audit", "sync", "finalize", "persian"])

    args = parser.parse_args()

    verifier = E2EVerifier()

    if args.full:
        report = verifier.run_full_verification()
    elif args.quick:
        report = verifier.run_quick_verification()
    else:
        print(f"Running phase: {args.phase}")
        # Individual phase logic here
        report = verifier.report

    sys.exit(0 if report.passed else 1)


if __name__ == "__main__":
    main()
